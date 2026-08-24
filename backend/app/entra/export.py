"""Multi-sheet .xlsx export of the entire Entra snapshot.

Everything under /entra in one file: every tab, every sub-tab, and the raw directory the tabs
only ever show counts of. Built from the snapshot object rather than by calling twenty-five
endpoints, so it is one read and it cannot drift from what the screens display.

Two things this format takes seriously.

**Blindness is not emptiness.** Entra is permission-gated in a way Azure RBAC is not: a missing
Graph scope blanks a whole domain, and on a real tenant three of nine domains are commonly
`partial`. A sheet that is empty because the collector was refused looks exactly like a sheet
that is empty because there is nothing wrong. Every sheet whose domain could not be read is
written with the reason instead of the rows, and the qualifiers that change how a number should
be read — `sampled`, `owners_known`, `assignment_known`, `fic_known`, `detail_known`,
`azure_link.stale` — travel in the same row as the number they qualify.

**Nested detail gets its own sheet.** A service principal carries its credentials, its granted
permissions and its reply URLs as lists. Joining those into one cell is how an expiring secret
becomes invisible, so each becomes a child sheet with one row per credential, per grant, per URL.
"""
from __future__ import annotations

import datetime as dt
from typing import Any

from app.core.xlsx import DATETIME_FMT, WorkbookBuilder, as_datetime, hyperlink

#: Domain statuses that mean "the rows are not trustworthy or not there".
BLIND_STATUSES = frozenset({"blind", "error", "not_collected", "unlicensed"})
#: Statuses that mean "we got rows, but something alongside them was degraded".
DEGRADED_STATUSES = frozenset({"partial", "stale"})

#: One tab color per PARENT screen, so the fifty sheets group visually the way the UI does —
#: every Conditional Access sheet purple, every Applications sheet cyan, and so on. Hues are
#: kept far apart rather than pretty: two adjacent shades would suggest a relationship between
#: sections that do not have one. The Index repeats the grouping as text, because a color is
#: not readable to everyone and does not survive a print.
SECTION_COLOURS: dict[str, str] = {
    "Overview": "44546A",              # slate — front matter and the blind-spot register
    "Posture": "2E75B6",               # blue
    "Findings & scanners": "C00000",   # red
    "Conditional Access": "7030A0",    # purple
    "Privileged Access": "ED7D31",     # orange
    "Applications": "00B0F0",          # cyan
    "Risk & sign-ins": "E91E63",       # pink
    "Directory": "548235",             # green
    "Governance": "BF8F00",            # gold
    "Blast radius": "833C00",          # brown
    "Setup & coverage": "808080",      # grey
}


#: Sign-in breakdowns, one sheet each. Titles are fixed so the workbook's shape does not
#: change with the data; a dimension the collector did not return simply has no sheet.
BREAKDOWN_SHEETS: dict[str, str] = {
    "by_day": "Sign-ins by day",
    "by_app": "Sign-ins by application",
    "by_user_top": "Sign-ins by user",
    "by_client_app": "Sign-ins by client app",
    "by_country": "Sign-ins by country",
    "by_ca_result": "Sign-ins by CA result",
    "by_failure_code": "Sign-ins by failure code",
}


def _label(key: str) -> str:
    text = str(key).replace("_", " ").strip()
    return text[:1].upper() + text[1:] if text else "Value"


def _explode(items: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]], set[str]]:
    """Headers, rows and the timestamp headers, from a list of flat records.

    Data-driven on purpose. The alternative — reading fixed key names off a payload whose
    shape was assumed rather than checked — is what produced several columns that were blank
    on every row, and a blank column reads as an answer.

    Because the columns are discovered rather than declared, the dates have to be discovered
    too. Detection is on the VALUES, never the key name: a column qualifies only when every
    value in it parses as an extended-format ISO timestamp. Guessing from a name like
    "package" is how a column of text gets mangled into 1900-era dates.
    """
    keys: list[str] = []
    for it in items:
        for k in it:
            if k not in keys:
                keys.append(k)
    rows = [[it.get(k, "") for k in keys] for it in items]

    dates: set[str] = set()
    for i, key in enumerate(keys):
        seen = [r[i] for r in rows if r[i] not in (None, "")]
        if seen and all(isinstance(v, str) and ("-" in v or ":" in v)
                        and as_datetime(v) is not None for v in seen):
            dates.add(_label(key))
    return [_label(k) for k in keys], rows, dates


def _session_controls(policy: dict[str, Any]) -> str:
    """The session controls a policy actually switches on, as words.

    The nested ``session`` block is a dozen booleans and their parameters; exporting the dict
    would be an unreadable cell and exporting a single flag would lose which control is on."""
    s = policy.get("session") or {}
    on: list[str] = []
    if s.get("sign_in_frequency"):
        on.append(f"sign-in frequency {s.get('sign_in_frequency_value', '')} "
                  f"{s.get('sign_in_frequency_type', '')}".strip())
    if s.get("persistent_browser"):
        on.append(f"persistent browser {s.get('persistent_browser_mode', '')}".strip())
    if s.get("app_enforced_restrictions"):
        on.append("app-enforced restrictions")
    if s.get("cloud_app_security"):
        on.append("cloud app security")
    if s.get("continuous_access_evaluation"):
        on.append(f"CAE {s['continuous_access_evaluation']}")
    return ", ".join(on)


def _domain(snapshot: dict[str, Any], name: str) -> dict[str, Any]:
    return (snapshot.get("domains") or {}).get(name) or {}


def _blind_reason(snapshot: dict[str, Any], name: str) -> str:
    """Why this domain cannot be shown, or "" when it can."""
    dom = _domain(snapshot, name)
    status = str(dom.get("status") or "")
    if status not in BLIND_STATUSES:
        return ""
    bits = [f"Domain '{name}' status is '{status}'."]
    missing = dom.get("missing_permissions") or []
    if missing:
        bits.append("Missing permission(s): " + ", ".join(str(m) for m in missing) + ".")
    if dom.get("error"):
        bits.append(str(dom["error"]))
    for note in (dom.get("notes") or [])[:3]:
        bits.append(str(note))
    return " ".join(bits)


def _caveat(snapshot: dict[str, Any], name: str) -> str:
    """A note for a domain that returned rows but is degraded. Empty when it is clean."""
    dom = _domain(snapshot, name)
    status = str(dom.get("status") or "")
    if status not in DEGRADED_STATUSES:
        return ""
    bits = [f"Domain '{name}' is '{status}' — these rows may be incomplete."]
    missing = dom.get("missing_permissions") or []
    if missing:
        bits.append("Missing: " + ", ".join(str(m) for m in missing) + ".")
    if dom.get("truncated"):
        bits.append("The collector truncated its result.")
    return " ".join(bits)


def _rows(snapshot: dict[str, Any], domain: str, key: str) -> list[dict[str, Any]]:
    data = ((snapshot.get("data") or {}).get(domain) or {}).get(key)
    return data if isinstance(data, list) else []


def _table(items: list[dict[str, Any]], fields: list[tuple[str, str]]) -> list[list[Any]]:
    return [[it.get(src, "") for _label, src in fields] for it in items]


def _heads(fields: list[tuple[str, str]]) -> list[str]:
    return [label for label, _src in fields]


def to_workbook(
    *,
    snapshot: dict[str, Any],
    escalations: list[dict[str, Any]] | None = None,
    history: list[dict[str, Any]] | None = None,
    setup_tiers: list[dict[str, Any]] | None = None,
    scanners: list[dict[str, Any]] | None = None,
    activations: list[dict[str, Any]] | None = None,
) -> bytes:
    """Serialise the whole Entra snapshot to a workbook.

    A pure formatter: everything it needs is passed in, so it is testable without a tenant and
    cannot itself trigger a Graph call. The few payloads that are computed by an endpoint rather
    than stored on the snapshot (escalations, the score history, the setup checklist, the
    scanner cards, the ledger-merged activations) are arguments for exactly that reason.

    ``activations`` matters more than it looks. The snapshot holds only what the current
    collection saw; the screen merges that with a durable ledger reaching past the 30 days
    Graph retains. Reading the snapshot alone lost 13 of 183 sessions on a live tenant — an
    export that quietly drops history is worse than none, because it looks complete."""
    data = snapshot.get("data") or {}
    analysis = snapshot.get("_analysis") or {}
    perms = snapshot.get("permissions") or {}
    wb = WorkbookBuilder()

    # ---------------------------------------------------------------- front matter
    wb.section("Overview", SECTION_COLOURS["Overview"])
    score = analysis.get("score") or {}
    ws = wb.first_sheet("Summary")
    ws.append(["Entra ID — identity review export"])
    from openpyxl.styles import Font

    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    tenant = (data.get("tenant") or {}).get("tenant") or {}
    for label, value in [
        ("Tenant", tenant.get("display_name", "") or snapshot.get("tenant_id", "")),
        ("Tenant id", snapshot.get("tenant_id", "")),
        ("Snapshot generated (UTC)", as_datetime(snapshot.get("generated_at", ""))
         or snapshot.get("generated_at", "")),
        ("Last full collection (UTC)", as_datetime(snapshot.get("last_full", ""))
         or snapshot.get("last_full", "")),
        ("Posture score", score.get("score", "")),
        ("Measured coverage", f"{round((score.get('coverage') or 0) * 100)}% of the weighted checks"),
        ("Findings", len(analysis.get("findings") or [])),
        ("Checks that could NOT run", len(analysis.get("not_measured") or {})),
        ("", ""),
        ("CONTAINS PERSONAL DATA", "Yes — the Users sheet carries UPNs, employee ids, "
                                   "departments, managers and sign-in times. Handle accordingly."),
        ("How to read a blank sheet", "Check the Index and the 'Coverage & blind spots' sheet "
                                      "first. A sheet marked NOT MEASURED is not a clean result."),
        ("How to read a date", "Every timestamp in this workbook is UTC and is a real Excel "
                               "date, so the date filters and sorting work on it directly."),
    ]:
        ws.append([label, value])
        if isinstance(value, dt.datetime):
            ws.cell(row=ws.max_row, column=2).number_format = DATETIME_FMT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    # The honesty register. Deliberately sheet 2 — before anything a reader might mistake for a
    # complete answer.
    domain_rows: list[list[Any]] = []
    for name, dom in sorted((snapshot.get("domains") or {}).items()):
        domain_rows.append([
            name, dom.get("status", ""), dom.get("generated_at", ""), dom.get("item_count", ""),
            "Yes" if dom.get("truncated") else "",
            ", ".join(str(m) for m in (dom.get("missing_permissions") or [])),
            dom.get("error", ""),
            " | ".join(str(n) for n in (dom.get("notes") or [])),
            " | ".join(str(b.get("text", b)) for b in (dom.get("blockers") or [])),
        ])
    wb.sheet(
        "Coverage & blind spots",
        ["Domain", "Status", "Collected at", "Items", "Truncated", "Missing permissions",
         "Error", "Notes", "Blockers"],
        domain_rows,
        note="Any domain whose status is not 'ok' limits every sheet derived from it.",
        dates={"Collected at"},
    )

    wb.sheet(
        "Permissions & licenses",
        ["Kind", "Name", "State", "Detail"],
        [["Graph scope", s, "granted", ""] for s in (perms.get("granted") or [])]
        + [["Graph scope list", "", "verified" if perms.get("granted_known") else "NOT VERIFIED",
            perms.get("claim_error", "") or perms.get("token_error", "")]]
        + [["License", k, "present" if v else "absent", ""]
           for k, v in sorted((snapshot.get("licences") or {}).items())]
        + [["Domain permission", k, "ok" if (v or {}).get("ok") else "blocked",
            ("license blocked; " if (v or {}).get("licence_blocked") else "")
            + ", ".join(str(m) for m in ((v or {}).get("missing") or []))]
           for k, v in sorted((perms.get("domains") or {}).items())],
        note="" if perms.get("granted_known") else
             "The granted-scope list could not be verified, so 'missing' may be understated.",
    )

    # ---------------------------------------------------------------- posture
    wb.section("Posture", SECTION_COLOURS["Posture"])
    wb.sheet(
        "Posture score",
        ["Pillar", "Key", "Score", "Weight", "State", "Why it is not complete"],
        # `coverage` is not a field a pillar has ever carried, so that column exported blank
        # for every row. `state` and `reason` are the two the scorer actually computes, and
        # they are the ones that say whether a low score is a finding or a blind spot.
        [[p.get("label", ""), p.get("key", ""),
          "not measured" if p.get("score") is None else p.get("score"),
          p.get("weight", ""), p.get("state", ""), p.get("reason", "")]
         for p in (score.get("pillars") or [])],
        highlight={"Score": "bar"},
    )
    if history:
        wb.sheet(
            "Posture trend",
            ["At", "Score", "Coverage"],
            [[h.get("at", ""), h.get("score", ""), h.get("coverage", "")] for h in history],
            dates={"At"},
            highlight={"Score": "bar"},
        )

    # Every registered check, whether it fired, and if not, why not. The catalog is what makes
    # the findings list interpretable: 0 findings for a signal that never ran is not a pass.
    from app.entra import signals as sig_mod

    not_measured = analysis.get("not_measured") or {}
    errors = analysis.get("errors") or {}
    by_signal = analysis.get("by_signal") or {}
    catalogue: list[list[Any]] = []
    for spec in sig_mod.registry():
        pub = spec.public() if hasattr(spec, "public") else {}
        sid = pub.get("id", "")
        catalogue.append([
            sid, pub.get("title", ""), pub.get("pillar", ""), pub.get("severity", ""),
            pub.get("weight", ""), by_signal.get(sid, 0),
            "no" if sid in not_measured else "yes",
            not_measured.get(sid, "") or errors.get(sid, ""),
            ", ".join(str(b) for b in (pub.get("benchmarks") or [])),
            pub.get("why", ""),
        ])
    wb.sheet(
        "Signal catalogue",
        ["Signal", "Title", "Pillar", "Severity", "Weight", "Findings", "Measured",
         "Why not measured", "Benchmarks", "Why it matters"],
        catalogue,
        highlight={"Severity": "severity"},
    )

    # ---------------------------------------------------------------- findings
    wb.section("Findings & scanners", SECTION_COLOURS["Findings & scanners"])
    findings = analysis.get("findings") or []
    wb.sheet(
        "Findings",
        ["Severity", "Pillar", "Signal", "Object kind", "Object", "Title", "Detail",
         "Portal link", "Fingerprint", "Evidence"],
        [[f.get("severity", ""), f.get("pillar", ""), f.get("signal_id", ""),
          f.get("object_kind", ""), f.get("object_name", "") or f.get("object_id", ""),
          f.get("title", ""), f.get("detail", ""),
          hyperlink("Open in Entra", f.get("portal_link", "")),
          f.get("fingerprint", ""), f.get("evidence", "")] for f in findings],
        highlight={"Severity": "severity"},
    )
    wb.sheet(
        "Findings - not measured",
        ["Signal", "Why it could not be checked"],
        sorted(([sid, reason] for sid, reason in not_measured.items()), key=lambda r: r[0]),
        note="These checks did not run. Their absence from the findings list is not a pass.",
    )
    if scanners is not None:
        wb.sheet(
            "Scanners",
            ["Scanner", "Cadence", "Severity floor", "Signals", "Can run", "Why it cannot run",
             "Last run", "New", "Resolved", "Persisting"],
            # "Blocked" used to read `s['blocked'] or 'no'` against a card that never carried
            # the key, so every scanner reported as runnable — a fabricated pass on the sheet
            # that says which checks are switched off.
            [[s.get("name", ""), s.get("cadence", ""), s.get("severity_floor", ""),
              s.get("signal_count", ""),
              "No" if s.get("blocked") else "Yes", s.get("blocked", ""),
              s.get("last_run", ""),
              (s.get("last_counts") or {}).get("new", ""),
              (s.get("last_counts") or {}).get("resolved", ""),
              (s.get("last_counts") or {}).get("persisting", "")] for s in scanners],
            dates={"Last run"},
            note="A blank 'Last run' means this scanner has never run for this tenant, which "
                 "is not the same as running and finding nothing.",
        )

    # ---------------------------------------------------------------- conditional access
    wb.section("Conditional Access", SECTION_COLOURS["Conditional Access"])
    ca_blind = _blind_reason(snapshot, "ca")
    ca = data.get("ca") or {}
    ca_analysis = data.get("_ca_analysis") or {}
    if ca_blind:
        wb.blind_sheet("CA policies", ca_blind)
    else:
        policies = ca.get("policies") or []
        wb.sheet(
            "CA policies",
            ["Policy", "Id", "State", "Created", "Modified", "Grant operator", "Grant controls",
             "Auth strength", "Session controls",
             "Client app types", "Include apps", "Exclude apps", "Include users", "Exclude users",
             "Include groups", "Exclude groups", "Include roles", "Exclude roles"],
            # The controls live in nested `grant` and `session` blocks. Reading them as
            # top-level `grant_controls`/`controls` exported blank for every policy — on the
            # two columns that say what a Conditional Access policy actually enforces.
            [[
                p.get("display_name", ""), p.get("id", ""), p.get("state", ""),
                p.get("created_at", ""), p.get("modified_at", ""),
                (p.get("grant") or {}).get("operator", ""),
                (p.get("grant") or {}).get("controls", ""),
                (p.get("grant") or {}).get("auth_strength_name", ""),
                _session_controls(p),
                (p.get("conditions") or {}).get("client_app_types", ""),
                (p.get("conditions") or {}).get("include_apps", ""),
                (p.get("conditions") or {}).get("exclude_apps", ""),
                (p.get("conditions") or {}).get("include_users", ""),
                (p.get("conditions") or {}).get("exclude_users", ""),
                (p.get("conditions") or {}).get("include_groups", ""),
                (p.get("conditions") or {}).get("exclude_groups", ""),
                (p.get("conditions") or {}).get("include_roles", ""),
                (p.get("conditions") or {}).get("exclude_roles", ""),
            ] for p in policies],
            note=_caveat(snapshot, "ca"),
            dates={"Created", "Modified"},
        )
        # One row per condition entry, so an exclusion can actually be searched for. A policy's
        # exclusions are where the exception that defeats it lives, and they are unreadable
        # squashed into a single cell.
        cond_rows: list[list[Any]] = []
        for p in policies:
            for bucket, values in (p.get("conditions") or {}).items():
                if isinstance(values, (list, tuple)):
                    for v in values:
                        cond_rows.append([p.get("display_name", ""), p.get("state", ""), bucket, v])
        wb.sheet("CA policy conditions", ["Policy", "State", "Condition", "Value"], cond_rows)

        coverage = ca_analysis.get("coverage") or {}
        class_labels = {c.get("id"): c.get("label", "") for c in (coverage.get("app_classes") or [])}
        control_labels = {c.get("key"): c.get("label", "") for c in (coverage.get("controls") or [])}
        cov_rows: list[list[Any]] = []
        for row in (coverage.get("matrix") or []):
            cohort = row.get("label", "") or row.get("cohort", "")
            for cell_key, cell in (row.get("cells") or {}).items():
                cell = cell if isinstance(cell, dict) else {}
                class_id, _, control_key = str(cell_key).partition("|")
                # These previously read `cell["count"]` and `cell["uncovered"]` — keys the cell
                # has never carried — so both columns exported blank for every row. An empty
                # "Uncovered" column in a spreadsheet reads as "nothing uncovered".
                cov_rows.append([
                    cohort,
                    class_labels.get(class_id, class_id),
                    control_labels.get(control_key, control_key),
                    cell.get("state", ""),
                    cell.get("users_covered", ""), cell.get("users_total", ""),
                    cell.get("apps_covered", ""), cell.get("apps_total", ""),
                    cell.get("uncovered_total", ""),
                    ", ".join(str(a) for a in (cell.get("apps_missing") or [])),
                    ", ".join(str(p) for p in (cell.get("policies") or [])),
                ])
        wb.sheet("CA coverage matrix",
                 ["Cohort", "Application class", "Control", "State",
                  "Users covered", "Users in cohort", "Apps covered", "Apps in class",
                  "Users uncovered", "Applications not reached", "Policies"],
                 cov_rows,
                 note="A cell is 'covered' only when the whole cohort AND every application in "
                      "the class is reached. 'n/a' means Entra does not offer that control for "
                      "that target — it is not a gap.",
                 highlight={"State": "severity"})

        derived = coverage.get("derived") or {}
        shadowed = derived.get("shadowed_classes") or {}
        unattributed = derived.get("unattributed_apps") or {}
        derived_rows: list[list[Any]] = [
            ["Shadowed class", class_labels.get(cid, cid),
             ", ".join((shadowed.get("detail") or {}).get(cid, []))]
            for cid in (shadowed.get("classes") or [])
        ]
        if unattributed.get("measured"):
            derived_rows += [["Unattributed application", a.get("name", ""), a.get("app_id", "")]
                             for a in (unattributed.get("apps") or [])]
        else:
            # NOT the same as "none found", and must not export as an empty section.
            derived_rows.append(["Unattributed applications", "NOT MEASURED",
                                 str(unattributed.get("reason") or "")])
        wb.sheet("CA derived exposure", ["Kind", "Subject", "Detail"], derived_rows)

        wb.sheet(
            "CA cohorts", ["Cohort", "Key", "Size"],
            [[c.get("label", ""), c.get("key", ""), c.get("size", "")]
             for c in (coverage.get("cohorts") or [])],
            note="; ".join(f"{k}={v}" for k, v in (coverage.get("headline") or {}).items()
                           if not isinstance(v, (list, dict))),
        )

        conflicts = ca_analysis.get("conflicts") or []
        wb.sheet(
            "CA conflicts",
            ["Kind", "Policy", "Policy state", "Other policy", "Affected", "Detail"],
            [[c.get("kind", ""), c.get("policy_name", "") or c.get("policy_id", ""),
              c.get("policy_state", ""), c.get("other_name", "") or c.get("other_id", ""),
              c.get("affected", ""), c.get("detail", "")]
             for c in conflicts],
        )
        bg = ca_analysis.get("breakglass") or {}
        wb.sheet(
            "CA break-glass",
            ["Account", "UPN", "User id", "Score", "Global admin", "Confirmed", "Note",
             "Excluded from", "Why it looks like break-glass"],
            [[b.get("display_name", ""), b.get("upn", ""), b.get("user_id", ""),
              b.get("score", ""), b.get("is_global_admin", ""), b.get("confirmed", ""),
              b.get("note", ""), b.get("excluded_from", ""), b.get("reasons", "")]
             for b in (bg.get("candidates") or [])],
            note=str(bg.get("heuristic_note") or "") + " These are CANDIDATES detected by "
                 "heuristic, not a confirmed list.",
        )
        wb.sheet(
            "CA named locations",
            ["Name", "Id", "Kind", "Trusted", "IP ranges", "Countries", "Unknown countries"],
            [[n.get("display_name", ""), n.get("id", ""), n.get("kind", ""),
              n.get("is_trusted", ""), n.get("ip_ranges", ""), n.get("countries", ""),
              n.get("include_unknown_countries", "")]
             for n in (ca.get("named_locations") or [])],
        )
        wb.sheet(
            "CA auth strengths",
            ["Name", "Id", "Policy type", "Requirements satisfied", "Combinations"],
            [[a.get("display_name", ""), a.get("id", ""), a.get("policy_type", ""),
              a.get("requirements_satisfied", ""), a.get("combinations", "")]
             for a in (ca.get("auth_strengths") or [])],
        )

    # ---------------------------------------------------------------- privileged access
    wb.section("Privileged Access", SECTION_COLOURS["Privileged Access"])
    roles_blind = _blind_reason(snapshot, "roles")
    roles = data.get("roles") or {}
    if roles_blind:
        wb.blind_sheet("Role assignments", roles_blind)
    else:
        wb.sheet(
            "Role definitions",
            ["Role", "Id", "Template id", "Tier", "Privileged", "MS privileged", "Built-in",
             "Enabled"],
            [[r.get("display_name", ""), r.get("id", ""), r.get("template_id", ""),
              r.get("tier", ""), r.get("privileged", ""), r.get("ms_privileged", ""),
              r.get("is_built_in", ""), r.get("is_enabled", "")]
             for r in (roles.get("definitions") or [])],
            note=_caveat(snapshot, "roles"),
            highlight={"Tier": "severity"},
        )
        assign_fields = [
            ("Role", "role_name"), ("Role id", "role_id"), ("Tier", "role_tier"),
            ("Privileged", "role_privileged"), ("Principal", "principal_name"),
            ("UPN", "principal_upn"), ("Principal id", "principal_id"),
            ("Principal type", "principal_type"), ("User type", "principal_user_type"),
            ("Enabled", "principal_enabled"), ("Scope", "scope"), ("Kind", "assignment_kind"),
            ("Source", "source"),
        ]
        active = (roles.get("assignments") or []) + (roles.get("group_derived") or [])
        wb.sheet("Role assignments", _heads(assign_fields), _table(active, assign_fields),
                 note=_caveat(snapshot, "roles"), highlight={"Tier": "severity"})
        # [:9] stops before "Enabled": the collector sets `principal_enabled` on ACTIVE
        # assignments only, so on an eligibility it is not "disabled", it is not a field.
        elig_fields = assign_fields[:9] + [
            ("Member type", "member_type"), ("Start", "start"), ("End", "end"),
            ("Permanent", "permanent"), ("Status", "status"),
        ]
        wb.sheet("Role eligibility (PIM)", _heads(elig_fields),
                 _table(roles.get("eligible") or [], elig_fields),
                 dates={"Start", "End"}, highlight={"Tier": "severity"},
                 note="A blank 'End' on a permanent eligibility is correct — there is no expiry "
                      "to record. Read it with the 'Permanent' column, not alone.")

    pim_blind = _blind_reason(snapshot, "pim")
    if pim_blind:
        wb.blind_sheet("PIM policy config", pim_blind)
    else:
        pim_fields = [
            ("Role id", "role_id"), ("Policy id", "policy_id"), ("Scope", "scope_id"),
            ("Score", "score"), ("MFA on activation", "mfa_on_activation"),
            ("Approval required", "approval_required"), ("Approvers", "approver_count"),
            ("Justification required", "justification_required"),
            ("Ticket required", "ticket_required"),
            ("Auth context required", "auth_context_required"),
            ("Max activation hours", "max_activation_hours"),
            ("Eligibility expires", "eligibility_expires"),
            ("Assignment expires", "assignment_expires"),
            ("Notification recipients", "notification_recipients"),
        ]
        wb.sheet("PIM policy config", _heads(pim_fields),
                 _table((data.get("pim") or {}).get("policies") or [], pim_fields),
                 note=_caveat(snapshot, "pim"))
        wb.sheet(
            "PIM group eligibility",
            ["Group", "Group id", "Principal id", "Access", "Member type", "Status",
             "Assignment id"],
            # No principal NAME is collected here, and an always-blank "Principal" column
            # reads as "nobody is eligible".
            [[g.get("group_name", ""), g.get("group_id", ""), g.get("principal_id", ""),
              g.get("access_id", ""), g.get("member_type", ""), g.get("status", ""),
              g.get("id", "")]
             for g in ((data.get("pim") or {}).get("group_eligibilities") or [])],
        )

    act_blind = _blind_reason(snapshot, "activations")
    if act_blind:
        wb.blind_sheet("Activation sessions", act_blind)
    else:
        act_fields = [
            ("Plane", "plane"), ("Source", "source"), ("Principal", "principal_name"),
            ("UPN", "principal_upn"), ("Principal id", "principal_id"),
            ("Principal type", "principal_type"), ("Role", "role_name"), ("Tier", "tier"),
            ("Scope", "scope_name"), ("Scope type", "scope_type"),
            ("Subscription", "subscription_id"), ("Start", "start"), ("End", "end"),
            ("Justification", "justification"), ("Justification quality", "justification_quality"),
            ("Detail known", "detail_known"), ("Ticket", "ticket_number"),
            ("Self service", "self_service"), ("Requestor", "requestor_id"),
        ]
        sessions = activations if activations is not None else (
            (data.get("activations") or {}).get("sessions") or []
        )
        wb.sheet(
            "Activation sessions", _heads(act_fields), _table(sessions, act_fields),
            note="'Detail known' = No means the justification was not returned by Graph, not "
                 "that none was given. History beyond Graph's 30-day window comes from the "
                 "durable ledger.",
            dates={"Start", "End"}, highlight={"Tier": "severity"},
        )

    link = data.get("_azure_link") or {}
    if not link.get("available"):
        wb.blind_sheet(
            "Cross-plane power",
            str(link.get("reason") or "The Azure control-plane join is unavailable, so Entra "
                                     "roles cannot be correlated with Azure RBAC."),
        )
    else:
        from app.entra import crossplane

        wb.sheet(
            "Cross-plane power",
            ["Principal", "Id", "Kind", "Entra roles", "Entra permissions",
             "Azure roles (powerful)", "Azure roles (all)", "Broad scopes", "Subscriptions",
             "Both planes"],
            [[r["name"], r["principal_id"], r["kind"], r["entra_roles"],
              r["entra_permissions"], r["azure_roles"], r["azure_all_roles"],
              r["azure_broad_scopes"], r["azure_subscriptions"], r["both_planes"]]
             for r in crossplane.rows(data)],
            note=("Join is stale — older than the Entra snapshot. " if link.get("stale") else "")
                 + "'Both planes' is the row that matters: Entra power and powerful Azure RBAC "
                   "on the same principal.",
            highlight={"Both planes": "severity"},
        )

    # ---------------------------------------------------------------- applications
    wb.section("Applications", SECTION_COLOURS["Applications"])
    apps_blind = _blind_reason(snapshot, "apps")
    apps = data.get("apps") or {}
    if apps_blind:
        wb.blind_sheet("Service principals", apps_blind)
    else:
        sps = apps.get("service_principals") or []
        sp_fields = [
            ("Name", "display_name"), ("Object id", "object_id"), ("App id", "app_id"),
            ("Type", "sp_type"), ("Enabled", "enabled"), ("External", "is_external"),
            ("First party", "is_first_party"), ("Publisher", "publisher_name"),
            ("Verified publisher", "verified_publisher"), ("Home tenant", "app_owner_tenant_id"),
            ("Owners known", "owners_known"), ("Assignment required", "assignment_required"),
            ("Assignment known", "assignment_known"), ("Assigned principals", "assigned_principals"),
            ("Orphaned", "orphaned"), ("SSO mode", "sso_mode"),
            ("Disabled by Microsoft", "disabled_by_microsoft"),
        ]
        sp_rows = _table(sps, sp_fields)
        for row, sp in zip(sp_rows, sps):
            row.append((sp.get("risk") or {}).get("score", ""))
        wb.sheet("Service principals", _heads(sp_fields) + ["Risk score"], sp_rows,
                 note=_caveat(snapshot, "apps"), highlight={"Risk score": "bar"})

        regs = apps.get("applications") or []
        reg_fields = [
            ("Name", "display_name"), ("Object id", "object_id"), ("App id", "app_id"),
            ("SP object id", "sp_object_id"), ("Sign-in audience", "sign_in_audience"),
            ("Multi-tenant", "multi_tenant"), ("Created", "created_at"),
            ("Owners known", "owners_known"), ("FIC known", "fic_known"),
            ("Verified publisher", "verified_publisher"), ("App roles", "app_roles"),
            ("Notes", "notes"),
        ]
        reg_rows = _table(regs, reg_fields)
        for row, app in zip(reg_rows, regs):
            row.append((app.get("risk") or {}).get("score", ""))
        wb.sheet("App registrations", _heads(reg_fields) + ["Risk score"], reg_rows,
                 dates={"Created"}, highlight={"Risk score": "bar"})

        # --- child sheets. One row per credential / grant / URL.
        perm_rows: list[list[Any]] = []
        for sp in sps:
            for g in (sp.get("granted_app_permissions") or []):
                if isinstance(g, dict):
                    perm_rows.append([
                        sp.get("display_name", ""), sp.get("app_id", ""), "Application",
                        g.get("permission", ""), g.get("resource", ""), g.get("tier", ""),
                        "", g.get("permission_id", ""),
                    ])
            for g in (sp.get("granted_delegated") or []):
                if isinstance(g, dict):
                    # A delegated grant carries a LIST of scopes and a consent type; the
                    # difference between AllPrincipals and a single user is the whole point.
                    for scope in (g.get("scopes") or [""]):
                        perm_rows.append([
                            sp.get("display_name", ""), sp.get("app_id", ""), "Delegated",
                            scope, g.get("resource", ""), g.get("max_tier", ""),
                            g.get("consent_type", "") or g.get("principal_id", ""), g.get("id", ""),
                        ])
        wb.sheet(
            "App permissions granted",
            ["Application", "App id", "Kind", "Permission", "Resource", "Tier", "Consent", "Id"],
            perm_rows,
            note="Consent 'AllPrincipals' is tenant-wide: it applies to every user, not to the "
                 "one who consented.",
            highlight={"Tier": "severity"},
        )
        wb.sheet(
            "App permissions requested",
            ["Application", "App id", "Permission", "Resource", "Kind", "Tier", "Known"],
            [[app.get("display_name", ""), app.get("app_id", ""), r.get("permission", ""),
              r.get("resource", ""), r.get("kind", ""), r.get("tier", ""), r.get("known", "")]
             for app in regs for r in (app.get("requested_permissions") or [])
             if isinstance(r, dict)],
            note="Requested on the registration. Compare with 'App permissions granted' to see "
                 "what was asked for but never consented.",
            highlight={"Tier": "severity"},
        )

        cred_rows: list[list[Any]] = []
        for src_label, coll in (("Service principal", sps), ("App registration", regs)):
            for item in coll:
                for c in (item.get("credentials") or []):
                    if isinstance(c, dict):
                        cred_rows.append([
                            item.get("display_name", ""), item.get("app_id", ""), src_label,
                            c.get("kind", ""), c.get("display_name", ""),
                            c.get("start", ""), c.get("end", ""),
                            c.get("expired", ""), c.get("days_left", ""),
                            c.get("lifetime_days", ""), c.get("id", ""),
                        ])
        wb.sheet(
            "App credentials",
            ["Application", "App id", "Held by", "Kind", "Name", "Start", "Expires",
             "Expired", "Days left", "Lifetime days", "Key id"],
            cred_rows,
            note="A negative 'Days left' is an ALREADY EXPIRED credential still on the object.",
            dates={"Start", "Expires"},
            highlight={"Days left": "days_left", "Expires": "expiry_date"},
        )

        fic_rows: list[list[Any]] = []
        for app in regs:
            for f in (app.get("federated_credentials") or []):
                if isinstance(f, dict):
                    fic_rows.append([
                        app.get("display_name", ""), app.get("app_id", ""),
                        f.get("name", ""), f.get("issuer", ""), f.get("subject", ""),
                        f.get("audiences", ""), f.get("description", ""),
                    ])
        wb.sheet(
            "Federated credentials",
            ["Application", "App id", "Name", "Issuer", "Subject", "Audiences", "Description"],
            fic_rows,
            note="A loose subject here lets any workload matching it authenticate as this app.",
        )

        url_rows: list[list[Any]] = []
        for app in regs:
            for u in (app.get("redirect_uris") or []):
                u = u if isinstance(u, dict) else {"uri": u}
                url_rows.append([app.get("display_name", ""), app.get("app_id", ""),
                                 "App registration", u.get("uri", ""), u.get("type", ""),
                                 u.get("risk", "")])
        for sp in sps:
            risks = {str(r.get("uri", "")): r for r in (sp.get("reply_url_risks") or [])
                     if isinstance(r, dict)}
            for u in (sp.get("reply_urls") or []):
                url_rows.append([sp.get("display_name", ""), sp.get("app_id", ""),
                                 "Service principal", u, "",
                                 (risks.get(str(u)) or {}).get("risk", "")])
        wb.sheet("Redirect / reply URLs",
                 ["Application", "App id", "Held by", "URL", "Type", "Risk"], url_rows,
                 highlight={"Risk": "severity"})

        wb.sheet(
            "App owners",
            ["Application", "App id", "Held by", "Owner id"],
            [[sp.get("display_name", ""), sp.get("app_id", ""), "Service principal", o]
             for sp in sps for o in (sp.get("owner_ids") or [])]
            + [[a.get("display_name", ""), a.get("app_id", ""), "App registration", o]
               for a in regs for o in (a.get("owner_ids") or [])],
            note="An application with no owner has nobody to ask about it at review time. "
                 "Check 'Owners known' on the parent sheet before reading an absence as one.",
        )
        wb.sheet(
            "Provisioning jobs",
            ["Service principal", "App id", "Job id", "Template", "Status", "Quarantined",
             "Last execution"],
            [[sp.get("display_name", ""), sp.get("app_id", ""), j.get("id", ""),
              j.get("template", ""), j.get("code", ""), j.get("quarantine", ""),
              j.get("last_execution", "")]
             for sp in sps for j in (sp.get("provisioning_jobs") or []) if isinstance(j, dict)],
            dates={"Last execution"}, highlight={"Last execution": "stale_date"},
        )

        tenant_data = data.get("tenant") or {}
        # One row per setting per property. The whole policy used to be stringified into a
        # single cell as `k=v; k=v`, which is the one shape a spreadsheet cannot filter.
        consent_rows: list[list[Any]] = []
        for label, payload in (
            [("Authorization policy", tenant_data.get("authorization_policy")),
             ("Admin consent policy", tenant_data.get("admin_consent_policy")),
             ("Cross-tenant default", tenant_data.get("cross_tenant_default"))]
            + [(f"Permission grant policy: {(p or {}).get('id', '')}", p)
               for p in (tenant_data.get("permission_grant_policies") or [])]
        ):
            if isinstance(payload, dict):
                consent_rows += [[label, k, v] for k, v in payload.items()]
            elif payload not in (None, ""):
                consent_rows.append([label, "", payload])
            else:
                consent_rows.append([label, "", "NOT MEASURED"])
        wb.sheet("Consent posture", ["Setting", "Property", "Value"], consent_rows)

    # ---------------------------------------------------------------- risk & sign-ins
    wb.section("Risk & sign-ins", SECTION_COLOURS["Risk & sign-ins"])
    risk_blind = _blind_reason(snapshot, "risk")
    risk = data.get("risk") or {}
    signins = risk.get("signins") or {}
    if risk_blind:
        wb.blind_sheet("Sign-in summary", risk_blind)
    else:
        sampled = bool(signins.get("sampled"))
        wb.sheet(
            "Sign-in summary",
            ["Metric", "Value"],
            [["Window start (UTC)", as_datetime(signins.get("window_start"))
              or signins.get("window_start", "")],
             ["Window end (UTC)", as_datetime(signins.get("window_end"))
              or signins.get("window_end", "")],
             ["Lookback days", signins.get("lookback_days", "")],
             ["Sampled", "YES — these totals are extrapolated from a sample" if sampled else "No"],
             ["Total sign-ins", signins.get("total", "")],
             ["Success", signins.get("success", "")],
             ["Failure", signins.get("failure", "")],
             ["Failure rate", signins.get("failure_rate", "")],
             ["Interactive", signins.get("interactive", "")],
             ["MFA challenged", signins.get("mfa_challenged", "")],
             ["Legacy auth sign-ins", signins.get("legacy", "")],
             ["Legacy auth successful users", signins.get("legacy_success_users", "")]],
            note="Sampled: totals are extrapolated, not counted." if sampled else _caveat(snapshot, "risk"),
        )
        # One sheet per dimension. Stacked into a single Dimension/Key/Count/Detail grid they
        # could not be pivoted or charted, and every figure beyond the count — success, failure,
        # MFA — was stringified into the Detail cell as `k=v; k=v`, which is the one shape a
        # spreadsheet cannot work with.
        any_breakdown = False
        for dim, title in BREAKDOWN_SHEETS.items():
            bucket = signins.get(dim)
            if isinstance(bucket, dict) and bucket:
                wb.sheet(title, ["Key", "Count"], [[k, v] for k, v in bucket.items()])
                any_breakdown = True
            elif isinstance(bucket, list) and bucket:
                heads, rows, dims_dates = _explode([b for b in bucket if isinstance(b, dict)])
                if heads:
                    wb.sheet(title, heads, rows, dates=dims_dates)
                    any_breakdown = True
        if not any_breakdown:
            wb.sheet("Sign-in breakdowns", ["Status", "Why"],
                     [["NOT MEASURED", "The sign-in collector returned no per-dimension "
                                       "breakdowns for this window."]])

        ru_fields = [("Name", "name"), ("UPN", "upn"), ("Id", "id"), ("Risk level", "level"),
                     ("State", "state"), ("Detail", "detail"), ("Last updated", "last_updated")]
        wb.sheet("Risky users", _heads(ru_fields), _table(risk.get("risky_users") or [], ru_fields),
                 dates={"Last updated"}, highlight={"Risk level": "severity", "State": "severity"})
        wb.sheet(
            "Risk detections",
            ["Detection id", "Type", "Risk level", "State", "User", "User id", "Detected at"],
            [[d.get("id", ""), d.get("type", ""), d.get("level", ""), d.get("state", ""),
              d.get("upn", ""), d.get("user_id", ""), d.get("detected_at", "")]
             for d in (risk.get("risk_detections") or [])],
            dates={"Detected at"},
            highlight={"Risk level": "severity", "State": "severity"},
        )
        wb.sheet(
            "Risky service principals",
            ["Name", "App id", "Object id", "Risk level", "State", "Detail", "Last updated"],
            [[s.get("display_name", "") or s.get("name", ""), s.get("app_id", ""),
              s.get("id", ""), s.get("level", ""), s.get("state", ""), s.get("detail", ""),
              s.get("last_updated", "")]
             for s in (risk.get("risky_service_principals") or [])],
            dates={"Last updated"}, highlight={"Risk level": "severity"},
        )
        # Data-driven columns: the pattern records were read through guessed key names and
        # exported a row whose label and description were both blank.
        pattern_heads, pattern_rows, pattern_dates = _explode(
            [p for p in (risk.get("patterns") or []) if isinstance(p, dict)])
        wb.sheet("Sign-in patterns", pattern_heads or ["Pattern", "Count", "Description"],
                 pattern_rows, dates=pattern_dates)

    # ---------------------------------------------------------------- people
    wb.section("Directory", SECTION_COLOURS["Directory"])
    people_blind = _blind_reason(snapshot, "people")
    people = data.get("people") or {}
    if people_blind:
        wb.blind_sheet("Users", people_blind)
    else:
        user_fields = [
            ("Display name", "display_name"), ("UPN", "mail"), ("Id", "id"),
            ("Enabled", "enabled"), ("External state", "external_user_state"),
            ("On-prem synced", "on_prem_synced"), ("MFA registered", "mfa_registered"),
            ("MFA capable", "mfa_capable"), ("Passwordless capable", "passwordless_capable"),
            ("Methods", "methods"), ("Last sign-in", "last_signin"),
            ("Last non-interactive", "last_noninteractive_signin"),
            ("Licenses", "licence_count"), ("Job title", "job_title"),
            ("Department", "department"), ("Company", "company_name"),
            ("Employee id", "employee_id"),
            ("Created", "created_at"), ("Admin reported", "is_admin_reported"),
        ]
        user_dates = {"Last sign-in", "Last non-interactive", "Created"}
        users = people.get("users") or []
        wb.sheet("Users", _heads(user_fields), _table(users, user_fields),
                 note="Personal data. " + (_caveat(snapshot, "people") or ""),
                 dates=user_dates, highlight={"Last sign-in": "stale_date"})
        # The MFA gap as its own sheet: it is the actionable subset and the tab caps it at 500.
        # The four method columns are dropped rather than carried: on a sheet defined as "has no
        # registered method" they are blank on every row by construction, and four blank columns
        # invite the reader to wonder what was not collected.
        gap_fields = [f for f in user_fields
                      if f[0] not in {"MFA registered", "MFA capable", "Passwordless capable",
                                      "Methods"}]
        gap = [u for u in users if u.get("enabled") and not u.get("mfa_registered")]
        wb.sheet(
            "MFA registration gap", _heads(gap_fields), _table(gap, gap_fields),
            note="Enabled users with no registered strong authentication method.",
            dates=user_dates, highlight={"Last sign-in": "stale_date"},
        )
        group_fields = [
            ("Name", "display_name"), ("Id", "id"), ("Description", "description"),
            ("Security enabled", "security_enabled"), ("Mail enabled", "mail_enabled"),
            ("Unified", "unified"), ("Dynamic", "dynamic"),
            ("Membership rule", "membership_rule"), ("Rule state", "membership_rule_state"),
            ("Role assignable", "is_assignable_to_role"), ("On-prem synced", "on_prem_synced"),
            ("Owners known", "owners_known"), ("Visibility", "visibility"), ("Created", "created_at"),
        ]
        wb.sheet("Groups", _heads(group_fields), _table(people.get("groups") or [], group_fields),
                 dates={"Created"})

    # ---------------------------------------------------------------- governance
    wb.section("Governance", SECTION_COLOURS["Governance"])
    # --- guest (B2B) hygiene -------------------------------------------------------
    # Exported here rather than under Directory because a review campaign is a governance
    # act: these two sheets ARE the working document for "which external access do we still
    # want", which is why the whole population is written rather than a capped page.
    people_blind = _blind_reason(snapshot, "people")
    if people_blind:
        wb.blind_sheet("Guests", people_blind)
    else:
        from app.entra import guests as guests_mod
        from app.entra import snapshot as snapshot_mod

        g = guests_mod.summarize(data.get("people") or {},
                                 stale_days=snapshot_mod.settings()["guest_stale_days"])
        g["domains"] = guests_mod.annotate_partners(
            g["domains"], (data.get("tenant") or {}).get("cross_tenant_partners") or {})
        wb.sheet(
            "Guests",
            ["Guest", "Sign-in address", "Organization", "Domain class", "Lifecycle",
             "Account", "Invited", "Invited (days)", "Accepted", "Last human sign-in",
             "Human (days)", "Last any activity", "Any (days)", "Sign-in measured",
             "Sponsors", "Company", "Licenses"],
            [[r["display_name"], r["mail"] or r["upn"], r["domain"], r["domain_class"],
              guests_mod.LIFECYCLE_LABEL.get(r["lifecycle"], r["lifecycle"]),
              "Enabled" if r["enabled"] else "Disabled",
              r["invited_at"], r["invited_days_ago"], r["accepted_at"],
              # "never" and "not measured" are DIFFERENT and must stay different — a reviewer
              # sorting this column would otherwise revoke access that was simply never looked
              # at. They are carried by 'Lifecycle' and 'Sign-in measured' rather than as words
              # in the date column, which keeps the date column a real date and therefore
              # sortable and filterable. A blank here is read WITH those two columns.
              r["last_human_signin"] if r["signin_known"] else "",
              r["last_human_days_ago"],
              r["last_any_signin"] if r["signin_known"] else "",
              r["last_any_days_ago"],
              "yes" if r["signin_known"] else "no",
              "; ".join(s.get("display_name", "") for s in r["sponsors"]),
              r["company_name"], r["licence_count"]]
             for r in g["guests"]],
            note="A blank sign-in date means EITHER never signed in OR never measured — read "
                 "'Sign-in measured' to tell which. " + (_caveat(snapshot, "people") or ""),
            dates={"Invited", "Accepted", "Last human sign-in", "Last any activity"},
            highlight={"Last human sign-in": "stale_date"},
        )
        wb.sheet(
            "Guest partner orgs",
            ["Organization", "Domain", "Partner tenant", "Domain class", "Guests",
             "Enabled", "Disabled", "Pending", "Never used", "Dormant", "Active",
             "Not measured", "Oldest invite (days)", "Cross-tenant policy", "Why"],
            [[d.get("partner_name") or d["domain"], d["domain"], d.get("partner_tenant_id", ""),
              d["domain_class"], d["guests"], d["enabled"], d["disabled"], d["pending"],
              d["never_used"], d["dormant"], d["active"], d["not_measured"],
              d["oldest_invite_days"], d.get("governance", ""), d.get("governance_reason", "")]
             for d in g["domains"]],
            highlight={"Guests": "bar"},
        )

    gov_blind = _blind_reason(snapshot, "governance")
    gov = data.get("governance") or {}
    if gov_blind:
        wb.blind_sheet("Access reviews", gov_blind)
    else:
        wb.sheet(
            "Access reviews",
            ["Review", "Id", "Status", "Created", "Last modified", "Scope kind", "Scope target",
             "Scope query"],
            [[r.get("display_name", ""), r.get("id", ""), r.get("status", ""),
              r.get("created_at", ""), r.get("last_modified", ""),
              (r.get("scope") or {}).get("kind", "") if isinstance(r.get("scope"), dict) else "",
              (r.get("scope") or {}).get("target", "") if isinstance(r.get("scope"), dict) else r.get("scope", ""),
              (r.get("scope") or {}).get("query", "") if isinstance(r.get("scope"), dict) else ""]
             for r in (gov.get("reviews") or [])],
            note=_caveat(snapshot, "governance"),
            dates={"Created", "Last modified"},
        )
        wb.sheet(
            "Entitlement packages",
            ["Package", "Id", "Description", "Catalog", "Hidden", "Created"],
            [[p.get("display_name", ""), p.get("id", ""), p.get("description", ""),
              p.get("catalog_id", ""), p.get("hidden", ""), p.get("created_at", "")]
             for p in (gov.get("packages") or [])],
            dates={"Created"},
        )
        assign_heads, assign_rows, assign_dates = _explode(
            [a for a in (gov.get("assignments") or []) if isinstance(a, dict)])
        wb.sheet("Entitlement assignments",
                 assign_heads or ["Package", "Principal", "State"], assign_rows,
                 dates=assign_dates)
        workflows = gov.get("workflows") or []
        wb.sheet(
            "Lifecycle workflows",
            ["Workflow", "Category", "Enabled", "Detail"],
            [[w.get("display_name", ""), w.get("category", ""), w.get("enabled", ""), w]
             for w in workflows]
            # A bare header row reads as a clean result. The distinction this workbook exists to
            # preserve is "nothing configured" versus "we could not look", so it is stated in
            # the sheet rather than only on the Index, where nobody scrolling tabs will see it.
            or [["NONE CONFIGURED", "", "",
                 "No lifecycle workflows exist in this tenant. The governance domain WAS "
                 "readable, so this is an absence of configuration, not an absence of data."]],
            note="" if workflows else
                 "No lifecycle workflows are configured. The governance domain WAS readable, so "
                 "this is an absence of configuration, not an absence of data.",
        )

    # ---------------------------------------------------------------- graph & setup
    wb.section("Blast radius", SECTION_COLOURS["Blast radius"])
    if escalations is not None:
        wb.sheet(
            "Escalations",
            ["Primitive", "Source", "Target", "Confidence", "Rule", "Reason"],
            [[e.get("primitive", ""), e.get("source", ""), e.get("target", ""),
              e.get("confidence", ""), e.get("rule", ""), e.get("reason", "")]
             for e in escalations],
            highlight={"Confidence": "severity"},
        )
    wb.section("Setup & coverage", SECTION_COLOURS["Setup & coverage"])
    if setup_tiers is not None:
        wb.sheet(
            "Setup checklist",
            ["Tier", "Name", "Complete", "Granted", "Missing", "What it unlocks"],
            # `label` is not a key a tier has ever had; the name column exported blank.
            [[t.get("tier", ""), t.get("name", ""), t.get("complete", ""),
              t.get("granted", ""), t.get("missing", ""), t.get("unlocks", "")]
             for t in setup_tiers],
        )
    wb.sheet(
        "Diagnostics",
        ["Signal", "Kind", "Detail"],
        [[sid, "error", msg] for sid, msg in sorted(errors.items())]
        + [[sid, "not measured", msg] for sid, msg in sorted(not_measured.items())],
    )

    wb.index_sheet(colour=SECTION_COLOURS["Overview"])
    return wb.to_bytes()
