"""Safe CSV, JSON, and XLSX exports for Essential Activity Log coverage."""
from __future__ import annotations

import csv
import io
import json
import re
from typing import Any

_SECRET_KEY = re.compile(r"(?:secret|token|password|credential|authorization|signature|sig|sas|key)$", re.IGNORECASE)
_URL_QUERY = re.compile(r"(https?://[^\s?\"']+)\?[^\s\"']+", re.IGNORECASE)
_COLUMNS = (
    "category", "status", "covered_subscriptions", "missing_subscriptions",
    "partial_subscriptions", "routing", "existing_rules", "issues", "recommended_action",
)


def _safe_scalar(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    value = _URL_QUERY.sub(r"\1?<redacted>", value)
    stripped = value.lstrip("\t\r\n ")
    return "'" + value if stripped and stripped[0] in "=+-@" else value


def sanitize(value: Any, *, key: str = "") -> Any:
    """Recursively remove secret-bearing fields and signed URL queries from an export."""
    if key and _SECRET_KEY.search(key):
        return "<redacted>"
    if isinstance(value, dict):
        return {str(name): sanitize(item, key=str(name)) for name, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [sanitize(item) for item in value]
    return _safe_scalar(value)


def export_rows(coverage: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten coverage into one complete, audit-friendly row per category."""
    scopes = list(coverage.get("scopes") or [])
    rows: list[dict[str, Any]] = []
    for summary in coverage.get("categories") or []:
        category = str(summary.get("category") or "")
        cells = [
            (scope, cell)
            for scope in scopes for cell in scope.get("categories") or []
            if str(cell.get("category") or "") == category
        ]
        def subscriptions(statuses: set[str]) -> list[str]:
            return [
                str(scope.get("subscription_display_name") or scope.get("subscription_id") or "")
                for scope, cell in cells if str(cell.get("status") or "") in statuses
            ]

        rows.append({
            "category": category,
            "status": summary.get("status", "unknown"),
            "covered_subscriptions": subscriptions({"covered"}),
            "missing_subscriptions": subscriptions({"missing", "disabled", "no_routing", "unknown"}),
            "partial_subscriptions": subscriptions({"partial"}),
            "routing": [
                {
                    "subscription_id": scope.get("subscription_id", ""),
                    "status": cell.get("status", ""),
                    "pending_effect": cell.get("pending_effect", "none"),
                    "projected_status": cell.get("projected_status", cell.get("status", "")),
                    "rules": [
                        {"rule_id": rule.get("id", ""), "routing": rule.get("routing", {})}
                        for rule in cell.get("rules") or []
                    ],
                }
                for scope, cell in cells
            ],
            "existing_rules": [
                {"subscription_id": scope.get("subscription_id", ""), **rule}
                for scope, cell in cells for rule in cell.get("rules") or []
            ],
            "issues": [
                {"subscription_id": scope.get("subscription_id", ""), **issue}
                for scope, cell in cells for issue in cell.get("issues") or []
            ],
            "recommended_action": [
                {"subscription_id": scope.get("subscription_id", ""), "action": cell.get("recommended_action", "")}
                for scope, cell in cells if cell.get("recommended_action")
            ],
        })
    return sanitize(rows)


def export_document(payload: dict[str, Any]) -> dict[str, Any]:
    document = {
        "connection_id": payload.get("connection_id", ""),
        "scope": payload.get("scope", {}),
        "coverage_percent": (payload.get("coverage") or {}).get("coverage_percent", 0),
        "complete": (payload.get("coverage") or {}).get("complete", False),
        "partial": (payload.get("coverage") or {}).get("partial", False),
        "rows": export_rows(payload.get("coverage") or {}),
    }
    return sanitize(document)


def _cell(value: Any) -> Any:
    safe = sanitize(value)
    if isinstance(safe, (dict, list)):
        return json.dumps(safe, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return safe


def to_json(payload: dict[str, Any]) -> str:
    return json.dumps(export_document(payload), indent=2, ensure_ascii=True, default=str)


def to_csv(payload: dict[str, Any]) -> str:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(_COLUMNS))
    writer.writeheader()
    for row in export_document(payload)["rows"]:
        writer.writerow({column: _cell(row.get(column, "")) for column in _COLUMNS})
    return buffer.getvalue()


def to_workbook(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Activity Log Coverage"
    sheet.append(list(_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F6CBD")
    for row in export_document(payload)["rows"]:
        sheet.append([_cell(row.get(column, "")) for column in _COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(70, max(16, len(column) + 2))
        for cell in list(sheet.columns)[index - 1]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
