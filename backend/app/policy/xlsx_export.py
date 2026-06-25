"""Generic multi-sheet Excel (.xlsx) export for the Policy screens.

The pivot/group logic lives in the frontend (one source of truth), so this module is a thin,
generic "sheets → styled workbook" writer: each sheet is a list of column headers + rows. Pivot
sheets additionally pass per-row ``outline_levels`` so the workbook gets **native Excel grouping**
(the expand/collapse outline on the left), reproducing the pivot's hierarchy in Excel itself.

Uses openpyxl — the repo's standard xlsx writer (see rbac/export.py, fmea/excel.py)."""
from __future__ import annotations

from io import BytesIO
from typing import Any


def _safe_title(title: str) -> str:
    """Excel sheet titles: ≤31 chars, no : \\ / ? * [ ]."""
    bad = ':\\/?*[]'
    t = "".join(c for c in (title or "Sheet") if c not in bad).strip() or "Sheet"
    return t[:31]


def _coerce(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (int, float, str)):
        return v
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v)


def build_workbook(sheets: list[dict[str, Any]]) -> bytes:
    """Build an .xlsx workbook from a list of sheet specs and return the raw bytes.

    Each sheet spec:
      - ``name``: sheet title.
      - ``columns``: list[str] header labels.
      - ``rows``: list[list[Any]] cell values (aligned to columns).
      - ``outline_levels`` (optional): list[int] parallel to ``rows`` — the Excel outline level for
        each row (0 = top). When present, row grouping is enabled so the sheet has the pivot's
        expand/collapse hierarchy natively in Excel.
      - ``bold_level0`` (optional, default True): bold rows at outline level 0 (group headers).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove the default empty sheet; we create our own.
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F6CBD")
    bold = Font(bold=True)

    if not sheets:
        sheets = [{"name": "Sheet1", "columns": ["(empty)"], "rows": []}]

    for spec in sheets:
        name = _safe_title(str(spec.get("name", "Sheet")))
        columns = [str(c) for c in (spec.get("columns") or [])]
        rows = spec.get("rows") or []
        outline = spec.get("outline_levels") or []
        bold_level0 = spec.get("bold_level0", True)

        ws = wb.create_sheet(name)
        ws.append(columns)
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        for i, r in enumerate(rows):
            ws.append([_coerce(v) for v in r])
            excel_row = i + 2  # 1 = header
            lvl = int(outline[i]) if i < len(outline) else 0
            if outline:
                if lvl > 0:
                    ws.row_dimensions[excel_row].outline_level = min(7, lvl)
                if lvl == 0 and bold_level0:
                    for c in range(1, len(columns) + 1):
                        ws.cell(row=excel_row, column=c).font = bold

        ws.freeze_panes = "A2"
        if outline:
            # Show outline above the data and summary on the left margin (Excel default for grouping).
            ws.sheet_properties.outlinePr.summaryBelow = False
        else:
            # Plain tables get an auto-filter for quick sorting/filtering in Excel.
            if rows:
                ws.auto_filter.ref = f"A1:{get_column_letter(max(1, len(columns)))}{len(rows) + 1}"

        # Autosize columns from the header + a sample of rows.
        for ci in range(1, len(columns) + 1):
            width = len(str(columns[ci - 1])) if ci - 1 < len(columns) else 8
            for r in rows[:300]:
                if ci - 1 < len(r):
                    width = max(width, len(str(_coerce(r[ci - 1]))))
            ws.column_dimensions[get_column_letter(ci)].width = min(70, max(10, width + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
