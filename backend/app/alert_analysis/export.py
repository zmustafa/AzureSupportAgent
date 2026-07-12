"""Safe flat CSV and JSON export for Alerts Manager snapshots."""
from __future__ import annotations

import csv
import io
import json
from typing import Any

_COLUMNS = (
    "row_kind",
    "finding_status",
    "finding_type",
    "risk_level",
    "overlap_group_ids",
    "overlap_confidence",
    "scope_kind",
    "scope_name",
    "subscription_id",
    "resource_group",
    "target_ids",
    "rule_id",
    "rule_name",
    "rule_type",
    "activity_category",
    "enabled",
    "severity",
    "signal_type",
    "signal_name",
    "operator",
    "threshold",
    "window",
    "frequency",
    "action_group_ids",
    "action_group_names",
    "recipients",
    "receiver_count",
    "firing_7d",
    "firing_30d",
    "last_fired",
    "cost_family",
    "cost_status",
    "cost_confidence",
    "cost_currency",
    "cost_period",
    "cost_catalog_version",
    "cost_monthly_usd",
    "cost_monthly_min_usd",
    "cost_monthly_max_usd",
    "cost_assumptions",
    "cost_components",
    "summary_current_monthly_usd",
    "summary_current_monthly_min_usd",
    "summary_current_monthly_max_usd",
    "summary_potential_disabled_monthly",
    "summary_potential_disabled_monthly_min",
    "summary_potential_disabled_monthly_max",
    "summary_priced_count",
    "summary_unknown_count",
    "issues",
    "related_rules",
    "explanation",
    "recommendation",
    "portal_url",
)


def _safe(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(str(item) for item in value)
    elif isinstance(value, dict):
        value = json.dumps(value, sort_keys=True, default=str)
    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip("\t\r\n ")
    return "'" + value if stripped and stripped[0] in "=+-@" else value


def _portal(resource_id: str) -> str:
    return f"https://portal.azure.com/#@/resource{resource_id}/overview" if resource_id.startswith("/") else ""


def _rule_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    overlaps = {item["id"]: item for item in snapshot.get("overlaps", [])}
    recipients = {item["fingerprint"]: item for item in snapshot.get("recipients", [])}
    rows: list[dict[str, Any]] = []
    summary = snapshot.get("cost_summary") or {}
    for rule in snapshot.get("rules", []):
        conditions = rule.get("conditions") or [{}]
        cost = rule.get("cost") or {}
        overlap_ids = rule.get("overlap_group_ids") or []
        overlap_confidence = "; ".join(
            sorted({str(overlaps[oid]["confidence"]) for oid in overlap_ids if oid in overlaps})
        )
        destinations = [recipients[fp].get("destination") or recipients[fp].get("masked", "") for fp in rule.get("receiver_fingerprints", []) if fp in recipients]
        for condition in conditions:
            rows.append(
                {
                    "row_kind": "rule",
                    "finding_status": rule.get("finding_status", "ok"),
                    "finding_type": "overlap" if overlap_ids else (rule.get("issues") or [""])[0],
                    "risk_level": "warning" if overlap_ids else ("error" if rule.get("issues") else ""),
                    "overlap_group_ids": overlap_ids,
                    "overlap_confidence": overlap_confidence,
                    "scope_kind": snapshot.get("scope_kind", ""),
                    "scope_name": snapshot.get("scope_name", ""),
                    "subscription_id": rule.get("subscription_id", ""),
                    "resource_group": rule.get("resource_group", ""),
                    "target_ids": [target.get("id", "") for target in rule.get("effective_targets", [])] or rule.get("scopes", []),
                    "rule_id": rule.get("id", ""),
                    "rule_name": rule.get("name", ""),
                    "rule_type": rule.get("type", ""),
                    "activity_category": rule.get("activity_category", ""),
                    "enabled": rule.get("enabled", False),
                    "severity": rule.get("severity_label", ""),
                    "signal_type": condition.get("signal_type", ""),
                    "signal_name": condition.get("signal_name", ""),
                    "operator": condition.get("operator", ""),
                    "threshold": condition.get("threshold", ""),
                    "window": condition.get("window", ""),
                    "frequency": condition.get("frequency", ""),
                    "action_group_ids": rule.get("action_group_ids", []),
                    "action_group_names": rule.get("action_group_names", []),
                    "recipients": destinations,
                    "receiver_count": rule.get("receiver_count", 0),
                    "firing_7d": rule.get("firing_7d", 0),
                    "firing_30d": rule.get("firing_30d", 0),
                    "last_fired": rule.get("last_fired", ""),
                    "cost_family": cost.get("family", ""),
                    "cost_status": cost.get("status", ""),
                    "cost_confidence": cost.get("confidence", ""),
                    "cost_currency": cost.get("currency", ""),
                    "cost_period": cost.get("period", ""),
                    "cost_catalog_version": cost.get("catalog_version", ""),
                    "cost_monthly_usd": cost.get("monthly_usd"),
                    "cost_monthly_min_usd": cost.get("monthly_min_usd"),
                    "cost_monthly_max_usd": cost.get("monthly_max_usd"),
                    "cost_assumptions": cost.get("assumptions", []),
                    "cost_components": cost.get("components", []),
                    "summary_current_monthly_usd": summary.get("monthly_usd"),
                    "summary_current_monthly_min_usd": summary.get("monthly_min_usd"),
                    "summary_current_monthly_max_usd": summary.get("monthly_max_usd"),
                    "summary_potential_disabled_monthly": summary.get("potential_disabled_monthly"),
                    "summary_potential_disabled_monthly_min": summary.get("potential_disabled_monthly_min"),
                    "summary_potential_disabled_monthly_max": summary.get("potential_disabled_monthly_max"),
                    "summary_priced_count": summary.get("priced_count", 0),
                    "summary_unknown_count": summary.get("unknown_count", 0),
                    "issues": rule.get("issues", []),
                    "related_rules": [name for oid in overlap_ids for name in overlaps.get(oid, {}).get("rule_names", []) if name != rule.get("name")],
                    "explanation": "; ".join(overlaps.get(oid, {}).get("explanation", "") for oid in overlap_ids),
                    "recommendation": "; ".join(overlaps.get(oid, {}).get("recommendation", "") for oid in overlap_ids),
                    "portal_url": _portal(str(rule.get("id", ""))),
                }
            )
    return rows


def _cost_summary_row(snapshot: dict[str, Any]) -> dict[str, Any]:
    summary = snapshot.get("cost_summary") or {}
    return {
        "row_kind": "cost_summary",
        "scope_kind": snapshot.get("scope_kind", ""),
        "scope_name": snapshot.get("scope_name", ""),
        "cost_currency": summary.get("currency", ""),
        "cost_period": summary.get("period", ""),
        "cost_catalog_version": summary.get("catalog_version", ""),
        "summary_current_monthly_usd": summary.get("monthly_usd"),
        "summary_current_monthly_min_usd": summary.get("monthly_min_usd"),
        "summary_current_monthly_max_usd": summary.get("monthly_max_usd"),
        "summary_potential_disabled_monthly": summary.get("potential_disabled_monthly"),
        "summary_potential_disabled_monthly_min": summary.get("potential_disabled_monthly_min"),
        "summary_potential_disabled_monthly_max": summary.get("potential_disabled_monthly_max"),
        "summary_priced_count": summary.get("priced_count", 0),
        "summary_unknown_count": summary.get("unknown_count", 0),
        "cost_assumptions": summary.get("assumptions", []),
        "cost_components": summary.get("by_family", {}),
    }


def _gap_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    rule_ids = {row.get("rule_id") for row in _rule_rows(snapshot)}
    for gap in snapshot.get("gaps", []):
        # Rule configuration gaps are already represented on rule rows. Add standalone
        # rows only for baseline/resource gaps that have no corresponding inventory row.
        if gap.get("rule_id") and gap.get("rule_id") in rule_ids:
            continue
        rows.append(
            {
                "row_kind": "gap",
                "finding_status": "gap",
                "finding_type": gap.get("type", ""),
                "risk_level": gap.get("risk", ""),
                "scope_kind": snapshot.get("scope_kind", ""),
                "scope_name": snapshot.get("scope_name", ""),
                "target_ids": gap.get("resource_id", ""),
                "rule_id": gap.get("rule_id", ""),
                "rule_name": gap.get("rule_name", ""),
                "signal_name": gap.get("signal", ""),
                "action_group_ids": gap.get("action_group_id", ""),
                "explanation": gap.get("explanation", ""),
                "recommendation": gap.get("recommendation", ""),
                "portal_url": _portal(str(gap.get("resource_id", "") or gap.get("rule_id", ""))),
            }
        )
    return rows


def to_csv(snapshot: dict[str, Any]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(_COLUMNS), extrasaction="ignore")
    writer.writeheader()
    for row in [_cost_summary_row(snapshot), *_rule_rows(snapshot), *_gap_rows(snapshot)]:
        writer.writerow({column: _safe(row.get(column, "")) for column in _COLUMNS})
    return buffer.getvalue()


def to_json(snapshot: dict[str, Any]) -> str:
    return json.dumps(snapshot, indent=2, default=str)


def to_workbook(snapshot: dict[str, Any], trend_points: list[dict[str, Any]] | None = None) -> bytes:
    """Build a human-readable highlighted workbook; no plaintext receiver destinations exist."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="0F6CBD")
    header_font = Font(bold=True, color="FFFFFF")
    fills = {
        "ok": PatternFill("solid", fgColor="E2F0D9"),
        "overlap": PatternFill("solid", fgColor="FFF2CC"),
        "gap": PatternFill("solid", fgColor="FCE4D6"),
        "error": PatternFill("solid", fgColor="F4CCCC"),
        "warning": PatternFill("solid", fgColor="FFF2CC"),
        "informational": PatternFill("solid", fgColor="D9EAF7"),
    }

    def sheet(name: str, headers: list[str], rows: list[list[Any]], row_tones: list[str] | None = None):
        ws = workbook.create_sheet(name[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top")
        for index, row in enumerate(rows, start=2):
            ws.append([_safe(value) for value in row])
            tone = row_tones[index - 2] if row_tones and index - 2 < len(row_tones) else ""
            if tone in fills:
                for cell in ws[index]:
                    cell.fill = fills[tone]
            for cell in ws[index]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        for column_index, header in enumerate(headers, start=1):
            width = len(header)
            for row in rows[:200]:
                if column_index - 1 < len(row):
                    width = max(width, len(str(_safe(row[column_index - 1]) or "")))
            ws.column_dimensions[get_column_letter(column_index)].width = min(60, max(11, width + 2))
        return ws

    summary_rows = [
        ["Scope", snapshot.get("scope_name", "")],
        ["Scope kind", snapshot.get("scope_kind", "")],
        ["Generated", snapshot.get("generated_at", "")],
        ["Rationalization score", snapshot.get("rationalization_score", "")],
        ["Source", snapshot.get("source", "")],
        ["Partial", bool(snapshot.get("partial"))],
        ["Cost currency", (snapshot.get("cost_summary") or {}).get("currency", "")],
        ["Cost catalog", (snapshot.get("cost_summary") or {}).get("catalog_version", "")],
        ["Current monthly estimate", (snapshot.get("cost_summary") or {}).get("monthly_usd")],
        ["Current monthly minimum", (snapshot.get("cost_summary") or {}).get("monthly_min_usd")],
        ["Current monthly maximum", (snapshot.get("cost_summary") or {}).get("monthly_max_usd")],
        ["Potential disabled monthly estimate", (snapshot.get("cost_summary") or {}).get("potential_disabled_monthly")],
        ["Potential disabled monthly minimum", (snapshot.get("cost_summary") or {}).get("potential_disabled_monthly_min")],
        ["Potential disabled monthly maximum", (snapshot.get("cost_summary") or {}).get("potential_disabled_monthly_max")],
        ["Priced rules", (snapshot.get("cost_summary") or {}).get("priced_count", 0)],
        ["Unknown-cost rules", (snapshot.get("cost_summary") or {}).get("unknown_count", 0)],
        ["Cost by family", (snapshot.get("cost_summary") or {}).get("by_family", {})],
        ["Cost assumptions", (snapshot.get("cost_summary") or {}).get("assumptions", [])],
        *[[str(key).replace("_", " ").title(), value] for key, value in (snapshot.get("kpis") or {}).items()],
        ["Recipient display", "Full email and phone destinations"],
    ]
    sheet("Summary", ["Metric", "Value"], summary_rows)

    rule_rows = _rule_rows(snapshot)
    rule_headers = list(_COLUMNS)
    sheet(
        "Rules",
        rule_headers,
        [[row.get(column, "") for column in rule_headers] for row in rule_rows],
        [str(row.get("finding_status", "")) for row in rule_rows],
    )

    action_rows: list[list[Any]] = []
    for group in snapshot.get("action_groups", []):
        receivers = group.get("receivers") or [{}]
        for receiver in receivers:
            action_rows.append(
                [
                    group.get("name", ""), group.get("id", ""), group.get("enabled", False),
                    group.get("active_receiver_count", 0), receiver.get("type", ""),
                    receiver.get("destination") or receiver.get("masked", ""), receiver.get("fingerprint", ""), receiver.get("enabled", False),
                ]
            )
    sheet(
        "ActionGroups",
        ["Action group", "Resource ID", "Group enabled", "Active receivers", "Receiver type", "Destination", "Fingerprint", "Receiver enabled"],
        action_rows,
    )

    overlap_rows = [
        [
            item.get("id", ""), item.get("type", ""), item.get("confidence", ""),
            item.get("target_id", ""), item.get("signal_type", ""), item.get("signal_name", ""),
            item.get("rule_names", []), item.get("shared_recipient_count", 0),
            item.get("notification_overlap", False), item.get("explanation", ""), item.get("recommendation", ""),
        ]
        for item in snapshot.get("overlaps", [])
    ]
    sheet(
        "Overlaps",
        ["Group", "Type", "Confidence", "Target", "Signal type", "Signal", "Rules", "Shared recipients", "Notification overlap", "Explanation", "Recommendation"],
        overlap_rows,
        ["error" if row[2] == "high" else "warning" for row in overlap_rows],
    )

    gap_rows = [
        [
            item.get("risk", ""), item.get("type", ""), item.get("resource_name", ""), item.get("resource_id", ""),
            item.get("rule_name", ""), item.get("rule_id", ""), item.get("action_group_id", ""),
            item.get("signal", ""), item.get("explanation", ""), item.get("recommendation", ""),
        ]
        for item in snapshot.get("gaps", [])
    ]
    sheet(
        "Gaps",
        ["Risk", "Gap type", "Resource", "Resource ID", "Rule", "Rule ID", "Action group", "Signal", "Explanation", "Recommendation"],
        gap_rows,
        [str(row[0]) for row in gap_rows],
    )

    def specialized(rule_type: str) -> list[list[Any]]:
        return [
            [
                rule.get("name", ""), rule.get("id", ""), rule.get("enabled", False), rule.get("severity_label", ""),
                (rule.get("conditions") or [{}])[0].get("signal_name", ""),
                (rule.get("conditions") or [{}])[0].get("query_fingerprint", ""),
                (rule.get("conditions") or [{}])[0].get("window", ""), rule.get("action_group_names", []),
                rule.get("finding_status", ""), rule.get("issues", []),
            ]
            for rule in snapshot.get("rules", []) if rule_type in rule.get("type", "")
        ]

    specialized_headers = ["Name", "Resource ID", "Enabled", "Severity", "Signal / detector", "Expression fingerprint", "Window", "Action groups", "Status", "Issues"]
    smart_rows = specialized("smartdetector")
    sheet("SmartDetector", specialized_headers, smart_rows, [str(row[8]) for row in smart_rows])
    prometheus_rows = specialized("prometheus")
    sheet("Prometheus", specialized_headers, prometheus_rows, [str(row[8]) for row in prometheus_rows])

    trend_rows = [
        [point.get("at", ""), point.get("pct", ""), point.get("extra", {}).get("overlap_groups", ""), point.get("extra", {}).get("gap_count", ""), point.get("extra", {}).get("notification_overlaps", ""), point.get("demo", False)]
        for point in (trend_points or [])
    ]
    sheet("Trends", ["At", "Score", "Overlap groups", "Gaps", "Duplicate paths", "Demo"], trend_rows)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
