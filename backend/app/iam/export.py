"""CSV / JSON / multi-sheet XLSX export of normalized access rows.

Two column sets: ``schema.COLUMNS`` (everything this product knows) and
``schema.SCANNER_COLUMNS`` (the frozen 46 the standalone all-azure-access scanner emits).
Passing ``columns=schema.SCANNER_COLUMNS`` projects an export back down to scanner shape so a
round trip through the import endpoint is byte-identical."""
from __future__ import annotations

import csv
import datetime as _dt
import io
import json
from typing import Any

from app.core.xlsx import DATETIME_FMT, WorkbookBuilder, as_datetime, cell_safe
from app.iam import schema

#: Kept as a module-local name because `app/api/assessments.py` imports it from here. The
#: behaviour is the shared one - the guard has exactly one definition, in app/core/xlsx.py.
_csv_safe = cell_safe


def to_csv(rows: list[dict[str, Any]], columns: tuple[str, ...] | None = None) -> str:
    """Serialize rows to CSV. Defaults to the full column set; pass
    ``schema.SCANNER_COLUMNS`` for scanner-shaped output."""
    cols = columns or schema.COLUMNS
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(cols), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: _csv_safe(row.get(c, "")) for c in cols})
    return buf.getvalue()


def to_json(rows: list[dict[str, Any]], columns: tuple[str, ...] | None = None) -> str:
    """Serialize rows to a pretty JSON array, optionally projected to a column subset."""
    if columns is not None:
        rows = [{c: r.get(c, "") for c in columns} for r in rows]
    return json.dumps(rows, indent=2, default=str)


# --------------------------------------------------------------------------- XLSX workbook
#: One tab color per section, so a twenty-odd sheet workbook groups visually the way the
#: screens do. Hues are kept far apart rather than pretty: two adjacent shades would imply a
#: relationship between sections that do not have one. The Index repeats the grouping as text.
SECTION_COLOURS: dict[str, str] = {
    "Overview": "44546A",        # slate — front matter and the coverage register
    "Access": "00B0F0",          # cyan — the grant lenses
    "Posture": "2E75B6",         # blue
    "Findings": "C00000",        # red
    "Right-sizing": "548235",    # green
    "Shadow access": "ED7D31",   # orange
    "Escalation": "7030A0",      # purple
    "Directory": "BF8F00",       # gold
    "Leavers": "833C00",         # brown
}

#: Columns whose value is a yes/no fact. The collectors speak four dialects for these —
#: `true`/`false`, `Yes`/blank, `yes`/`no` and `unknown`/`notApplicable` — so filtering "not
#: privileged" meant selecting blanks in one column, `false` in another and `no` in a third.
#: Normalised on the way into the workbook only; the CSV and JSON paths stay byte-faithful
#: because a round trip through the import endpoint has to reproduce the source exactly.
_TRI_COLUMNS: frozenset[str] = frozenset({
    "principalExists", "roleIsPrivileged", "roleHasDataActions", "isInherited",
    "pimManaged", "isPermanentEligible", "requiresApproval", "requiresMfa",
    "requiresJustification", "doNotApplyToChildScopes", "imported",
    "principalAccountEnabled", "principalOnPremSynced", "membershipGroupOnPremSynced",
    "membershipGroupRoleAssignable", "membershipGroupDynamic",
})

#: Schema columns holding a timestamp.
_DATE_COLUMNS: frozenset[str] = frozenset({
    "assignmentCreatedOn", "assignmentUpdatedOn",
    "eligibilityStartDateTime", "eligibilityEndDateTime", "activationExpiresOn",
})

_TRI_TRUE = {"true", "yes"}
_TRI_FALSE = {"false", "no"}


def _tri(value: Any) -> Any:
    """One vocabulary for every yes/no column: Yes / No / Not measured / n/a / blank.

    ``unknown`` becomes "Not measured" rather than staying a lowercase word that reads like a
    property of the account. On this data it is the value of `principalAccountEnabled` for most
    rows, and "we could not check whether this person still works here" is the single most
    important caveat in the file."""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    text = str(value).strip().lower()
    if text in _TRI_TRUE:
        return "Yes"
    if text in _TRI_FALSE:
        return "No"
    if text == "unknown":
        return "Not measured"
    if text in {"notapplicable", "n/a", "na"}:
        return "n/a"
    return value


def _scope_label(row: dict[str, Any]) -> str:
    """A readable scope.

    The collectors set ``scopeDisplayName`` to the ARM id itself for subscription- and
    management-group-scoped rows, which produced two 60-wide columns of the same
    ``/subscriptions/…`` string side by side. Falls back to the id, so nothing is ever lost."""
    name = str(row.get("scopeDisplayName") or "")
    scope = str(row.get("scope") or "")
    if name and name != scope:
        return name
    parts = [p for p in (row.get("subscriptionName"), row.get("resourceGroup"),
                         row.get("resourceName")) if p]
    return " / ".join(str(p) for p in parts) or scope


# A friendlier subset/order of the schema columns for the human-readable access lenses. The
# COMPLETE record is a sheet of its own ("Access — all columns"), so nothing is only available
# outside the workbook.
#
# `effect` leads deliberately. It used to be absent entirely, which meant the eight Deny
# assignments at production scale sat in a sheet called "Effective Access" among 5,506 Allow rows
# with nothing to tell them apart — anyone filtering or pivoting that sheet counted a control as
# a grant. `principalExists` is here for the same reason: 100 rows on that tenant point at
# principals that no longer exist, and without the column they read as live access.
# `assignmentCreatedOn` is here because "how long has this person held Owner?" is a review
# question, and it used to be answerable only by pivoting to the 71-column raw sheet.
_ACCESS_HEADERS: tuple[str, ...] = (
    "effect",
    "effectivePrincipalName",
    "effectivePrincipalUserPrincipalName",
    "effectivePrincipalType",
    "principalExists",
    "roleName",
    "roleIsPrivileged",
    "roleHasDataActions",
    "surface",
    "accessPath",
    "sourceGroupName",
    "scopeDisplayName",
    "scope",
    "scopeType",
    "subscriptionName",
    "subscriptionId",
    "resourceGroup",
    "assignmentState",
    "assignmentCreatedOn",
    "principalDisplayName",
    "principalId",
    "condition",
    "collector",
    # Identifiers, so a row can be traced back to Azure or fed to a remediation script.
    "assignmentId",
    "roleDefinitionId",
)

#: Applied to every lens sheet and to the raw-column sheet.
_ACCESS_HIGHLIGHT = {"effect": "severity", "roleIsPrivileged": "severity"}


def _access_row(row: dict[str, Any], headers: tuple[str, ...] | list[str]) -> list[Any]:
    """One spreadsheet row: the schema record projected, with the yes/no columns normalised."""
    out: list[Any] = []
    for h in headers:
        value = row.get(h, "")
        if h == "scopeDisplayName":
            value = _scope_label(row)
        elif h in _TRI_COLUMNS:
            value = _tri(value)
        out.append(value)
    return out



def to_workbook(
    *,
    rows: list[dict[str, Any]],
    overview: dict[str, Any],
    pivots: dict[str, list[dict[str, Any]]],
    pivot_labels: dict[str, str],
    directory: dict[str, Any],
    findings: dict[str, Any] | None = None,
    rightsizing: dict[str, Any] | None = None,
    bypass: dict[str, Any] | None = None,
    escalation: dict[str, Any] | None = None,
    scanners: list[dict[str, Any]] | None = None,
    score: dict[str, Any] | None = None,
    dataplane: list[dict[str, Any]] | None = None,
    leavers: dict[str, Any] | None = None,
) -> bytes:
    """Build the multi-sheet ``.xlsx`` workbook: everything the review screen can show.

    The access lenses are only half of it. A workbook that carries the grid but not the
    FINDINGS, the right-sizing, the shadow-access sweep or the escalation paths hands somebody
    a list of role assignments and calls it an access review — the analysis is the part a
    reviewer cannot reproduce themselves from the portal.

    Every analysis sheet distinguishes "nothing found" from "not measured". A sheet that is
    empty because the sweep never ran is written with the reason instead of the rows, because
    an empty sheet titled "Shadow Access" reads as a clean result.

    ``rows`` is the (optionally filtered) master row set. The optional payloads are passed in
    rather than computed here so this stays a pure formatter.
    """
    from openpyxl.styles import Font

    wb = WorkbookBuilder()
    access_dates = {h for h in _ACCESS_HEADERS if h in _DATE_COLUMNS}

    def _sheet(title: str, headers: list[str], data: list[list[Any]], **kw: Any) -> Any:
        return wb.sheet(title, headers, data, **kw)

    def _rows_for(predicate) -> list[list[Any]]:
        return [_access_row(r, _ACCESS_HEADERS) for r in rows if predicate(r)]

    def _blind_sheet(title: str, reason: str) -> None:
        """A sheet for something that could NOT be measured.

        Deliberately not an empty grid. "Shadow Access" with no rows reads as "no shadow access",
        which is the opposite of "the sweep never ran"."""
        wb.blind_sheet(title, reason)

    # 1. Summary — KPIs + generation metadata.
    wb.section("Overview", SECTION_COLOURS["Overview"])
    kpis = overview.get("kpis", {})
    collectors = overview.get("collectors", []) or []
    # A collector that never ran cannot produce a zero. `PIM eligible: 0` next to a skipped PIM
    # collector is the single most misleading cell this workbook can print.
    blocked = {
        str(c.get("collector", "")): str(c.get("message", "") or "")
        for c in collectors if str(c.get("status", "")) not in ("Succeeded", "")
    }
    pim_blocked = next((why for name, why in blocked.items() if "pim" in name.lower()), "")
    label_map = [
        ("Total grants", "total_assignments"),
        ("Unique principals", "unique_principals"),
        ("Privileged", "privileged"),
        ("Data-plane", "data_plane"),
        ("Group-derived", "group_derived"),
        ("Service-principal owners", "owners"),
        ("Entra directory roles", "entra_roles"),
        ("PIM eligible", "eligible"),
        ("Scopes", "scopes"),
        ("Subscriptions", "subscriptions"),
    ]
    ws0 = wb.first_sheet("Summary")
    ws0.append(["RBAC — Access Review export"])
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    generated = as_datetime(overview.get("generated_at")) or overview.get("generated_at", "")
    ws0.append(["Generated (UTC)", generated])
    if isinstance(generated, _dt.datetime):
        ws0.cell(row=ws0.max_row, column=2).number_format = DATETIME_FMT
    ws0.append(["Tenant", overview.get("tenant_id", "")])
    ws0.append(["Demo dataset", "Yes" if overview.get("demo") else "No"])
    if blocked:
        ws0.append(["COVERAGE", f"{len(blocked)} collector run(s) did not succeed — see "
                                "'Coverage & blind spots'. Counts below are not complete."])
    ws0.append([])
    ws0.append(["Metric", "Value"])
    hdr_row = ws0.max_row
    for c in range(1, 3):
        ws0.cell(row=hdr_row, column=c).font = wb._header_font
        ws0.cell(row=hdr_row, column=c).fill = wb._header_fill
    for label, key in label_map:
        value = kpis.get(key, 0)
        if label == "PIM eligible" and pim_blocked:
            # NOT a zero. Written as text so nobody can sum or chart it as one.
            value = f"NOT MEASURED — {pim_blocked}"
        ws0.append([label, value])
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 60

    # 2. Coverage & blind spots — deliberately before any sheet a reader might mistake for a
    # complete answer. Where PIM is unlicensed a large share of collector runs are skipped, and
    # a workbook that reports the result without saying so misrepresents itself.
    rollup: dict[tuple[str, str], dict[str, Any]] = {}
    for c in collectors:
        key = (str(c.get("collector", "")), str(c.get("status", "")))
        entry = rollup.setdefault(key, {"scopes": 0, "rows": 0, "message": ""})
        entry["scopes"] += 1
        entry["rows"] += int(c.get("rowsAdded", 0) or 0)
        if not entry["message"] and c.get("message"):
            entry["message"] = str(c.get("message"))
    _sheet(
        "Coverage & blind spots",
        ["Collector", "Status", "Scopes", "Rows", "Why"],
        # Anything that did not succeed sorts first: it is the reason to read this sheet.
        [[name, status, v["scopes"], v["rows"], v["message"]]
         for (name, status), v in sorted(
             rollup.items(), key=lambda kv: (kv[0][1] == "Succeeded", kv[0][0]))],
        note="A skipped or partial collector means the sheets derived from it are incomplete. "
             "'No findings' and 'we could not look' are opposite facts.",
        highlight={"Status": "severity"},
    )

    # 3–8. Access lenses.
    wb.section("Access", SECTION_COLOURS["Access"])
    _sheet("Effective Access", list(_ACCESS_HEADERS), _rows_for(lambda r: True),
           dates=access_dates, highlight=_ACCESS_HIGHLIGHT)
    # The complete record, every schema column. The lenses above are for reading; this is for
    # anyone who needs the field the lens left out.
    _sheet(
        "Access - all columns",
        list(schema.COLUMNS),
        [_access_row(r, schema.COLUMNS) for r in rows],
        dates={h for h in schema.COLUMNS if h in _DATE_COLUMNS},
        highlight=_ACCESS_HIGHLIGHT,
        note="Every schema column. A blank PIM column here means PIM was not readable, not "
             "that the assignment is unmanaged — check 'Coverage & blind spots'.",
    )
    _sheet("Privileged", list(_ACCESS_HEADERS), _rows_for(lambda r: bool(r.get("roleIsPrivileged"))),
           dates=access_dates, highlight=_ACCESS_HIGHLIGHT)
    _sheet("Group-Derived", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("accessPath") == schema.PATH_GROUP),
           dates=access_dates, highlight=_ACCESS_HIGHLIGHT)
    _sheet("SP Owners", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("accessPath") == schema.PATH_OWNER),
           dates=access_dates, highlight=_ACCESS_HIGHLIGHT)
    _sheet("Entra Roles", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("surface") == schema.SURFACE_ENTRA),
           dates=access_dates, highlight=_ACCESS_HIGHLIGHT)
    kv_rows = _rows_for(lambda r: r.get("surface") == schema.SURFACE_KEY_VAULT)
    if kv_rows:
        _sheet("Key Vault", list(_ACCESS_HEADERS), kv_rows,
               dates=access_dates, highlight=_ACCESS_HIGHLIGHT)

    # 9. Scopes freshness.
    wb.section("Overview", SECTION_COLOURS["Overview"])
    scope_headers = ["displayName", "scopeType", "status", "row_count", "collectors_attention", "generated_at", "demo"]
    _sheet(
        "Scopes",
        ["Scope", "Type", "Status", "Grants", "Attention", "Generated", "Demo"],
        [[s.get(h, "") for h in scope_headers] for s in overview.get("scopes", [])],
        dates={"Generated"}, highlight={"Status": "severity"},
    )

    # 10. Role definitions (directory reference).
    wb.section("Directory", SECTION_COLOURS["Directory"])
    rd = directory.get("role_defs", []) or []
    if rd:
        # `actionsCount`/`dataActionsCount` are keys only the demo fixture writes; the real
        # index carries the permission LISTS. Reading the counts exported blank on every row.
        def _count(rdef: dict[str, Any], list_key: str, count_key: str) -> Any:
            if isinstance(rdef.get(list_key), (list, tuple)):
                return len(rdef[list_key])
            return rdef.get(count_key, "")

        _sheet("Role Definitions",
               ["Role", "Category", "Privileged", "Data actions", "Actions #", "Data actions #",
                "Description"],
               [[r.get("roleName", ""), r.get("roleCategory", ""),
                 _tri(r.get("roleIsPrivileged")), _tri(r.get("roleHasDataActions")),
                 _count(r, "actions", "actionsCount"),
                 _count(r, "dataActions", "dataActionsCount"),
                 r.get("description", "")] for r in rd],
               highlight={"Privileged": "severity"})

    # 11. Principal directory (the resolved GUID → name map).
    pr = directory.get("principals", []) or []
    if pr:
        pr_headers = ["displayName", "principalType", "userPrincipalName", "appId", "principalId", "source"]
        _sheet("Principals", ["Name", "Type", "UPN", "App ID", "Object ID", "Source"],
               [[p.get(h, "") for h in pr_headers] for p in pr])

    # 12. Insights — every pivot flattened.
    insight_data: list[list[Any]] = []
    for key, items in pivots.items():
        title = pivot_labels.get(key, key)
        for it in items:
            insight_data.append([title, it.get("label", ""), it.get("count", 0)])
    _sheet("Insights", ["Pivot", "Label", "Count"], insight_data, highlight={"Count": "bar"})

    # 13. Diagnostics — collector statuses + any errors, one row per scope.
    wb.section("Overview", SECTION_COLOURS["Overview"])
    diag_headers = ["collector", "scopeLabel", "status", "rowsAdded", "message"]
    _sheet(
        "Diagnostics",
        ["Collector", "Scope", "Status", "Rows", "Message"],
        [[c.get(h, "") for h in diag_headers] for c in collectors],
        note="The per-scope detail behind 'Coverage & blind spots'.",
        highlight={"Status": "severity"},
    )


    # ---- 13+. The analysis. Everything below is a judgment the reader cannot reconstruct
    # from a list of role assignments, which is exactly why leaving it out made the workbook a
    # data dump rather than a review.
    if score is not None:
        wb.section("Posture", SECTION_COLOURS["Posture"])
        pillar_rows: list[list[Any]] = [[
            "OVERALL", score.get("grade", ""), score.get("score", ""),
            f"{round((score.get('coverage') or 0) * 100)}% of the weighted checks could be measured",
            "", "",
        ]]
        for p in score.get("pillars", []) or []:
            pillar_rows.append([
                p.get("label", p.get("key", "")),
                p.get("state", ""),
                # A blind pillar has no score. Writing 0 would read as "scored zero".
                "not measured" if p.get("score") is None else p.get("score"),
                p.get("reason", "") or p.get("desc", ""),
                p.get("findings", 0),
                f"{p.get('signals_measured', 0)}/{p.get('signals', 0)} checks measured",
            ])
        _sheet("Posture Score", ["Pillar", "State", "Score", "Notes", "Findings", "Coverage"],
               pillar_rows, highlight={"Score": "bar", "State": "severity"})

    if findings is not None:
        wb.section("Findings", SECTION_COLOURS["Findings"])
        items = findings.get("findings") or []
        _sheet(
            "Findings",
            ["Severity", "Pillar", "Signal", "Title", "Subject", "Detail", "Count", "State",
             "Remediation", "Frameworks", "Fingerprint"],
            [[
                f.get("severity", ""), f.get("pillar", ""), f.get("signal_id", ""),
                f.get("title", ""), f.get("subject_label", "") or f.get("subject", ""),
                f.get("detail", ""), f.get("count", 0), f.get("state", ""),
                f.get("remediation", ""), ", ".join(f.get("frameworks") or []), f.get("id", ""),
            ] for f in items],
            highlight={"Severity": "severity"},
        )
        # The checks that could NOT run, as their own sheet. Folding them into the findings list
        # would let "we could not look" disappear into "nothing found".
        unmeasured = findings.get("unmeasured") or []
        _sheet(
            "Findings - not measured",
            ["Signal", "Title", "Pillar", "Why it could not be checked"],
            [[u.get("signal_id", ""), u.get("title", ""), u.get("pillar", ""), u.get("reason", "")]
             for u in unmeasured]
            or [["", "Every registered check ran.", "", "Nothing was withheld."]],
            note="A check that could not run is not a pass.",
        )

    if scanners is not None:
        _sheet(
            "Scanners",
            ["Scanner", "Cadence", "Severity floor", "Can run", "Why it cannot run", "Total",
             "New", "Resolved", "Persisting", "Last run", "Due"],
            [[
                s.get("name", ""), s.get("cadence", ""), s.get("severity_floor", ""),
                "No" if s.get("blocked") else "Yes", "; ".join(s.get("blocked") or []),
                *( [ (s.get("counts") or {}).get(k, "") for k in ("total", "new", "resolved", "persisting") ]
                   if s.get("counts") else ["not measured"] * 4 ),
                s.get("last_run_at", ""),
                _tri(s.get("due")),
            ] for s in scanners],
            dates={"Last run"},
            note="A blank 'Last run' means this scanner has never run for this tenant, which is "
                 "not the same as running and finding nothing.",
            highlight={"Severity floor": "severity"},
        )

    if rightsizing is not None:
        wb.section("Right-sizing", SECTION_COLOURS["Right-sizing"])
        if not rightsizing.get("measured"):
            _blind_sheet(
                "Right-sizing",
                "; ".join(rightsizing.get("limitations") or ["Usage has not been collected for this tenant."]),
            )
        else:
            _sheet(
                "Right-sizing",
                ["Principal", "Type", "Scope", "Current roles", "Granted actions", "Used actions",
                 "Unused %", "Confidence", "Recommendation", "Note", "Window (days)",
                 "Window clamped"],
                # `window` is a dict. It used to reach the cell through str(), printing the
                # Python literal {'days': 60, 'clamped': True}. 'clamped' is load-bearing: it
                # means the usage evidence is shorter than intended, so a "100% unused" verdict
                # rests on less than it appears to.
                [[
                    r.get("principalName", "") or r.get("principalId", ""), r.get("principalType", ""),
                    r.get("scopeName", "") or r.get("scope", ""),
                    ", ".join(r.get("currentRoles") or []),
                    r.get("grantedActionCount", 0), r.get("usedActionCount", 0),
                    round((r.get("unusedRatio") or 0) * 100),
                    r.get("confidence", ""), r.get("recommendation", ""), r.get("note", ""),
                    (r.get("window") or {}).get("days", "") if isinstance(r.get("window"), dict) else r.get("window", ""),
                    _tri((r.get("window") or {}).get("clamped")) if isinstance(r.get("window"), dict) else "",
                ] for r in (rightsizing.get("recommendations") or [])],
                highlight={"Unused %": "bar", "Confidence": "severity"},
            )

    if bypass is not None:
        wb.section("Shadow access", SECTION_COLOURS["Shadow access"])
        if bypass.get("never_loaded"):
            _blind_sheet(
                "Shadow Access",
                "The RBAC-bypass sweep has not run, so shared keys, local authentication, admin "
                "users and SQL logins have NOT been checked. This is not a clean result.",
            )
        else:
            _sheet(
                "Shadow Access",
                ["Severity", "Family", "Resource", "Type", "Subscription", "Resource group",
                 "Bypass kind", "Finding", "Open", "RBAC-only possible", "Who can reach it",
                 "Detail", "Remediation"],
                [[
                    b.get("severity", ""), b.get("family", ""), b.get("resourceName", ""),
                    b.get("resourceType", ""), b.get("subscriptionId", ""), b.get("resourceGroup", ""),
                    b.get("bypassKind", ""), b.get("title", ""),
                    _tri(b.get("enabled")),
                    _tri(b.get("rbacOnlyPossible")),
                    b.get("reachableCount", "") if b.get("reachabilityAvailable") else "not determined",
                    b.get("detail", ""), b.get("remediation", ""),
                ] for b in (bypass.get("rows") or [])],
                highlight={"Severity": "severity", "Who can reach it": "bar"},
            )

    if escalation is not None:
        wb.section("Escalation", SECTION_COLOURS["Escalation"])
        paths = escalation.get("paths") or []
        _sheet(
            "Escalation Paths",
            ["From", "To", "Hops", "Min confidence", "Route", "Primitives"],
            [[
                p.get("fromLabel", "") or p.get("from", ""),
                (p.get("hops") or [{}])[-1].get("targetLabel", "") or p.get("to", ""),
                p.get("length", len(p.get("hops") or [])),
                p.get("min_confidence", ""),
                " -> ".join(
                    [str((p.get("hops") or [{}])[0].get("sourceLabel", ""))]
                    + [str(h.get("targetLabel", "")) for h in (p.get("hops") or [])]
                ),
                ", ".join(dict.fromkeys(str(h.get("primitive", "")) for h in (p.get("hops") or []))),
            ] for p in paths],
            highlight={"Min confidence": "severity"},
        )
        # What the map could not see travels WITH it. A path list without its blind spots is an
        # invitation to read "no path" as "no path exists".
        limits = escalation.get("limitations") or []
        if limits:
            _sheet("Escalation - blind spots", ["Limitation", "What this map cannot see"],
                   [[f"#{i}", x] for i, x in enumerate(limits, start=1)])

    if dataplane is not None:
        wb.section("Shadow access", SECTION_COLOURS["Shadow access"])
        _sheet(
            "Data-plane Coverage",
            ["Service", "Azure RBAC is the whole picture", "Why not", "Other doors"],
            [[
                s.get("label", ""),
                "yes" if s.get("rbac_is_complete") else "NO",
                s.get("blind_reason", ""),
                "; ".join(s.get("doors") or []),
            ] for s in dataplane],
        )

    if leavers is not None:
        wb.section("Leavers", SECTION_COLOURS["Leavers"])
        # Disabled accounts that still hold access. A blind sheet when account state was never
        # collected — an empty "Disabled Access" grid is the most reassuring page in the whole
        # workbook and must never be produced by not having looked.
        if not leavers.get("measured"):
            _blind_sheet("Disabled Access", str(leavers.get("reason") or "Account state not collected."))
        else:
            flat = flatten_identities(leavers.get("identities") or [], leavers.get("tiers") or {})
            keys = [k for k, _ in IDENTITY_HEADERS]
            _sheet(
                "Disabled Access",
                [label for _, label in IDENTITY_HEADERS],
                [[_tri(row.get(k, "")) if k in _IDENTITY_TRI else row.get(k, "") for k in keys]
                 for row in flat],
                dates=_IDENTITY_DATE_LABELS,
                highlight={"Exposure": "severity"},
            )

    wb.index_sheet(colour=SECTION_COLOURS["Overview"])
    return wb.to_bytes()


# ------------------------------------------------------- disabled-but-entitled export
# One row per PERSON. The main access export is one row per grant, which answers "what is
# granted"; this answers "who should not still be here", and a leaver holding Contributor on
# four subscriptions is one offboarding task rather than four findings.
IDENTITY_HEADERS: tuple[tuple[str, str], ...] = (
    ("displayName", "Name"),
    ("userPrincipalName", "User principal name"),
    ("principalType", "Type"),
    ("userType", "Member or guest"),
    ("accountEnabled", "Account enabled"),
    ("onPremSynced", "Synced from on-prem AD"),
    ("softDeleted", "In the recycle bin"),
    ("deletedDateTime", "Deleted at"),
    ("tierLabel", "Exposure"),
    ("lastSignIn", "Last sign-in"),
    ("signInInteractive", "Last interactive sign-in"),
    ("signInNonInteractive", "Last non-interactive sign-in"),
    ("signInSuccessful", "Last successful sign-in"),
    ("signInServicePrincipal", "Owned app last sign-in"),
    ("dormancyLabel", "Dormancy"),
    ("dormancyDays", "Days since last sign-in"),
    ("lastSignInSource", "Sign-in source"),
    ("lastActivity", "Last used (Activity Log)"),
    ("activityEvents", "Operations recorded"),
    ("oldestGrantAt", "Oldest grant"),
    ("newestGrantAt", "Newest grant"),
    ("grants", "Grants"),
    ("privilegedGrants", "Privileged grants"),
    ("highestRole", "Highest role"),
    ("planesText", "Planes"),
    ("directGrants", "Held directly"),
    ("groupGrants", "Held via group"),
    ("groupsText", "Groups granting access"),
    ("pimEligible", "PIM eligible"),
    ("permanentlyEligible", "Permanently eligible"),
    ("ownedText", "Owns service principals"),
    ("subscriptionsText", "Subscriptions"),
    ("scopesText", "Scopes"),
    ("principalId", "Object id"),
)

#: Identity columns holding a timestamp, keyed by the LABEL the sheet shows.
_IDENTITY_DATE_LABELS: frozenset[str] = frozenset({
    "Deleted at", "Last sign-in", "Last interactive sign-in", "Last non-interactive sign-in",
    "Last successful sign-in", "Owned app last sign-in", "Last used (Activity Log)",
    "Oldest grant", "Newest grant",
})
#: Identity fields carrying a yes/no fact, keyed by the SOURCE key.
_IDENTITY_TRI: frozenset[str] = frozenset({
    "accountEnabled", "onPremSynced", "softDeleted",
})


def flatten_identities(identities: list[dict[str, Any]], tiers: dict[str, Any]) -> list[dict[str, Any]]:
    """Project the report's identity records into flat, spreadsheet-shaped dicts.

    The list-valued fields are joined here rather than in the caller so the CSV and the XLSX
    cannot drift apart in what they show."""
    from app.iam.leavers import DORMANCY_LABELS

    out: list[dict[str, Any]] = []
    for i in identities:
        flat = dict(i)
        flat["tierLabel"] = str((tiers.get(i.get("tier"), {}) or {}).get("label") or i.get("tier") or "")
        flat["planesText"] = "; ".join(i.get("planes") or [])
        flat["groupsText"] = "; ".join(i.get("groupsGrantingAccess") or [])
        flat["ownedText"] = "; ".join(i.get("ownedServicePrincipals") or [])
        flat["subscriptionsText"] = "; ".join(i.get("subscriptions") or [])
        flat["scopesText"] = "; ".join(i.get("scopes") or [])
        sign = i.get("signIn") or {}
        flat["signInInteractive"] = sign.get("interactive", "")
        flat["signInNonInteractive"] = sign.get("nonInteractive", "")
        flat["signInSuccessful"] = sign.get("successful", "")
        flat["signInServicePrincipal"] = sign.get("servicePrincipal", "")
        # The LABEL, not the key: "Not measured" has to survive into the file, because a blank
        # cell in a spreadsheet reads as zero and this one means the opposite.
        flat["dormancyLabel"] = DORMANCY_LABELS.get(str(i.get("dormancyBucket") or ""), "")
        flat["dormancyDays"] = i.get("dormancyDays") if i.get("dormancyDays") is not None else ""
        if not i.get("activityMeasured"):
            flat["lastActivity"] = "not measured"
            flat["activityEvents"] = ""
        out.append(flat)
    return out


def to_identity_csv(identities: list[dict[str, Any]], tiers: dict[str, Any]) -> str:
    """One row per disabled identity, formula-injection neutralised like every other export."""
    flat = flatten_identities(identities, tiers)
    keys = [k for k, _ in IDENTITY_HEADERS]
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([label for _, label in IDENTITY_HEADERS])
    for row in flat:
        writer.writerow([_csv_safe(row.get(k, "")) for k in keys])
    return buf.getvalue()


def to_disabled_workbook(
    *,
    report: dict[str, Any],
    grants: list[dict[str, Any]],
    tenant_id: str = "",
    filters: dict[str, Any] | None = None,
) -> bytes:
    """The disabled-access workbook.

    The **Not measured** sheet is mandatory and is written even when it is empty of problems.
    A spreadsheet gets forwarded, filtered and pasted into a ticket long after it has left the
    screen that produced it, and every caveat the UI renders is gone by then. If the limits of
    the data are not IN the file, the file misrepresents itself the moment it is downloaded —
    so the denominator, the collection date and everything this report cannot see travel with
    the rows rather than beside them."""
    wb = WorkbookBuilder()

    def _sheet(title: str, headers: list[str], data: list[list[Any]], **kw: Any) -> None:
        wb.sheet(title, headers, data, **kw)

    tiers = report.get("tiers") or {}
    identities = report.get("identities") or []
    denom = report.get("denominator") or {}
    totals = report.get("totals") or {}
    counts = report.get("tier_counts") or {}
    measured = bool(report.get("measured"))

    # 1. Summary — the headline, its denominator, and what the tiers actually mean.
    wb.section("Overview", SECTION_COLOURS["Overview"])
    summary: list[list[Any]] = [
        ["Report", "Disabled accounts that still hold access"],
        ["Tenant", tenant_id],
        ["Generated (UTC)", as_datetime(report.get("generated_at")) or report.get("generated_at", "")],
        ["Account state collected", "yes" if measured else "NO — see Not measured"],
        ["", ""],
    ]
    if measured:
        summary += [
            ["Disabled identities holding access", totals.get("identities", 0)],
            ["Grants they hold", totals.get("grants", 0)],
            ["…of which privileged", totals.get("privileged_grants", 0)],
            ["Subscriptions touched", totals.get("subscriptions_touched", 0)],
            ["Held only through a group", totals.get("via_group_only", 0)],
            ["Synced from on-prem AD (fix in AD)", totals.get("on_prem_synced", 0)],
            ["Still PIM-eligible", totals.get("pim_eligible", 0)],
            ["", ""],
        ]
        for key, meta in tiers.items():
            summary.append([f"{meta.get('label', key)} — identities", counts.get(key, 0)])
            summary.append(["", meta.get("detail", "")])
        summary.append(["", ""])
    summary += [
        ["DENOMINATOR", ""],
        ["Principals holding access", denom.get("principals_with_access", 0)],
        ["…account state resolved", denom.get("state_resolved", 0)],
        ["…could NOT be checked", denom.get("state_unknown", 0)],
        ["…no account state applies (groups etc.)", denom.get("not_applicable", 0)],
    ]
    if filters:
        summary.append(["", ""])
        summary.append(["FILTERS APPLIED", ""])
        for k, v in filters.items():
            if v not in (None, "", False):
                summary.append([k, v])
    _sheet("Summary", ["Field", "Value"], summary)

    # 2. Identities.
    wb.section("Leavers", SECTION_COLOURS["Leavers"])
    flat = flatten_identities(identities, tiers)
    id_keys = [k for k, _ in IDENTITY_HEADERS]
    _sheet(
        "Identities",
        [label for _, label in IDENTITY_HEADERS],
        [[_tri(row.get(k, "")) if k in _IDENTITY_TRI else row.get(k, "") for k in id_keys]
         for row in flat],
        dates=_IDENTITY_DATE_LABELS,
        highlight={"Exposure": "severity"},
    )

    # 3. Grants — the row-level record, in the same friendly column order the main access
    #    workbook uses so the two are comparable, plus the account-state columns.
    wb.section("Access", SECTION_COLOURS["Access"])
    grant_headers = (*_ACCESS_HEADERS, "principalAccountEnabled", "principalOnPremSynced")
    _sheet(
        "Grants",
        list(grant_headers),
        [_access_row(r, grant_headers) for r in grants],
        dates={h for h in grant_headers if h in _DATE_COLUMNS},
        highlight=_ACCESS_HIGHLIGHT,
    )

    # 4. Via groups — its own sheet because its remediation is the opposite of the others':
    #    remove the member, never the assignment.
    wb.section("Leavers", SECTION_COLOURS["Leavers"])
    group_rows: list[list[Any]] = []
    for i in identities:
        for g in i.get("groupsGrantingAccess") or []:
            group_rows.append([
                g, i.get("displayName", ""), i.get("userPrincipalName", ""),
                i.get("groupGrants", 0), i.get("privilegedGrants", 0),
            ])
    group_rows.sort(key=lambda r: (str(r[0]).lower(), str(r[1]).lower()))
    _sheet(
        "Via groups",
        ["Group", "Disabled member", "User principal name", "Grants via this path", "Privileged grants"],
        group_rows,
    )

    # 5. Owns credentials — the only tier that is exploitable today.
    owner_rows = [
        [
            i.get("displayName", ""), i.get("userPrincipalName", ""), o.get("name", ""),
            o.get("appId", ""),
            o.get("lastSignIn") or ("not measured" if not o.get("lastSignInKnown") else "not seen in window"),
            i.get("privilegedGrants", 0),
        ]
        for i in identities
        for o in (i.get("ownedDetail") or [])
    ]
    _sheet(
        "Owns credentials",
        ["Disabled owner", "User principal name", "Service principal", "App id",
         "App last sign-in", "Privileged grants"],
        owner_rows,
    )
    # 6. Resources — one row per (person, scope), keeping the ARM structure. The Identities
    #    sheet joins scopes into one cell, which is unusable the moment somebody holds access
    #    on forty of them; this is the sheet you filter and pivot.
    resource_rows = [
        [
            i.get("displayName", ""), i.get("userPrincipalName", ""),
            r.get("scopeType", ""), r.get("subscriptionName", ""), r.get("resourceGroup", ""),
            r.get("resourceType", ""), r.get("resourceName", "") or r.get("scopeDisplayName", ""),
            "; ".join(r.get("roles") or []), r.get("grants", 0), r.get("privileged", 0),
            "direct" if r.get("direct") else "; ".join(r.get("viaGroups") or []),
            r.get("scope", ""),
        ]
        for i in identities
        for r in (i.get("resources") or [])
    ]
    _sheet(
        "Resources",
        ["Person", "User principal name", "Scope type", "Subscription", "Resource group",
         "Resource type", "Resource", "Roles", "Grants", "Privileged", "Held via", "Scope id"],
        resource_rows,
    )

    # 7. Not measured — MANDATORY, always written.
    limits: list[list[Any]] = []
    if not measured:
        limits.append(["Account state", report.get("reason", "Not collected.")])
    for text in report.get("limitations") or []:
        limits.append(["Limitation", text])
    if not limits:
        limits.append(["", "Nothing was withheld: every principal holding access was checked."])
    wb.section("Overview", SECTION_COLOURS["Overview"])
    _sheet("Not measured", ["Scope", "What this report cannot tell you"], limits)

    wb.index_sheet(position=2, colour=SECTION_COLOURS["Overview"])
    return wb.to_bytes()

