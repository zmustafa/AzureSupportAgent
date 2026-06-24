"""Ownership owners — sheet (CSV / Excel) export and import parsing.

* EXPORT — render the owner directory as CSV or a styled XLSX for download.
* IMPORT PARSE — read an uploaded CSV or XLSX into ``(columns, rows)`` regardless of layout,
  so the AI column-mapper (:mod:`app.ownership.importer`) can infer which column means what.

Pure / offline; no Azure. Excel reading uses openpyxl (already a dependency via the RBAC +
identity workbook exporters)."""
from __future__ import annotations

import csv
import io
from typing import Any

# The canonical owner columns we export (and the import target fields mirror these).
EXPORT_COLUMNS = ["display_name", "email", "department", "kind", "source", "notes", "tags", "assignment_count"]


def _csv_safe(value: Any) -> str:
    """Neutralize ``= + - @`` spreadsheet formula-injection vectors."""
    s = "" if value is None else str(value)
    stripped = s.lstrip("\t\r\n ")
    if stripped and stripped[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def owners_to_rows(owners: list[dict[str, Any]], assignment_counts: dict[str, int] | None = None) -> list[dict[str, Any]]:
    counts = assignment_counts or {}
    rows: list[dict[str, Any]] = []
    for o in owners:
        rows.append({
            "display_name": o.get("display_name", ""),
            "email": o.get("email", ""),
            "department": o.get("department", "") or (o.get("link", {}) or {}).get("department", ""),
            "kind": o.get("kind", "person"),
            "source": o.get("source", "manual"),
            "notes": o.get("notes", ""),
            "tags": ", ".join(o.get("tags", []) or []),
            "assignment_count": counts.get(o.get("id", ""), 0),
        })
    return rows


def owners_to_csv(owners: list[dict[str, Any]], assignment_counts: dict[str, int] | None = None) -> str:
    rows = owners_to_rows(owners, assignment_counts)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({c: _csv_safe(r.get(c, "")) for c in EXPORT_COLUMNS})
    return buf.getvalue()


def owners_to_xlsx(owners: list[dict[str, Any]], assignment_counts: dict[str, int] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = owners_to_rows(owners, assignment_counts)
    wb = Workbook()
    ws = wb.active
    ws.title = "Owners"
    ws.append([c.replace("_", " ").title() for c in EXPORT_COLUMNS])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F6CBD")
    for ci in range(1, len(EXPORT_COLUMNS) + 1):
        ws.cell(row=1, column=ci).font = header_font
        ws.cell(row=1, column=ci).fill = header_fill
    widths = [len(c) for c in EXPORT_COLUMNS]
    for r in rows:
        vals = [_csv_safe(r.get(c, "")) for c in EXPORT_COLUMNS]
        ws.append(vals)
        for i, v in enumerate(vals):
            widths[i] = max(widths[i], min(60, len(str(v))))
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(10, w + 2)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{len(rows) + 1}"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def blank_template_csv() -> str:
    """A blank import template with the recommended columns + one example row."""
    headers = ["name", "email", "department", "kind", "workload", "subscription", "resource_ids", "role", "notes"]
    example = ["Jane Smith", "jane@contoso.com", "Platform", "person", "Payments API", "", "", "technical", "Primary on-call"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerow(example)
    return buf.getvalue()


# ------------------------------------------------------------------- import parsing
def list_sheet_names(filename: str, content: bytes) -> list[str]:
    """Return the sheet/tab names of an uploaded workbook (CSV → single synthetic sheet).

    Cheap: opens the workbook just to read the names, without iterating rows. Used so the UI can
    let the user pick a sheet BEFORE the (more expensive) parse + AI column-mapping runs."""
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or name.endswith(".xlsm") or (
        not name.endswith((".csv", ".tsv", ".txt")) and content[:2] == b"PK"
    )
    if not is_xlsx:
        return ["csv"]
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Could not read the Excel file: {exc}") from exc
    names = list(wb.sheetnames)
    wb.close()
    return names or ["Sheet1"]


def parse_sheet(filename: str, content: bytes, sheet: str | None = None) -> dict[str, Any]:
    """Parse an uploaded CSV or XLSX into ``{columns, rows, sheet, row_count, sheet_names}``.

    Rows are dicts keyed by the header column names (first row). Values are strings. For an XLSX,
    ``sheet`` selects a named tab (default: the active/first sheet). Raises ValueError on an
    empty/garbage file or an unknown sheet name."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content, sheet)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _parse_csv(content, tab=name.endswith(".tsv"))
    # Sniff: try xlsx (zip magic 'PK'), else csv.
    if content[:2] == b"PK":
        return _parse_xlsx(content, sheet)
    return _parse_csv(content)


def _decode(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _parse_csv(content: bytes, *, tab: bool = False) -> dict[str, Any]:
    text = _decode(content)
    if not text.strip():
        raise ValueError("The file is empty.")
    # Sniff the delimiter (comma / tab / semicolon) from the first non-empty line.
    sample = "\n".join(text.splitlines()[:5])
    delim = "\t" if tab else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delim = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not all_rows:
        raise ValueError("No rows found in the file.")
    headers = [_clean_header(h, i) for i, h in enumerate(all_rows[0])]
    rows: list[dict[str, Any]] = []
    for raw in all_rows[1:]:
        row = {headers[i]: (raw[i].strip() if i < len(raw) else "") for i in range(len(headers))}
        if any(v for v in row.values()):
            rows.append(row)
    return {"columns": headers, "rows": rows, "sheet": "csv", "row_count": len(rows),
            "sheet_names": ["csv"]}


def _parse_xlsx(content: bytes, sheet: str | None = None) -> dict[str, Any]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Could not read the Excel file: {exc}") from exc
    sheet_names = list(wb.sheetnames)
    if sheet:
        if sheet not in sheet_names:
            raise ValueError(f"Sheet '{sheet}' not found in the workbook.")
        ws = wb[sheet]
    else:
        ws = wb.active
    if ws is None:
        raise ValueError("The workbook has no sheets.")
    it = ws.iter_rows(values_only=True)
    header_row: tuple[Any, ...] | None = None
    for r in it:
        if r and any((str(c).strip() if c is not None else "") for c in r):
            header_row = r
            break
    if header_row is None:
        raise ValueError("No header row found in the sheet.")
    headers = [_clean_header("" if c is None else str(c), i) for i, c in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    for r in it:
        if not r or not any(c is not None and str(c).strip() for c in r):
            continue
        row = {headers[i]: ("" if (i >= len(r) or r[i] is None) else str(r[i]).strip()) for i in range(len(headers))}
        rows.append(row)
    return {"columns": headers, "rows": rows, "sheet": ws.title, "row_count": len(rows),
            "sheet_names": sheet_names}


def _clean_header(h: str, idx: int) -> str:
    h = (h or "").strip()
    return h if h else f"column_{idx + 1}"
