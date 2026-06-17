"""CSV / JSON / multi-sheet XLSX export of normalized access rows (46-column parity with the
scanner output)."""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from app.rbac import schema


# Excel / LibreOffice interpret a cell that starts with one of these characters as a
# formula — `=cmd|'...'!A1` is the classic vector for CSV-injection RCE on the
# admin's workstation. Prefixing with a single quote forces literal-text interpretation.
_FORMULA_TRIGGERS = ("=", "+", "-", "@")
# Tab / CR / LF can also kick off formula interpretation in some spreadsheet apps.
_FORMULA_LEADING_WS = ("\t", "\r", "\n")


def _csv_safe(value: Any) -> Any:
    """Neutralize CSV / Excel formula-injection vectors in a single cell value.

    Strings that begin with ``= + - @`` (or with leading whitespace followed by
    one of those) are prefixed with a leading apostrophe so the spreadsheet
    treats them as plain text. Non-string values pass through unchanged.
    """
    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip("\t\r\n ")
    if stripped and stripped[0] in _FORMULA_TRIGGERS:
        return "'" + value
    if value[0] in _FORMULA_LEADING_WS and stripped and stripped[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def to_csv(rows: list[dict[str, Any]]) -> str:
    """Serialize rows to CSV with the canonical 46-column header order."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(schema.COLUMNS), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: _csv_safe(row.get(c, "")) for c in schema.COLUMNS})
    return buf.getvalue()


def to_json(rows: list[dict[str, Any]]) -> str:
    """Serialize rows to a pretty JSON array (each row already a 46-key dict)."""
    return json.dumps(rows, indent=2, default=str)


# --------------------------------------------------------------------------- XLSX workbook
# A friendlier subset/order of the 46 columns for the human-readable access sheets (the full
# 46-column fidelity stays available via the CSV/JSON export).
_ACCESS_HEADERS: tuple[str, ...] = (
    "effectivePrincipalName",
    "effectivePrincipalUserPrincipalName",
    "effectivePrincipalType",
    "roleName",
    "roleIsPrivileged",
    "roleHasDataActions",
    "surface",
    "accessPath",
    "sourceGroupName",
    "scopeDisplayName",
    "scope",
    "scopeType",
    "subscriptionName",
    "subscriptionId",
    "resourceGroup",
    "assignmentState",
    "principalDisplayName",
    "principalId",
    "condition",
    "collector",
)


def _safe_sheet_title(title: str) -> str:
    """Excel sheet titles: ≤31 chars, none of ``[]:*?/\\``."""
    for ch in "[]:*?/\\":
        title = title.replace(ch, " ")
    return title.strip()[:31] or "Sheet"


def _coerce(value: Any) -> Any:
    if isinstance(value, bool):
        return "Yes" if value else ""
    if value is None:
        return ""
    if isinstance(value, (int, float, str)):
        # Apply the same formula-injection neutralization as the CSV path. Excel
        # interprets `=...` / `+...` / `-...` / `@...` cells as formulas even in
        # an .xlsx workbook.
        return _csv_safe(value)
    return _csv_safe(str(value))


def to_workbook(
    *,
    rows: list[dict[str, Any]],
    overview: dict[str, Any],
    pivots: dict[str, list[dict[str, Any]]],
    pivot_labels: dict[str, str],
    directory: dict[str, Any],
) -> bytes:
    """Build a comprehensive multi-sheet ``.xlsx`` workbook dumping every RBAC view:
    a Summary, the effective-access grid (+ privileged / group-derived / owner / Entra / Key
    Vault lenses), per-scope freshness, role definitions, the principal directory, the Insights
    pivots and the collector diagnostics. ``rows`` is the (optionally filtered) master row set."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F6CBD")

    def _sheet(title: str, headers: list[str], data: list[list[Any]]) -> None:
        ws = wb.create_sheet(_safe_sheet_title(title))
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for r in data:
            ws.append([_coerce(v) for v in r])
        # Freeze the header + size columns to their content (bounded).
        ws.freeze_panes = "A2"
        for ci, h in enumerate(headers, start=1):
            width = len(str(h))
            for r in data[:200]:
                if ci - 1 < len(r):
                    width = max(width, len(str(_coerce(r[ci - 1]))))
            ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, width + 2))
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    def _rows_for(predicate) -> list[list[Any]]:
        return [[r.get(h, "") for h in _ACCESS_HEADERS] for r in rows if predicate(r)]

    # 1. Summary — KPIs + generation metadata.
    kpis = overview.get("kpis", {})
    summary: list[list[Any]] = [["Metric", "Value"]]
    label_map = [
        ("Total grants", "total_assignments"),
        ("Unique principals", "unique_principals"),
        ("Privileged", "privileged"),
        ("Data-plane", "data_plane"),
        ("Group-derived", "group_derived"),
        ("Service-principal owners", "owners"),
        ("Entra directory roles", "entra_roles"),
        ("PIM eligible", "eligible"),
        ("Scopes", "scopes"),
        ("Subscriptions", "subscriptions"),
    ]
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["RBAC — Access Review export"])
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws0.append(["Generated", overview.get("generated_at", "")])
    ws0.append(["Tenant", overview.get("tenant_id", "")])
    ws0.append(["Demo dataset", "Yes" if overview.get("demo") else "No"])
    ws0.append([])
    ws0.append(["Metric", "Value"])
    hdr_row = ws0.max_row
    for c in range(1, 3):
        ws0.cell(row=hdr_row, column=c).font = header_font
        ws0.cell(row=hdr_row, column=c).fill = header_fill
    for label, key in label_map:
        ws0.append([label, kpis.get(key, 0)])
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 40

    # 2–7. Access lenses.
    _sheet("Effective Access", list(_ACCESS_HEADERS), _rows_for(lambda r: True))
    _sheet("Privileged", list(_ACCESS_HEADERS), _rows_for(lambda r: bool(r.get("roleIsPrivileged"))))
    _sheet("Group-Derived", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("accessPath") == schema.PATH_GROUP))
    _sheet("SP Owners", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("accessPath") == schema.PATH_OWNER))
    _sheet("Entra Roles", list(_ACCESS_HEADERS), _rows_for(lambda r: r.get("surface") == schema.SURFACE_ENTRA))
    kv_rows = _rows_for(lambda r: r.get("surface") == schema.SURFACE_KEY_VAULT)
    if kv_rows:
        _sheet("Key Vault", list(_ACCESS_HEADERS), kv_rows)

    # 8. Scopes freshness.
    scope_headers = ["displayName", "scopeType", "status", "row_count", "collectors_attention", "generated_at", "demo"]
    _sheet(
        "Scopes",
        ["Scope", "Type", "Status", "Grants", "Attention", "Generated", "Demo"],
        [[s.get(h, "") for h in scope_headers] for s in overview.get("scopes", [])],
    )

    # 9. Role definitions (directory reference).
    rd = directory.get("role_defs", []) or []
    if rd:
        rd_headers = ["roleName", "roleCategory", "roleIsPrivileged", "roleHasDataActions", "actionsCount", "dataActionsCount", "description"]
        _sheet("Role Definitions", ["Role", "Category", "Privileged", "Data actions", "Actions", "Data actions #", "Description"],
               [[r.get(h, "") for h in rd_headers] for r in rd])

    # 10. Principal directory (the resolved GUID → name map).
    pr = directory.get("principals", []) or []
    if pr:
        pr_headers = ["displayName", "principalType", "userPrincipalName", "appId", "principalId", "source"]
        _sheet("Principals", ["Name", "Type", "UPN", "App ID", "Object ID", "Source"],
               [[p.get(h, "") for h in pr_headers] for p in pr])

    # 11. Insights — every pivot flattened.
    insight_data: list[list[Any]] = []
    for key, items in pivots.items():
        title = pivot_labels.get(key, key)
        for it in items:
            insight_data.append([title, it.get("label", ""), it.get("count", 0)])
    _sheet("Insights", ["Pivot", "Label", "Count"], insight_data)

    # 12. Diagnostics — collector statuses + any errors.
    diag_headers = ["collector", "scopeLabel", "status", "rowsAdded", "message"]
    _sheet(
        "Diagnostics",
        ["Collector", "Scope", "Status", "Rows", "Message"],
        [[c.get(h, "") for h in diag_headers] for c in overview.get("collectors", [])],
    )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

