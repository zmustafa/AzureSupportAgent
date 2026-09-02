"""Shared .xlsx workbook builder.

Extracted from the IAM export when the Entra one landed, so both produce the same artefact:
same header style, same frozen top row, same auto-filter, same formula-injection neutralisation
and — most importantly — the same way of saying "this was not measured".

The formula-injection guard is the reason this is not just a thin wrapper around openpyxl.
Excel and LibreOffice interpret a cell beginning ``= + - @`` as a formula, and
``=cmd|'/c calc'!A1`` in a display name is a working remote-code-execution vector against the
reviewer's workstation. Every cell that reaches a sheet goes through :func:`cell_safe`.

The second reason is *types*. A timestamp written as a string sorts lexically and gets Excel's
text filter, so "everything that expired last month" is not a question the file can answer.
Columns named in ``dates=`` are converted to real datetimes; see :func:`as_datetime`.
"""
from __future__ import annotations

import datetime as dt
import re
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Any

_FORMULA_TRIGGERS = ("=", "+", "-", "@")
_FORMULA_LEADING_WS = ("\t", "\r", "\n")

#: Excel's hard limit on a sheet title.
MAX_TITLE = 31

#: Sub-second precision is noise in a review, and Excel's serial format cannot hold the
#: 7-digit fraction Graph emits anyway.
DATETIME_FMT = "yyyy-mm-dd hh:mm:ss"
INTEGER_FMT = "#,##0"
#: Widths. A clamped column is one whose content did not fit, and that is what earns wrapping.
MAX_WIDTH = 60
MIN_WIDTH = 10


@dataclass(frozen=True)
class HyperlinkCell:
    """Displayed text plus a validated HTTPS target for a real OOXML hyperlink."""

    text: str
    url: str


def hyperlink(text: Any, url: Any) -> HyperlinkCell | str:
    """Create a safe hyperlink value, falling back to plain text for an invalid target."""
    label = str(text or "")
    target = str(url or "")
    if not target.lower().startswith("https://") or any(ch in target for ch in ("\r", "\n", "\t")):
        return label
    return HyperlinkCell(label, target)


def cell_safe(value: Any) -> Any:
    """Neutralise CSV / Excel formula-injection vectors in a single cell value.

    Strings beginning ``= + - @`` (or leading whitespace then one of those) are prefixed with an
    apostrophe so the spreadsheet treats them as literal text. Everything else passes through.
    """
    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip("\t\r\n ")
    if stripped and stripped[0] in _FORMULA_TRIGGERS:
        return "'" + value
    if value[0] in _FORMULA_LEADING_WS and stripped and stripped[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def as_datetime(value: Any) -> dt.datetime | None:
    """A timestamp as a NAIVE UTC datetime, or ``None`` when the value is not one.

    Returning ``None`` instead of raising is deliberate: an unparseable value is left in the
    sheet as its original text rather than disappearing, and a column that quietly emptied
    itself is the failure this whole change exists to remove.

    Graph emits three shapes — ``...Z``, ``...+00:00`` and a 7-digit fractional second from
    .NET. ``fromisoformat`` accepts all three on the supported interpreter, so there is no
    hand-rolled parser here to drift.

    Stripping the offset is not cosmetic. openpyxl refuses a timezone-aware datetime outright
    ("Excel does not support timezones in datetimes"), so the conversion to UTC has to happen
    here — and because the offset is then unrecoverable, the column has to say UTC in its
    header. :meth:`WorkbookBuilder.sheet` does that renaming itself so the two can never
    become separated.
    """
    if isinstance(value, dt.datetime):
        parsed = value
    elif isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    elif isinstance(value, str) and value.strip():
        try:
            parsed = dt.datetime.fromisoformat(value.strip())
        except ValueError:
            return None
    else:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(dt.timezone.utc).replace(tzinfo=None)
    return parsed.replace(microsecond=0)


def coerce(value: Any) -> Any:
    """A python value as a spreadsheet cell.

    Booleans are tri-state: ``True``/``False`` become "Yes"/"No" and only a *missing* value is
    blank. An earlier version rendered ``False`` as blank to keep the grid quiet, which made
    "no", "not applicable" and "we never looked" indistinguishable — on a review sheet those
    are three different answers and only one of them is reassuring.
    """
    if isinstance(value, HyperlinkCell):
        return cell_safe(value.text)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if value is None:
        return ""
    # Before the str() fallback, which would turn a real date back into text.
    if isinstance(value, dt.datetime):
        return value.astimezone(dt.timezone.utc).replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float, str)):
        return cell_safe(value)
    if isinstance(value, (list, tuple)):
        return cell_safe(", ".join(str(v) for v in value))
    if isinstance(value, dict):
        return cell_safe("; ".join(f"{k}={v}" for k, v in value.items()))
    return cell_safe(str(value))


def safe_title(title: str, used: set[str] | None = None) -> str:
    """A legal, UNIQUE sheet title: ≤31 chars, none of ``[]:*?/\\``.

    Uniqueness matters as much as legality. openpyxl silently renames a duplicate to
    "Name1", so two sheets truncated to the same 31 characters would produce a workbook where
    one of them is not the sheet its name claims.
    """
    for ch in "[]:*?/\\":
        title = title.replace(ch, " ")
    # Collapse the gap the substitution just opened: "Redirect / reply URLs" became
    # "Redirect   reply URLs", which reads as a typo rather than as a sanitised character.
    title = re.sub(r"\s{2,}", " ", title)
    base = title.strip()[:MAX_TITLE] or "Sheet"
    if used is None:
        return base
    candidate, n = base, 2
    while candidate.lower() in used:
        suffix = f" {n}"
        candidate = base[: MAX_TITLE - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def argb(colour: str) -> str:
    """A color as OPAQUE 8-digit ARGB.

    The alpha channel is the whole point. OOXML stores colors as aRGB, and openpyxl pads a
    6-digit value by prepending ``00`` — alpha zero, fully transparent — so Excel dutifully
    renders nothing at all. Every tab color looked correct in the file and in an openpyxl
    round-trip, and was invisible when opened.
    """
    value = (colour or "").lstrip("#").upper()
    if not value:
        return ""
    if len(value) == 6:
        return "FF" + value
    return value


def _unique_headers(headers: list[str]) -> list[str]:
    """Header labels made unique and non-empty.

    Excel repairs — and silently discards — a table whose header row has a duplicate or a
    blank. Nothing in this repository writes one deliberately, but the failure is invisible
    until somebody opens the file, so it is cheaper to guarantee than to audit."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, h in enumerate(headers):
        label = str(h).strip() or f"Column {i + 1}"
        n = seen.get(label.lower(), 0) + 1
        seen[label.lower()] = n
        out.append(label if n == 1 else f"{label} ({n})")
    return out


#: Conditional-format vocabularies. Applied by NAME from the call site rather than sniffed
#: from the header, because a colour that means "critical" appearing on the wrong column is a
#: false statement, not a cosmetic slip.
_SEVERITY_COLOURS = {
    "critical": "FFC7CE", "high": "FFD8B0", "medium": "FFEB9C", "low": "E2EFDA",
    "uncovered": "FFC7CE", "partial": "FFEB9C", "covered": "C6EFCE",
    "atrisk": "FFC7CE", "dismissed": "E7E6E6", "remediated": "C6EFCE",
    "tier0": "FFC7CE", "tier1": "FFD8B0", "tier2": "FFEB9C",
    "yes": "FFC7CE", "no": "C6EFCE",
}


def _apply_highlight(ws: Any, letter: str, last_row: int, rule: str) -> None:
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    from openpyxl.styles import Font, PatternFill

    span = f"{letter}2:{letter}{last_row}"
    if rule == "severity":
        for word, colour in _SEVERITY_COLOURS.items():
            ws.conditional_formatting.add(span, CellIsRule(
                operator="equal", formula=[f'"{word}"'],
                fill=PatternFill("solid", bgColor=colour)))
    elif rule == "days_left":
        # Negative is already-expired and still attached to the object; 0-30 is the window in
        # which somebody still has time to act.
        ws.conditional_formatting.add(span, CellIsRule(
            operator="lessThan", formula=["0"],
            fill=PatternFill("solid", bgColor="FFC7CE"), font=Font(color="9C0006")))
        ws.conditional_formatting.add(span, CellIsRule(
            operator="between", formula=["0", "30"],
            fill=PatternFill("solid", bgColor="FFEB9C")))
    elif rule == "expiry_date":
        ws.conditional_formatting.add(span, CellIsRule(
            operator="lessThan", formula=["TODAY()"],
            fill=PatternFill("solid", bgColor="FFC7CE"), font=Font(color="9C0006")))
        ws.conditional_formatting.add(span, CellIsRule(
            operator="between", formula=["TODAY()", "TODAY()+30"],
            fill=PatternFill("solid", bgColor="FFEB9C")))
    elif rule == "bar":
        ws.conditional_formatting.add(span, DataBarRule(
            start_type="num", start_value=0, end_type="max", color="638EC6"))
    elif rule == "stale_date":
        # The mirror of expiry: here a date going far into the PAST is the problem.
        ws.conditional_formatting.add(span, CellIsRule(
            operator="lessThan", formula=["TODAY()-90"],
            fill=PatternFill("solid", bgColor="FFEB9C")))


class WorkbookBuilder:
    """Accumulates sheets, then serialises once.

    Deliberately tiny. The value is not the abstraction, it is that `sheet` and `blind_sheet`
    exist side by side so writing "we could not look" is exactly as easy as writing rows —
    the moment it is harder, somebody writes an empty sheet instead and it reads as a pass.
    """

    def __init__(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        self.wb = Workbook()
        # openpyxl always starts with one sheet; the caller fills it via `first_sheet`.
        self._default = self.wb.active
        self._default_claimed = False
        self._used: set[str] = set()
        #: Table display names are workbook-scoped, not sheet-scoped; a collision is a repair.
        self._tables: set[str] = set()
        self._header_font = Font(bold=True, color="FFFFFF")
        self._header_fill = PatternFill("solid", fgColor="0F6CBD")
        #: The section subsequent sheets belong to: (label, tab color). Set with `section`.
        self._section: tuple[str, str] = ("", "")
        #: (title, section, rows, note) for every sheet written, so an index can be built after.
        self.manifest: list[tuple[str, str, int, str]] = []

    def section(self, label: str, colour: str = "") -> None:
        """Group the sheets that follow, and give their tabs a shared color.

        Carried on the builder rather than passed to every `sheet` call: a fifty-sheet workbook
        has fifty chances to pass the wrong color, and one sheet tinted like the wrong parent
        tab is worse than no color at all — it is a wrong label."""
        self._section = (label, argb(colour))

    # ---------------------------------------------------------------- sheets
    def sheet(
        self,
        title: str,
        headers: list[str],
        data: list[list[Any]],
        *,
        note: str = "",
        dates: Iterable[str] = (),
        highlight: dict[str, str] | None = None,
    ) -> Any:
        """One sheet. ``dates`` and ``highlight`` are keyed on the labels in ``headers``.

        A column named in ``dates`` is written as a real datetime and its header gains a
        "(UTC)" suffix — the suffix is added here rather than at the call site so a converted
        column cannot end up unlabelled."""
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        ws = self.wb.create_sheet(safe_title(title, self._used))
        label, colour = self._section
        if colour:
            ws.sheet_properties.tabColor = colour

        date_idx = {i for i, h in enumerate(headers) if h in set(dates)}
        rules = {i: (highlight or {}).get(h, "") for i, h in enumerate(headers)}
        shown = _unique_headers([f"{h} (UTC)" if i in date_idx else h
                                 for i, h in enumerate(headers)])
        ncols = len(shown)

        ws.append(shown)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        widths = [len(h) for h in shown]
        # A column is formatted as a number only when EVERY value in it is one; one stray
        # "n/a" and a thousands separator on the rest would be a lie about the column's type.
        numeric = [True] * ncols
        populated = [False] * ncols
        # Counted, never read back from the sheet: openpyxl's `max_row` is `max(self._cells)`,
        # a full scan of every cell written so far, so asking it once per row made this method
        # O(rows^2 x cols). A 16,000-row sheet took ~35 s and a whole workbook outlived the
        # 240 s Azure Container Apps request timeout; counting makes the same sheet 0.6 s.
        for offset, r in enumerate(data):
            row_number = offset + 2
            for ci in range(ncols):
                raw = r[ci] if ci < len(r) else None
                if ci in date_idx:
                    stamp = as_datetime(raw)
                    if stamp is not None:
                        cell = ws.cell(row=row_number, column=ci + 1, value=stamp)
                        cell.number_format = DATETIME_FMT
                        widths[ci] = max(widths[ci], len(DATETIME_FMT))
                        populated[ci] = True
                        numeric[ci] = False
                        continue
                value = coerce(raw)
                cell = ws.cell(row=row_number, column=ci + 1, value=value)
                if isinstance(raw, HyperlinkCell) and raw.url:
                    cell.hyperlink = raw.url
                    cell.style = "Hyperlink"
                if isinstance(value, (dt.datetime, dt.date)):
                    # A caller that handed us a datetime directly (a key/value sheet, where
                    # `dates=` cannot address a column) still gets a formatted date.
                    cell.number_format = DATETIME_FMT
                    widths[ci] = max(widths[ci], len(DATETIME_FMT))
                    populated[ci] = True
                    numeric[ci] = False
                elif value != "":
                    populated[ci] = True
                    # bool is a subclass of int, and it has already become "Yes"/"No".
                    if isinstance(raw, bool) or not isinstance(raw, (int, float)):
                        numeric[ci] = False
                    widths[ci] = max(widths[ci], len(str(value)))

        ws.freeze_panes = "A2"
        for ci in range(ncols):
            letter = get_column_letter(ci + 1)
            width = min(MAX_WIDTH, max(MIN_WIDTH, widths[ci] + 2))
            ws.column_dimensions[letter].width = width
            if numeric[ci] and populated[ci] and ci not in date_idx:
                for row_number in range(2, len(data) + 2):
                    cell = ws.cell(row=row_number, column=ci + 1)
                    if isinstance(cell.value, int):
                        cell.number_format = INTEGER_FMT
            # Clamped means the content did not fit. Without wrapping it is simply cut off,
            # which on the narrative columns is where the reason for a finding lives.
            if widths[ci] + 2 > MAX_WIDTH and data:
                for row_number in range(2, len(data) + 2):
                    ws.cell(row=row_number, column=ci + 1).alignment = Alignment(
                        vertical="top", wrap_text=True)
            if data and rules[ci]:
                _apply_highlight(ws, letter, len(data) + 1, rules[ci])

        if data:
            self._listobject(ws, ncols, len(data))
        else:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
        self.manifest.append((ws.title, label, len(data), note))
        return ws

    def _listobject(self, ws: Any, ncols: int, nrows: int) -> None:
        """Register the range as a real Excel table.

        A table carries its own filter, banded rows, structured references and — the reason it
        is worth the trouble — it grows when rows are added, so a reviewer's own working
        columns stay attached to the right record."""
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title).strip("_") or "t"
        name = f"tbl_{name}"[:250]
        candidate, n = name, 2
        while candidate.lower() in self._tables:
            candidate = f"{name}_{n}"
            n += 1
        self._tables.add(candidate.lower())
        table = Table(displayName=candidate,
                      ref=f"A1:{get_column_letter(ncols)}{nrows + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    def blind_sheet(self, title: str, reason: str) -> Any:
        """A sheet for something that could NOT be measured.

        Never an empty grid. A sheet called "Risky users" with no rows reads as "no risky
        users"; that is a different statement from "we were not permitted to look", and only
        one of them is true."""
        ws = self.sheet(title, ["Status", "Why"], [["NOT MEASURED", reason]], note=reason)
        return ws

    def first_sheet(self, title: str) -> Any:
        """Take over the workbook's default sheet for the front matter."""
        self._default.title = safe_title(title, self._used)
        self._default_claimed = True
        _label, colour = self._section
        if colour:
            self._default.sheet_properties.tabColor = colour
        return self._default

    def _drop_unused_default(self) -> None:
        """Remove the sheet openpyxl creates for us when nobody claimed it.

        A caller that builds its front matter with `sheet()` instead of `first_sheet()` would
        otherwise ship an empty tab called "Sheet", and the index's position arithmetic would
        count it."""
        if self._default_claimed or self._default is None:
            return
        self.wb.remove(self._default)
        self._default = None
        self._default_claimed = True

    def index_sheet(self, title: str = "Index", *, position: int = 1, colour: str = "") -> None:
        """A contents page. Written last, moved to the front.

        A fifty-sheet workbook without one is a filing cabinet with no labels. `Section` names
        the parent screen each sheet came from — the same grouping the tab colors show, spelled
        out, because color alone is not readable to everyone.

        The entries link. There is deliberately no link back on each sheet: it would have to
        live in the header row, where it becomes a stray column for every reader that consumes
        the file programmatically."""
        from openpyxl.styles import Alignment, Font
        from openpyxl.worksheet.hyperlink import Hyperlink

        self._drop_unused_default()
        ws = self.wb.create_sheet(safe_title(title, self._used))
        if colour:
            ws.sheet_properties.tabColor = argb(colour)
        ws.append(["Sheet", "Section", "Rows", "Note"])
        for c in range(1, 5):
            cell = ws.cell(row=1, column=c)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(vertical="center")
        link_font = Font(color="0563C1", underline="single")
        for offset, (name, section, count, note) in enumerate(self.manifest):
            row_number = offset + 2
            ws.append([name, section, count, coerce(note)])
            cell = ws.cell(row=row_number, column=1)
            # `location`, not `target`: a target would be written as an EXTERNAL relationship
            # and Excel would try to open it as a file. An apostrophe in a sheet name has to be
            # doubled or the quoted reference terminates early.
            quoted = str(name).replace("'", "''")
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{quoted}'!A1")
            cell.font = link_font
            ws.cell(row=row_number, column=4).alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 90
        self.wb.move_sheet(ws, offset=-(len(self.wb.sheetnames) - position))

    # ---------------------------------------------------------------- output
    def to_bytes(self) -> bytes:
        import io

        self._drop_unused_default()
        bio = io.BytesIO()
        self.wb.save(bio)
        return bio.getvalue()
