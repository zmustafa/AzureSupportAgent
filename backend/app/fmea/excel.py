"""Rich Excel (.xlsx) export for an FMEA document.

Renders the FMEA worksheet the way a reliability engineer expects to see it in Excel:
a Summary sheet with the risk roll-up, then ONE sheet per table laid out exactly like the
on-screen grid — grouped "Current Controls" and "FMEA Results" header bands, a green→amber→
red colour scale on the Severity/Occurrence/Detection factor columns, risk-band conditional
formatting on the RPN columns, and **live RPN formulas** (=Severity×Occurrence×Detection) so
the spreadsheet recalculates as a user edits the factors — matching the classic FMEA template
note "RPN columns contain a formula to auto-calculate".

Uses openpyxl (the repo's standard xlsx writer; see rbac/export.py, ownership/sheet.py).
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from typing import Any

from app.fmea import compute

# ⟦TODO: … | key=…⟧ placeholder tokens are human-only hints — never export the raw token.
_TODO_RE = re.compile(r"⟦\s*TODO[\s\S]*?⟧", re.IGNORECASE)

# Column layout (A..U), mirroring the grid. ``kind`` drives styling/formulas.
#   text   — free text (wrapped)
#   factor — a 1-10 score (colour-scaled)
#   rpn    — derived; written as a live Excel formula + band conditional formatting
#   date   — ISO date (real Excel date when parseable)
#   id     — the display row number
_COLUMNS: list[dict[str, Any]] = [
    {"key": "__id", "header": "ID", "group": "", "width": 5, "kind": "id"},
    {"key": "item", "header": "System / Item / Process Step", "group": "", "width": 26, "kind": "text"},
    {"key": "function", "header": "Function", "group": "", "width": 18, "kind": "text"},
    {"key": "failure_mode", "header": "Potential Failure Mode", "group": "", "width": 26, "kind": "text"},
    {"key": "effects", "header": "Effects of Failure", "group": "", "width": 26, "kind": "text"},
    {"key": "severity", "header": "Severity", "group": "", "width": 9, "kind": "factor"},
    {"key": "causes", "header": "Causes", "group": "", "width": 24, "kind": "text"},
    {"key": "occurrence", "header": "Occurrence", "group": "", "width": 9, "kind": "factor"},
    {"key": "control_prevention", "header": "Prevention", "group": "Current Controls", "width": 22, "kind": "text"},
    {"key": "control_detection", "header": "Detection", "group": "Current Controls", "width": 22, "kind": "text"},
    {"key": "detection", "header": "Detection", "group": "", "width": 9, "kind": "factor"},
    {"key": "rpn", "header": "RPN", "group": "", "width": 8, "kind": "rpn"},
    {"key": "recommended_actions", "header": "Recommended Actions", "group": "", "width": 30, "kind": "text"},
    {"key": "owner", "header": "Owner", "group": "", "width": 18, "kind": "text"},
    {"key": "date_due", "header": "Date Due", "group": "", "width": 14, "kind": "date"},
    {"key": "action_results", "header": "Action Results", "group": "FMEA Results", "width": 24, "kind": "text"},
    {"key": "date_completed", "header": "Date Completed", "group": "FMEA Results", "width": 14, "kind": "date"},
    {"key": "severity_post", "header": "Severity", "group": "FMEA Results", "width": 9, "kind": "factor"},
    {"key": "occurrence_post", "header": "Occurrence", "group": "FMEA Results", "width": 9, "kind": "factor"},
    {"key": "detection_post", "header": "Detection", "group": "FMEA Results", "width": 9, "kind": "factor"},
    {"key": "rpn_post", "header": "RPN", "group": "FMEA Results", "width": 8, "kind": "rpn"},
]

# 0-based indices used to build the RPN formulas (=factor*factor*factor).
_IDX = {c["key"]: i for i, c in enumerate(_COLUMNS)}


def _strip_todo(value: Any) -> str:
    if value is None:
        return ""
    return _TODO_RE.sub("", str(value)).strip()


def _as_date(value: Any) -> date | str:
    """Return a real ``date`` (so Excel treats it as a date) or "" if not parseable."""
    s = _strip_todo(value)
    if not s:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        try:
            return date.fromisoformat(s)
        except ValueError:
            return ""
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return ""


def _safe_sheet_title(title: str, used: set[str]) -> str:
    """Excel sheet titles: ≤31 chars, none of ``[]:*?/\\``, and unique within the book."""
    clean = re.sub(r"[\[\]:*?/\\]", " ", title or "Table").strip()[:31] or "Table"
    base = clean
    n = 2
    while clean.lower() in used:
        suffix = f" ({n})"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(clean.lower())
    return clean


def build_fmea_xlsx(doc: dict[str, Any], workload_name: str) -> bytes:
    """Build the styled FMEA workbook and return the raw .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    compute.recompute_doc(doc)
    title = doc.get("title") or (f"FMEA — {workload_name}" if workload_name else "Failure Mode and Effects Analysis")

    # ---- shared styles ----
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14, color="1F3864")
    group_fill = PatternFill("solid", fgColor="2E75B6")   # blue band for grouped headers
    group_font = Font(bold=True, color="FFFFFF")
    results_fill = PatternFill("solid", fgColor="2F8F83")  # teal band for "FMEA Results"
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    head_font = Font(bold=True, color="1F3864")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_mid = Alignment(horizontal="center", vertical="center")

    # Conditional-format helpers (applied per data range, recolour live in Excel).
    def _factor_scale() -> ColorScaleRule:
        return ColorScaleRule(
            start_type="num", start_value=1, start_color="FF63BE7B",   # green
            mid_type="num", mid_value=5, mid_color="FFFFEB84",          # yellow
            end_type="num", end_value=10, end_color="FFF8696B",         # red
        )

    def _rpn_rules() -> list[CellIsRule]:
        # Most-severe first; stopIfTrue so only one band paints a cell.
        return [
            CellIsRule(operator="greaterThanOrEqual", formula=["200"], stopIfTrue=True,
                       fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)),
            CellIsRule(operator="greaterThanOrEqual", formula=["120"], stopIfTrue=True,
                       fill=PatternFill("solid", fgColor="FFD9A0"), font=Font(color="9C5700", bold=True)),
            CellIsRule(operator="greaterThanOrEqual", formula=["40"], stopIfTrue=True,
                       fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(color="9C6500", bold=True)),
            CellIsRule(operator="greaterThanOrEqual", formula=["1"], stopIfTrue=True,
                       fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)),
        ]

    wb = Workbook()

    # ============================================================ Summary sheet
    summary = compute.summarize(doc)
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.append(["Workload", workload_name or "—"])
    ws.append(["Status", str(doc.get("status", "draft")).replace("_", " ").title()])
    ws.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    ws.append(["Risk band", "Count"])
    hdr = ws.max_row
    for c in (1, 2):
        ws.cell(row=hdr, column=c).font = head_font
        ws.cell(row=hdr, column=c).fill = head_fill
    band_colors = {"critical": "FFC7CE", "high": "FFD9A0", "medium": "FFEB9C", "low": "C6EFCE", "none": "F2F2F2"}
    for band in ("critical", "high", "medium", "low", "none"):
        ws.append([band.title(), summary["counts"].get(band, 0)])
        ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=band_colors[band])
    ws.append([])
    ws.append(["Total failure modes", summary["total_rows"]])
    ws.append(["Scored", summary["scored_rows"]])
    ws.append(["Highest RPN", summary["top_rpn"]])
    ws.append(["Mitigated (RPN reduced)", summary["mitigated_rows"]])
    ws.append(["Open actions", summary["open_actions"]])
    ws.append([])
    ws.append(["Table", "Rows", "Top RPN"])
    thdr = ws.max_row
    for c in (1, 2, 3):
        ws.cell(row=thdr, column=c).font = head_font
        ws.cell(row=thdr, column=c).fill = head_fill
    for t in doc.get("tables", []) or []:
        rows = t.get("rows", []) or []
        top = max((compute.rpn(r.get("severity"), r.get("occurrence"), r.get("detection")) or 0) for r in rows) if rows else 0
        ws.append([t.get("name", "Table"), len(rows), top])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    # ============================================================ per-table sheets
    used_titles: set[str] = {"summary"}
    ncols = len(_COLUMNS)
    last_col = get_column_letter(ncols)
    for table in doc.get("tables", []) or []:
        ws = wb.create_sheet(_safe_sheet_title(table.get("name", "Table"), used_titles))

        # Row 1: title (workload · table). Row 2: scope_ref. Then 2 header rows.
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = f"{title}  ·  {table.get('name', 'Table')}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")
        scope = str(table.get("scope_ref") or "").strip()
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = f"Scope: {scope}" if scope else ""
        ws["A2"].font = Font(italic=True, color="6B7280")

        hr1, hr2 = 3, 4
        # Sub-header row (hr2) cells + group bands (hr1).
        for i, col in enumerate(_COLUMNS):
            letter = get_column_letter(i + 1)
            if col["group"]:
                # Group columns: band label on hr1 (merged later), sub-label on hr2.
                cell = ws.cell(row=hr2, column=i + 1, value=col["header"])
            else:
                # Single columns: merge hr1:hr2 and label once.
                ws.merge_cells(f"{letter}{hr1}:{letter}{hr2}")
                cell = ws.cell(row=hr1, column=i + 1, value=col["header"])
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = center
            cell.border = border
            # Make sure the merged partner cell also carries a border/fill.
            ws.cell(row=hr2, column=i + 1).border = border
            if not col["group"]:
                ws.cell(row=hr2, column=i + 1).fill = head_fill

        # Group header bands on hr1 (merge across each contiguous group).
        def _band(group: str, fill: PatternFill) -> None:
            members = [i for i, c in enumerate(_COLUMNS) if c["group"] == group]
            if not members:
                return
            a, b = get_column_letter(members[0] + 1), get_column_letter(members[-1] + 1)
            ws.merge_cells(f"{a}{hr1}:{b}{hr1}")
            cell = ws.cell(row=hr1, column=members[0] + 1, value=group)
            cell.font = group_font
            cell.fill = fill
            cell.alignment = center_mid
            for i in members:
                ws.cell(row=hr1, column=i + 1).border = border

        _band("Current Controls", group_fill)
        _band("FMEA Results", results_fill)

        # ---- data rows ----
        first_data = hr2 + 1
        rows = table.get("rows", []) or []
        for ri, row in enumerate(rows):
            r = first_data + ri
            for i, col in enumerate(_COLUMNS):
                cell = ws.cell(row=r, column=i + 1)
                cell.border = border
                cell.alignment = wrap_top
                kind = col["kind"]
                if kind == "id":
                    cell.value = ri + 1
                    cell.alignment = center_mid
                elif kind == "factor":
                    n = compute.normalize_factor(row.get(col["key"]))
                    cell.value = n if n > 0 else None
                    cell.alignment = center_mid
                elif kind == "rpn":
                    if col["key"] == "rpn":
                        sev, occ, det = "severity", "occurrence", "detection"
                    else:
                        sev, occ, det = "severity_post", "occurrence_post", "detection_post"
                    fc = get_column_letter(_IDX[sev] + 1)
                    oc = get_column_letter(_IDX[occ] + 1)
                    dc = get_column_letter(_IDX[det] + 1)
                    cell.value = f'=IF(OR({fc}{r}=0,{oc}{r}=0,{dc}{r}=0),"",{fc}{r}*{oc}{r}*{dc}{r})'
                    cell.alignment = center_mid
                    cell.font = Font(bold=True)
                elif kind == "date":
                    d = _as_date(row.get(col["key"]))
                    cell.value = d
                    if isinstance(d, date):
                        cell.number_format = "yyyy-mm-dd"
                    cell.alignment = center_mid
                else:  # text
                    cell.value = _strip_todo(row.get(col["key"]))

        # ---- widths, freeze, autofilter ----
        for i, col in enumerate(_COLUMNS):
            ws.column_dimensions[get_column_letter(i + 1)].width = col["width"]
        ws.freeze_panes = ws.cell(row=first_data, column=1)
        last_row = max(first_data - 1, first_data + len(rows) - 1)
        ws.auto_filter.ref = f"A{hr2}:{last_col}{last_row}"

        # ---- conditional formatting (only when there are data rows) ----
        if rows:
            data_lo, data_hi = first_data, first_data + len(rows) - 1
            for i, col in enumerate(_COLUMNS):
                letter = get_column_letter(i + 1)
                rng = f"{letter}{data_lo}:{letter}{data_hi}"
                if col["kind"] == "factor":
                    ws.conditional_formatting.add(rng, _factor_scale())
                elif col["kind"] == "rpn":
                    for rule in _rpn_rules():
                        ws.conditional_formatting.add(rng, rule)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
