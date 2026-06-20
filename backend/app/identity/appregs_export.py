"""Multi-sheet XLSX export for Entra ID app registrations.

Produces a workbook that breaks the app-registrations snapshot into one sheet per concern:
  - Summary        — KPIs + generation metadata
  - Applications   — one row per app (the main inventory grid)
  - Credentials    — one row per secret/certificate (expiry pivot)
  - API Permissions— one row per granted permission (Application/Delegated + risk)
  - Owners         — one row per (app, owner); ownerless apps flagged
  - High Risk      — only the apps flagged high-risk, with the reason signals
  - Permission Pivot — count of apps per permission (the facet rollup)

Mirrors app.rbac.export.to_workbook (same styling + CSV-injection neutralization)."""
from __future__ import annotations

from typing import Any

# Excel / LibreOffice treat a cell starting with one of these as a formula — the classic
# CSV-injection vector. Prefix with a single quote so it's read as literal text.
_FORMULA_TRIGGERS = ("=", "+", "-", "@")


def _csv_safe(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip("\t\r\n ")
    if stripped and stripped[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _coerce(value: Any) -> Any:
    if isinstance(value, bool):
        return "Yes" if value else ""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return _csv_safe(str(value))


def _safe_sheet_title(title: str) -> str:
    for ch in "[]:*?/\\":
        title = title.replace(ch, " ")
    return title.strip()[:31] or "Sheet"


def _high_risk_reasons(app: dict[str, Any]) -> str:
    """Human-readable signals behind the high-risk flag."""
    reasons: list[str] = []
    high_perms = [p for p in (app.get("permissions") or []) if p.get("risk") == "high"]
    if high_perms:
        names = ", ".join(sorted({p.get("value", "") for p in high_perms if p.get("value")}))
        reasons.append(f"High-risk permission(s): {names}")
    if app.get("ownerless"):
        reasons.append("No owners")
    if app.get("expiredCredentials"):
        reasons.append(f"{app['expiredCredentials']} expired credential(s)")
    return "; ".join(reasons)


def to_workbook(snap: dict[str, Any]) -> bytes:
    """Build the multi-sheet app-registrations workbook from a snapshot dict."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    apps: list[dict[str, Any]] = list(snap.get("apps") or [])
    summary: dict[str, Any] = snap.get("summary") or {}
    facets: dict[str, Any] = snap.get("facets") or {}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F6CBD")

    def _sheet(title: str, headers: list[str], data: list[list[Any]]) -> None:
        ws = wb.create_sheet(_safe_sheet_title(title))
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for r in data:
            ws.append([_coerce(v) for v in r])
        ws.freeze_panes = "A2"
        for ci, h in enumerate(headers, start=1):
            width = len(str(h))
            for r in data[:200]:
                if ci - 1 < len(r):
                    width = max(width, len(str(_coerce(r[ci - 1]))))
            ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, width + 2))
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    # 1. Summary — KPIs + metadata.
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["Application Registrations export"])
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws0.append(["Generated", snap.get("generated_at", "")])
    ws0.append(["Tenant", snap.get("tenant_id", "")])
    ws0.append(["Source", "Demo dataset" if snap.get("source") == "demo_dummy_data" else "Microsoft Graph"])
    ws0.append([])
    ws0.append(["Metric", "Value"])
    hdr_row = ws0.max_row
    for c in range(1, 3):
        ws0.cell(row=hdr_row, column=c).font = header_font
        ws0.cell(row=hdr_row, column=c).fill = header_fill
    for label, key in [
        ("Total applications", "total"),
        ("With secrets", "withSecrets"),
        ("With certificates", "withCerts"),
        ("Credentials expiring soon", "expiringSoon"),
        ("Expired credentials", "expired"),
        ("High risk", "highRisk"),
        ("Ownerless", "ownerless"),
        ("Application permissions", "applicationPerms"),
        ("Delegated permissions", "delegatedPerms"),
    ]:
        ws0.append([label, summary.get(key, 0)])
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 24

    # 2. Applications — the main inventory grid.
    _sheet(
        "Applications",
        ["Name", "App ID", "Object ID", "Sign-in audience", "Publisher domain",
         "Secrets", "Certs", "Next expiry (days)", "Expired creds",
         "App permissions", "Delegated permissions", "High risk", "Ownerless", "Owners", "Created"],
        [
            [
                a.get("displayName", ""), a.get("appId", ""), a.get("id", ""),
                a.get("signInAudience", ""), a.get("publisherDomain", ""),
                a.get("secretsCount", 0), a.get("certsCount", 0),
                a.get("nextExpiryDays"), a.get("expiredCredentials", 0),
                a.get("applicationPermissionsCount", 0), a.get("delegatedPermissionsCount", 0),
                bool(a.get("highRisk")), bool(a.get("ownerless")),
                "; ".join(a.get("owners") or []), a.get("createdDateTime", ""),
            ]
            for a in apps
        ],
    )

    # 3. Credentials — one row per secret / certificate.
    cred_rows: list[list[Any]] = []
    for a in apps:
        for c in (a.get("credentials") or []):
            cred_rows.append([
                a.get("displayName", ""), a.get("appId", ""),
                c.get("type", ""), c.get("displayName", ""),
                c.get("endDateTime", ""), c.get("daysUntilExpiry"),
                "Expired" if (c.get("daysUntilExpiry") is not None and c["daysUntilExpiry"] < 0) else "Active",
            ])
    _sheet(
        "Credentials",
        ["Application", "App ID", "Type", "Credential name", "Expires", "Days until expiry", "Status"],
        cred_rows,
    )

    # 4. API Permissions — one row per granted permission.
    perm_rows: list[list[Any]] = []
    for a in apps:
        for p in (a.get("permissions") or []):
            perm_rows.append([
                a.get("displayName", ""), a.get("appId", ""),
                p.get("api", ""), p.get("value", ""), p.get("type", ""), p.get("risk", ""),
            ])
    _sheet(
        "API Permissions",
        ["Application", "App ID", "API", "Permission", "Type", "Risk"],
        perm_rows,
    )

    # 5. Owners — one row per (app, owner); ownerless apps surfaced explicitly.
    owner_rows: list[list[Any]] = []
    for a in apps:
        owners = a.get("owners") or []
        if owners:
            for o in owners:
                owner_rows.append([a.get("displayName", ""), a.get("appId", ""), o, ""])
        else:
            owner_rows.append([a.get("displayName", ""), a.get("appId", ""), "", "Ownerless"])
    _sheet(
        "Owners",
        ["Application", "App ID", "Owner", "Flag"],
        owner_rows,
    )

    # 6. High Risk — only flagged apps, with the reason signals.
    _sheet(
        "High Risk",
        ["Name", "App ID", "Reasons", "App permissions", "Delegated permissions", "Owners"],
        [
            [
                a.get("displayName", ""), a.get("appId", ""), _high_risk_reasons(a),
                a.get("applicationPermissionsCount", 0), a.get("delegatedPermissionsCount", 0),
                "; ".join(a.get("owners") or []),
            ]
            for a in apps if a.get("highRisk")
        ],
    )

    # 7. Permission Pivot — apps-per-permission rollup (from the facet, else recomputed).
    perm_facet = facets.get("permissions") or []
    if not perm_facet:
        counts: dict[str, int] = {}
        for a in apps:
            for v in {p.get("value", "") for p in (a.get("permissions") or []) if p.get("value")}:
                counts[v] = counts.get(v, 0) + 1
        perm_facet = [{"value": k, "count": v} for k, v in sorted(counts.items(), key=lambda kv: -kv[1])]
    _sheet(
        "Permission Pivot",
        ["Permission", "Applications granting it"],
        [[f.get("value", ""), f.get("count", 0)] for f in perm_facet],
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
