"""CSV exports and the complete Backup Manager Excel review pack.

CSV stays deliberately simple and stable.  The workbook is richer but follows the same rule
as the UI: it serialises the last completed snapshot and never calls Azure while exporting.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from app.backup_manager import service
from app.core.azure_portal import resource_url_for_host
from app.core.xlsx import WorkbookBuilder, cell_safe, hyperlink

INSTANCE_COLUMNS: Sequence[tuple[str, str]] = (
    ("friendly_name", "Item"),
    ("datasource_type", "Datasource type"),
    ("datasource_id", "Datasource id"),
    ("vault_name", "Vault"),
    ("vault_kind", "Vault kind"),
    ("policy_name", "Policy"),
    ("protection_state", "Protection state"),
    ("last_backup_status", "Last backup status"),
    ("latest_recovery_point", "Latest recovery point"),
    ("recovery_point_age_hours", "Recovery point age (h)"),
    ("orphaned", "Orphaned"),
    ("subscription_id", "Subscription"),
)

JOB_COLUMNS: Sequence[tuple[str, str]] = (
    ("start_time", "Started"),
    ("entity_name", "Item"),
    ("operation", "Operation"),
    ("status", "Status"),
    ("duration_seconds", "Duration (s)"),
    ("error_code", "Error code"),
    ("failure_title", "Cause"),
    ("failure_remediation", "Remediation"),
    ("vault_name", "Vault"),
    ("subscription_id", "Subscription"),
)

POLICY_COLUMNS: Sequence[tuple[str, str]] = (
    ("name", "Policy"),
    ("vault_name", "Vault"),
    ("backup_management_type", "Management type"),
    ("schedule_summary", "Schedule"),
    ("retention_days", "Retention (days)"),
    ("in_use_count", "Protected items"),
    ("below_floor", "Below baseline"),
    ("unused", "Unused"),
)

GAP_COLUMNS: Sequence[tuple[str, str]] = (
    ("resource_name", "Resource"),
    ("display_type", "Type"),
    ("resource_id", "Resource id"),
    ("resource_group", "Resource group"),
    ("subscription_id", "Subscription"),
    ("location", "Region"),
    ("severity", "Severity"),
    ("reason", "Reason"),
)

POSTURE_COLUMNS: Sequence[tuple[str, str]] = (
    ("vault_name", "Vault"),
    ("vault_kind", "Kind"),
    ("subscription_id", "Subscription"),
    ("score", "Score"),
    ("band", "Band"),
    ("instance_count", "Protected items"),
)

DRILL_COLUMNS: Sequence[tuple[str, str]] = (
    ("name", "Drill"),
    ("kind", "Kind"),
    ("target_name", "Target"),
    ("status", "Status"),
    ("due_at", "Due"),
    ("executed_at", "Executed"),
    ("executed_by", "Executed by"),
    ("rto_minutes", "RTO (min)"),
    ("outcome_notes", "Notes"),
)


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, float):
        return f"{value:.1f}"
    return str(cell_safe(str(value)))


def to_csv(rows: Iterable[dict[str, Any]], columns: Sequence[tuple[str, str]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([label for _key, label in columns])
    for row in rows:
        writer.writerow([_cell(row.get(key)) for key, _label in columns])
    return buffer.getvalue()


EXPORTS = {
    "instances": INSTANCE_COLUMNS,
    "jobs": JOB_COLUMNS,
    "policies": POLICY_COLUMNS,
    "gaps": GAP_COLUMNS,
    "posture": POSTURE_COLUMNS,
    "drills": DRILL_COLUMNS,
}


def export(kind: str, rows: Iterable[dict[str, Any]]) -> str:
    columns = EXPORTS.get(kind)
    if columns is None:
        raise ValueError(f"Unknown Backup Manager export '{kind}'.")
    return to_csv(rows, columns)


def evidence_payload(
    *, estate: dict[str, Any], posture: dict[str, Any], compliance: dict[str, Any],
    rpo: dict[str, Any], drills: list[dict[str, Any]], scope: dict[str, Any],
) -> dict[str, Any]:
    """The recoverability evidence bundle handed to the Evidence Locker for hash-stamping.

    Deliberately a summary rather than a raw dump: an evidence snapshot must stay readable
    years later and must not embed anything an auditor should not see."""
    return {
        "kind": "backup_manager.recoverability",
        "generated_at": estate.get("generated_at"),
        "scope": scope,
        "estate": {
            "vaults": len(estate.get("vaults", [])),
            "protected_items": len(estate.get("instances", [])),
            "replicated_items": len(estate.get("replication", [])),
            "policies": len(estate.get("policies", [])),
        },
        "posture": {
            "average_score": posture.get("average_score"),
            "red_vaults": posture.get("red_vaults"),
            "by_check": posture.get("by_check", []),
        },
        "compliance": {
            "total": compliance.get("total"),
            "compliant": compliance.get("compliant"),
            "compliance_pct": compliance.get("compliance_pct"),
        },
        "rpo": {
            "attainment_pct": rpo.get("attainment_pct"),
            "breached": rpo.get("breached"),
            "at_risk": rpo.get("at_risk"),
        },
        "drills": [
            {
                "name": d.get("name"), "kind": d.get("kind"), "status": d.get("status"),
                "executed_at": d.get("executed_at"), "executed_by": d.get("executed_by"),
                "rto_minutes": d.get("rto_minutes"), "target_name": d.get("target_name"),
            }
            for d in drills
        ],
    }


# --------------------------------------------------------------------------- workbook
SECTION_COLOURS = {
    "Overview": "44546A",
    "Protection": "2E75B6",
    "Jobs": "ED7D31",
    "Policies": "BF8F00",
    "Vault posture": "548235",
    "Gaps": "C00000",
    "Disaster recovery": "833C00",
    "Cost": "808080",
    "Managed operations": "7030A0",
}

SNAPSHOT_STALE_SECONDS = 24 * 60 * 60
_WINDOWS_PATH = re.compile(r"(?i)\b[a-z]:\\(?:[^\s\\]+\\)*[^\s\\]+")
_PYTHON_PATH = re.compile(r"(?<![\w-])/(?:[^\s/:]+/)+(?:[^\s/:]+\.(?:py|pyc|pyo))(?::\d+)?")
_SECRET_VALUE = re.compile(
    r"(?i)\b(authorization|bearer|access[_ -]?token|client[_ -]?secret|password|sas)\b\s*[:=]\s*[^\s,;]+"
)


def _list(value: Any) -> list[dict[str, Any]]:
    return [row for row in value if isinstance(row, dict)] if isinstance(value, list) else []


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _portal(resource_id: Any, host: str, label: str = "Open in Azure") -> Any:
    url = resource_url_for_host(resource_id, host)
    return hyperlink(label, url) if url else ""


def _join(value: Any) -> str:
    if isinstance(value, dict):
        return "; ".join(f"{key}={item}" for key, item in value.items())
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(item) for item in value)
    return str(value or "")


def _bounded_text(value: Any, limit: int = 1000) -> str:
    """A single-line, bounded operational detail suitable for a review workbook."""
    text = service.safe_error(str(value or ""))
    text = _WINDOWS_PATH.sub("<redacted path>", text)
    text = _PYTHON_PATH.sub("<redacted path>", text)
    text = _SECRET_VALUE.sub(lambda match: f"{match.group(1)}=<redacted>", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _paint_status(ws: Any, column: int, values: dict[str, str]) -> None:
    """Subtle row tint keyed by a status/severity column; text remains the primary signal."""
    from openpyxl.styles import PatternFill

    fills = {key: PatternFill("solid", fgColor=colour) for key, colour in values.items()}
    for row_number in range(2, ws.max_row + 1):
        token = str(ws.cell(row=row_number, column=column).value or "").lower()
        fill = fills.get(token)
        if fill:
            for cell in ws[row_number]:
                cell.fill = fill


def _limitations(snapshot: dict[str, Any], *, changes_truncated: bool, change_count: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    generated = snapshot.get("generated_at", "")
    for source, detail in sorted(_dict(snapshot.get("errors")).items()):
        rows.append([
            source, "failed", "All dependent sections", 0, "",
            f"NOT MEASURED — {_bounded_text(detail, 500)}", generated,
        ])
    for source, detail in sorted(_dict(snapshot.get("warnings")).items()):
        rows.append([
            source, "partial", "All dependent sections", "", "",
            _bounded_text(detail, 500), generated,
        ])

    if snapshot.get("partial"):
        rows.append([
            "analysis", "partial", "All sheets", "", "",
            "The completed analysis contains one or more failed sources; use the source rows below before interpreting totals.",
            generated,
        ])

    age = snapshot.get("age_seconds")
    if isinstance(age, (int, float)) and age >= SNAPSHOT_STALE_SECONDS:
        rows.append([
            "snapshot", "stale", "All sheets", "", "",
            f"The completed analysis is {round(age / 3600, 1)} hours old. Run Analyze again before making a current-state decision.",
            generated,
        ])

    for label, section in (("Protection inventory", "inventory"), ("Backup jobs", "jobs"), ("Protection gaps", "gaps")):
        value = _dict(snapshot.get(section))
        if value.get("truncated"):
            exported = len(_list(value.get("rows") if section != "gaps" else value.get("gaps")))
            rows.append([
                section, "truncated", label, exported, value.get("total_count", ""),
                "The completed snapshot retained a bounded subset; totals describe the analyzed estate.", generated,
            ])
    for section, detail in sorted(_dict(snapshot.get("truncation")).items()):
        if section in {"Protected items", "Backup jobs", "Protection gaps"}:
            continue
        value = _dict(detail)
        rows.append([
            section.lower(), "truncated", section, value.get("exported", ""), value.get("known_total", ""),
            "The completed snapshot retained a bounded subset; totals describe the analyzed estate.", generated,
        ])

    cost = _dict(snapshot.get("cost"))
    if cost.get("rate_error"):
        rows.append([
            "pricing", "partial", "Cost", "", "", _bounded_text(cost.get("rate_error"), 500),
            cost.get("as_of", generated),
        ])
    coverage_status = _dict(_dict(snapshot.get("gaps")).get("coverage_status"))
    if coverage_status.get("not_measured"):
        rows.append([
            "Backup & DR Coverage cache", "failed", "Coverage findings", 0,
            coverage_status.get("missing_snapshots", ""),
            "NOT MEASURED — no cached subscription-level coverage scans exist for this management group.", generated,
        ])
    elif coverage_status.get("partial"):
        rows.append([
            "Backup & DR Coverage cache", "partial", "Coverage findings",
            coverage_status.get("available_snapshots", ""),
            int(coverage_status.get("available_snapshots") or 0) + int(coverage_status.get("missing_snapshots") or 0),
            f"{coverage_status.get('missing_snapshots', 0)} subscription coverage snapshot(s) were missing.", generated,
        ])
    if cost.get("confidence") in ("assumed", "partial"):
        rows.append([
            "storage consumption", cost.get("confidence"), "Cost", cost.get("measured_instances", 0),
            cost.get("instance_count", 0),
            f"Unmeasured rows use the stated {cost.get('assumed_instance_gb', 0)} GB assumption.", cost.get("as_of", generated),
        ])
    if int(cost.get("unpriced_instances") or 0):
        rows.append([
            "retail meters", "unpriced", "Cost", cost.get("instance_count", 0) - cost.get("unpriced_instances", 0),
            cost.get("instance_count", 0), "Some protected-instance meters could not be matched.", cost.get("as_of", generated),
        ])
    if len(_list(cost.get("top_rows"))) < int(cost.get("instance_count") or 0):
        rows.append([
            "cost detail", "truncated", "Cost by protected item", len(_list(cost.get("top_rows"))),
            cost.get("instance_count", 0), "Cost detail is a bounded top-cost list.", cost.get("as_of", generated),
        ])
    actuals = _dict(cost.get("actuals"))
    if actuals.get("mixed_currency"):
        rows.append([
            "cost management", "partial", "Actual cost", len(_dict(actuals.get("totals_by_currency"))),
            len(_dict(actuals.get("totals_by_currency"))),
            "Multiple billing currencies are reported separately; no combined total, allocation, or variance is calculated.",
            generated,
        ])
    if actuals and not actuals.get("available", False):
        rows.append([
            "cost management", "failed", "Actual cost", 0, "",
            f"NOT MEASURED — {_bounded_text(actuals.get('reason', ''), 500)}", generated,
        ])
    if actuals.get("available") and actuals.get("partial"):
        rows.append([
            "cost management", "partial", "Actual cost",
            actuals.get("subscriptions_succeeded", 0), len(actuals.get("subscriptions") or []),
            _bounded_text(actuals.get("reason") or "Some subscription cost queries failed.", 500), generated,
        ])
    if actuals.get("available") and actuals.get("partial_period"):
        rows.append(["cost management", "partial", "Actual cost", "", "", "The selected billing period is incomplete.", generated])

    if changes_truncated:
        rows.append([
            "managed changes", "truncated", "Managed changes", change_count, "more than export limit",
            "The connection-wide managed-change ledger exceeded the workbook safety limit.", generated,
        ])
    if snapshot.get("demo"):
        rows.append(["scope", "demo", "All sheets", "", "", "Synthetic demonstration data — not an Azure observation.", generated])
    rows.append([
        "job history", "ok", "Jobs", len(_list(_dict(snapshot.get("jobs")).get("rows"))),
        "", f"Resource Graph job history covers approximately {snapshot.get('job_window_days', 0)} day(s).", generated,
    ])
    return rows


def to_workbook(
    *,
    snapshot: dict[str, Any],
    changes: list[dict[str, Any]] | None = None,
    drills: list[dict[str, Any]] | None = None,
    portal_host: str = "portal.azure.com",
    connection_label: str = "",
    ledger_generated_at: str = "",
    changes_truncated: bool = False,
) -> bytes:
    """Build the complete Backup Manager review pack from already-read data."""
    wb = WorkbookBuilder()
    generated_now = datetime.now(timezone.utc).isoformat()
    summary = _dict(snapshot.get("summary"))
    scope = _dict(snapshot.get("scope"))
    protection = _dict(summary.get("protection"))
    jobs_summary = _dict(summary.get("jobs"))
    rpo_summary = _dict(summary.get("rpo"))
    posture_summary = _dict(summary.get("posture"))
    dr_summary = _dict(summary.get("dr"))
    cost_summary = _dict(summary.get("cost"))
    cost = _dict(snapshot.get("cost"))

    # ---------------------------------------------------------------- front matter
    wb.section("Overview", SECTION_COLOURS["Overview"])
    ws = wb.first_sheet("Summary")
    from openpyxl.styles import Font

    ws.append(["Backup Manager — recoverability review pack"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    for label, value in [
        ("Scope kind", scope.get("scope_kind", "")),
        ("Scope", scope.get("scope_name", "") or scope.get("scope_id", "")),
        ("Scope id", scope.get("scope_id", "")),
        ("Subscriptions", scope.get("subscription_count", len(scope.get("subscriptions") or []))),
        ("Connection", connection_label),
        ("Snapshot generated", snapshot.get("generated_at", "")),
        ("Workbook generated", generated_now),
        ("Snapshot age (seconds)", snapshot.get("age_seconds", "")),
        ("Snapshot schema", snapshot.get("schema_version", "")),
        ("Demo data", bool(snapshot.get("demo"))),
        ("Analysis state", "Partial" if snapshot.get("partial") else "Complete"),
        ("", ""),
        ("Vaults", protection.get("vaults", 0)),
        ("Protected items", protection.get("protected_items", 0)),
        ("Stopped protection", protection.get("stopped", 0)),
        ("Orphaned protection", protection.get("orphaned", 0)),
        ("Policies", protection.get("policies", 0)),
        ("Jobs", jobs_summary.get("total", 0)),
        ("Failed jobs", jobs_summary.get("failed", 0)),
        ("Protection gaps", len(_list(_dict(snapshot.get("gaps")).get("gaps")))),
        ("Chronic failures", summary.get("chronic_failures", 0)),
        ("RPO attainment (%)", rpo_summary.get("attainment_pct", 0)),
        ("RPO breached", rpo_summary.get("breached", 0)),
        ("RPO at risk", rpo_summary.get("at_risk", 0)),
        ("RPO unknown", rpo_summary.get("unknown", 0)),
        ("Ransomware readiness", posture_summary.get("average_score", 0)),
        ("Ransomware readiness band", posture_summary.get("band", "")),
        ("Site Recovery health (%)", dr_summary.get("health_pct", 0)),
        ("Stale drills", dr_summary.get("stale_drills", 0)),
        ("Estimated monthly cost", cost_summary.get("monthly_total", 0)),
        ("Estimated annual cost", cost.get("annual_total", 0)),
        ("Currency", cost_summary.get("currency", "")),
        ("Cost confidence", cost_summary.get("confidence", "")),
        ("Recoverable monthly waste", cost_summary.get("recoverable_monthly", 0)),
        ("Actionable managed changes", summary.get("actionable_changes", 0)),
        ("", ""),
        ("Snapshot notice", "This workbook reflects the last completed analysis; it does not trigger or include a newer Azure collection."),
        ("Live ledgers read", ledger_generated_at or generated_now),
        ("How to read this file", "Start with Coverage & limitations. A failed or truncated source is not a clean result."),
        ("Data handling", "Contains operational Azure resource identifiers and backup posture. Handle as governance evidence."),
    ]:
        ws.append([cell_safe(label), cell_safe(value)])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100

    limits = _limitations(
        snapshot, changes_truncated=changes_truncated, change_count=len(changes or []),
    )
    wb.sheet(
        "Coverage & limitations",
        ["Source", "State", "Affected section", "Exported", "Known total", "Detail", "As of"],
        limits,
        note="Read this before interpreting empty sheets or zero counts.",
    )

    # ---------------------------------------------------------------- protection
    if scope.get("scope_kind") == "management_group":
        wb.sheet(
            "Scope subscriptions", ["Subscription id"],
            [[subscription] for subscription in scope.get("subscriptions") or []],
            note="Resolved descendants of the selected management group at analysis time.",
        )

    wb.section("Protection", SECTION_COLOURS["Protection"])
    instances = _list(_dict(snapshot.get("inventory")).get("rows"))
    protected_rows: list[list[Any]] = []
    for row in instances:
        source_link = "Source deleted" if row.get("orphaned") else _portal(row.get("datasource_id"), portal_host, "Open source")
        protected_rows.append([
            row.get("friendly_name", ""), row.get("datasource_type", ""), row.get("datasource_id", ""), source_link,
            row.get("id", ""), _portal(row.get("id"), portal_host, "Open protected item"),
            row.get("vault_name", ""), row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("policy_name", ""), row.get("policy_id", ""), _portal(row.get("policy_id"), portal_host, "Open policy"),
            row.get("protection_state", ""), row.get("protection_status", "") or row.get("health_status", ""),
            row.get("last_backup_status", ""), row.get("last_backup_time", ""),
            row.get("latest_recovery_point", ""), row.get("recovery_point_age_hours", ""),
            row.get("recovery_point_source", ""), row.get("protection_stopped", False),
            row.get("retain_data_only", False), row.get("orphaned", False), row.get("subscription_id", ""),
        ])
    wb.sheet(
        "Protected items",
        ["Item", "Datasource type", "Datasource id", "Source portal", "Protected item id", "Protected item portal",
         "Vault", "Vault id", "Vault portal", "Policy", "Policy id", "Policy portal", "Protection state",
         "Protection status", "Last backup status", "Last backup", "Latest recovery point", "Recovery point age (h)",
         "Recovery point source", "Protection stopped", "Retain data only", "Orphaned", "Subscription"],
        protected_rows,
        note="Orphaned rows intentionally have no source-resource link.",
    )

    rpo_rows = _list(_dict(_dict(snapshot.get("dr")).get("rpo")).get("rows"))
    rpo_ws = wb.sheet(
        "RPO attainment",
        ["Item", "Datasource type", "Datasource id", "Source portal", "Protected item id", "Protected item portal",
         "Vault", "Tier", "RPO target (h)", "Recovery point age (h)", "Latest recovery point",
         "Recovery point source", "Status", "Subscription"],
        [[
            row.get("name", ""), row.get("datasource_type", ""), row.get("datasource_id", ""),
            _portal(row.get("datasource_id"), portal_host, "Open source"), row.get("instance_id", ""),
            _portal(row.get("instance_id"), portal_host, "Open protected item"), row.get("vault_name", ""),
            row.get("tier", ""), row.get("rpo_target_hours", ""), row.get("recovery_point_age_hours", ""),
            row.get("latest_recovery_point", ""), row.get("recovery_point_source", ""), row.get("status", ""),
            row.get("subscription_id", ""),
        ] for row in rpo_rows],
    )
    _paint_status(rpo_ws, 13, {"breached": "FCE8E6", "at_risk": "FFF2CC", "met": "E2F0D9", "unknown": "E7E6E6"})
    wb.sheet(
        "Orphaned protection",
        ["Item", "Datasource id", "Source", "Protected item id", "Protected item portal", "Vault", "Vault portal",
         "Policy", "Latest recovery point", "Monthly billing continues"],
        [[
            row.get("friendly_name", ""), row.get("datasource_id", ""), "Source deleted", row.get("id", ""),
            _portal(row.get("id"), portal_host, "Open protected item"), row.get("vault_name", ""),
            _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("policy_name", ""),
            row.get("latest_recovery_point", ""), True,
        ] for row in instances if row.get("orphaned")],
        note="Backup data is retained and billed; destructive removal remains portal-only.",
    )

    # ---------------------------------------------------------------- jobs
    wb.section("Jobs", SECTION_COLOURS["Jobs"])
    jobs = _list(_dict(snapshot.get("jobs")).get("rows"))
    jobs_ws = wb.sheet(
        "Backup jobs",
        ["Started", "Ended", "Duration (s)", "Item", "Operation", "Status", "Status bucket", "Error code",
         "Cause", "Remediation", "Job id", "Job portal", "Datasource id", "Source portal", "Vault", "Vault portal",
         "Subscription"],
        [[
            row.get("start_time", ""), row.get("end_time", ""), row.get("duration_seconds", ""), row.get("entity_name", ""),
            row.get("operation", ""), row.get("status", ""), row.get("status_bucket", ""), row.get("error_code", ""),
            _bounded_text(row.get("failure_title", "") or row.get("error_message", "")),
            _bounded_text(row.get("failure_remediation", "")), row.get("id", ""),
            _portal(row.get("id"), portal_host, "Open job"), row.get("datasource_id", ""),
            _portal(row.get("datasource_id"), portal_host, "Open source"), row.get("vault_name", ""),
            _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("subscription_id", ""),
        ] for row in jobs],
        note=f"Resource Graph job history covers approximately {snapshot.get('job_window_days', 0)} day(s).",
    )
    _paint_status(jobs_ws, 7, {"failed": "FCE8E6", "running": "DDEBF7", "succeeded": "E2F0D9", "unknown": "E7E6E6"})

    job_analysis = _dict(snapshot.get("job_analysis"))
    wb.sheet(
        "Failure clusters",
        ["Error code", "Title", "Category", "Severity", "Known", "Retryable", "Jobs", "Items", "Subscriptions",
         "Vaults", "Latest", "Cause", "Remediation", "Affected items", "Sample message"],
        [[
            row.get("error_code", ""), row.get("title", ""), row.get("category", ""), row.get("severity", ""),
            row.get("known", False), row.get("retryable", False), row.get("job_count", 0), row.get("entity_count", 0),
            row.get("subscription_count", 0), row.get("vault_count", 0), row.get("latest_at", ""),
            _bounded_text(row.get("cause", "")),
            _bounded_text(row.get("remediation", "")), _join(row.get("entities")),
            _bounded_text(row.get("sample_message", "")),
        ] for row in _list(job_analysis.get("clusters"))],
    )
    wb.sheet(
        "Chronic failures",
        ["Item", "Datasource type", "Datasource id", "Source portal", "Protected item id", "Protected item portal",
         "Vault", "Vault portal", "Policy", "Recovery point age (days)", "Latest recovery point", "Error code",
         "Error", "Severity", "Subscription"],
        [[
            row.get("name", ""), row.get("datasource_type", ""), row.get("datasource_id", ""),
            _portal(row.get("datasource_id"), portal_host, "Open source"), row.get("instance_id", ""),
            _portal(row.get("instance_id"), portal_host, "Open protected item"), row.get("vault_name", ""),
            _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("policy_name", ""),
            "never" if row.get("age_days") is None else row.get("age_days"), row.get("latest_recovery_point", ""),
            row.get("error_code", ""), _bounded_text(row.get("error_message", "")), row.get("severity", ""),
            row.get("subscription_id", ""),
        ] for row in _list(job_analysis.get("chronic"))],
    )
    wb.sheet(
        "Job congestion",
        ["Hour", "Jobs", "Failed", "Average duration (s)"],
        [[row.get("hour", ""), row.get("total", 0), row.get("failed", 0), row.get("avg_duration_s", 0)]
         for row in _list(job_analysis.get("congestion"))],
    )

    # ---------------------------------------------------------------- policies
    wb.section("Policies", SECTION_COLOURS["Policies"])
    policies = _list(_dict(snapshot.get("policies")).get("policies"))
    wb.sheet(
        "Policies",
        ["Policy", "Policy id", "Policy portal", "Vault", "Vault id", "Vault portal", "Management type", "Workload type",
         "Schedule", "Time zone", "Retention (days)", "Instant RP (days)", "Protected items", "Below baseline", "Unused",
         "Duplicate count", "Fingerprint", "Subscription"],
        [[
            row.get("name", ""), row.get("arm_id", "") or row.get("id", ""),
            _portal(row.get("arm_id", "") or row.get("id", ""), portal_host, "Open policy"), row.get("vault_name", ""),
            row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("backup_management_type", ""), row.get("workload_type", ""), row.get("schedule_summary", ""),
            row.get("time_zone", ""), row.get("retention_days", ""), row.get("instant_rp_days", ""),
            row.get("in_use_count", 0), row.get("below_floor", False), row.get("unused", False),
            len(row.get("duplicate_of") or []), row.get("fingerprint", ""), row.get("subscription_id", ""),
        ] for row in policies],
    )
    wb.sheet(
        "Duplicate policies",
        ["Fingerprint", "Copies", "Vaults", "Protected items", "Retention (days)", "Schedule", "Management type",
         "Policy names", "Vault names"],
        [[
            row.get("fingerprint", ""), row.get("policy_count", 0), row.get("vault_count", 0), row.get("protected_items", 0),
            row.get("retention_days", ""), row.get("schedule_summary", ""), row.get("backup_management_type", ""),
            _join(row.get("names")), _join(row.get("vaults")),
        ] for row in _list(_dict(snapshot.get("policies")).get("duplicate_groups"))],
    )
    compliance = _dict(snapshot.get("compliance"))
    wb.sheet(
        "Policy compliance",
        ["Item", "Datasource type", "Vault", "Tier", "Tier label", "RPO target (h)", "Recovery age (h)", "RPO met",
         "Retention (days)", "Retention target", "Retention met", "Offsite required", "Offsite met", "Compliant",
         "Protected item id", "Protected item portal"],
        [[
            row.get("name", ""), row.get("datasource_type", ""), row.get("vault_name", ""), row.get("tier", ""),
            row.get("tier_label", ""), row.get("rpo_target_hours", ""), row.get("recovery_point_age_hours", ""),
            row.get("rpo_ok", False), row.get("retention_days", ""), row.get("retention_target_days", ""),
            row.get("retention_ok", False), row.get("offsite_required", False), row.get("offsite_ok", False),
            row.get("compliant", False), row.get("instance_id", ""), _portal(row.get("instance_id"), portal_host, "Open protected item"),
        ] for row in _list(compliance.get("rows"))],
    )
    wb.sheet(
        "Compliance tiers",
        ["Tier", "Label", "RPO (h)", "Retention (days)", "Offsite required", "Drill interval (days)"],
        [[row.get("id", ""), row.get("label", ""), row.get("rpo_hours", ""), row.get("retention_days", ""),
          row.get("require_offsite", False), row.get("drill_days", "")] for row in _list(compliance.get("tiers"))],
    )

    # ---------------------------------------------------------------- vault posture
    wb.section("Vault posture", SECTION_COLOURS["Vault posture"])
    vaults = _list(_dict(snapshot.get("vaults")).get("vaults"))
    wb.sheet(
        "Vaults",
        ["Vault", "Vault id", "Vault portal", "Kind", "Region", "Resource group", "Subscription", "SKU",
         "Provisioning", "Public network access", "Private endpoints", "Soft delete", "Soft-delete retention",
         "Immutability", "Redundancy", "Cross-region restore", "Cross-subscription restore", "Built-in alerts",
         "Customer-managed key", "MUA", "Resource Guard", "Diagnostics", "Diagnostic workspaces",
         "Primary diagnostic workspace id", "Primary diagnostic workspace portal",
         "Protected items", "Policies", "Replicated items", "Empty", "Enrichment error"],
        [[
            row.get("name", ""), row.get("id", ""), _portal(row.get("id"), portal_host, "Open vault"), row.get("kind", ""),
            row.get("location", ""), row.get("resource_group", ""), row.get("subscription_id", ""), row.get("sku", ""),
            row.get("provisioning_state", ""), row.get("public_network_access", ""), row.get("private_endpoints", 0),
            row.get("soft_delete_state", ""), row.get("soft_delete_retention_days", ""), row.get("immutability_state", ""),
            row.get("redundancy", ""), row.get("cross_region_restore", ""), row.get("cross_subscription_restore", ""),
            row.get("monitor_alerts", ""), row.get("cmk", False), row.get("mua_enabled", ""), row.get("mua_resource_guard_id", ""),
            row.get("diagnostics_enabled", ""), _join(row.get("diagnostics_workspaces")),
            (row.get("diagnostics_workspaces") or [""])[0],
            _portal((row.get("diagnostics_workspaces") or [""])[0], portal_host, "Open diagnostic workspace"),
            row.get("instance_count", 0),
            row.get("policy_count", 0), row.get("replicated_item_count", 0), row.get("empty", False),
            _bounded_text(row.get("enrichment_error", "")),
        ] for row in vaults],
    )
    posture = _dict(snapshot.get("posture"))
    posture_rows = _list(posture.get("vaults"))
    posture_ws = wb.sheet(
        "Vault posture",
        ["Vault", "Vault id", "Vault portal", "Kind", "Region", "Subscription", "Score", "Band", "Worst status",
         "Failing checks", "Warning checks", "Available actions", "Portal-only gaps", "Protected items", "Replicated items"],
        [[
            row.get("vault_name", ""), row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("vault_kind", ""), row.get("location", ""), row.get("subscription_id", ""), row.get("score", 0),
            row.get("band", ""), row.get("status", ""), _join(row.get("failing")), _join(row.get("warning")),
            _join(row.get("actionable")), _join([gap.get("label", "") for gap in _list(row.get("portal_only_gaps"))]),
            row.get("instance_count", 0), row.get("replicated_item_count", 0),
        ] for row in posture_rows],
    )
    _paint_status(posture_ws, 8, {"red": "FCE8E6", "amber": "FFF2CC", "green": "E2F0D9"})
    check_rows: list[list[Any]] = []
    for vault in posture_rows:
        for check in _list(vault.get("checks")):
            check_rows.append([
                vault.get("vault_name", ""), vault.get("vault_id", ""), _portal(vault.get("vault_id"), portal_host, "Open vault"),
                check.get("id", ""), check.get("label", ""), check.get("status", ""), check.get("value", ""),
                check.get("detail", ""), check.get("severity", ""), check.get("weight", 0), check.get("action", ""),
                check.get("portal_only", False), check.get("portal_reason", ""), check.get("why", ""),
            ])
    checks_ws = wb.sheet(
        "Vault checks",
        ["Vault", "Vault id", "Vault portal", "Check id", "Check", "Status", "Observed value", "Detail",
         "Severity", "Weight", "Available action", "Portal-only", "Portal-only reason", "Why it matters"],
        check_rows,
    )
    _paint_status(checks_ws, 6, {"fail": "FCE8E6", "warn": "FFF2CC", "pass": "E2F0D9", "na": "E7E6E6"})
    wb.sheet(
        "Vault capacity",
        ["Vault", "Vault id", "Vault portal", "Kind", "Subscription", "Instances", "Instance limit", "Instance %",
         "Policies", "Policy limit", "Policy %", "At risk"],
        [[
            row.get("vault_name", ""), row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("vault_kind", ""), row.get("subscription_id", ""), row.get("instances", 0), row.get("instance_limit", 0),
            row.get("instance_pct", 0), row.get("policies", 0), row.get("policy_limit", 0), row.get("policy_pct", 0),
            row.get("at_risk", False),
        ] for row in _list(_dict(snapshot.get("vaults")).get("capacity"))],
    )

    # ---------------------------------------------------------------- gaps
    wb.section("Gaps", SECTION_COLOURS["Gaps"])
    gaps = _dict(snapshot.get("gaps"))
    gap_fields = ["severity", "resource_name", "display_type", "resource_type", "resource_id", "resource_group",
                  "location", "subscription_id", "reason", "source", "mechanism", "target_vault_kind"]
    gap_headers = ["Severity", "Resource", "Display type", "Resource type", "Resource id", "Resource group",
                   "Region", "Subscription", "Reason", "Source", "Mechanism", "Target vault kind", "Resource portal"]
    def gap_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
        rank = {"critical": 0, "error": 1, "warning": 2, "info": 3}
        ordered = sorted(items, key=lambda row: (rank.get(str(row.get("severity") or "").lower(), 4),
                                                 str(row.get("display_type") or "").lower(),
                                                 str(row.get("resource_name") or "").lower(), str(row.get("gap_id") or "")))
        return [[*(row.get(field, "") for field in gap_fields), _portal(row.get("resource_id"), portal_host, "Open resource")]
                for row in ordered]
    gaps_ws = wb.sheet("Protection gaps", gap_headers, gap_rows(_list(gaps.get("gaps"))))
    _paint_status(gaps_ws, 1, {"critical": "FCE8E6", "error": "FCE8E6", "warning": "FFF2CC", "info": "DDEBF7"})
    wb.sheet("Coverage findings", gap_headers, gap_rows(_list(gaps.get("coverage_gaps"))),
             note="Findings ingested from the read-only Backup & DR Coverage detector.")
    wb.sheet(
        "Native-only protections", ["Resource type", "Why it is not a vault gap"],
        [[row.get("type", ""), row.get("note", "")] for row in _list(gaps.get("native_only"))],
    )

    # ---------------------------------------------------------------- disaster recovery
    wb.section("Disaster recovery", SECTION_COLOURS["Disaster recovery"])
    dr = _dict(snapshot.get("dr"))
    dr_ws = wb.sheet(
        "Replicated items",
        ["Item", "Item id", "Item portal", "Vault", "Vault portal", "Source id", "Source portal", "Protection state",
         "Replication health", "Status", "Primary region", "Recovery region", "RPO (seconds)", "Last test failover",
         "Test age (days)", "Stale drill", "Test active", "Issues"],
        [[
            row.get("friendly_name", ""), row.get("id", ""), _portal(row.get("id"), portal_host, "Open replicated item"),
            row.get("vault_name", ""), _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("datasource_id", ""),
            _portal(row.get("datasource_id"), portal_host, "Open source"), row.get("protection_state", ""),
            row.get("replication_health", ""), row.get("status", ""), row.get("primary_region", ""), row.get("recovery_region", ""),
            row.get("rpo_seconds", ""), row.get("last_test_failover", ""), row.get("last_test_failover_age_days", ""),
            row.get("stale_drill", False), row.get("test_failover_active", False),
            _bounded_text(_join(row.get("issues"))),
        ] for row in _list(dr.get("items"))],
    )
    _paint_status(dr_ws, 10, {"red": "FCE8E6", "amber": "FFF2CC", "green": "E2F0D9"})
    wb.sheet(
        "Recovery plans",
        ["Plan", "Plan id", "Plan portal", "Vault", "Vault portal", "Primary region", "Recovery region",
         "Protected items", "Scenario", "Scenario status", "Last test failover", "Test age (days)", "Stale drill"],
        [[
            row.get("friendly_name", ""), row.get("id", ""), _portal(row.get("id"), portal_host, "Open recovery plan"),
            row.get("vault_name", ""), _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("primary_region", ""),
            row.get("recovery_region", ""), row.get("protected_item_count", 0), row.get("current_scenario", ""),
            row.get("current_scenario_status", ""), row.get("last_test_failover", ""), row.get("last_test_failover_age_days", ""),
            row.get("stale_drill", False),
        ] for row in _list(dr.get("recovery_plans"))],
    )
    wb.sheet(
        "Drill register",
        ["Drill", "Kind", "Scope kind", "Scope id", "Target", "Target id", "Target portal", "Status", "Cadence (days)",
         "Due", "Overdue", "Executed", "Executed by", "RTO (min)", "Outcome", "Change id", "Evidence id", "Created"],
        [[
            row.get("name", ""), row.get("kind", ""), row.get("scope_kind", ""), row.get("scope_id", ""),
            row.get("target_name", ""), row.get("target_id", ""), _portal(row.get("target_id"), portal_host, "Open target"),
            row.get("status", ""), row.get("cadence_days", 0), row.get("due_at", ""), row.get("overdue", False),
            row.get("executed_at", ""), row.get("executed_by", ""), row.get("rto_minutes", ""), row.get("outcome_notes", ""),
            row.get("change_id", ""), row.get("evidence_id", ""), row.get("created_at", ""),
        ] for row in (drills or [])],
        note=f"Current connection-wide drill register read at {ledger_generated_at or generated_now}.",
    )

    # ---------------------------------------------------------------- cost
    wb.section("Cost", SECTION_COLOURS["Cost"])
    actuals = _dict(cost.get("actuals"))
    allocation = _dict(cost.get("allocation"))
    variance = _dict(cost.get("variance"))
    waste = _dict(cost.get("waste"))
    wb.sheet(
        "Cost summary", ["Metric", "Value", "Currency / context"],
        [["Rate source", cost.get("rate_source", ""), _bounded_text(cost.get("rate_error", ""), 500)],
         ["Region", cost.get("region", ""), ""], ["As of", cost.get("as_of", ""), ""],
         ["Estimate confidence", cost.get("confidence", ""), ""], ["Estimate only", cost.get("estimate_only", True), ""],
         ["Protected-instance cost", cost.get("protected_instance_cost", 0), cost.get("currency", "")],
         ["Storage cost", cost.get("storage_cost", 0), cost.get("currency", "")],
         ["Site Recovery cost", cost.get("site_recovery_cost", 0), cost.get("currency", "")],
         ["Estimated monthly", cost.get("monthly_total", 0), cost.get("currency", "")],
         ["Estimated annual", cost.get("annual_total", 0), cost.get("currency", "")],
         ["Protected items", cost.get("instance_count", 0), ""], ["Measured items", cost.get("measured_instances", 0), ""],
         ["Unpriced items", cost.get("unpriced_instances", 0), ""], ["Assumed GB per unmeasured item", cost.get("assumed_instance_gb", 0), "GB"],
         ["Actuals available", actuals.get("available", False), actuals.get("reason", "")],
         ["Actual total", actuals.get("total", 0), actuals.get("currency", "")],
         ["Actual totals by currency", _join(actuals.get("totals_by_currency")), "not summed" if actuals.get("mixed_currency") else ""],
         ["Actual period", _join(actuals.get("period")), "partial" if actuals.get("partial_period") else "complete"],
         ["Variance comparable", variance.get("comparable", False), variance.get("reason", "")],
         ["Variance", variance.get("delta", 0), variance.get("actual_currency", "")],
         ["Recoverable monthly waste", waste.get("recoverable_monthly", 0), waste.get("currency", "")],
         ["Waste basis", waste.get("basis", ""), ""]],
    )
    cost_rows = _list(cost.get("top_rows"))
    wb.sheet(
        "Cost by protected item",
        ["Item", "Protected item id", "Protected item portal", "Datasource id", "Source portal", "Type", "Vault", "Vault id",
         "Vault portal", "Redundancy", "Stored GB", "Measured", "Instance cost", "Storage cost", "Monthly cost",
         "Meter", "Priced", "Note", "Currency"],
        [[
            row.get("name", ""), row.get("instance_id", ""), _portal(row.get("instance_id"), portal_host, "Open protected item"),
            row.get("datasource_id", ""), _portal(row.get("datasource_id"), portal_host, "Open source"), row.get("datasource_type", ""),
            row.get("vault_name", ""), row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("redundancy", ""), row.get("stored_gb", 0), row.get("measured", False), row.get("instance_cost", 0),
            row.get("storage_cost", 0), row.get("monthly_cost", 0), row.get("meter", ""), row.get("priced", False),
            row.get("note", ""), cost.get("currency", ""),
        ] for row in cost_rows],
        note="This is a bounded top-cost list when the analyzed estate exceeds the detail limit.",
    )
    wb.sheet(
        "Cost allocation",
        ["Item", "Protected item id", "Protected item portal", "Type", "Vault", "Vault id", "Vault portal",
         "Allocated actual", "Vault actual", "Weight", "Weight basis", "Currency"],
        [[
            row.get("name", ""), row.get("instance_id", ""), _portal(row.get("instance_id"), portal_host, "Open protected item"),
            row.get("datasource_type", ""), row.get("vault_name", ""), row.get("vault_id", ""),
            _portal(row.get("vault_id"), portal_host, "Open vault"), row.get("allocated_cost", 0), row.get("vault_total", 0),
            row.get("weight", 0), row.get("weight_basis", ""), allocation.get("currency", ""),
        ] for row in _list(allocation.get("rows"))],
        note=allocation.get("note", ""),
    )
    wb.sheet(
        "Recoverable waste",
        ["Severity", "Kind", "Finding", "Detail", "Item", "Target id", "Target portal", "Datasource id", "Source portal",
         "Vault", "Vault id", "Vault portal", "Monthly cost", "Currency", "Action", "Basis"],
        [[
            row.get("severity", ""), row.get("kind", ""), row.get("title", ""),
            _bounded_text(row.get("detail", "")), row.get("name", ""),
            row.get("instance_id", ""), _portal(row.get("instance_id"), portal_host, "Open target"), row.get("datasource_id", ""),
            "Source deleted" if row.get("kind") == "orphaned_protection" else _portal(row.get("datasource_id"), portal_host, "Open source"),
            row.get("vault_name", ""), row.get("vault_id", ""), _portal(row.get("vault_id"), portal_host, "Open vault"),
            row.get("monthly_cost", 0), waste.get("currency", ""), row.get("action", ""), waste.get("basis", ""),
        ] for row in _list(waste.get("findings"))],
    )

    # ---------------------------------------------------------------- managed operations
    wb.section("Managed operations", SECTION_COLOURS["Managed operations"])
    wb.sheet(
        "Managed changes",
        ["Change id", "Target type", "Target", "Target id", "Target portal", "Operation", "Status", "Risk", "Summary",
         "Requested by", "Requested at", "Decided by", "Decided at", "Decision reason", "Dual approval", "Second approver",
         "Applied by", "Applied at", "Error code", "Error", "Rollback of", "Evidence id", "Plan id", "Depends on",
         "Azure job id", "Poll attempts", "Async", "Can rollback"],
        [[
            row.get("id", ""), row.get("target_label", "") or row.get("target_type", ""), row.get("target_name", ""),
            row.get("target_id", ""), _portal(row.get("target_id"), portal_host, "Open target"), row.get("operation", ""),
            row.get("status", ""), row.get("risk", ""), _bounded_text(_join(row.get("summary"))), row.get("requested_by", ""),
            row.get("requested_at", ""), row.get("decided_by", ""), row.get("decided_at", ""), row.get("decision_reason", ""),
            row.get("requires_dual_approval", False), row.get("second_approver", ""), row.get("applied_by", ""),
            row.get("applied_at", ""), row.get("error_code", ""), _bounded_text(row.get("error_message", "")),
            row.get("rollback_of", ""),
            row.get("evidence_id", ""), row.get("plan_id", ""), _join(row.get("depends_on")), row.get("azure_job_id", ""),
            row.get("poll_attempts", 0), row.get("is_async", False), row.get("can_rollback", False),
        ] for row in (changes or [])],
        note=f"Current connection-wide public change ledger read at {ledger_generated_at or generated_now}. Encrypted payloads are never exported.",
    )

    wb.index_sheet(position=2, colour=SECTION_COLOURS["Overview"])
    return wb.to_bytes()
