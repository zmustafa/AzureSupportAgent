"""Management-group scope is exact, complete, isolated, and analysis-only."""
from __future__ import annotations

from io import BytesIO
from typing import Any

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.api import backup_manager as api
from app.api import workloads as workloads_api
from app.backup_manager import costmgmt, export, gaps, inventory, reports, service
from app.backup_manager import snapshot as snapshot_store
from app.core.security import Principal
from app.workloads import discovery

TENANT = "tenant-management-group"
CONNECTION = {"id": "connection-management-group", "display_name": "Synthetic connection"}
MG = "mg-synthetic-root"
SUB_A = "10000000-0000-0000-0000-000000000001"
SUB_B = "10000000-0000-0000-0000-000000000002"


def _principal() -> Principal:
    return Principal("operator@example.test", "operator@example.test", TENANT, "operator", frozenset({"backup_manager.read"}))


@pytest.fixture(autouse=True)
def isolated_stores(tmp_path, monkeypatch):
    monkeypatch.setattr(snapshot_store, "_PATH", tmp_path / "backup-manager-snapshots.json")
    monkeypatch.setattr(snapshot_store, "_locks", {})
    from app.backupdr import cache as coverage_cache

    monkeypatch.setattr(coverage_cache, "_PATH", tmp_path / "backupdr-snapshots.json")
    monkeypatch.setattr(coverage_cache, "_locks", {})
    return tmp_path


def test_scope_contract_requires_exactly_one_scope() -> None:
    assert service.scope_identity("workload-a", "", "") == ("workload", "workload-a")
    assert service.scope_identity("", SUB_A, "") == ("subscription", SUB_A)
    assert service.scope_identity("", "", f"/providers/Microsoft.Management/managementGroups/{MG}") == (
        "management_group", MG,
    )
    with pytest.raises(ValueError, match="Select a workload"):
        service.scope_identity("", "", "")
    with pytest.raises(ValueError, match="exactly one"):
        service.scope_identity("workload-a", SUB_A, "")
    with pytest.raises(ValueError, match="identifier is invalid"):
        service.scope_identity("", "", "../other")


@pytest.mark.asyncio
async def test_management_group_picker_does_not_fall_back_from_unknown_connection(monkeypatch) -> None:
    monkeypatch.setattr(workloads_api, "get_connection", lambda _connection_id: None)

    def no_fallback(_connection_id):
        raise AssertionError("an explicit unknown connection must never fall back")

    monkeypatch.setattr(workloads_api, "resolve_connection", no_fallback)
    with pytest.raises(HTTPException) as excinfo:
        await workloads_api.tree_endpoint(
            workloads_api.TreeRequest(connection_id="missing-connection", group_by="mg_flat"),
            _principal(),
        )
    assert excinfo.value.status_code == 404


@pytest.mark.asyncio
async def test_management_group_resolution_recurses_dedupes_and_handles_cycles(monkeypatch) -> None:
    async def token(_connection):
        return "token", None

    async def groups(_token):
        return [
            {"id": MG, "name": "Synthetic root", "depth": 0},
            {"id": "mg-child", "name": "Synthetic child", "depth": 1},
        ], None

    async def children(_token, group):
        if group == MG:
            return [
                {"kind": "mg", "id": "mg-child", "name": "Synthetic child"},
                {"kind": "subscription", "id": SUB_A, "name": "Subscription A"},
            ], None
        return [
            {"kind": "mg", "id": MG, "name": "Synthetic root"},
            {"kind": "subscription", "id": SUB_A, "name": "Subscription A"},
            {"kind": "subscription", "id": SUB_B, "name": "Subscription B"},
        ], None

    monkeypatch.setattr(discovery, "get_arm_token", token)
    monkeypatch.setattr(discovery, "list_all_management_groups", groups)
    monkeypatch.setattr(discovery, "get_management_group_children", children)
    result = await discovery.resolve_management_group_scope(CONNECTION, MG)
    assert result["management_group_name"] == "Synthetic root"
    assert result["subscriptions"] == [SUB_A, SUB_B]
    assert result["subscription_count"] == 2
    assert result["descendant_management_group_count"] == 1
    assert result["resolution_complete"] is True


@pytest.mark.asyncio
@pytest.mark.parametrize("mode", ["missing", "empty", "branch_error"])
async def test_management_group_resolution_fails_closed(monkeypatch, mode: str) -> None:
    async def token(_connection):
        return "token", None

    async def groups(_token):
        found = [] if mode == "missing" else [{"id": MG, "name": "Synthetic root", "depth": 0}]
        return found, None

    async def children(_token, _group):
        if mode == "branch_error":
            return [], "access denied"
        return [], None

    monkeypatch.setattr(discovery, "get_arm_token", token)
    monkeypatch.setattr(discovery, "list_all_management_groups", groups)
    monkeypatch.setattr(discovery, "get_management_group_children", children)
    expected = LookupError if mode == "missing" else PermissionError if mode == "branch_error" else ValueError
    with pytest.raises(expected):
        await discovery.resolve_management_group_scope(CONNECTION, MG)


@pytest.mark.asyncio
async def test_arg_batches_large_management_group_and_preserves_completeness(monkeypatch) -> None:
    subscriptions = [f"10000000-0000-0000-0000-{index:012d}" for index in range(205)]
    calls: list[list[str]] = []

    async def token(_connection):
        return "token"

    async def query(_token, _query, batch, **_kwargs):
        calls.append(list(batch or []))
        rows = [{"id": f"resource-{subscription}"} for subscription in batch or []]
        return rows, None, True, len(rows)

    monkeypatch.setattr(service, "token_for", token)
    monkeypatch.setattr("app.azure.arm.query_resource_graph_paged", query)
    rows, metadata = await service.arg(CONNECTION, "resources | project id", subscriptions, max_rows=500)
    assert [len(batch) for batch in calls] == [100, 100, 5]
    assert len(rows) == 205
    assert metadata == {
        "partial": False,
        "source_total": 205,
        "source_count": 205,
        "source_limit": 500,
        "subscription_count": 205,
        "batch_count": 3,
        "successful_batches": 3,
        "failed_batches": 0,
        "errors": [],
    }


@pytest.mark.asyncio
async def test_arg_partial_batch_is_truthful_and_globally_capped(monkeypatch) -> None:
    subscriptions = [f"10000000-0000-0000-0000-{index:012d}" for index in range(205)]

    async def token(_connection):
        return "token"

    async def query(_token, _query, batch, **_kwargs):
        if batch and batch[0].endswith("000000000100"):
            return [], "denied", False, None
        rows = [{"id": f"resource-{subscription}"} for subscription in batch or []]
        return rows, None, True, len(rows)

    monkeypatch.setattr(service, "token_for", token)
    monkeypatch.setattr("app.azure.arm.query_resource_graph_paged", query)
    rows, metadata = await service.arg(CONNECTION, "resources | project id", subscriptions, max_rows=80)
    assert len(rows) == 80
    assert metadata["partial"] is True
    assert metadata["failed_batches"] == 1
    assert metadata["successful_batches"] == 2
    assert metadata["source_total"] == 105


@pytest.mark.asyncio
async def test_inventory_uses_resolved_scope_and_never_empty_all_visible(monkeypatch) -> None:
    captured: dict[str, Any] = {}

    async def resolved(*_args, **_kwargs):
        return {
            "scope_kind": "management_group", "scope_id": MG, "scope_name": "Synthetic root",
            "management_group_id": MG, "management_group_name": "Synthetic root",
            "subscriptions": [SUB_A, SUB_B], "subscription_count": 2,
            "resolution_complete": True, "resolution_warnings": [],
        }

    async def collect(_connection, *, subscriptions, scope, **_kwargs):
        captured["subscriptions"] = subscriptions
        captured["scope"] = scope
        return {"scope": scope, "vaults": [], "instances": [], "jobs": [], "policies": [], "errors": {}}

    async def direct(_key, producer):
        return await producer()

    monkeypatch.setattr(service, "resolve_scope", resolved)
    monkeypatch.setattr(inventory, "_collect_uncached", collect)
    monkeypatch.setattr(inventory.inventory_cache, "get_or_create", direct)
    result = await inventory.collect_estate(CONNECTION, tenant_id=TENANT, management_group_id=MG, force=False)
    assert captured["subscriptions"] == {SUB_A, SUB_B}
    assert captured["scope"]["subscription_count"] == 2
    assert result["scope"]["management_group_name"] == "Synthetic root"


@pytest.mark.asyncio
async def test_inventory_reuses_preflight_management_group_boundary(monkeypatch) -> None:
    resolved = {
        "scope_kind": "management_group", "scope_id": MG, "scope_name": "Synthetic root",
        "management_group_id": MG, "management_group_name": "Synthetic root",
        "subscriptions": [SUB_A, SUB_B], "subscription_count": 2,
        "resolution_complete": True, "resolution_warnings": [],
    }

    async def no_second_resolution(*_args, **_kwargs):
        raise AssertionError("the preflight management-group boundary must be reused")

    async def collect(_connection, *, subscriptions, scope, **_kwargs):
        assert subscriptions == {SUB_A, SUB_B}
        return {"scope": scope, "vaults": [], "instances": [], "jobs": [], "policies": [], "errors": {}}

    async def direct(_key, producer):
        return await producer()

    monkeypatch.setattr(service, "resolve_scope", no_second_resolution)
    monkeypatch.setattr(inventory, "_collect_uncached", collect)
    monkeypatch.setattr(inventory.inventory_cache, "get_or_create", direct)
    result = await inventory.collect_estate(
        CONNECTION, tenant_id=TENANT, management_group_id=MG, resolved_scope=resolved,
    )
    assert result["scope"]["subscription_count"] == 2


@pytest.mark.asyncio
async def test_management_group_coverage_merges_subscription_caches() -> None:
    from app.backupdr import cache as coverage_cache

    coverage_cache.write_snapshot(TENANT, "subscription", SUB_A, {
        "gaps": [{"resource_id": f"/subscriptions/{SUB_A}/resourceGroups/rg/providers/Microsoft.Compute/virtualMachines/vm-a", "resource_name": "VM A", "resource_type": "microsoft.compute/virtualmachines", "subscription_id": SUB_A, "failed_checks": ["backup"]}],
    })
    rows, status = gaps.ingest_coverage_gaps_for_scope(TENANT, "management_group", MG, [SUB_A, SUB_B])
    assert len(rows) == 1
    assert status == {"available_snapshots": 1, "missing_snapshots": 1, "partial": True, "not_measured": False}


@pytest.mark.asyncio
async def test_multi_workspace_report_merges_scope_safe_rows_and_marks_partial(monkeypatch) -> None:
    estate = {
        "scope": {"scope_kind": "management_group", "subscriptions": [SUB_A, SUB_B]},
        "instances": [{"friendly_name": "Synthetic VM A"}, {"friendly_name": "Synthetic VM B"}],
        "vaults": [
            {"diagnostics_enabled": True, "diagnostics_workspaces": ["workspace-a"]},
            {"diagnostics_enabled": True, "diagnostics_workspaces": ["workspace-b"]},
        ],
    }

    async def query(_connection, workspace, kql, **_kwargs):
        if workspace == "workspace-b" and kql == reports.SLA_KQL:
            raise reports.ReportsUnavailable("denied")
        if kql == reports.STORAGE_KQL:
            return [{"BackupItemUniqueId": f"Synthetic VM {'A' if workspace == 'workspace-a' else 'B'}", "StorageConsumedInMBs": 1024, "StorageType": "Vault"}]
        if kql == reports.SLA_KQL:
            return [{"BackupItemUniqueId": "Synthetic VM A", "Total": 1, "Succeeded": 1, "SuccessRate": 100}]
        return [{"TimeGenerated": "2026-08-12", "Total": 1, "Failed": 0, "Succeeded": 1, "JobFailureCode": "x", "Failures": 1, "Items": 1}]

    monkeypatch.setattr(reports, "query", query)
    result = await reports.build_report(CONNECTION, estate)
    assert result["available"] is True
    assert result["workspaces_total"] == 2
    assert result["workspaces_succeeded"] == 1
    assert result["workspaces_failed"] == 1
    assert result["partial"] is True
    assert result["job_trend"] == []
    assert result["failure_history"] == []
    assert set(result["storage_by_item"]) == {"Synthetic VM A"}


@pytest.mark.asyncio
async def test_cost_management_keeps_multiple_currencies_separate(monkeypatch) -> None:
    async def token(_connection):
        return "token"

    async def query(_token, scope, _body):
        currency = "EUR" if scope.endswith(SUB_A) else "CAD"
        return ["Cost", "Currency", "ResourceId", "Meter", "MeterSubCategory", "UsageDate"], [
            [10, currency, f"{scope}/resourceGroups/rg/providers/Microsoft.RecoveryServices/vaults/vault", "Storage", "Backup", "20260801"]
        ], ""

    monkeypatch.setattr(service, "token_for", token)
    monkeypatch.setattr(costmgmt, "_query", query)
    result = await costmgmt.backup_actuals(CONNECTION, [SUB_A, SUB_B])
    assert result["mixed_currency"] is True
    assert result["currency"] == ""
    assert result["total"] == 0.0
    assert result["totals_by_currency"] == {"CAD": 10.0, "EUR": 10.0}
    assert "not summed" in result["reason"]


def test_management_group_snapshot_isolated_and_workbook_has_subscription_manifest() -> None:
    snapshot = snapshot_store.empty_snapshot("management_group", MG)
    snapshot.update({
        "report_exists": True,
        "scope": {
            "scope_kind": "management_group", "scope_id": MG, "scope_name": "Synthetic root",
            "management_group_id": MG, "management_group_name": "Synthetic root",
            "subscriptions": [SUB_A, SUB_B], "subscription_count": 2,
            "resolution_complete": True, "resolution_warnings": [],
        },
    })
    snapshot_store.write_snapshot(TENANT, CONNECTION["id"], "management_group", MG, snapshot)
    assert snapshot_store.read_snapshot(TENANT, CONNECTION["id"], "management_group", MG) is not None
    assert snapshot_store.read_snapshot(TENANT, CONNECTION["id"], "subscription", MG) is None
    stored = snapshot_store.list_scopes(TENANT)
    assert stored[0]["scope_name"] == "Synthetic root"
    assert stored[0]["subscription_count"] == 2
    workbook = load_workbook(BytesIO(export.to_workbook(snapshot=snapshot)), read_only=False, data_only=False)
    assert "Scope subscriptions" in workbook.sheetnames
    assert [row[0] for row in workbook["Scope subscriptions"].iter_rows(min_row=2, values_only=True)] == [SUB_A, SUB_B]
    summary = {row[0].value: row[1].value for row in workbook["Summary"].iter_rows(min_col=1, max_col=2) if row[0].value}
    assert summary["Scope"] == "Synthetic root"
    assert summary["Scope id"] == MG
    assert summary["Subscriptions"] == 2
    workbook.close()


def test_management_group_scale_detail_sections_are_bounded_and_reported() -> None:
    snapshot = snapshot_store.empty_snapshot("management_group", MG)
    snapshot["policies"]["policies"] = [
        {"id": f"policy-{index}"} for index in range(snapshot_store.MAX_ROWS["policies"] + 1)
    ]
    snapshot["dr"]["items"] = [
        {"id": f"replication-{index}"} for index in range(snapshot_store.MAX_ROWS["replicated_items"] + 2)
    ]
    snapshot_store.bound(snapshot)
    assert len(snapshot["policies"]["policies"]) == snapshot_store.MAX_ROWS["policies"]
    assert len(snapshot["dr"]["items"]) == snapshot_store.MAX_ROWS["replicated_items"]
    assert snapshot["truncation"]["Policies"]["known_total"] == snapshot_store.MAX_ROWS["policies"] + 1
    assert snapshot["truncation"]["Replicated items"]["known_total"] == snapshot_store.MAX_ROWS["replicated_items"] + 2
    workbook = load_workbook(BytesIO(export.to_workbook(snapshot=snapshot)), read_only=False, data_only=False)
    limitations = list(workbook["Coverage & limitations"].iter_rows(min_row=2, values_only=True))
    assert any(row[2] == "Policies" and row[1] == "truncated" for row in limitations)
    assert any(row[2] == "Replicated items" and row[1] == "truncated" for row in limitations)
    workbook.close()


@pytest.mark.asyncio
async def test_management_group_mutations_are_server_rejected_before_estate_or_azure(monkeypatch) -> None:
    body = api.AdhocBackupRequest(connection_id=CONNECTION["id"], management_group_id=MG, instance_id="synthetic-item")

    async def forbidden(*_args, **_kwargs):
        raise AssertionError("management-group mutation must not read or write Azure")

    monkeypatch.setattr(api, "_estate", forbidden)
    with pytest.raises(HTTPException) as excinfo:
        await api.backup_now(body, principal=_principal(), db=object())
    assert excinfo.value.status_code == 400
    assert "analysis-only" in str(excinfo.value.detail)


@pytest.mark.asyncio
async def test_refresh_start_empty_management_group_creates_no_job(monkeypatch) -> None:
    monkeypatch.setattr(api, "_connection", lambda *_args, **_kwargs: CONNECTION)
    monkeypatch.setattr(api, "_refresh_jobs", type(api._refresh_jobs)("mg-empty"))

    async def empty(*_args, **_kwargs):
        raise ValueError("No visible subscriptions were found under the selected management group.")

    async def no_arg(*_args, **_kwargs):
        raise AssertionError("empty management-group resolution must perform zero ARG queries")

    monkeypatch.setattr(api.service, "resolve_scope", empty)
    monkeypatch.setattr(api.service, "arg", no_arg)
    with pytest.raises(HTTPException) as excinfo:
        await api.refresh_start(
            connection_id=CONNECTION["id"], workload_id="", subscription_id="",
            management_group_id=MG, principal=_principal(),
        )
    assert excinfo.value.status_code == 409
    assert api._refresh_jobs.jobs_with_prefix("") == []
