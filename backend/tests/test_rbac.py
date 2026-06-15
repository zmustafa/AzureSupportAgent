"""Unit tests for the RBAC (access review) feature: schema, cache, demo, compose, pivots,
export and the read-only agent tools.

The cache is isolated to a tmp dir per test (monkeypatching the module paths) so tests never
touch the real ``.data`` snapshot."""
from __future__ import annotations

import gzip
import json

import pytest

from app.rbac import cache, compose, demo, export, pivots, schema, scopes
from app.rbac.agent_tool import build_rbac_tools
from app.rbac.store import _privileged_keys


@pytest.fixture()
def isolated_cache(tmp_path, monkeypatch):
    """Point the cache index + blob dir at a tmp location."""
    monkeypatch.setattr(cache, "_DATA", tmp_path)
    monkeypatch.setattr(cache, "_INDEX", tmp_path / "rbac_cache.json")
    monkeypatch.setattr(cache, "_BLOBS", tmp_path / "rbac")
    return tmp_path


# --------------------------------------------------------------------------- schema
def test_make_row_has_all_columns():
    row = schema.make_row(roleName="Owner", roleIsPrivileged=True)
    assert set(row.keys()) == set(schema.COLUMNS)
    assert len(schema.COLUMNS) == 46
    assert row["roleIsPrivileged"] is True
    assert row["roleHasDataActions"] is False  # boolean default
    assert row["principalId"] == ""  # string default
    # None coerces to "" for string columns
    assert schema.make_row(principalId=None)["principalId"] == ""


def test_parse_scope_levels():
    assert schema.parse_scope("/")["scopeType"] == schema.SCOPE_TENANT
    mg = schema.parse_scope("/providers/Microsoft.Management/managementGroups/contoso")
    assert mg["scopeType"] == schema.SCOPE_MANAGEMENT_GROUP and mg["managementGroupId"] == "contoso"
    sub = schema.parse_scope("/subscriptions/abc")
    assert sub["scopeType"] == schema.SCOPE_SUBSCRIPTION and sub["subscriptionId"] == "abc"
    rg = schema.parse_scope("/subscriptions/abc/resourceGroups/rg1")
    assert rg["scopeType"] == schema.SCOPE_RESOURCE_GROUP and rg["resourceGroup"] == "rg1"
    res = schema.parse_scope("/subscriptions/abc/resourceGroups/rg1/providers/Microsoft.Storage/storageAccounts/sa1")
    assert res["scopeType"] == schema.SCOPE_RESOURCE
    assert res["resourceType"] == "Microsoft.Storage/storageAccounts" and res["resourceName"] == "sa1"


def test_role_privilege_classification():
    assert schema.role_is_privileged("Owner") is True
    assert schema.role_is_privileged("Contributor") is True
    assert schema.role_is_privileged("Reader") is False
    # Data-plane owner/contributor roles are privileged on the data path.
    assert schema.role_is_privileged("Storage Blob Data Contributor", has_data_actions=True) is True
    assert schema.role_is_privileged("Storage Blob Data Reader", has_data_actions=True) is False
    # Entra surface uses the directory-role set.
    assert schema.role_is_privileged("Global Administrator", surface=schema.SURFACE_ENTRA) is True
    assert schema.role_is_privileged("Owner", surface=schema.SURFACE_ENTRA) is False
    assert schema.role_category(True) == "DataPlane"
    assert schema.role_category(False) == "ControlPlane"


# --------------------------------------------------------------------------- cache
def test_cache_scope_roundtrip_gzip(isolated_cache):
    rows = [schema.make_row(principalId="u1", roleName="Owner")]
    meta = {"scopeType": "subscription", "displayName": "Prod", "demo": True}
    entry = cache.write_scope("t1", "/subscriptions/s1", meta=meta, rows=rows)
    assert entry["row_count"] == 1 and entry["rows_ref"]
    # Sidecar is gzip-compressed JSON.
    blob = cache._blob_path("t1", "/subscriptions/s1")
    assert blob.exists()
    decoded = json.loads(gzip.decompress(blob.read_bytes()).decode())
    assert decoded["rows"][0]["roleName"] == "Owner"
    # Read back via the API.
    assert cache.read_scope_rows("t1", "/subscriptions/s1")[0]["principalId"] == "u1"
    assert len(cache.list_scope_meta("t1")) == 1
    assert len(cache.all_scope_rows("t1")) == 1
    assert cache.is_demo("t1") is True


def test_cache_directory_and_delete(isolated_cache):
    cache.write_directory(
        "t1",
        meta={"status": "Succeeded"},
        rows=[schema.make_row(principalId="u1")],
        role_defs=[{"roleName": "Owner"}],
        principals=[{"principalId": "u1"}],
        groups={"g1": {"name": "G", "members": []}},
    )
    d = cache.read_directory("t1")
    assert len(d["rows"]) == 1 and len(d["role_defs"]) == 1 and "g1" in d["groups"]
    meta = cache.read_directory_meta("t1")
    assert meta["role_def_count"] == 1 and meta["group_count"] == 1
    # delete a scope + whole tenant
    cache.write_scope("t1", "/subscriptions/s1", meta={}, rows=[])
    assert cache.delete_scope("t1", "/subscriptions/s1") is True
    assert cache.has_any("t1") is True  # directory still present
    assert cache.delete_tenant("t1") >= 0
    assert cache.has_any("t1") is False


def test_purge_demo_preserves_real_scans(isolated_cache):
    # A real scan slice + directory (no demo flag) alongside demo-flagged ones.
    cache.write_scope("t1", "/subscriptions/real", meta={"demo": False, "displayName": "Real"}, rows=[schema.make_row(principalId="r1")])
    cache.write_scope("t1", "/subscriptions/demo", meta={"demo": True, "displayName": "Demo"}, rows=[schema.make_row(principalId="d1")])
    cache.write_directory("t1", meta={"status": "Succeeded", "demo": False}, rows=[schema.make_row(principalId="r1")])

    removed = cache.purge_demo("t1")
    assert removed == 1  # only the demo scope dropped
    assert cache.read_scope_meta("t1", "/subscriptions/real") is not None  # real scan preserved
    assert cache.read_scope_meta("t1", "/subscriptions/demo") is None
    assert cache.read_directory_meta("t1")  # a non-demo directory is kept
    assert cache.is_demo("t1") is False

    # When the directory itself is demo, purge clears it too.
    cache.write_directory("t1", meta={"status": "Succeeded", "demo": True}, rows=[])
    cache.purge_demo("t1")
    assert cache.read_directory_meta("t1") == {}
    assert cache.read_scope_meta("t1", "/subscriptions/real") is not None  # still preserved


def test_cache_age_helpers():
    from datetime import datetime, timedelta, timezone

    assert cache.age_seconds(None) is None
    assert cache.age_seconds("not-a-date") is None
    past = (datetime.now(timezone.utc) - timedelta(seconds=120)).isoformat()
    age = cache.age_seconds(past)
    assert age is not None and 110 < age < 130
    assert cache.is_fresh(past, 3600) is True
    assert cache.is_fresh(past, 60) is False


# --------------------------------------------------------------------------- demo + compose
def test_demo_seed_and_master_rows(isolated_cache):
    summary = demo.seed_demo("t1")
    assert summary["scopes"] == 3 and summary["directory_rows"] == 4
    master = compose.build_master_rows("t1")
    # 13 direct scope rows + 4 directory rows + 4 group-expanded = 21
    assert len(master) == 21
    assert all(set(r.keys()) == set(schema.COLUMNS) for r in master)
    group_rows = [r for r in master if r["accessPath"] == schema.PATH_GROUP]
    assert len(group_rows) == 4  # 2 groups x 2 members
    # group-derived rows carry the member as effective principal + the group as source
    assert all(r["effectivePrincipalId"] and r["sourceGroupName"] for r in group_rows)
    owners = [r for r in master if r["accessPath"] == schema.PATH_OWNER]
    assert len(owners) == 1
    eligible = [r for r in master if r["assignmentState"] == schema.STATE_ELIGIBLE]
    assert len(eligible) == 1


def test_expand_group_rows_only_known_groups():
    groups = {"g1": {"name": "Admins", "members": [{"principalId": "m1", "principalType": "User", "principalDisplayName": "M1"}]}}
    scope_rows = [
        schema.make_row(principalId="g1", principalType="Group", roleName="Owner", scope="/subscriptions/s"),
        schema.make_row(principalId="g-unknown", principalType="Group", roleName="Reader"),
        schema.make_row(principalId="u1", principalType="User", roleName="Reader"),
    ]
    expanded = compose.expand_group_rows(scope_rows, groups)
    assert len(expanded) == 1  # only g1 is known; the user row and unknown group are not expanded
    assert expanded[0]["effectivePrincipalId"] == "m1"
    assert expanded[0]["accessPath"] == schema.PATH_GROUP


def test_compute_overview_kpis(isolated_cache):
    demo.seed_demo("t1")
    ov = compose.compute_overview("t1")
    k = ov["kpis"]
    assert k["total_assignments"] == 21
    assert k["privileged"] >= 1 and k["owners"] == 1 and k["group_derived"] == 4
    assert k["subscriptions"] == 2 and k["scopes"] == 3
    assert ov["group_severity"]["privileged"] == "error"
    assert ov["never_loaded"] is False and ov["demo"] is True
    # the Prod scope has an Unauthorized reservation collector → attention count
    prod = next(s for s in ov["scopes"] if s["subscriptionId"] == demo.SUB_PROD)
    assert prod["collectors_attention"] >= 1


def test_overview_never_loaded(isolated_cache):
    ov = compose.compute_overview("empty-tenant")
    assert ov["never_loaded"] is True and ov["kpis"]["total_assignments"] == 0


# --------------------------------------------------------------------------- pivots
def test_pivots_thirteen_sections(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    piv = pivots.compute_pivots(master)
    assert len(piv) == 13
    assert {"label": "Owner", "count": 5} in piv["by_role"]
    # PIM eligible vs active counts
    pim = {d["label"]: d["count"] for d in piv["pim_eligible_vs_active"]}
    assert pim["Eligible"] == 1 and pim["Active"] >= 1
    # privileged-by-principal only counts privileged grants
    assert all(isinstance(d["count"], int) for d in piv["privileged_by_principal"])


# --------------------------------------------------------------------------- export
def test_export_csv_and_json(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    csv_text = export.to_csv(master)
    header = csv_text.splitlines()[0]
    assert header.split(",")[0] == "surface" and "errorMessage" in header
    assert len(header.split(",")) == 46
    parsed = json.loads(export.to_json(master))
    assert len(parsed) == len(master) and set(parsed[0].keys()) == set(schema.COLUMNS)


# --------------------------------------------------------------------------- name resolution
def test_principal_index_and_apply_names():
    directory = {
        "principals": [
            {"principalId": "u-1", "displayName": "Alice Admin", "userPrincipalName": "alice@x", "principalType": "User"},
            {"principalId": "sp-1", "displayName": "deploy-sp", "appId": "app-9", "principalType": "ServicePrincipal"},
            {"principalId": "g-1", "displayName": "Platform Admins", "principalType": "Group"},
        ],
        "groups": {"g-1": {"name": "Platform Admins", "members": [{"principalId": "u-2", "principalDisplayName": "Bob", "principalType": "User"}]}},
        "rows": [],
    }
    idx = compose._principal_index(directory, [])
    assert idx["u-1"]["name"] == "Alice Admin" and idx["u-1"]["upn"] == "alice@x"
    assert idx["sp-1"]["name"] == "deploy-sp" and idx["sp-1"]["type"] == "ServicePrincipal"
    assert idx["u-2"]["name"] == "Bob"  # from group members

    rows = [
        schema.make_row(principalId="u-1", effectivePrincipalId="u-1", roleName="Owner"),
        schema.make_row(principalId="sp-1", effectivePrincipalId="sp-1", roleName="Contributor"),
        schema.make_row(principalId="unknown", effectivePrincipalId="unknown", principalDisplayName="", roleName="Reader"),
    ]
    compose._apply_names(rows, idx)
    assert rows[0]["principalDisplayName"] == "Alice Admin" and rows[0]["effectivePrincipalName"] == "Alice Admin"
    assert rows[0]["principalUserPrincipalName"] == "alice@x"
    assert rows[1]["principalDisplayName"] == "deploy-sp" and rows[1]["effectivePrincipalType"] == "ServicePrincipal"
    # An unresolved GUID is left untouched (stays a GUID for the UI to show).
    assert rows[2]["principalDisplayName"] == ""


def test_build_master_rows_backfills_guid_names(isolated_cache):
    # A scope assignment carrying ONLY a principal GUID (the ARM reality), plus a directory that
    # resolves that GUID to a name — build_master_rows must backfill it.
    cache.write_scope(
        "t1",
        "/subscriptions/s1",
        meta={"scopeType": "subscription", "displayName": "Prod"},
        rows=[schema.make_row(surface=schema.SURFACE_AZURE_RBAC, principalId="guid-x", effectivePrincipalId="guid-x", principalType="User", roleName="Owner", scope="/subscriptions/s1")],
    )
    cache.write_directory(
        "t1",
        meta={"status": "Succeeded"},
        rows=[],
        principals=[{"principalId": "guid-x", "displayName": "Xavier User", "userPrincipalName": "xavier@x", "principalType": "User"}],
        groups={},
    )
    master = compose.build_master_rows("t1")
    row = next(r for r in master if r.get("principalId") == "guid-x")
    assert row["principalDisplayName"] == "Xavier User"
    assert row["effectivePrincipalName"] == "Xavier User"


def test_normalize_principal_type():
    from app.rbac.collectors import _normalize_principal_type

    assert _normalize_principal_type("#microsoft.graph.servicePrincipal") == "ServicePrincipal"
    assert _normalize_principal_type("#microsoft.graph.user") == "User"
    assert _normalize_principal_type("#microsoft.graph.group") == "Group"
    assert _normalize_principal_type("") == ""


# --------------------------------------------------------------------------- xlsx workbook
def test_to_workbook_multi_sheet(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    overview = compose.compute_overview("t1")
    piv = pivots.compute_pivots(master)
    directory = cache.read_directory("t1")
    content = export.to_workbook(
        rows=master, overview=overview, pivots=piv, pivot_labels=pivots.PIVOT_LABELS, directory=directory
    )
    assert isinstance(content, bytes) and content[:2] == b"PK"  # xlsx is a zip
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True)
    names = set(wb.sheetnames)
    assert {"Summary", "Effective Access", "Privileged", "Insights", "Diagnostics", "Scopes", "Principals"} <= names
    # The Effective Access sheet has a header row + one row per master row.
    ws = wb["Effective Access"]
    assert ws.max_row == len(master) + 1
    wb.close()



    rows = [
        schema.make_row(effectivePrincipalId="u1", roleName="Owner", roleIsPrivileged=True, scope="/subscriptions/s"),
        schema.make_row(effectivePrincipalId="u2", roleName="Reader", roleIsPrivileged=False, scope="/subscriptions/s"),
    ]
    keys = _privileged_keys(rows)
    assert keys == ["u1|Owner|/subscriptions/s"]


# --------------------------------------------------------------------------- agent tools
async def test_agent_tools_read_cache(isolated_cache):
    demo.seed_demo("t1")
    tools = {t.name: t for t in build_rbac_tools("t1")}
    assert set(tools) == {"who_can_access", "privileged_access_review", "effective_access_for_principal"}
    assert all(t.kind == "read" for t in tools.values())

    # who_can_access filters by scope substring
    res = await tools["who_can_access"].handler({}, {"scope": "Production"})
    assert res["isError"] is False and "Production" in res["content"][0] or "grant" in res["content"][0]

    # privileged_access_review summarizes privileged grants incl group + owner
    res = await tools["privileged_access_review"].handler({}, {})
    assert res["isError"] is False and "privileged grant" in res["content"][0]

    # effective_access_for_principal resolves a group member's transitive access
    res = await tools["effective_access_for_principal"].handler({}, {"principal": "Eve"})
    assert res["isError"] is False and "Eve" in res["content"][0]
    # missing principal arg errors
    res = await tools["effective_access_for_principal"].handler({}, {})
    assert res["isError"] is True


async def test_who_can_access_privileged_only(isolated_cache):
    demo.seed_demo("t1")
    tools = {t.name: t for t in build_rbac_tools("t1")}
    res = await tools["who_can_access"].handler({}, {"scope": "", "privileged_only": True})
    assert res["isError"] is False
    assert "Reader" not in res["content"][0]  # non-privileged filtered out


# --------------------------------------------------------------------------- scope tree + filtering
def test_build_scope_tree_demo(isolated_cache):
    demo.seed_demo("t1")
    tree = scopes.build_scope_tree("t1")
    assert tree["demo"] is True
    assert tree["subscription_count"] == 2 and tree["mg_count"] == 1
    root = tree["root"]
    assert root["type"] == "root" and root["count"] == 21
    # Single management group → both subscriptions nested under it (inferred).
    assert len(root["children"]) == 1
    mg = root["children"][0]
    assert mg["type"] == "managementGroup" and mg["inferred"] is True
    assert mg["count"] == 17  # 1 MG-level grant + 13 (prod) + 3 (dev)
    sub_ids = {c["id"]: c for c in mg["children"]}
    assert scopes.sub_scope_id(demo.SUB_PROD) in sub_ids
    prod = sub_ids[scopes.sub_scope_id(demo.SUB_PROD)]
    dev = sub_ids[scopes.sub_scope_id(demo.SUB_DEV)]
    assert prod["count"] == 13 and dev["count"] == 3  # prod incl 4 group-expanded rows
    assert set(mg["subscriptionIds"]) == {demo.SUB_PROD, demo.SUB_DEV}


def test_build_scope_tree_empty(isolated_cache):
    tree = scopes.build_scope_tree("empty")
    assert tree["root"]["children"] == [] and tree["root"]["count"] == 0
    assert tree["subscription_count"] == 0 and tree["mg_count"] == 0


def test_scope_tree_resolves_mg_name_from_directory(isolated_cache):
    # A subscription scope whose only MG reference is an INHERITED assignment carrying the MG
    # GUID but no name — the resolved management_groups map must name the MG node.
    mg_id = "mg-guid-001"
    cache.write_scope(
        "t1",
        "/subscriptions/s1",
        meta={"scopeType": "subscription", "displayName": "Prod"},
        rows=[
            schema.make_row(principalId="u1", roleName="Owner", scope="/subscriptions/s1", subscriptionId="s1", subscriptionName="Prod"),
            # Inherited MG-level assignment: managementGroupId set, name empty.
            schema.make_row(
                principalId="u2", roleName="Reader", scopeType=schema.SCOPE_MANAGEMENT_GROUP,
                scope=f"/providers/Microsoft.Management/managementGroups/{mg_id}", managementGroupId=mg_id,
            ),
        ],
    )
    # Before resolution: the MG node falls back to its GUID.
    tree = scopes.build_scope_tree("t1")
    mg_node = next(c for c in tree["root"]["children"] if c["type"] == "managementGroup")
    assert mg_node["name"] == mg_id

    # After a directory refresh populates the id→name map: the node shows the friendly name.
    cache.write_directory("t1", meta={"status": "Succeeded"}, rows=[], principals=[], groups={}, management_groups={mg_id: "Corp Platform"})
    tree2 = scopes.build_scope_tree("t1")
    mg_node2 = next(c for c in tree2["root"]["children"] if c["type"] == "managementGroup")
    assert mg_node2["name"] == "Corp Platform"
    # And MG-scoped rows get the name backfilled too.
    master = compose.build_master_rows("t1")
    mg_row = next(r for r in master if r.get("managementGroupId") == mg_id)
    assert mg_row["managementGroupName"] == "Corp Platform"
    assert mg_row["scopeDisplayName"] == "Corp Platform"


async def test_filter_rows_by_subscription(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    # By subscription scope prefix (Dev): only the 3 Dev subscription-level grants.
    dev = await scopes.filter_rows(master, scope_id=scopes.sub_scope_id(demo.SUB_DEV))
    assert len(dev) == 3 and all(r["subscriptionId"] == demo.SUB_DEV for r in dev)
    # By subscription id list (Prod): 9 direct + 4 group-expanded = 13.
    prod = await scopes.filter_rows(master, subscription_ids=[demo.SUB_PROD])
    assert len(prod) == 13 and all(r["subscriptionId"] == demo.SUB_PROD for r in prod)
    # Directory (Entra/owner) rows have no subscription → excluded by a scope filter.
    assert all(r["surface"] != schema.SURFACE_ENTRA for r in prod)


async def test_filter_rows_by_management_group(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    # MG selection sends the MG scope (inherited rows) + its descendant subscription ids.
    rows = await scopes.filter_rows(
        master,
        scope_id=scopes.mg_scope_id(demo.MG_ID),
        subscription_ids=[demo.SUB_PROD, demo.SUB_DEV],
    )
    assert len(rows) == 17  # 1 MG-level + 13 prod + 3 dev
    # The management-group inherited grant is included.
    assert any(schema.SCOPE_MANAGEMENT_GROUP == r["scopeType"] for r in rows)


async def test_filter_rows_no_filter_is_identity(isolated_cache):
    demo.seed_demo("t1")
    master = compose.build_master_rows("t1")
    same = await scopes.filter_rows(master)
    assert same is master  # no scope/workload → returned unchanged


def test_row_in_workload_matcher():
    f = {
        "direct_subs": {"sub1"},
        "effective_subs": {"sub1", "sub2"},
        "rg_pairs": {("sub2", "rg-a")},
        "resource_ids": ["/subscriptions/sub3/resourcegroups/rg/providers/microsoft.storage/storageaccounts/sa"],
        "name": "wl",
    }
    # Whole-subscription membership.
    assert scopes._row_in_workload(schema.make_row(subscriptionId="sub1", scope="/subscriptions/sub1/resourceGroups/x"), f)
    # Resource-group membership.
    assert scopes._row_in_workload(schema.make_row(subscriptionId="sub2", resourceGroup="rg-a", scope="/subscriptions/sub2/resourceGroups/rg-a"), f)
    # Subscription-level inherited grant for an effective subscription.
    assert scopes._row_in_workload(schema.make_row(subscriptionId="sub2", scopeType=schema.SCOPE_SUBSCRIPTION, scope="/subscriptions/sub2"), f)
    # Exact resource membership.
    assert scopes._row_in_workload(schema.make_row(scope=f["resource_ids"][0]), f)
    # A different resource group in an effective subscription is NOT included.
    assert not scopes._row_in_workload(schema.make_row(subscriptionId="sub2", resourceGroup="rg-b", scopeType=schema.SCOPE_RESOURCE_GROUP, scope="/subscriptions/sub2/resourceGroups/rg-b"), f)
    # Unrelated subscription is excluded.
    assert not scopes._row_in_workload(schema.make_row(subscriptionId="sub9", scope="/subscriptions/sub9"), f)


async def test_resolve_workload_filter_unknown_returns_none(isolated_cache):
    assert await scopes.resolve_workload_filter("no-such-workload", None) is None

