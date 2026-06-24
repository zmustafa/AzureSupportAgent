"""Tests for the FMEA feature: RPN scoring, registry CRUD + Trash, revisions, prune-guard,
and the AI generator (stubbed LLM)."""
import asyncio

import pytest

from app.agent.provider import StreamEvent
from app.fmea import compute
from app.fmea import generator as fgen
from app.fmea import registry as freg
from app.fmea import revisions as frev


@pytest.fixture()
def _isolate(monkeypatch, tmp_path):
    """Point the FMEA + revision stores at a temp dir so tests never touch real ``.data``."""
    monkeypatch.setattr(freg, "_PATH", tmp_path / "fmea.json")
    monkeypatch.setattr(frev, "_PATH", tmp_path / "fmea_revisions.json")


# ----------------------------------------------------------------- compute (RPN scoring)
def test_normalize_factor_clamps_to_1_10():
    assert compute.normalize_factor(0) == 0
    assert compute.normalize_factor("") == 0
    assert compute.normalize_factor(None) == 0
    assert compute.normalize_factor("abc") == 0
    assert compute.normalize_factor(5) == 5
    assert compute.normalize_factor(11) == 10
    assert compute.normalize_factor(-3) == 0
    assert compute.normalize_factor("7") == 7
    assert compute.normalize_factor(4.6) == 5


def test_rpn_is_product_or_none():
    assert compute.rpn(10, 10, 10) == 1000
    assert compute.rpn(2, 3, 4) == 24
    # Any unset factor → no RPN (a half-filled row never shows a misleading number).
    assert compute.rpn(5, 5, 0) is None
    assert compute.rpn(0, 5, 5) is None
    assert compute.rpn(5, None, 5) is None


def test_risk_band_thresholds():
    assert compute.risk_band(None) == "none"
    assert compute.risk_band(1000) == "critical"
    assert compute.risk_band(200) == "critical"
    assert compute.risk_band(199) == "high"
    assert compute.risk_band(120) == "high"
    assert compute.risk_band(119) == "medium"
    assert compute.risk_band(40) == "medium"
    assert compute.risk_band(39) == "low"
    assert compute.risk_band(1) == "low"


def test_recompute_row_sets_rpn_and_bands():
    row = {"severity": 10, "occurrence": 10, "detection": 10,
           "severity_post": 2, "occurrence_post": 2, "detection_post": 2}
    compute.recompute_row(row)
    assert row["rpn"] == 1000
    assert row["risk_band"] == "critical"
    assert row["rpn_post"] == 8
    assert row["risk_band_post"] == "low"


def test_summarize_counts_and_mitigation():
    doc = {"tables": [
        {"rows": [
            {"severity": 10, "occurrence": 10, "detection": 10,
             "severity_post": 2, "occurrence_post": 2, "detection_post": 2,
             "recommended_actions": "fix it"},
            {"severity": 5, "occurrence": 2, "detection": 2, "recommended_actions": "do later"},
            {"severity": 1, "occurrence": 1, "detection": 1},
        ]},
    ]}
    s = compute.summarize(doc)
    assert s["total_rows"] == 3
    assert s["scored_rows"] == 3
    assert s["top_rpn"] == 1000
    assert s["counts"]["critical"] == 1
    assert s["counts"]["low"] >= 1
    assert s["mitigated_rows"] == 1
    # Row 2 has an action but no post-RPN → an open action.
    assert s["open_actions"] == 1


# ----------------------------------------------------------------- registry CRUD + Trash
def test_create_and_get(_isolate):
    doc = freg.create_fmea(architecture_id="arch-1", workload_name="Shop", tenant_id="t1", actor="alice")
    assert doc["status"] == "draft"
    assert doc["title"] == "FMEA — Shop"
    got = freg.get_fmea(doc["id"])
    assert got is not None and got["id"] == doc["id"]


def test_update_recomputes_rpn_server_side(_isolate):
    doc = freg.create_fmea(architecture_id="arch-1", tenant_id="t1", actor="alice")
    saved = freg.update_fmea(
        doc["id"], tenant_id="t1", actor="alice",
        tables=[{
            "name": "Data tier",
            "rows": [{
                "item": "SQL DB", "failure_mode": "Compute quota exhaustion",
                "severity": 9, "occurrence": 4, "detection": 6,
                # A bogus client-supplied rpn must be ignored/overwritten.
                "rpn": 1,
            }],
        }],
    )
    row = saved["tables"][0]["rows"][0]
    assert row["rpn"] == 9 * 4 * 6
    assert row["risk_band"] == "critical"


def test_soft_delete_restore_purge(_isolate):
    doc = freg.create_fmea(architecture_id="arch-1", tenant_id="t1", actor="alice")
    fid = doc["id"]
    assert freg.soft_delete(fid, "alice") is True
    assert freg.get_fmea(fid)["deleted_at"]  # still present, flagged
    # Hidden from the default list, visible in only_deleted.
    assert all(d["id"] != fid for d in freg.list_fmea("t1"))
    assert any(d["id"] == fid for d in freg.list_fmea("t1", only_deleted=True))
    assert freg.restore(fid) is not None
    assert not freg.get_fmea(fid)["deleted_at"]
    assert freg.purge(fid) is True
    assert freg.get_fmea(fid) is None


def test_empty_trash(_isolate):
    a = freg.create_fmea(architecture_id="arch-1", tenant_id="t1", actor="x")
    b = freg.create_fmea(architecture_id="arch-2", tenant_id="t1", actor="x")
    freg.soft_delete(a["id"])
    freg.soft_delete(b["id"])
    assert freg.empty_trash("t1") == 2
    assert freg.list_fmea("t1", only_deleted=True) == []


def test_revisions_snapshot_and_restore(_isolate):
    doc = freg.create_fmea(architecture_id="arch-1", tenant_id="t1", actor="alice")
    fid = doc["id"]
    freg.update_fmea(fid, tenant_id="t1", actor="alice", title="V1",
                     tables=[{"name": "T", "rows": [{"item": "A", "severity": 1, "occurrence": 1, "detection": 1}]}])
    freg.update_fmea(fid, tenant_id="t1", actor="alice", title="V2",
                     tables=[{"name": "T", "rows": [{"item": "B", "severity": 9, "occurrence": 9, "detection": 9}]}])
    revs = frev.list_revisions(fid)
    assert len(revs) >= 2
    # Restore the oldest revision (title V1) and confirm content reverts.
    oldest = revs[-1]
    restored = freg.restore_revision(fid, oldest["id"], "alice")
    assert restored["title"] == "V1"
    assert restored["tables"][0]["rows"][0]["item"] == "A"


def test_prune_orphans_guard(_isolate):
    """Pruning with the real architecture set drops orphans; the endpoint guards empty sets."""
    freg.create_fmea(architecture_id="arch-keep", tenant_id="t1", actor="x")
    orphan = freg.create_fmea(architecture_id="arch-gone", tenant_id="t1", actor="x")
    pruned = freg.prune_orphans({"arch-keep"})
    assert pruned == 1
    assert freg.get_fmea(orphan["id"]) is None


# ----------------------------------------------------------------- generator (stubbed LLM)
class _FakeProvider:
    def __init__(self, payload: str):
        self._payload = payload

    async def stream(self, messages, tools=None, max_tokens=None):
        # Emit the whole JSON payload as a single token event.
        yield StreamEvent(type="token", text=self._payload)


def test_generator_parses_tables_and_owner_is_todo(_isolate, monkeypatch):
    payload = (
        '{"tables": [{"name": "Ingress", "scope_ref": "rg-edge", "rows": ['
        '{"item": "Front Door", "function": "Global ingress", "failure_mode": "Origin unreachable",'
        ' "effects": "Site down", "causes": "Backend health probe fails",'
        ' "control_prevention": "Multi-origin", "control_detection": "FD health alert",'
        ' "recommended_actions": "Add second origin",'
        ' "owner": "\u27e6TODO: Owner | key=owner\u27e7", "date_due": "\u27e6TODO: Target date | key=date_due\u27e7",'
        ' "severity": 9, "occurrence": 3, "detection": 4}]}], "confidence": 0.8}'
    )
    monkeypatch.setattr(fgen, "build_provider", lambda: _FakeProvider(payload))
    result = asyncio.run(fgen.generate_fmea(
        workload_name="Shop", memory={"sections": [{"key": "overview", "content": "A shop."}]},
        facts={"subscriptions": [], "resource_groups": [], "regions": [], "resources": []},
        two_pass=False,
    ))
    assert result is not None
    table = result["tables"][0]
    assert table["name"] == "Ingress"
    row = table["rows"][0]
    assert "TODO" in row["owner"]
    assert row["severity"] == 9


def test_generator_returns_none_on_garbage(_isolate, monkeypatch):
    monkeypatch.setattr(fgen, "build_provider", lambda: _FakeProvider("not json at all"))
    result = asyncio.run(fgen.generate_fmea(
        workload_name="Shop", memory={"sections": [{"key": "overview", "content": "x"}]},
        facts={"subscriptions": [], "resource_groups": [], "regions": [], "resources": []},
        two_pass=False,
    ))
    assert result is None


def test_no_banned_words_in_prompt():
    """Guard against reintroducing banned tokens in the system prompt."""
    banned = ("SfMC", "CSAM", "DocuSign")
    for word in banned:
        assert word not in fgen.SYSTEM_PROMPT


# ----------------------------------------------------------------- Excel export
def test_build_fmea_xlsx_is_valid_workbook():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.fmea import excel as fexcel

    doc = {
        "title": "FMEA — Shop",
        "status": "draft",
        "tables": [{
            "id": "t1", "name": "Data tier", "scope_ref": "rg-data",
            "rows": [{
                "id": "r1", "item": "SQL DB", "function": "OLTP",
                "failure_mode": "Quota exhaustion", "effects": "Outage",
                "causes": "Spike", "control_prevention": "Autoscale",
                "control_detection": "Alert", "recommended_actions": "Add capacity",
                "owner": "⟦TODO: Owner | key=owner⟧", "date_due": "⟦TODO: date⟧",
                "action_results": "", "date_completed": "2026-01-15",
                "severity": 9, "occurrence": 4, "detection": 6,
                "severity_post": 0, "occurrence_post": 0, "detection_post": 0,
            }],
        }],
    }
    data = fexcel.build_fmea_xlsx(doc, "Shop")
    assert data[:2] == b"PK"  # xlsx is a zip
    wb = load_workbook(BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Data tier" in wb.sheetnames
    ws = wb["Data tier"]
    # The RPN column (L) carries a live formula, not a literal number, and the TODO owner
    # token was stripped to blank.
    formula_cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    assert any("*" in f for f in formula_cells), "expected an RPN multiplication formula"
    all_text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "TODO" not in all_text


def test_build_fmea_xlsx_handles_empty_tables():
    from app.fmea import excel as fexcel

    data = fexcel.build_fmea_xlsx({"title": "Empty", "tables": []}, "WL")
    assert data[:2] == b"PK"

