"""The Investigate API handlers, in-process, over the demo snapshot.

Focused on the question the screen could not answer until this landed: which groups does
this user belong to. The demo tenant is the right fixture because it carries both cached
membership sources — a role-granting group (`grp-admins` -> bob, carol) and a
Conditional-Access-targeted one (`grp-ca-exclude` -> carol) — so the union, the source
tagging and the ordering are all exercised on data that was not written for this test.

Calling handlers directly bypasses FastAPI's dependency resolution, so every ``Query``
default has to be passed explicitly or the parameter arrives as a ``Query`` object.
"""
from __future__ import annotations

import asyncio

import pytest

from app.api import entra as entra_api
from app.entra import cache, demo, investigate
from app.entra import investigate_members as inv_members
from app.entra import snapshot as snapshot_mod


class _FakeDB:
    def __init__(self) -> None:
        self.rows: list[object] = []

    def add(self, row: object) -> None:
        self.rows.append(row)

    async def commit(self) -> None:
        pass


class _Principal:
    tenant_id = demo.DEMO_TENANT
    subject = "dev"


@pytest.fixture(autouse=True)
def _demo_tenant(tmp_path, monkeypatch):
    cache.set_root_for_tests(tmp_path / "entra")
    snapshot_mod._analysis_memo.clear()  # noqa: SLF001 - test isolation

    import app.core.azure_connections as ac

    monkeypatch.setattr(
        ac, "resolve_connection",
        lambda cid: {"id": "conn-demo", "tenant_id": demo.DEMO_TENANT} if cid == "conn-demo" else None,
    )
    # The IAM directory belongs to another product surface and is absent here. Absent is a
    # legitimate state: the other two sources must carry the answer on their own.
    from app.iam import cache as iam_cache

    monkeypatch.setattr(iam_cache, "read_directory", lambda _t: {"groups": {}})
    demo.seed()
    yield
    cache.clear_memo()


def _run(coro):
    return asyncio.run(coro)


def _dossier(principal_id: str):
    return _run(entra_api.investigate_dossier(
        principal_id=principal_id, connection_id="conn-demo",
        principal=_Principal(), db=_FakeDB()))


def _members(principal_id: str, **kw):
    body = entra_api.InvestigateMembersBody(**kw)
    return _run(entra_api.investigate_members(
        principal_id=principal_id, body=body, connection_id="conn-demo",
        principal=_Principal(), db=_FakeDB()))


# --------------------------------------------------------------------------- capabilities
def test_a_user_can_have_group_memberships():
    """The defect this feature fixes: a user dossier had no way to say which groups the
    user was in, and the capability list is what the screen renders from."""
    assert investigate.CAP_MEMBERSHIPS in _dossier("u-carol")["capabilities"]


def test_every_kind_that_can_be_in_a_group_reports_the_capability():
    for pid in ("u-carol", "g-partner1", "grp-eng"):
        caps = _dossier(pid)["capabilities"]
        assert investigate.CAP_MEMBERSHIPS in caps, pid


def test_a_group_still_reports_both_directions():
    caps = _dossier("grp-admins")["capabilities"]
    assert investigate.CAP_MEMBERS in caps
    assert investigate.CAP_MEMBERSHIPS in caps


# --------------------------------------------------------------------------- the section
def test_a_users_groups_are_listed_with_why_each_one_matters():
    section = _dossier("u-carol")["sections"]["memberships"]
    data = section["data"]
    assert data["readable"] is True
    by_id = {g["id"]: g for g in data["groups"]}
    assert set(by_id) == {"grp-admins", "grp-ca-exclude"}
    assert by_id["grp-admins"]["sources"] == [inv_members.SOURCE_DIRECTORY_ROLE]
    assert by_id["grp-ca-exclude"]["sources"] == [inv_members.SOURCE_CA_TARGET]


def test_the_group_name_is_resolved_not_left_as_a_guid():
    groups = _dossier("u-carol")["sections"]["memberships"]["data"]["groups"]
    assert {g["display_name"] for g in groups} == {"Tenant Admins", "CA Exclusions"}


def test_a_role_assignable_group_is_flagged_and_leads():
    data = _dossier("u-carol")["sections"]["memberships"]["data"]
    assert data["groups"][0]["id"] == "grp-admins"
    assert data["groups"][0]["role_assignable"] is True
    assert data["role_assignable_count"] == 1


def test_the_count_is_always_published_as_a_floor():
    """Nobody expands a whole directory. A short list that reads as a complete one is how
    somebody concludes 'this account is in no privileged group'."""
    section = _dossier("u-bob")["sections"]["memberships"]
    assert "floor, not the complete list" in section["provenance"]["reason"]


def test_a_user_in_no_expanded_group_is_readable_not_unreadable():
    section = _dossier("u-erin")["sections"]["memberships"]
    assert section["data"]["groups"] == []
    assert section["data"]["readable"] is True
    assert section["provenance"]["unreadable"] is False


def test_the_client_never_has_to_invent_the_wording_for_a_source():
    labels = _dossier("u-carol")["sections"]["memberships"]["data"]["source_labels"]
    assert set(labels) == {inv_members.SOURCE_AZURE_RBAC, inv_members.SOURCE_DIRECTORY_ROLE,
                           inv_members.SOURCE_CA_TARGET}


def test_the_azure_platform_has_no_membership_section():
    """A capability the subject cannot have must be absent, not empty."""
    sections = _dossier("00000000-0000-0000-0000-000000000000")["sections"]
    assert "memberships" not in sections


# ------------------------------------------------------------------------ the live endpoint
def test_the_upward_read_is_no_longer_refused_for_a_user(monkeypatch):
    """The gate this feature lifts. It used to answer every non-group with a sentence, which
    made 'which groups is this user in' unanswerable from the live directory too."""
    seen: dict = {}

    async def _fake_expand(_conn, root_id, *, expand_ids, direction, root_kind, transitive):
        seen.update(root_id=root_id, direction=direction, root_kind=root_kind,
                    transitive=transitive, expand_ids=expand_ids)
        return {"nodes": {root_id: []}, "notes": [], "truncated": False}

    monkeypatch.setattr(inv_members, "expand", _fake_expand)
    out = _members("u-carol", direction="up", transitive=True)
    assert seen == {"root_id": "u-carol", "direction": "up", "root_kind": "user",
                    "transitive": True, "expand_ids": []}
    assert out["notes"] == []


def test_the_kind_reaches_graph_so_the_right_collection_is_read(monkeypatch):
    kinds: list[str] = []

    async def _fake_expand(_conn, root_id, *, expand_ids, direction, root_kind, transitive):  # noqa: ARG001
        kinds.append(root_kind)
        return {"nodes": {}, "notes": [], "truncated": False}

    monkeypatch.setattr(inv_members, "expand", _fake_expand)
    _members("g-partner1", direction="up")
    _members("grp-eng", direction="up")
    assert kinds == ["guest", "group"]


def test_asking_a_user_for_its_members_is_still_answered_with_a_sentence():
    out = _members("u-carol", direction="down")
    assert out["nodes"] == {}
    assert any("only groups have members" in n for n in out["notes"])


def test_transitive_is_ignored_downward(monkeypatch):
    """It is an upward-only distinction; honouring it downward would silently return the
    leaves with the tree shape thrown away."""
    seen: dict = {}

    async def _fake_expand(_conn, root_id, *, expand_ids, direction, root_kind, transitive):  # noqa: ARG001
        seen["transitive"] = transitive
        return {"nodes": {}, "notes": [], "truncated": False}

    monkeypatch.setattr(inv_members, "expand", _fake_expand)
    _members("grp-eng", direction="down", transitive=True)
    assert seen["transitive"] is False


def test_an_unresolvable_principal_is_told_why_rather_than_read(monkeypatch):
    called = False

    async def _fake_expand(*_a, **_kw):
        nonlocal called
        called = True
        return {"nodes": {}, "notes": [], "truncated": False}

    monkeypatch.setattr(inv_members, "expand", _fake_expand)
    out = _members("no-such-principal", direction="up")
    assert called is False
    assert any("not a claim that it belongs to none" in n for n in out["notes"])


def test_the_upward_read_is_audited_with_its_shape(monkeypatch):
    async def _fake_expand(*_a, **_kw):
        return {"nodes": {}, "notes": [], "truncated": False}

    monkeypatch.setattr(inv_members, "expand", _fake_expand)
    db = _FakeDB()
    body = entra_api.InvestigateMembersBody(direction="up", transitive=True)
    _run(entra_api.investigate_members(
        principal_id="u-carol", body=body, connection_id="conn-demo",
        principal=_Principal(), db=db))
    meta = db.rows[0].metadata_json
    assert meta["direction"] == "up"
    assert meta["kind"] == "user"
    assert meta["transitive"] is True


# ------------------------------------------------------------------------------- export
def _export(principal_id: str):
    return _run(entra_api.investigate_export(
        principal_id=principal_id, connection_id="conn-demo",
        principal=_Principal(), db=_FakeDB()))


def _sheets(body: bytes) -> list[str]:
    import io

    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(body), read_only=True).sheetnames


def test_the_export_carries_the_group_memberships_sheet():
    resp = _export("u-carol")
    assert resp.status_code == 200
    names = _sheets(resp.body)
    assert "Group memberships" in names
    # The caveat travels with the data. An auditor reading a two-row sheet must not conclude
    # the account is in two groups.
    assert "Provenance" in names


def test_the_export_still_works_for_a_group_and_carries_both_sheets():
    names = _sheets(_export("grp-admins").body)
    assert "Members" in names
    assert "Group memberships" in names


def test_the_export_of_an_unresolvable_principal_does_not_crash():
    """Its assignments outlive it, and that is usually the finding."""
    resp = _export("00000000-0000-0000-0000-000000000000")
    assert resp.status_code == 200
    assert "Group memberships" not in _sheets(resp.body)
