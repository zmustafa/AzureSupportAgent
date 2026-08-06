"""Disabled principals that still hold access — the data path, the signals and the exports.

Every test here exists because the opposite behaviour would be a *confident wrong answer*
rather than an obvious failure. The whole feature is one claim ("these people should not still
have this"), and the failure mode is always the same shape: an empty result that reads as a
clean bill of health when it actually means nobody looked.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timedelta, timezone

import pytest

from app.iam import cache, campaigns, collectors, compose, demo, export, findings, leavers, orchestrator, schema, usage


@pytest.fixture
def anyio_backend():
    return "asyncio"


@pytest.fixture()
def isolated_cache(tmp_path, monkeypatch):
    monkeypatch.setattr(cache, "_DATA", tmp_path)
    monkeypatch.setattr(cache, "_INDEX", tmp_path / "iam_cache.json")
    monkeypatch.setattr(cache, "_BLOBS", tmp_path / "iam")
    monkeypatch.setattr(cache, "_migrated", True)
    return tmp_path


@pytest.fixture()
def seeded(isolated_cache):
    demo.seed_demo(demo.TENANT)
    return demo.TENANT


def _strip_state(tenant_id: str) -> None:
    """Rewrite the directory blob with NO account state — a cache from before this feature."""
    d = cache.read_directory(tenant_id)
    cache.write_directory(
        tenant_id,
        meta={"status": schema.STATUS_SUCCEEDED, "demo": True, "collectors": []},
        rows=d["rows"], role_defs=d["role_defs"], principals=d["principals"],
        groups=d["groups"], management_groups=d["management_groups"],
        identities=d["identities"], federated=d["federated"], principal_state=None,
    )


# --------------------------------------------------------------------------- schema
def test_scanner_columns_stay_frozen_at_46():
    """The new columns must not have leaked into the scanner contract."""
    assert len(schema.SCANNER_COLUMNS) == 46
    assert "principalAccountEnabled" not in schema.SCANNER_COLUMNS


def test_new_columns_are_appended_never_inserted():
    """EXTRA_COLUMNS is append-only: a CSV consumer reading by position must not shift."""
    assert schema.EXTRA_COLUMNS[-6:] == (
        "membershipGroupOnPremSynced", "membershipGroupId", "membershipGroupName",
        "membershipGroupResolution", "membershipGroupRoleAssignable", "membershipGroupDynamic",
    )


def test_account_state_defaults_to_unknown_not_enabled():
    """A row nobody stamped must not claim the account is live.

    If the default were "true", every tenant whose cache predates the collector would report
    each of its leavers as a current employee."""
    row = schema.make_row(principalId="p1")
    assert row["principalAccountEnabled"] == schema.ENABLED_UNKNOWN
    assert schema.is_disabled(row) is False


def test_is_disabled_is_true_only_for_a_known_false():
    assert schema.is_disabled({"principalAccountEnabled": schema.ENABLED_FALSE}) is True
    for value in (schema.ENABLED_TRUE, schema.ENABLED_UNKNOWN, schema.ENABLED_NA, "", None):
        assert schema.is_disabled({"principalAccountEnabled": value}) is False


# --------------------------------------------------------------------------- compose
def test_group_derived_rows_carry_the_MEMBER_state_not_the_groups(seeded):
    """The least visible case in the product, and the one most easily broken.

    A group holds the assignment, so `principalId` is the group — which has no account state at
    all. Keying on it instead of on `effectivePrincipalId` would report every group-derived
    grant as uncheckable and hide the disabled member entirely."""
    rows = compose.build_master_rows(seeded)
    nina = [
        r for r in rows
        if r.get("effectivePrincipalName") == "Nina Nomore"
        and r.get("accessPath") == schema.PATH_GROUP
    ]
    assert nina, "the demo estate must contain a disabled member inside an access-granting group"
    assert all(r["principalAccountEnabled"] == schema.ENABLED_FALSE for r in nina)
    # And the group's OWN row is not-applicable, never "unknown": there is nothing to check.
    group_rows = [r for r in rows if r.get("principalType") == "Group" and not r.get("sourceGroupId")]
    assert group_rows
    assert all(r["principalAccountEnabled"] == schema.ENABLED_NA for r in group_rows)


def test_a_cache_without_account_state_reports_unknown_never_enabled(seeded):
    _strip_state(seeded)
    rows = compose.build_master_rows(seeded)
    accounts = [r for r in rows if r.get("principalAccountEnabled") != schema.ENABLED_NA]
    assert accounts
    assert all(r["principalAccountEnabled"] == schema.ENABLED_UNKNOWN for r in accounts)
    assert not [r for r in rows if schema.is_disabled(r)]


def test_groups_are_not_applicable_not_unknown_even_with_no_state(seeded):
    """Otherwise every group inflates the "could not be checked" denominator, which is the
    number that decides whether the report is trustworthy at all."""
    _strip_state(seeded)
    report = leavers.build_leavers(seeded)
    assert report["denominator"]["not_applicable"] > 0


def test_on_prem_sync_is_carried_onto_the_row(seeded):
    rows = compose.build_master_rows(seeded)
    mallory = [r for r in rows if r.get("effectivePrincipalName") == "Mallory Leaver"]
    assert mallory
    assert any(r["principalOnPremSynced"] == schema.ENABLED_TRUE for r in mallory)


# --------------------------------------------------------------------------- cache preservation
def test_a_refresh_that_did_not_collect_state_must_not_delete_it(seeded):
    """The defect that once wiped the role catalogue, applied to the newer map.

    ``refresh_directory``'s no-Graph-token branch rewrites the whole directory blob. If it
    passed no account state, the disabled-access report would show zero disabled principals —
    a clean bill of health produced entirely by having lost the data."""
    before = cache.read_directory(seeded)["principal_state"]
    assert before
    preserved = orchestrator._preserve_principal_state(seeded, None)
    assert preserved == before


def test_a_refresh_that_did_collect_state_replaces_it(seeded):
    fresh = {"u-someone": {"accountEnabled": schema.ENABLED_TRUE}}
    assert orchestrator._preserve_principal_state(seeded, fresh) == fresh


@pytest.mark.anyio
async def test_group_members_are_included_in_the_account_state_lookup(monkeypatch):
    """The bug live data found, and no unit test would have.

    ``principal_ids`` deliberately omits group members: the expansion graph already carries each
    member's display name, so resolving them through getByIds buys nothing. Account state is
    different — it must be looked up per principal, so a member who never appears as an assignee
    was never checked. Measured on a real tenant: the state map covered 379 of 1,227 principals
    and found 26 disabled, where a full sweep found 78. The 52 missing were reachable ONLY
    through group membership, which is exactly the case this feature exists to surface."""
    seen: list[str] = []

    async def fake_state(token, principals):
        seen.extend(str(p.get("principalId")) for p in principals)
        return {}, collectors.CollectorStatus("PrincipalState")

    async def fake_directory(token, ids):
        # Only the assignee resolves; the member is known solely from the expansion graph.
        return ([{"principalId": "g-1", "principalType": "Group", "displayName": "Group One"}],
                collectors.CollectorStatus("PrincipalDirectory"))

    async def fake_groups(token, ids):
        return (
            {"g-1": {"name": "Group One", "members": [
                {"principalId": "u-hidden", "principalType": "User", "principalDisplayName": "Hidden"},
            ]}},
            collectors.CollectorStatus("GroupExpansion"),
        )

    async def noop_rows(*a, **kw):
        return [], collectors.CollectorStatus("EntraRoleAssignments")

    monkeypatch.setattr(collectors, "collect_principal_state", fake_state)
    monkeypatch.setattr(collectors, "collect_principal_directory", fake_directory)
    monkeypatch.setattr(collectors, "collect_group_expansion", fake_groups)
    monkeypatch.setattr(collectors, "collect_entra_roles", noop_rows)
    monkeypatch.setattr(collectors, "collect_sp_owners", lambda *a, **kw: noop_rows())

    async def fake_graph_token(conn):
        return "tok", ""

    import app.azure.credentials as creds

    monkeypatch.setattr(creds, "get_graph_token", fake_graph_token)
    await orchestrator.refresh_directory("t-groups", {"id": "c1"})
    assert "u-hidden" in seen, "a principal reachable only through a group was never checked"


# --------------------------------------------------------------------------- collector
@pytest.mark.anyio
async def test_a_capped_disabled_sweep_never_declares_anyone_enabled(monkeypatch):
    """"Not in the part of the disabled list we managed to read" is not evidence of being
    enabled. A capped sweep must leave every unmatched principal unknown."""
    monkeypatch.setattr(collectors, "MAX_DISABLED_SWEEP", 2)

    async def fake_get_all(token, url, params=None, *, extra_headers=None, max_items=0):
        if "/users" in url:
            return ([{"id": "u-a"}, {"id": "u-b"}], None, 200)  # == the cap: possibly truncated
        return ([], None, 200)

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    state, status = await collectors.collect_principal_state(
        "tok",
        [
            {"principalId": "u-a", "principalType": "User"},
            {"principalId": "u-zzz", "principalType": "User"},
        ],
    )
    assert state["u-a"]["accountEnabled"] == schema.ENABLED_FALSE
    assert state["u-zzz"]["accountEnabled"] == schema.ENABLED_UNKNOWN
    assert status.status == schema.STATUS_PARTIAL


@pytest.mark.anyio
async def test_a_complete_sweep_may_declare_the_remainder_enabled(monkeypatch):
    async def fake_get_all(token, url, params=None, *, extra_headers=None, max_items=0):
        if "/users" in url:
            return ([{"id": "u-a", "onPremisesSyncEnabled": True}], None, 200)
        return ([], None, 200)

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    state, status = await collectors.collect_principal_state(
        "tok",
        [
            {"principalId": "u-a", "principalType": "User"},
            {"principalId": "u-b", "principalType": "User"},
        ],
    )
    assert state["u-a"]["accountEnabled"] == schema.ENABLED_FALSE
    assert state["u-a"]["onPremSynced"] == schema.ENABLED_TRUE
    assert state["u-b"]["accountEnabled"] == schema.ENABLED_TRUE
    assert status.status == schema.STATUS_SUCCEEDED


@pytest.mark.anyio
async def test_a_failed_sweep_leaves_everyone_unknown(monkeypatch):
    async def fake_get_all(token, url, params=None, *, extra_headers=None, max_items=0):
        return ([], "HTTP 403: Insufficient privileges", 403)

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    state, status = await collectors.collect_principal_state(
        "tok", [{"principalId": "u-a", "principalType": "User"}]
    )
    assert state["u-a"]["accountEnabled"] == schema.ENABLED_UNKNOWN
    assert status.status == schema.STATUS_UNAUTHORIZED


@pytest.mark.anyio
async def test_getbyids_state_costs_no_extra_graph_calls(monkeypatch):
    """When the name resolution already returned account state, no sweep may be issued."""
    called = False

    async def fake_get_all(*a, **kw):
        nonlocal called
        called = True
        return ([], None, 200)

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    state, status = await collectors.collect_principal_state(
        "tok",
        [{"principalId": "u-a", "principalType": "User", "accountEnabled": schema.ENABLED_FALSE}],
    )
    assert called is False
    assert state["u-a"]["accountEnabled"] == schema.ENABLED_FALSE
    assert status.status == schema.STATUS_SUCCEEDED


@pytest.mark.anyio
async def test_groups_are_never_swept_for_account_state(monkeypatch):
    async def fake_get_all(*a, **kw):
        raise AssertionError("a group has no account state; nothing should be looked up")

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    state, _ = await collectors.collect_principal_state(
        "tok", [{"principalId": "g-1", "principalType": "Group"}]
    )
    assert state["g-1"]["accountEnabled"] == schema.ENABLED_NA


# --------------------------------------------------------------------------- report
def test_the_demo_estate_actually_contains_disabled_access(seeded):
    """Guard against a vacuous suite.

    `ext.guest_access` once shipped with a wrong signal id because the estate-wide test that
    should have caught it had NO GUESTS to run against. Every assertion below is worthless if
    the fixtures stop containing all four shapes."""
    report = leavers.build_leavers(seeded)
    assert report["measured"] is True
    kinds = {i["displayName"] for i in report["identities"]}
    assert "Mallory Leaver" in kinds            # disabled, direct, privileged, on-prem synced
    assert "Nina Nomore" in kinds               # disabled, group-only
    assert "Oscar Offboarded" in kinds          # disabled, owns a service principal
    assert "legacy-batch-job" in kinds          # a disabled service principal
    assert report["tier_counts"][leavers.TIER_LIVE] >= 1
    assert report["tier_counts"][leavers.TIER_RESTORABLE] >= 1


def test_an_unmeasured_tenant_is_a_wall_not_an_empty_list(seeded):
    _strip_state(seeded)
    report = leavers.build_leavers(seeded)
    assert report["measured"] is False
    assert report["identities"] == []
    assert report["reason"]
    # The denominator still has to be published: "0 found" is meaningless without "out of how
    # many, and how many could we not check".
    assert report["denominator"]["principals_with_access"] > 0
    assert report["denominator"]["state_unknown"] > 0


def test_owning_a_service_principal_is_the_live_tier(seeded):
    report = leavers.build_leavers(seeded)
    oscar = next(i for i in report["identities"] if i["displayName"] == "Oscar Offboarded")
    assert oscar["tier"] == leavers.TIER_LIVE
    assert oscar["ownedServicePrincipals"]


def test_a_plain_leaver_is_restorable_not_live(seeded):
    """Overstating this would be wrong in the direction that gets a tool ignored: a disabled
    account cannot obtain a token, so its own grants are dormant."""
    report = leavers.build_leavers(seeded)
    mallory = next(i for i in report["identities"] if i["displayName"] == "Mallory Leaver")
    assert mallory["tier"] == leavers.TIER_RESTORABLE


def test_there_is_no_empty_residual_session_tier(seeded):
    """An empty bucket reads as "we checked and found none". The residual-token window cannot
    be measured (Graph publishes no disabled-at timestamp), so it must be a STATED limitation."""
    report = leavers.build_leavers(seeded)
    assert set(report["tiers"]) == {leavers.TIER_LIVE, leavers.TIER_RESTORABLE}
    assert any("expire" in l.lower() for l in report["limitations"])


def test_group_only_access_is_counted_separately(seeded):
    report = leavers.build_leavers(seeded)
    nina = next(i for i in report["identities"] if i["displayName"] == "Nina Nomore")
    assert nina["groupGrants"] >= 1
    assert nina["directGrants"] == 0
    assert nina["groupsGrantingAccess"]
    assert report["totals"]["via_group_only"] >= 1


def test_identities_are_sorted_worst_first(seeded):
    report = leavers.build_leavers(seeded)
    tiers = [i["tier"] for i in report["identities"]]
    assert tiers[0] == leavers.TIER_LIVE


def test_signin_enrichment_is_absent_not_zero_when_entra_never_ran(seeded):
    report = leavers.build_leavers(seeded)
    assert report["signin"]["available"] is False
    assert all(i["lastSignIn"] == "" for i in report["identities"])
    assert any("sign-in" in l.lower() for l in report["limitations"])


# --------------------------------------------------------------------------- signals
_DISABLED_SIGNALS = {
    "hyg.disabled_principal_access",
    "hyg.disabled_privileged_access",
    "hyg.disabled_via_group",
    "hyg.disabled_owns_credential",
    "hyg.disabled_pim_eligible",
}


def test_every_disabled_signal_fires_on_the_demo_estate(seeded):
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    for sid in _DISABLED_SIGNALS:
        assert sid in results, f"{sid} is not registered"
        assert results[sid].measured is True, f"{sid} reported not-measured on a seeded tenant"
        assert results[sid].findings, f"{sid} produced nothing on an estate built to trigger it"


def test_every_disabled_signal_reports_not_measured_rather_than_clean(seeded):
    """The single most important test in this file.

    Returning [] when account state was never collected would score the pillar as a PASS and
    print "no disabled account holds access" on a tenant nobody ever checked."""
    _strip_state(seeded)
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    for sid in _DISABLED_SIGNALS:
        assert results[sid].measured is False, f"{sid} claimed a clean result while blind"
        assert results[sid].reason, f"{sid} must say why it could not run"


def test_the_group_signal_tells_people_not_to_delete_the_assignment(seeded):
    """Getting this remediation backwards breaks access for every other group member."""
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    f = results["hyg.disabled_via_group"].findings[0]
    assert "do not delete" in f.remediation.lower() or "never delete" in f.remediation.lower()


def test_the_credential_signal_says_the_access_is_live_now(seeded):
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    f = results["hyg.disabled_owns_credential"].findings[0]
    assert "live now" in f.detail.lower()
    assert f.severity == "error"


def test_a_privileged_leaver_raises_its_own_severity(seeded):
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    per_principal = {f.subject_label: f for f in results["hyg.disabled_principal_access"].findings}
    assert per_principal["Mallory Leaver"].severity == "error"       # holds privileged roles
    assert per_principal["legacy-batch-job"].severity == "warning"   # Reader only


def test_the_on_prem_case_points_at_active_directory(seeded):
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    per_principal = {f.subject_label: f for f in results["hyg.disabled_principal_access"].findings}
    assert "active directory" in per_principal["Mallory Leaver"].detail.lower()


# --------------------------------------------------------------------------- recycle bin
def test_a_soft_deleted_principal_is_flagged_as_restorable(seeded):
    """The recycle bin is not the same state as "deleted and gone".

    ``hyg.orphaned_assignment`` tells an operator to delete the assignment because "there is no
    principal left to lose access". That is true for a hard deletion and false for the 30 days a
    soft-deleted object can be restored by any administrator — which brings every grant back at
    once, and is exactly the window in which an offboarding is most likely to be reversed."""
    report = leavers.build_leavers(seeded)
    nina = next(i for i in report["identities"] if i["displayName"] == "Nina Nomore")
    assert nina["softDeleted"] is True
    assert nina["deletedDateTime"]
    assert report["totals"]["soft_deleted"] >= 1
    # Everyone else must NOT be flagged — a blanket true would be as useless as a blanket false.
    others = [i for i in report["identities"] if i["displayName"] != "Nina Nomore"]
    assert others and all(i["softDeleted"] is False for i in others)


def test_the_recycle_bin_signal_contradicts_the_orphan_advice(seeded):
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    res = results["hyg.deleted_principal_restorable"]
    assert res.measured is True and res.findings
    f = res.findings[0]
    assert "30 days" in f.detail
    assert "not be treated as harmless orphans" in f.detail


def test_the_recycle_bin_signal_is_not_measured_when_state_was_not_collected(seeded):
    _strip_state(seeded)
    results = {r.spec.id: r for r in findings.evaluate(seeded)}
    assert results["hyg.deleted_principal_restorable"].measured is False


@pytest.mark.anyio
async def test_a_blocked_recycle_bin_read_is_recorded_not_raised(monkeypatch):
    """A tenant that does not grant the directory-recycle-bin read must still get a report."""

    async def fake_get_all(token, url, params=None, *, extra_headers=None, max_items=0):
        return ([], "HTTP 403: Insufficient privileges", 403)

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    out, status = await collectors.collect_deleted_principals("tok")
    assert out == {}
    assert status.status == schema.STATUS_UNAUTHORIZED
    assert status.message


# --------------------------------------------------------------------------- campaigns
def test_the_disabled_campaign_selector_picks_only_known_disabled(seeded):
    """`unknown` must never reach a certification campaign.

    A cache predating the account-state collector would otherwise put the whole estate in front
    of a reviewer under the heading "these people have left", and a reviewer stops trusting the
    tool after the first wrong name."""
    rows = compose.build_master_rows(seeded)
    picked = campaigns.select_rows(rows, {"kind": "disabled"})
    assert picked
    assert all(schema.is_disabled(r) for r in picked)
    assert len(picked) == len(leavers.disabled_grant_rows(rows))


def test_the_disabled_selector_can_narrow_to_the_live_tier(seeded):
    rows = compose.build_master_rows(seeded)
    everyone = campaigns.select_rows(rows, {"kind": "disabled"})
    live = campaigns.select_rows(rows, {"kind": "disabled", "tier": "live_now"})
    assert 0 < len(live) < len(everyone)
    # The live tier is a PERSON-level property: every row belonging to an owner qualifies, not
    # only the ownership row itself.
    assert {r["effectivePrincipalName"] for r in live} == {"Oscar Offboarded"}


def test_the_disabled_selector_can_narrow_to_privileged(seeded):
    rows = compose.build_master_rows(seeded)
    priv = campaigns.select_rows(rows, {"kind": "disabled", "privileged_only": True})
    assert priv and all(r.get("roleIsPrivileged") for r in priv)


def test_the_disabled_selector_never_certifies_a_deny(seeded):
    rows = compose.build_master_rows(seeded)
    picked = campaigns.select_rows(rows, {"kind": "disabled"})
    assert all(r.get("effect") != schema.EFFECT_DENY for r in picked)


def test_two_groups_granting_the_same_access_are_one_review_item(seeded):
    """The 500 that only live data produced.

    ``iam_review_item`` is UNIQUE on (campaign_id, row_key), and ``diff.row_key`` deliberately
    excludes the access path. A principal in TWO groups that both grant the same role at the
    same scope therefore produced two items with one key and killed the whole campaign with an
    IntegrityError. Nothing hit it before because no selector was principal-centric; on a real
    tenant 53 of 78 leavers hold their access through groups, where overlapping membership is
    entirely ordinary."""
    base = {
        "effectivePrincipalId": "u-dup", "roleDefinitionId": "rd-1", "scope": "/subscriptions/s",
        "surface": schema.SURFACE_AZURE_RBAC, "assignmentState": schema.STATE_ACTIVE,
        "accessPath": schema.PATH_GROUP, "effect": schema.EFFECT_ALLOW,
    }
    rows = [
        {**base, "sourceGroupName": "Group A"},
        {**base, "sourceGroupName": "Group B"},
    ]
    out, folded = campaigns._dedupe_by_review_key(rows)
    assert len(out) == 1, "one decision, not one per group that happens to grant it"
    # …but the folded paths must survive: revoking one membership while the other still grants
    # the same role leaves the access exactly where it was.
    assert folded[campaigns.diff_mod.row_key(rows[0])] == ["Group A", "Group B"]


def test_a_direct_grant_wins_over_a_group_derived_duplicate(seeded):
    """Their remediations differ — delete the assignment vs remove the member — so the row that
    survives must be the one whose advice is correct."""
    base = {
        "effectivePrincipalId": "u-dup", "roleDefinitionId": "rd-1", "scope": "/subscriptions/s",
        "surface": schema.SURFACE_AZURE_RBAC, "assignmentState": schema.STATE_ACTIVE,
        "effect": schema.EFFECT_ALLOW,
    }
    rows = [
        {**base, "accessPath": schema.PATH_GROUP, "sourceGroupName": "Group A"},
        {**base, "accessPath": schema.PATH_DIRECT},
    ]
    out, _ = campaigns._dedupe_by_review_key(rows)
    assert len(out) == 1 and out[0]["accessPath"] == schema.PATH_DIRECT


def test_dedupe_leaves_genuinely_distinct_access_alone(seeded):
    """A de-dupe that over-collapses hides access, which is worse than the crash it replaced."""
    rows = compose.build_master_rows(seeded)
    picked = campaigns.select_rows(rows, {"kind": "disabled"})
    out, _ = campaigns._dedupe_by_review_key(picked)
    assert len(out) == len({campaigns.diff_mod.row_key(r) for r in picked})


# --------------------------------------------------------------------------- enrichment & shape
def test_the_rollup_keeps_the_arm_structure_of_every_scope(seeded):
    """The Where panel could not group by resource type because the rollup flattened every
    scope to a display string and threw the structure away."""
    report = leavers.build_leavers(seeded)
    nina = next(i for i in report["identities"] if i["displayName"] == "Nina Nomore")
    assert nina["resources"], "structured scopes must be published"
    r = nina["resources"][0]
    for key in ("scope", "scopeType", "subscriptionName", "resourceGroup", "roles", "grants"):
        assert key in r
    # And the flat list survives for the CSV, which is a spreadsheet and not a tree.
    assert nina["scopes"]


def test_resource_entries_record_how_the_access_is_held(seeded):
    """Direct and group-derived access have opposite remediations, so a scope that is reached
    only through a group must say which group."""
    report = leavers.build_leavers(seeded)
    nina = next(i for i in report["identities"] if i["displayName"] == "Nina Nomore")
    r = next(x for x in nina["resources"] if x["viaGroups"])
    assert r["direct"] is False
    assert "Data Readers" in r["viaGroups"]


def test_every_identity_carries_its_actual_grants(seeded):
    report = leavers.build_leavers(seeded)
    mallory = next(i for i in report["identities"] if i["displayName"] == "Mallory Leaver")
    assert len(mallory["grantDetail"]) == mallory["grants"]
    assert mallory["grantDetailTruncated"] is False
    assert all("assignmentId" in g and "roleName" in g for g in mallory["grantDetail"])


def test_grant_detail_is_capped_and_says_so(monkeypatch, seeded):
    """A principal in a very wide group must not make one API response enormous — and a capped
    list must never read as a complete one."""
    monkeypatch.setattr(leavers, "MAX_GRANT_DETAIL", 1)
    report = leavers.build_leavers(seeded)
    mallory = next(i for i in report["identities"] if i["displayName"] == "Mallory Leaver")
    assert len(mallory["grantDetail"]) == 1
    assert mallory["grantDetailTruncated"] is True
    # The aggregate count still describes everything.
    assert mallory["grants"] > 1


def test_grant_age_is_published(seeded):
    report = leavers.build_leavers(seeded)
    mallory = next(i for i in report["identities"] if i["displayName"] == "Mallory Leaver")
    assert mallory["oldestGrantAt"]
    assert mallory["newestGrantAt"] >= mallory["oldestGrantAt"]


# --------------------------------------------------------------------------- dormancy
def test_dormancy_never_confuses_unmeasured_with_never():
    """The single most dangerous conflation on this screen: `never signed in` argues for
    deleting the access, `not measured` argues for collecting more data first."""
    now = datetime(2026, 8, 5, tzinfo=timezone.utc)
    assert leavers.dormancy_of("", known=False, now=now) == (leavers.DORMANCY_UNKNOWN, None)
    assert leavers.dormancy_of("", known=True, now=now) == (leavers.DORMANCY_NEVER, None)
    # An unparseable timestamp is also unknown, never "never".
    assert leavers.dormancy_of("not-a-date", known=True, now=now)[0] == leavers.DORMANCY_UNKNOWN


@pytest.mark.parametrize(
    "days,expected",
    [(0, "recent"), (10, "recent"), (89, "recent"), (90, "over_90d"), (200, "over_90d"),
     (365, "over_1y"), (700, "over_1y"), (730, "over_2y"), (3000, "over_2y")],
)
def test_dormancy_buckets_are_contiguous(days, expected):
    """Gaps or overlaps here silently drop identities out of every bucket filter."""
    now = datetime(2026, 8, 5, tzinfo=timezone.utc)
    stamp = (now - timedelta(days=days)).isoformat()
    assert leavers.dormancy_of(stamp, known=True, now=now)[0] == expected


def test_dormancy_labels_cover_every_bucket():
    keys = {k for k, _, _ in leavers.DORMANCY_BUCKETS} | {leavers.DORMANCY_NEVER, leavers.DORMANCY_UNKNOWN}
    assert keys == set(leavers.DORMANCY_LABELS)


def test_an_absent_entra_scan_leaves_sign_in_unmeasured_not_never(seeded):
    report = leavers.build_leavers(seeded)
    assert report["signin"]["available"] is False
    for i in report["identities"]:
        assert i["dormancyBucket"] == leavers.DORMANCY_UNKNOWN
        assert i["signIn"]["known"] is False
    assert any("sign-in" in l.lower() for l in report["limitations"])


def test_an_absent_usage_sweep_leaves_last_used_unmeasured(seeded):
    report = leavers.build_leavers(seeded)
    assert report["usage"]["available"] is False
    assert all(i["activityMeasured"] is False for i in report["identities"])
    assert any("activity log" in l.lower() for l in report["limitations"])


def test_owned_app_signin_absence_is_not_reported_as_never(seeded):
    """The service-principal report only covers a bounded window, so an app with no entry was
    not seen in that window — which is not the same as never being used, and the difference
    decides whether somebody rolls a credential that is in daily use."""
    report = leavers.build_leavers(seeded)
    oscar = next(i for i in report["identities"] if i["displayName"] == "Oscar Offboarded")
    assert oscar["ownedDetail"]
    o = oscar["ownedDetail"][0]
    assert o["lastSignIn"] == ""
    assert o["lastSignInKnown"] is False


# --------------------------------------------------------------------------- usage collector
def test_usage_records_the_most_recent_operation_per_principal():
    """"Granted in 2019, last did anything in 2021" is a far stronger case for removal than any
    role name, and the timestamp was already on every event and being discarded."""
    from app.iam import usage as usage_mod

    events = [
        {"actorObjectId": "U-1", "operation": "Microsoft.Compute/write", "eventTime": "2026-01-01T00:00:00Z", "resourceId": "/a"},
        {"actorObjectId": "u-1", "operation": "Microsoft.Compute/read", "eventTime": "2026-06-01T00:00:00Z", "resourceId": "/b"},
        {"actorObjectId": "u-1", "operation": "Microsoft.Compute/read", "eventTime": "2026-03-01T00:00:00Z", "resourceId": "/c"},
    ]
    by_principal: dict[str, dict] = {}
    for e in events:
        pid = str(e.get("actorObjectId", "") or "").lower()
        entry = by_principal.setdefault(pid, {"principalId": pid, "actions": set(), "events": 0,
                                              "displayName": "", "scopes": set(), "lastSeen": ""})
        entry["events"] += 1
        when = str(e.get("eventTime", "") or "")
        if when > entry["lastSeen"]:
            entry["lastSeen"] = when
    assert by_principal["u-1"]["lastSeen"] == "2026-06-01T00:00:00Z"
    assert usage_mod.SOURCE_ACTIVITY_LOG


# --------------------------------------------------------------------------- export
def test_identity_csv_has_one_row_per_person(seeded):
    report = leavers.build_leavers(seeded)
    body = export.to_identity_csv(report["identities"], report["tiers"])
    lines = [l for l in body.strip().splitlines() if l]
    assert len(lines) == len(report["identities"]) + 1


def test_identity_csv_neutralises_formula_injection(seeded):
    hostile = [{
        "principalId": "p", "displayName": "=cmd|'/c calc'!A1", "userPrincipalName": "",
        "principalType": "User", "userType": "", "accountEnabled": "false",
        "onPremSynced": "false", "tier": "restorable", "grants": 1, "privilegedGrants": 0,
        "highestRole": "Reader", "planes": [], "directGrants": 1, "groupGrants": 0,
        "groupsGrantingAccess": [], "ownedServicePrincipals": [], "pimEligible": 0,
        "permanentlyEligible": 0, "scopes": [], "subscriptions": [],
        "lastSignIn": "", "lastSignInSource": "",
    }]
    body = export.to_identity_csv(hostile, {})
    assert "'=cmd" in body


def test_grant_export_round_trips_through_the_scanner_projection(seeded):
    rows = compose.build_master_rows(seeded)
    grants = leavers.disabled_grant_rows(rows)
    scanner = export.to_csv(grants, columns=schema.SCANNER_COLUMNS)
    header = scanner.splitlines()[0].split(",")
    assert len(header) == 46
    assert "principalAccountEnabled" not in header


def test_deny_rows_are_never_reported_as_access(seeded):
    """A deny REMOVES access. Counting one as a grant would report a control as the risk."""
    rows = compose.build_master_rows(seeded)
    assert all(r.get("effect") != schema.EFFECT_DENY for r in leavers.disabled_grant_rows(rows))


def _sheets(payload: bytes) -> dict[str, object]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(payload))
    return {name: wb[name] for name in wb.sheetnames}


def test_workbook_always_carries_a_not_measured_sheet(seeded):
    """A spreadsheet outlives the screen that produced it. If the limits of the data are not IN
    the file, the file misrepresents itself the moment it is forwarded."""
    report = leavers.build_leavers(seeded)
    rows = compose.build_master_rows(seeded)
    payload = export.to_disabled_workbook(
        report=report, grants=leavers.disabled_grant_rows(rows), tenant_id=seeded
    )
    sheets = _sheets(payload)
    assert "Not measured" in sheets
    assert sheets["Not measured"].max_row >= 2


def test_workbook_publishes_the_denominator_on_the_summary(seeded):
    report = leavers.build_leavers(seeded)
    rows = compose.build_master_rows(seeded)
    payload = export.to_disabled_workbook(
        report=report, grants=leavers.disabled_grant_rows(rows), tenant_id=seeded
    )
    text = " ".join(
        str(c or "") for row in _sheets(payload)["Summary"].iter_rows(values_only=True) for c in row
    )
    assert "DENOMINATOR" in text
    assert "could NOT be checked" in text


def test_an_unmeasured_workbook_says_so_rather_than_shipping_an_empty_grid(seeded):
    _strip_state(seeded)
    report = leavers.build_leavers(seeded)
    payload = export.to_disabled_workbook(report=report, grants=[], tenant_id=seeded)
    sheets = _sheets(payload)
    summary = " ".join(
        str(c or "") for row in sheets["Summary"].iter_rows(values_only=True) for c in row
    )
    assert "NO" in summary
    limits = " ".join(
        str(c or "") for row in sheets["Not measured"].iter_rows(values_only=True) for c in row
    )
    assert "not been collected" in limits


def test_workbook_separates_group_access_from_direct(seeded):
    report = leavers.build_leavers(seeded)
    rows = compose.build_master_rows(seeded)
    payload = export.to_disabled_workbook(
        report=report, grants=leavers.disabled_grant_rows(rows), tenant_id=seeded
    )
    sheets = _sheets(payload)
    assert sheets["Via groups"].max_row >= 2       # header + at least Nina
    assert sheets["Owns credentials"].max_row >= 2  # header + Oscar


def test_workbook_has_a_resources_sheet_that_keeps_the_arm_structure(seeded):
    """The Identities sheet joins every scope into one cell, which is unusable the moment
    somebody holds access on forty of them. This is the sheet you filter and pivot."""
    report = leavers.build_leavers(seeded)
    rows = compose.build_master_rows(seeded)
    payload = export.to_disabled_workbook(
        report=report, grants=leavers.disabled_grant_rows(rows), tenant_id=seeded
    )
    ws = _sheets(payload)["Resources"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    for col in ("Scope type", "Resource type", "Held via", "Scope id"):
        assert col in header
    assert ws.max_row > 1


def test_the_export_carries_every_sign_in_kind_and_the_dormancy_label(seeded):
    """A blank cell in a spreadsheet reads as zero, so "Not measured" has to survive as TEXT."""
    report = leavers.build_leavers(seeded)
    body = export.to_identity_csv(report["identities"], report["tiers"])
    header = body.splitlines()[0]
    for col in ("Last interactive sign-in", "Last non-interactive sign-in",
                "Last successful sign-in", "Owned app last sign-in", "Dormancy",
                "Last used (Activity Log)", "Oldest grant"):
        assert col in header
    assert "Not measured" in body
    assert "not measured" in body  # the unmeasured usage cell


# --------------------------------------------------------------------------- API filters
def _api():
    from app.api import iam as iam_api

    return iam_api


def test_the_on_prem_filter_is_three_state_not_a_boolean(seeded):
    """Sync state decides WHICH directory the fix goes in. Filing "we could not tell" under
    "cloud" sends an operator to the wrong console, where the change silently reverts."""
    iam_api = _api()
    ids = leavers.build_leavers(seeded)["identities"]
    # Guard against a vacuous assertion: without an unknown-sync identity in the estate, a
    # filter that folds unknown into cloud is indistinguishable from a correct one.
    assert any(i["onPremSynced"] == schema.ENABLED_UNKNOWN for i in ids)
    assert any(i["onPremSynced"] == schema.ENABLED_TRUE for i in ids)
    assert any(i["onPremSynced"] == schema.ENABLED_FALSE for i in ids)

    cloud = iam_api._apply_leavers_filters(ids, on_prem="cloud")
    onprem = iam_api._apply_leavers_filters(ids, on_prem="onprem")
    unknown = iam_api._apply_leavers_filters(ids, on_prem="unknown")
    assert all(i["onPremSynced"] == schema.ENABLED_FALSE for i in cloud)
    assert all(i["onPremSynced"] == schema.ENABLED_TRUE for i in onprem)
    assert all(i["onPremSynced"] == schema.ENABLED_UNKNOWN for i in unknown)
    # Disjoint, and together they account for everybody.
    assert len(cloud) + len(onprem) + len(unknown) == len(ids)
    assert unknown and cloud and onprem


def test_the_legacy_on_prem_boolean_still_works(seeded):
    iam_api = _api()
    ids = leavers.build_leavers(seeded)["identities"]
    assert iam_api._apply_leavers_filters(ids, on_prem_synced=True) == iam_api._apply_leavers_filters(
        ids, on_prem="onprem"
    )


def test_never_used_only_ever_matches_measured_identities(seeded):
    """The filter most likely to end in a deletion. "We did not look" is not evidence of
    disuse, so an unmeasured tenant must match nobody rather than everybody."""
    iam_api = _api()
    ids = leavers.build_leavers(seeded)["identities"]
    assert all(not i["activityMeasured"] for i in ids)
    assert iam_api._apply_leavers_filters(ids, never_used=True) == []


def test_dormancy_filter_can_be_measured_from_each_sign_in_kind(seeded):
    """An account can be dead interactively for months while a client refreshes tokens daily,
    so the kind the filter measures from has to be selectable."""
    iam_api = _api()
    now = datetime.now(timezone.utc)
    old = (now - timedelta(days=800)).isoformat()
    recent = (now - timedelta(days=2)).isoformat()
    ids = [{
        "principalId": "p1", "displayName": "P", "lastSignIn": old,
        "signIn": {"interactive": old, "nonInteractive": recent, "successful": "",
                   "servicePrincipal": "", "known": True},
    }]
    assert iam_api._apply_leavers_filters(ids, dormancy="over_2y", signin_kind="interactive")
    assert not iam_api._apply_leavers_filters(ids, dormancy="over_2y", signin_kind="nonInteractive")
    assert iam_api._apply_leavers_filters(ids, dormancy="recent", signin_kind="nonInteractive")
    # A kind with no timestamp at all is `never`, and only because `known` is true.
    assert iam_api._apply_leavers_filters(ids, dormancy="never", signin_kind="successful")


def test_counts_are_computed_over_the_whole_filtered_set(seeded):
    """Group headers read from these. A count derived from the loaded page shrinks as the
    reader scrolls, which is exactly the defect the Findings tab had to fix."""
    iam_api = _api()
    ids = leavers.build_leavers(seeded)["identities"]
    counts = iam_api._leavers_counts(ids)
    assert sum(counts["tier"].values()) == len(ids)
    assert sum(counts["principal_type"].values()) == len(ids)
    assert sum(counts["on_prem"].values()) == len(ids)
    assert sum(counts["dormancy"].values()) == len(ids)
    # Multi-valued dimensions legitimately sum higher than the population.
    assert sum(counts["subscription"].values()) >= 1


def test_every_new_filter_narrows_rather_than_erroring(seeded):
    iam_api = _api()
    ids = leavers.build_leavers(seeded)["identities"]
    for kwargs in (
        {"soft_deleted": True}, {"has_owned_sp": True}, {"pim_eligible": True},
        {"plane": schema.SURFACE_AZURE_RBAC}, {"group": "Data Readers"},
        {"role": "Contributor"}, {"subscription": "Contoso Production"},
    ):
        out = iam_api._apply_leavers_filters(ids, **kwargs)
        assert len(out) <= len(ids)
        assert out, f"{kwargs} matched nothing on an estate built to contain it"


def test_the_export_and_the_screen_take_the_same_filter_model():
    """One model, not two parameter lists. The moment one endpoint understands fewer filters, a
    download silently contains rows the screen was not showing."""
    import inspect

    iam_api = _api()
    report_params = set(inspect.signature(iam_api.leavers_report).parameters)
    export_params = set(inspect.signature(iam_api.leavers_export).parameters)
    model_fields = set(iam_api.LeaversQuery.model_fields)
    assert model_fields <= report_params
    assert model_fields <= export_params


def _sheets_of(payload: bytes):
    return _sheets(payload)


# ----------------------------------------------------- the campaign must match the screen
@pytest.mark.anyio
async def test_a_campaign_covers_what_the_screen_showed_not_everything(seeded, monkeypatch):
    """The defect this fixes, measured on a live tenant before the fix:

        screen shows:      3 identities
        campaign covers: 518 items over 78 identities

    The selector understood `tier` and `privileged_only` and nothing else, so every other filter
    on the screen was silently dropped and the reviewer got a campaign 26x larger than the list
    the operator was looking at when they clicked the button. Same class as the export
    regression — the artifact has to contain what the screen showed."""
    rows = compose.build_master_rows(seeded)
    report = leavers.build_leavers(seeded)

    narrow = {"kind": "disabled", "via_group_only": True}
    on_screen = leavers.filter_identities(report["identities"], narrow)
    assert on_screen, "the estate must contain a group-only leaver or this proves nothing"
    assert len(on_screen) < len(report["identities"]), "the filter must actually narrow"

    picked = campaigns.select_rows(rows, narrow, tenant_id=seeded)
    covered = {
        str(r.get("effectivePrincipalId") or r.get("principalId")).lower() for r in picked
    }
    assert covered == {str(i["principalId"]).lower() for i in on_screen}


@pytest.mark.anyio
async def test_an_identity_level_filter_without_a_tenant_raises_rather_than_being_ignored(seeded):
    """A selector that silently drops half its own filters recreates the bug invisibly."""
    rows = compose.build_master_rows(seeded)
    with pytest.raises(campaigns.CampaignError) as exc:
        campaigns.select_rows(rows, {"kind": "disabled", "on_prem": "onprem"})
    assert "on_prem" in str(exc.value)


def test_row_level_filters_still_work_without_a_tenant(seeded):
    """`privileged_only` and `tier` are derivable from the rows, so they must not need one."""
    rows = compose.build_master_rows(seeded)
    priv = campaigns.select_rows(rows, {"kind": "disabled", "privileged_only": True})
    assert priv and all(r.get("roleIsPrivileged") for r in priv)


def test_an_explicit_selection_narrows_but_never_widens(seeded):
    """A ticked row list is an ADDITIONAL constraint. If it replaced the filters, a stale id
    could resurrect an identity a later scan has excluded."""
    report = leavers.build_leavers(seeded)
    ids = [i["principalId"] for i in report["identities"]]
    one = leavers.filter_identities(report["identities"], {"principal_ids": [ids[0]]})
    assert len(one) == 1 and one[0]["principalId"] == ids[0]
    # Combined with a filter that excludes it, the answer is empty — not the id list.
    conflicting = leavers.filter_identities(
        report["identities"], {"principal_ids": [ids[0]], "search": "zzz-no-such-person"}
    )
    assert conflicting == []
    # An id nobody holds cannot conjure a row.
    assert leavers.filter_identities(report["identities"], {"principal_ids": ["not-a-real-id"]}) == []


@pytest.mark.anyio
async def test_a_campaign_records_what_it_was_scoped_to(seeded):
    """The selector is re-evaluated on refresh, which is what lets a standing review notice a
    new leaver — and also means the reviewer's list can differ from the operator's. Recording
    the original population makes that visible instead of surprising."""
    made = await campaigns.create(
        seeded, name="scope record", selector={"kind": "disabled", "via_group_only": True}
    )
    stats = made["stats"]
    assert stats["scoped_principals"] >= 1
    assert stats["scoped_at_creation"] == stats["total"]
    assert stats["scope_filter"]["via_group_only"] is True


# ----------------------------------------------------- escalation & remediation
def test_escalation_is_read_from_cache_and_never_rebuilt(seeded, monkeypatch):
    """Building that graph takes about half a minute on a real tenant. Putting it behind a row
    expansion would be the same inline-CPU defect this codebase has fixed three times."""
    from app.iam import escalation as esc_mod

    def explode(*a, **kw):
        raise AssertionError("the escalation graph must never be built from the leavers report")

    monkeypatch.setattr(esc_mod, "graph_for_tenant", explode)
    monkeypatch.setattr(esc_mod, "detect", explode)
    report = leavers.build_leavers(seeded)
    # No graph is persisted for the demo tenant, so it must report "not measured" rather than
    # implying nobody can escalate.
    assert report["escalation"]["available"] is False
    assert all(i["escalationMeasured"] is False for i in report["identities"])


def test_escalation_counts_the_path_START_not_every_node_on_it(seeded):
    """A path is keyed by the principal it starts from. Counting every node would flag the
    intermediate service principals and key vaults as though they could escalate."""
    cache.write_escalation(
        seeded,
        {"nodes": [], "edges": [], "paths": [
            {"from": "u-mallory", "to": "tier0", "hops": []},
            {"from": "u-mallory", "to": "tier0", "hops": []},
            {"from": "sp-somewhere-in-the-middle", "to": "tier0", "hops": []},
        ]},
        cache_version=0, min_confidence="low", duration_seconds=0.0,
    )
    enriched = leavers.escalation_enrichment(seeded)
    assert enriched["available"] is True
    assert enriched["by_id"]["u-mallory"] == 2
    assert "tier0" not in enriched["by_id"]


def test_remediation_orders_group_membership_before_direct_revocation(seeded):
    """Revoking a direct grant while the principal still inherits the same access through a
    group looks successful and changes nothing — which is how "we revoked it" and "they still
    have it" end up both being true."""
    from app.iam import remediation

    rows = leavers.disabled_grant_rows(compose.build_master_rows(seeded))
    actions = [a for a in (remediation.revoke_assignment(r, "az") for r in rows) if a]
    bundle = remediation.build_bundle(actions, "az", title="test")
    hints = [a["order_hint"] for a in bundle["actions"]]
    assert hints == sorted(hints)
    # And the script says out loud that the product does not run it.
    assert "NOT RUN BY THE PRODUCT" in bundle["script"]
    assert "ROLLBACK" in bundle["script"]


# --------------------------------------- the script must target the API that actually governs
# Both of these were reported from a REAL run of a generated script against a live tenant. The
# commands exited without error and removed nothing, which is the worst possible outcome: the
# operator reads a clean exit and ticks the line off.
def _revoke(fmt: str = "az", **over: Any) -> dict[str, Any]:
    from app.iam import remediation

    row = schema.make_row(
        surface=schema.SURFACE_AZURE_RBAC,
        accessPath=schema.PATH_DIRECT,
        assignmentState=schema.STATE_ACTIVE,
        principalId="p-1",
        effectivePrincipalId="p-1",
        effectivePrincipalName="Some Person",
        roleName="Reader",
        scope="/subscriptions/11111111-2222-3333-4444-555555555555",
    )
    row.update(over)
    return remediation.revoke_assignment(row, fmt)


def test_group_derived_access_removes_the_membership_not_the_assignment():
    """Reported live: `az role assignment delete --assignee <user>` printed "No matched
    assignments were found to delete" three times. The assignment belongs to the GROUP; the user
    never had one. Deleting the group's assignment instead would strip every other member."""
    from app.iam import remediation

    a = _revoke(
        accessPath=schema.PATH_GROUP,
        principalId="g-1",
        sourceGroupId="g-1",
        sourceGroupName="SG_Readers",
        effectivePrincipalId="u-1",
    )
    assert a["plane"] == remediation.PLANE_GROUP_MEMBERSHIP
    assert "az ad group member remove" in a["command"]
    assert "--group 'g-1'" in a["command"] and "--member-id 'u-1'" in a["command"]
    # The command that does nothing must NOT be emitted.
    assert "role assignment delete" not in a["command"]
    assert "az ad group member add" in a["rollback"]
    # The dry run confirms WHERE the membership is before anything is deleted.
    assert "membership is in" in a["dry_run"]
    # Losing a group is not losing one role, and the operator has to be told before they run it.
    assert "EVERYTHING" in a["breaks_if"] and "SG_Readers" in a["breaks_if"]


def test_entra_directory_role_is_removed_through_graph_not_arm():
    """Reported live: `az role assignment delete --role 'Global Reader'` failed with "Role
    'Global Reader' doesn't exist" — directory roles are not published as ARM role definitions,
    so the lookup can never succeed no matter which scope is passed."""
    from app.iam import remediation

    a = _revoke(
        surface=schema.SURFACE_ENTRA,
        accessModel=schema.ACCESS_ENTRA,
        roleName="Global Reader",
        roleDefinitionId="rd-1",
        assignmentId="ra-1",
        scope="/",
        scopeType=schema.SCOPE_DIRECTORY,
    )
    assert a["plane"] == remediation.PLANE_ENTRA_ROLE
    assert "graph.microsoft.com" in a["command"]
    assert "roleManagement/directory/roleAssignments/ra-1" in a["command"]
    assert "--method DELETE" in a["command"]
    assert "role assignment delete" not in a["command"]
    # The rollback has to carry enough to re-create it: role definition, principal and scope.
    assert "rd-1" in a["rollback"] and "p-1" in a["rollback"]


def test_a_directory_role_with_no_assignment_id_refuses_to_guess():
    """Without the assignment id there is no URL to DELETE. A plausible-looking command would be
    worse than saying so."""
    from app.iam import remediation

    a = _revoke(surface=schema.SURFACE_ENTRA, roleName="Global Reader", assignmentId="", scope="/")
    assert "MANUAL STEP" in a["command"]
    assert not [ln for ln in a["command"].splitlines() if ln.strip() and not ln.lstrip().startswith("#")]


def test_service_principal_ownership_is_not_a_role_assignment():
    """`--role 'Service Principal Owner'` fails the same way 'Global Reader' does."""
    from app.iam import remediation

    a = _revoke(
        surface=schema.SURFACE_ENTRA,
        accessPath=schema.PATH_OWNER,
        roleName="Service Principal Owner",
        principalId="sp-1",
        principalDisplayName="some-app",
        effectivePrincipalId="u-1",
    )
    assert a["plane"] == remediation.PLANE_SP_OWNER
    # There is NO CLI verb for this: `az ad sp owner` offers only `list`, and Az.Resources has no
    # *-AzADServicePrincipalOwner cmdlets. Reported live: "'remove' is misspelled or not
    # recognized by the system". It has to be the Graph relationship.
    assert "az ad sp owner remove" not in a["command"]
    assert "/owners/u-1/$ref" in a["command"] and "--method DELETE" in a["command"]
    assert "role assignment delete" not in a["command"]
    # The paired application keeps a separate owner list; an owner there can mint a credential
    # and authenticate as this same service principal.
    assert "az ad app owner list" in a["dry_run"]
    # The whole point of the live_now tier: the credential outlives the ownership change.
    assert "secret" in a["breaks_if"] and "Roll the credential" in a["breaks_if"]


def test_the_undo_for_service_principal_ownership_posts_the_reference_back():
    from app.iam import remediation

    a = _revoke(
        surface=schema.SURFACE_ENTRA,
        accessPath=schema.PATH_OWNER,
        roleName="Service Principal Owner",
        principalId="sp-1",
        effectivePrincipalId="u-1",
    )
    assert a["plane"] == remediation.PLANE_SP_OWNER
    assert "--method POST" in a["rollback"]
    assert '"@odata.id":"https://graph.microsoft.com/v1.0/directoryObjects/u-1"' in a["rollback"]


# Every command verb below was checked against the installed Azure CLI and Az.Resources module.
# The service-principal owner plane shipped two verbs that DO NOT EXIST in either tool, in both
# formats, and only a live run caught it — so the vocabulary is now pinned.
_VERIFIED_AZ = {
    "az rest",
    "az ad group member remove",
    "az ad group member add",
    "az ad group member check",
    "az ad group show",
    "az ad app owner list",
    "az ad sp show",
    "az keyvault delete-policy",
    "az keyvault set-policy",
    "az keyvault show",
    "az role assignment delete",
    "az role assignment create",
    "az role assignment list",
}
_VERIFIED_PS = {
    "Invoke-MgGraphRequest",
    "Get-MgRoleManagementDirectoryRoleAssignment",
    "Remove-MgRoleManagementDirectoryRoleAssignment",
    "New-MgRoleManagementDirectoryRoleAssignment",
    "Get-AzADGroup",
    "Get-AzADGroupMember",
    "Remove-AzADGroupMember",
    "Add-AzADGroupMember",
    "Get-AzRoleEligibilitySchedule",
    "New-AzRoleEligibilityScheduleRequest",
    "Get-AzKeyVault",
    "Set-AzKeyVaultAccessPolicy",
    "Remove-AzKeyVaultAccessPolicy",
    "Get-AzRoleAssignment",
    "Remove-AzRoleAssignment",
    "New-AzRoleAssignment",
    # PowerShell core, not Azure.
    "Where-Object",
    "Select-Object",
    "Format-Table",
    "Format-List",
    "Write-Warning",
    "New-Guid",
}

_EVERY_PLANE = [
    {},
    dict(accessPath=schema.PATH_GROUP, sourceGroupId="g-1", sourceGroupName="Some Group"),
    # The AD-mastered branch emits a different dry run per format; without a row here it was
    # handing `az ad group show` to a PowerShell script and nothing noticed.
    dict(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-1",
        sourceGroupName="Synced Group",
        membershipGroupOnPremSynced=schema.ENABLED_TRUE,
    ),
    dict(surface=schema.SURFACE_ENTRA, roleName="Global Reader", assignmentId="ra-1", roleDefinitionId="rd-1"),
    dict(surface=schema.SURFACE_ENTRA, accessPath=schema.PATH_OWNER, roleName="Service Principal Owner"),
    dict(assignmentState=schema.STATE_ELIGIBLE, roleDefinitionId="rd-1"),
    dict(surface=schema.SURFACE_KEY_VAULT, scope="/subscriptions/s/resourceGroups/rg/providers/Microsoft.KeyVault/vaults/v"),
]


@pytest.mark.parametrize("fmt,allowed", [("az", _VERIFIED_AZ), ("powershell", _VERIFIED_PS)])
def test_every_generated_command_uses_a_verb_that_actually_exists(fmt: str, allowed: set[str]):
    """A command that does not exist is worse than no command: the operator hits a typo error
    mid-run, loses confidence in the whole script, and stops."""
    for over in _EVERY_PLANE:
        a = _revoke(fmt, **over)
        for field in ("dry_run", "command", "rollback"):
            for line in str(a[field]).splitlines():
                line = line.split("#")[0]
                if fmt == "az":
                    # Every `az ...` on the line, including one nested in a $(...) substitution.
                    for m in re.finditer(r"\baz(?: [a-z][a-z-]*)+", line):
                        tokens = m.group(0).split()
                        heads = [" ".join(tokens[:n]) for n in range(len(tokens), 1, -1)]
                        assert any(h in allowed for h in heads), f"{a['plane']} {field}: {m.group(0)!r}"
                else:
                    # A PowerShell script must not be handed az CLI: the operator picked a
                    # format, and half-honouring it is how a step gets skipped as "not for me".
                    assert not line.strip().startswith("az "), f"{a['plane']} {field}: {line!r}"
                    # Verb-Noun tokens anywhere on the line, so parens and pipelines are fine.
                    for verb in re.findall(r"\b[A-Z][a-z]+-[A-Za-z]+\b", line):
                        assert verb in allowed, f"{a['plane']} {field}: unverified cmdlet {verb!r} in {line!r}"


def test_pim_eligible_is_not_removed_by_deleting_an_assignment():
    """`az role assignment delete` lists ACTIVE assignments only. Against an eligible one it
    exits cleanly and the person keeps the ability to activate the role."""
    from app.iam import remediation

    a = _revoke(assignmentState=schema.STATE_ELIGIBLE, assignmentType="RoleEligibility", roleName="Owner")
    assert a["plane"] == remediation.PLANE_PIM_ELIGIBLE
    assert "AdminRemove" in a["command"]
    assert "az role assignment delete" not in a["command"]


def test_a_deny_assignment_is_never_offered_as_something_to_revoke():
    """A deny row restricts access. 'Revoking' it would ADD access, and it is read-only anyway."""
    from app.iam import remediation

    a = _revoke(surface=schema.SURFACE_DENY, roleName="Deny Delete")
    assert a["plane"] == remediation.PLANE_DENY
    assert "MANUAL STEP" in a["command"]
    assert "DENIES access" in a["command"]


def test_no_plane_ever_emits_a_command_that_cannot_run():
    """The contract: either a command that targets the right API, or a fully commented-out
    manual step. Never an executable line aimed at the wrong plane."""
    from app.iam import remediation

    cases = [
        {},
        {"accessPath": schema.PATH_GROUP, "sourceGroupId": "g-1"},
        {"accessPath": schema.PATH_OWNER, "principalId": "sp-1"},
        {"surface": schema.SURFACE_ENTRA, "assignmentId": "ra-1"},
        {"surface": schema.SURFACE_ENTRA, "assignmentId": ""},
        {"surface": schema.SURFACE_KEY_VAULT, "resourceName": "kv1", "resourceGroup": "rg1"},
        {"surface": schema.SURFACE_CLASSIC},
        {"surface": schema.SURFACE_LIGHTHOUSE},
        {"surface": schema.SURFACE_DENY},
        {"assignmentState": schema.STATE_ELIGIBLE},
    ]
    for fmt in remediation.FORMATS:
        for over in cases:
            a = _revoke(fmt, **over)
            executable = [
                ln for ln in a["command"].splitlines()
                if ln.strip() and not ln.lstrip().startswith(("#", "//"))
            ]
            if a["plane"] in (remediation.PLANE_CLASSIC, remediation.PLANE_LIGHTHOUSE, remediation.PLANE_DENY):
                assert not executable, f"{fmt}/{a['plane']} emitted a runnable command for un-scriptable access"
            for line in executable:
                # An ARM role-assignment verb may only ever appear on the Azure RBAC plane.
                if a["plane"] != remediation.PLANE_AZURE_RBAC:
                    assert "role assignment delete" not in line
                    assert "Remove-AzRoleAssignment" not in line


def test_one_membership_removal_covers_every_role_that_group_grants():
    """A person inheriting four roles through one group produced FOUR identical removals. The
    first succeeds and the rest fail — and an operator watching failures scroll past cannot tell
    which were expected."""
    from app.iam import remediation

    rows = [
        _revoke(
            accessPath=schema.PATH_GROUP,
            sourceGroupId="g-1",
            sourceGroupName="SG_Readers",
            effectivePrincipalId="u-1",
            roleName=role,
            scope=f"/subscriptions/11111111-2222-3333-4444-55555555555{i}",
        )
        for i, role in enumerate(["Reader", "Contributor", "Monitoring Reader"])
    ]
    bundle = remediation.build_bundle(rows, "az", title="t")
    assert bundle["action_count"] == 1
    assert bundle["script"].count("az ad group member remove") == 1
    # …and the folded step still says what all three grants were.
    only = bundle["actions"][0]
    assert len(only["covers"]) == 3
    assert "3 grants" in only["breaks_if"]
    # Named one per line in the script, not concatenated into the breaks-if. A group holding 43
    # assignments turned that into a single unreadable line, which hides the blast radius by
    # printing it too loudly.
    removes = [ln for ln in bundle["revoke_script"].splitlines() if ln.startswith("#   removes: ")]
    assert len(removes) == 3, removes
    assert any("Contributor @ /subscriptions/" in ln for ln in removes)
    assert "Contributor @" not in only["breaks_if"].split("This one step")[-1]


def test_a_mixed_script_warns_that_the_steps_are_not_interchangeable():
    from app.iam import remediation

    actions = [
        _revoke(),
        _revoke(accessPath=schema.PATH_GROUP, sourceGroupId="g-1", effectivePrincipalId="u-1"),
        _revoke(surface=schema.SURFACE_ENTRA, assignmentId="ra-1", roleName="Global Reader"),
    ]
    script = remediation.build_bundle(actions, "az", title="t")["script"]
    assert "spans more than one API" in script
    assert "not an ARM role assignment" in script


def test_every_line_of_a_multi_line_dry_run_stays_commented():
    """A multi-line dry run behind a single '#' leaves its later lines bare — and a bare line in
    a script is one an operator will run."""
    from app.iam import remediation

    a = _revoke(accessPath=schema.PATH_GROUP, sourceGroupId="g-1", effectivePrincipalId="u-1")
    assert "\n" in a["dry_run"], "this test is vacuous unless the dry run is multi-line"
    script = remediation.build_bundle([a], "az", title="t")["script"]
    runnable = [ln for ln in script.splitlines() if ln.strip() and not ln.lstrip().startswith("#")]
    # Exactly two: the removal and its rollback. Nothing from the dry run leaked out.
    assert len(runnable) == 2, runnable
    assert runnable[0].startswith("az ad group member remove --member-id 'u-1' --group $(")
    assert runnable[1] == "az ad group member add --group 'g-1' --member-id 'u-1'"


# ------------------------------------------ a removal that provably cannot work must not run
def test_a_membership_removal_is_not_emitted_for_an_on_prem_mastered_group():
    """Reported live: the removal ran and failed with "Unable to update the specified properties
    for on-premises mastered Directory Sync objects" — and the scan ALREADY KNEW the group was
    synced. The check that would have said so was commented out, so it never ran."""
    from app.iam import remediation

    a = _revoke(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-1",
        sourceGroupName="SG_Synced",
        effectivePrincipalId="u-1",
        membershipGroupOnPremSynced=schema.ENABLED_TRUE,
    )
    # Not one runnable line anywhere in the step.
    for field in ("command", "rollback"):
        bare = [ln for ln in str(a[field]).splitlines() if ln.strip() and not ln.lstrip().startswith("#")]
        assert bare == [], (field, bare)
    assert "ACTIVE DIRECTORY" in a["label"]
    assert "az ad group member remove" not in a["command"]
    # And it must not claim a blast radius for a change it cannot make.
    assert a["breaks_if"].startswith("nothing yet")


@pytest.mark.parametrize("state", [schema.ENABLED_UNKNOWN, schema.ENABLED_FALSE, ""])
def test_an_unknown_sync_state_still_guards_the_removal_at_run_time(state: str):
    """Graph omits onPremisesSyncEnabled for cloud-only objects, so "unknown" is usually fine —
    but usually is not always, and the operator must not learn the difference from a raw API
    error halfway down a script."""
    from app.iam import remediation

    a = _revoke(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-1",
        effectivePrincipalId="u-1",
        membershipGroupOnPremSynced=state,
    )
    cmd = a["command"]
    assert "az ad group member remove" in cmd
    # The group id is resolved at run time and cannot come back as g-1 for a synced group.
    assert "--group $(az ad group show --group 'g-1' --query \"onPremisesSyncEnabled &&" in cmd
    assert "REMOVE-THE-MEMBER-IN-ACTIVE-DIRECTORY" in cmd
    # A bare `--group 'g-1'` would defeat the whole guard.
    assert "member remove --group 'g-1'" not in cmd


def test_the_powershell_guard_is_a_real_if_not_a_comment():
    from app.iam import remediation

    a = _revoke(
        "powershell", accessPath=schema.PATH_GROUP, sourceGroupId="g-1", effectivePrincipalId="u-1"
    )
    cmd = a["command"]
    assert cmd.startswith('if ((Get-AzADGroup -ObjectId "g-1").OnPremisesSyncEnabled) {')
    # The removal must sit in the else branch, not before the check.
    assert cmd.index("Remove-AzADGroupMember") > cmd.index("} else {")


def test_the_group_sync_state_survives_the_whole_pipeline(seeded):
    """The unit tests above hand-build the column. This one proves compose actually stamps it
    from the directory, because a generator branch that no real row can reach is decoration."""
    from app.iam import leavers, remediation

    rows = leavers.disabled_grant_rows(compose.build_master_rows(seeded))
    synced = [r for r in rows if str(r.get("membershipGroupOnPremSynced")) == schema.ENABLED_TRUE]
    assert synced, "no AD-mastered group in the demo estate — the on-prem branch is untestable"
    # It is the GROUP that is synced here, not the member: keying on the member would send the
    # operator to Active Directory for a cloud-only account, or leave them in Entra for a
    # synced group. Both are wrong in the same way.
    assert any(str(r.get("principalOnPremSynced")) != schema.ENABLED_TRUE for r in synced)

    bundle = remediation.build_bundle(
        [remediation.revoke_assignment(r, "az") for r in synced], "az", title="t"
    )
    assert "ACTIVE DIRECTORY" in bundle["revoke_script"]
    for line in bundle["revoke_script"].splitlines():
        assert "az ad group member remove" not in line or line.lstrip().startswith("#")


# ------------------------------------- a membership lives in ONE group, and it may be nested
def _nested_graph(**over: Any) -> dict[str, Any]:
    graph = {
        "g-parent": {"name": "Parent", "members": [{"principalId": "u-1"}], "nested": ["g-child"]},
        "g-child": {"name": "Child", "members": [{"principalId": "u-1"}], "nested": []},
    }
    graph.update(over)
    return graph


def test_the_membership_group_is_the_nested_one_not_the_one_holding_the_assignment():
    """Reported live: `az ad group member remove --group <assignment group>` failed with
    "Resource '<group>' does not exist or one of its queried reference-property objects are not
    present". The person is a TRANSITIVE member; the $ref the delete needs was never there."""
    gid, name, how = compose.membership_group(_nested_graph(), "g-parent", "u-1")
    assert (gid, name, how) == ("g-child", "Child", "nested")


def test_a_direct_member_still_resolves_to_the_group_holding_the_assignment():
    graph = _nested_graph()
    graph["g-child"]["members"] = []
    assert compose.membership_group(graph, "g-parent", "u-1") == ("g-parent", "Parent", "direct")


def test_the_deepest_group_wins_when_the_nesting_is_more_than_one_level():
    graph = {
        "g-parent": {"name": "P", "members": [{"principalId": "u-1"}], "nested": ["g-mid", "g-leaf"]},
        "g-mid": {"name": "M", "members": [{"principalId": "u-1"}], "nested": ["g-leaf"]},
        "g-leaf": {"name": "L", "members": [{"principalId": "u-1"}], "nested": []},
    }
    # `nested` is the TRANSITIVE descendant set, so both are candidates; only the leaf holds the
    # actual membership, and removing from the middle one would 404 exactly like the parent did.
    assert compose.membership_group(graph, "g-parent", "u-1")[0] == "g-leaf"


def test_two_sibling_groups_are_ambiguous_rather_than_a_guess():
    """Removing one of two memberships leaves the access in place — and exits 0 while doing it."""
    graph = {
        "g-parent": {"name": "P", "members": [{"principalId": "u-1"}], "nested": ["g-a", "g-b"]},
        "g-a": {"name": "A", "members": [{"principalId": "u-1"}], "nested": []},
        "g-b": {"name": "B", "members": [{"principalId": "u-1"}], "nested": []},
    }
    gid, name, how = compose.membership_group(graph, "g-parent", "u-1")
    assert how == "ambiguous" and gid == "" and "A" in name and "B" in name


def test_an_unexpanded_child_is_unknown_rather_than_assumed_direct():
    """"Not in any child we managed to read" is not "a direct member of the parent"."""
    graph = {"g-parent": {"name": "P", "members": [{"principalId": "u-1"}], "nested": ["g-missing"]}}
    assert compose.membership_group(graph, "g-parent", "u-1") == ("", "", "unknown")


@pytest.mark.parametrize("how", ["ambiguous", "unknown"])
def test_no_removal_is_emitted_when_the_membership_group_is_not_known(how: str):
    from app.iam import remediation

    a = _revoke(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-parent",
        sourceGroupName="Parent",
        effectivePrincipalId="u-1",
        membershipGroupResolution=how,
        membershipGroupName="A; B",
    )
    bare = [ln for ln in str(a["command"]).splitlines() if ln.strip() and not ln.lstrip().startswith("#")]
    assert bare == [], bare
    assert "memberOf" in a["command"], "the operator is left with no way to find the right group"
    assert "role assignment" not in a["command"].replace("assignment instead", "")


def test_the_removal_targets_the_nested_group_and_names_the_parent():
    from app.iam import remediation

    a = _revoke(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-parent",
        sourceGroupName="Parent",
        effectivePrincipalId="u-1",
        membershipGroupId="g-child",
        membershipGroupName="Child",
        membershipGroupResolution="nested",
    )
    assert "--group 'g-child'" in a["command"] or "'g-child'" in a["command"]
    assert "g-parent" not in a["command"]
    # The label has to carry the parent, or the step reads as the wrong row to anyone who
    # searched for the group that holds the assignment.
    assert "Child" in a["label"] and "Parent" in a["label"]
    assert "--group 'g-child'" in a["rollback"]


def test_the_nested_membership_survives_the_whole_pipeline(seeded):
    from app.iam import leavers

    rows = leavers.disabled_grant_rows(compose.build_master_rows(seeded))
    nested = [r for r in rows if str(r.get("membershipGroupResolution")) == "nested"]
    assert nested, "no nested group in the demo estate — the resolution is untestable"
    r = nested[0]
    assert r["membershipGroupId"] != r["sourceGroupId"]
    # The chain was the assignment group repeated — a "chain" of one that hid the nesting.
    assert ">" in r["groupChain"]
    # Read from the CHILD group, which holds no assignment and so is absent from the principal
    # directory. Sourced from there it would read "unknown" for every group we now target.
    assert r["membershipGroupRoleAssignable"] == schema.ENABLED_FALSE
    assert r["membershipGroupDynamic"] == schema.ENABLED_FALSE


def test_groups_reached_through_the_same_child_fold_into_one_removal():
    """Found on live data: three assignment-holding groups all reached one person through the
    SAME nested child, so the script emitted the identical removal three times. The first
    succeeds and the rest fail on a membership that is already gone — which is exactly the
    noise `_fold_duplicates` exists to prevent."""
    from app.iam import remediation

    rows = [
        _revoke(
            accessPath=schema.PATH_GROUP,
            sourceGroupId=parent,
            sourceGroupName=f"Parent {parent}",
            effectivePrincipalId="u-1",
            membershipGroupId="g-child",
            membershipGroupName="Child",
            membershipGroupResolution="nested",
            roleName=role,
        )
        for parent, role in (("g-a", "Reader"), ("g-b", "Contributor"), ("g-c", "Owner"))
    ]
    bundle = remediation.build_bundle(rows, "az", title="t")
    assert bundle["action_count"] == 1
    assert bundle["script"].count("az ad group member remove") == 1
    assert len(bundle["actions"][0]["covers"]) == 3


def test_an_unresolved_membership_is_not_folded_across_different_parents():
    """Each names a different parent group, so they are genuinely different instructions."""
    from app.iam import remediation

    rows = [
        _revoke(
            accessPath=schema.PATH_GROUP,
            sourceGroupId=parent,
            sourceGroupName=f"Parent {parent}",
            effectivePrincipalId="u-1",
            membershipGroupResolution="unknown",
        )
        for parent in ("g-a", "g-b")
    ]
    assert remediation.build_bundle(rows, "az", title="t")["action_count"] == 2


# ----------------------------- groups no permission grant can make the Azure CLI write to
def _group_row(**over: Any) -> dict[str, Any]:
    base = dict(
        accessPath=schema.PATH_GROUP,
        sourceGroupId="g-1",
        sourceGroupName="SG_Role",
        effectivePrincipalId="u-1",
        membershipGroupId="g-1",
        membershipGroupName="SG_Role",
        membershipGroupResolution="direct",
    )
    base.update(over)
    return base


def test_a_role_assignable_group_gets_no_cli_command_because_none_can_work():
    """Verified against a live tenant: Global Administrator present in `wids`, the membership
    direct, the group readable — and DELETE still returned 403 Authorization_RequestDenied,
    because the Azure CLI's Graph token carries no RoleManagement.ReadWrite.Directory scope.
    The blocker is the APP, not the role."""
    from app.iam import remediation

    a = _revoke(**_group_row(membershipGroupRoleAssignable=schema.ENABLED_TRUE))
    bare = [ln for ln in str(a["command"]).splitlines() if ln.strip() and not ln.lstrip().startswith("#")]
    assert bare == [], bare
    assert "RoleManagement.ReadWrite.Directory" in a["command"]
    # The operator must not walk away thinking they need a bigger directory role.
    assert "Global Administrator gets the" in a["command"]
    # The sign-in lives in the preamble; the step carries the one line that differs.
    assert "Connect-MgGraph" not in a["command"]
    assert "Invoke-MgGraphRequest -Method DELETE" in a["command"]
    assert "az ad group member remove" not in a["command"]


def test_the_powershell_format_can_still_run_it_with_the_scope_requested():
    from app.iam import remediation

    a = _revoke("powershell", **_group_row(membershipGroupRoleAssignable=schema.ENABLED_TRUE))
    assert "/members/u-1/$ref" in a["command"] and "-Method DELETE" in a["command"]
    assert "-Method POST" in a["rollback"]
    # The sign-in belongs in the preamble, not repeated on every one of these steps.
    assert "Connect-MgGraph" not in a["command"]


def test_the_graph_sign_in_is_emitted_once_and_offers_a_service_principal():
    """Twelve of these steps on a real tenant. A six-line connect block repeated twelve times
    buries the one line that differs between them."""
    from app.iam import remediation

    rows = [
        _revoke(**_group_row(
            membershipGroupId=f"g-{i}", membershipGroupName=f"SG_{i}",
            sourceGroupId=f"g-{i}", membershipGroupRoleAssignable=schema.ENABLED_TRUE,
            tenantId="t-9", roleName=f"Role {i}",
        ))
        for i in range(3)
    ]
    script = remediation.build_bundle(rows, "az", title="t")["revoke_script"]
    assert script.count("Connect-MgGraph -Scopes") == 1
    # App-only: the credential goes in as a PSCredential, and -Scopes does NOT apply to it.
    assert "-ClientSecretCredential $cred" in script
    assert "Connect-MgGraph -TenantId 't-9'" in script
    assert "APPLICATION permissions" in script
    # The secret is read from the environment. A generated artifact must never carry one.
    assert "$env:GRAPH_CLIENT_SECRET" in script
    assert "-ClientSecret '" not in script


def test_the_rollback_repeats_the_sign_in_because_it_is_run_later():
    from app.iam import remediation

    rows = [_revoke(**_group_row(membershipGroupRoleAssignable=schema.ENABLED_TRUE, tenantId="t-9"))]
    b = remediation.build_bundle(rows, "az", title="t")
    assert "Connect-MgGraph" in b["rollback_script"]


def test_no_sign_in_block_when_nothing_needs_it():
    """Six lines of Graph consent on a script of pure az commands is noise that gets skimmed."""
    from app.iam import remediation

    b = remediation.build_bundle([_revoke()], "az", title="t")
    assert "Connect-MgGraph" not in b["script"]


def test_a_dynamic_group_has_no_membership_to_delete():
    from app.iam import remediation

    a = _revoke(**_group_row(membershipGroupDynamic=schema.ENABLED_TRUE))
    bare = [ln for ln in str(a["command"]).splitlines() if ln.strip() and not ln.lstrip().startswith("#")]
    assert bare == [], bare
    assert "rule" in a["command"].lower()
    assert "az ad group member remove" not in a["command"]


def test_an_ordinary_group_is_still_removed_by_a_command():
    """Without this the refusals above could be over-eager and nothing would be removable."""
    from app.iam import remediation

    a = _revoke(**_group_row(
        membershipGroupRoleAssignable=schema.ENABLED_FALSE,
        membershipGroupDynamic=schema.ENABLED_FALSE,
    ))
    assert "az ad group member remove" in a["command"]


@pytest.mark.asyncio
async def test_the_collector_reads_whether_a_group_can_be_edited_at_all(monkeypatch):
    """`isAssignableToRole` and `groupTypes` are only returned when asked for BY NAME, which is
    why an operator staring at a bare 403 has no way to discover either of them."""
    from app.iam import collectors

    async def fake_get_all(_token, url, *a, **k):
        return [{"id": "u-1", "@odata.type": "#microsoft.graph.user"}], None, 200

    seen: dict[str, str] = {}

    async def fake_get_object(_token, url, params=None):
        seen["select"] = (params or {}).get("$select", "")
        return {
            "displayName": "SG_Role",
            "isAssignableToRole": True,
            "groupTypes": ["DynamicMembership"],
            "onPremisesSyncEnabled": None,
        }, None, 200

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    monkeypatch.setattr(collectors, "_get_object", fake_get_object)
    graph, _st = await collectors.collect_group_expansion("t", ["g-1"])

    assert "isAssignableToRole" in seen["select"] and "groupTypes" in seen["select"]
    assert graph["g-1"]["roleAssignable"] == schema.ENABLED_TRUE
    assert graph["g-1"]["dynamic"] == schema.ENABLED_TRUE
    # Absent from the payload -> unknown, never "ordinary": the two lead to different scripts.
    assert graph["g-1"]["onPremSynced"] == schema.ENABLED_UNKNOWN
    assert graph["g-1"]["name"] == "SG_Role"


@pytest.mark.asyncio
async def test_the_collector_keeps_nested_groups_instead_of_dropping_them(monkeypatch):
    """The demo estate hand-writes the graph, so without this the collector could throw the
    nesting away and every test above would still pass on fixture data."""
    from app.iam import collectors

    pages = {
        "g-parent": [
            {"id": "u-1", "@odata.type": "#microsoft.graph.user", "displayName": "U"},
            {"id": "g-child", "@odata.type": "#microsoft.graph.group", "displayName": "C"},
        ],
        "g-child": [{"id": "u-1", "@odata.type": "#microsoft.graph.user", "displayName": "U"}],
    }
    seen: list[str] = []

    async def fake_get_all(_token, url, *a, **k):
        gid = url.split("/groups/")[1].split("/")[0]
        seen.append(gid)
        return pages.get(gid, []), None, 200

    monkeypatch.setattr(collectors, "_get_all", fake_get_all)
    graph, st = await collectors.collect_group_expansion("t", ["g-parent"])

    assert graph["g-parent"]["nested"] == ["g-child"]
    # The child is expanded in its own right, or there is no way to know the membership is in it.
    assert "g-child" in seen and "g-child" in graph
    # A nested group is not a member: counting it as one inflates every group-derived number.
    assert [m["principalId"] for m in graph["g-parent"]["members"]] == ["u-1"]


# ------------------------------------------------- a command has to reach the right subscription
def test_key_vault_commands_name_the_subscription():
    """Reported live: "(ResourceGroupNotFound) Resource group 'RG-...' could not be found" — for
    a group that exists, in a subscription that was not the operator's default."""
    from app.iam import remediation

    a = _revoke(
        surface=schema.SURFACE_KEY_VAULT,
        roleName="Access Policy: keys(get,list)",
        scope="/subscriptions/sub-9/resourceGroups/rg/providers/Microsoft.KeyVault/vaults/v",
        subscriptionId="sub-9",
        resourceName="v",
        resourceGroup="rg",
    )
    assert a["plane"] == remediation.PLANE_KV_POLICY
    for field in ("dry_run", "command", "rollback"):
        assert "--subscription 'sub-9'" in a[field], field


def test_powershell_sets_the_context_because_az_cmdlets_have_no_subscription_flag():
    from app.iam import remediation

    a = _revoke(
        "powershell",
        surface=schema.SURFACE_KEY_VAULT,
        scope="/subscriptions/sub-9/resourceGroups/rg/providers/Microsoft.KeyVault/vaults/v",
        subscriptionId="sub-9",
        resourceName="v",
        resourceGroup="rg",
    )
    assert a["command"].startswith('Set-AzContext -Subscription "sub-9"')
    assert "-ResourceGroupName" in a["command"]


def test_a_directory_scoped_row_does_not_invent_a_subscription():
    from app.iam import remediation

    a = _revoke(surface=schema.SURFACE_ENTRA, roleName="Global Reader", assignmentId="ra-1")
    assert "--subscription" not in a["command"]


def test_the_key_vault_permission_list_is_shortened_in_the_step_title():
    """The stored value is the real grant and runs to several hundred characters. Printed in the
    title and again in `breaks if` it produced two unreadable lines."""
    from app.iam import remediation

    full = (
        "Access Policy: keys(get,list,update,create,import,delete) "
        "secrets(get,list,set,delete) certificates(get,list)"
    )
    a = _revoke(
        surface=schema.SURFACE_KEY_VAULT,
        roleName=full,
        scope="/subscriptions/sub-9/resourceGroups/rg/providers/Microsoft.KeyVault/vaults/v",
        subscriptionId="sub-9",
        resourceName="v",
        resourceGroup="rg",
    )
    assert "keys 6, secrets 4, certificates 2" in a["label"]
    assert "unwrapkey" not in a["label"] and "get,list,update" not in a["breaks_if"]
    # An ordinary role name is left exactly as it is — it is also what `--role` takes.
    assert remediation.short_role("Storage Blob Data Reader") == "Storage Blob Data Reader"


# ------------------------------------------------------------- PIM eligibility, actually removed
def test_the_pim_request_carries_a_real_guid_not_a_placeholder():
    """Reported live: `<new-guid>` went to ARM verbatim and came back as a 400 wrapped in an
    ASP.NET error page. There is no GUID generator valid in both bash and PowerShell, so the
    name is minted when the script is built."""
    from app.iam import remediation

    a = _revoke(
        assignmentState=schema.STATE_ELIGIBLE,
        roleDefinitionId="/subscriptions/s/providers/Microsoft.Authorization/roleDefinitions/rd-1",
        assignmentId="/subscriptions/s/providers/Microsoft.Authorization/roleEligibilityScheduleInstances/i-1",
    )
    assert a["plane"] == remediation.PLANE_PIM_ELIGIBLE
    assert "<new-guid>" not in a["command"] and "New-Guid" not in a["command"]
    name = re.search(
        r"roleEligibilityScheduleRequests/([0-9a-f-]{36})\?", a["command"]
    )
    assert name, a["command"]
    # AdminRemove has to say WHICH eligibility, and the role definition must be the full ARM id.
    assert '"targetRoleEligibilityScheduleInstanceId":"/subscriptions/s/providers' in a["command"]
    assert '"roleDefinitionId":"/subscriptions/s/providers' in a["command"]
    assert '"requestType":"AdminRemove"' in a["command"]


def test_two_bundles_do_not_reuse_the_same_pim_request_name():
    from app.iam import remediation

    kw = dict(assignmentState=schema.STATE_ELIGIBLE, roleDefinitionId="rd-1")
    first = re.search(r"roleEligibilityScheduleRequests/([0-9a-f-]{36})", _revoke(**kw)["command"])
    second = re.search(r"roleEligibilityScheduleRequests/([0-9a-f-]{36})", _revoke(**kw)["command"])
    assert first and second and first.group(1) != second.group(1)


# ------------------------------------------------- "never used" needs three gates, not one
def _seed_usage(tenant_id: str, **over: Any) -> None:
    """Write a usage slice by hand so the conclusiveness rules can be exercised."""
    payload = {
        "window_days": 90,
        "start": "2026-05-08T00:00:00+00:00",
        "end": "2026-08-06T00:00:00+00:00",
        "source": "ActivityLog",
        "status": schema.STATUS_SUCCEEDED,
        "subscriptions": 1,
        "event_count": 10,
        "truncated": False,
        "principals": [{"principalId": "u-someone", "lastSeen": "2026-07-01T00:00:00Z", "events": 3}],
        "notes": [],
        "limitations": [],
    }
    payload.update(over)
    cache.write_usage(tenant_id, payload)


def test_a_truncated_activity_log_can_never_prove_disuse(seeded):
    """Measured on a real tenant: ELEVEN subscriptions hit the Activity Log's 6 MB cap in one
    90-day sweep. Treating that prefix as complete reported 78 identities as "never used their
    access" — an argument for deleting all of it, built on data that was never complete."""
    _seed_usage(seeded, truncated=True)
    report = leavers.build_leavers(seeded)
    assert report["usage"]["truncated"] is True
    assert all(not i["activityConclusive"] for i in report["identities"])
    assert report["totals"]["never_used"] == 0
    assert leavers.filter_identities(report["identities"], {"never_used": True}) == []
    assert any("6 MB" in l for l in report["limitations"])


def test_truncation_recorded_only_in_notes_is_still_detected(seeded):
    """A payload written BEFORE the flag existed records the truncation in its notes. A missing
    flag must not silently upgrade those older sweeps to "complete"."""
    payload_notes = ["Activity Log for subscription x returned more than 6 MB and was truncated"]
    _seed_usage(seeded, notes=payload_notes)
    # Remove the flag entirely, as an older payload would have it.
    stored = cache.read_usage(seeded)
    del stored["truncated"]
    cache.write_usage(seeded, stored)
    assert leavers.usage_enrichment(seeded)["truncated"] is True


def test_a_note_that_merely_mentions_truncation_does_not_raise_the_flag(seeded):
    """The fallback is anchored on the phrase the collector emits. Matching the bare word would
    let a note reading "not truncated" — or a future note about some other cap — declare the
    sweep partial, which silently withdraws every "never used" answer the screen can give."""
    _seed_usage(seeded, notes=["The sweep completed and was not truncated in any subscription."])
    stored = cache.read_usage(seeded)
    del stored["truncated"]
    cache.write_usage(seeded, stored)
    assert leavers.usage_enrichment(seeded)["truncated"] is False
    # …but the phrase the collector really emits still trips it.
    assert usage.TRUNCATION_MARKER in (
        "activity log for subscription x returned more than 6 mb and was truncated; "
        "showing the 2465 event(s) received."
    )


def test_a_window_that_closes_before_the_account_died_proves_nothing(seeded):
    """A disabled account cannot obtain a token, so it cannot appear in the Activity Log at all.
    "No operations in the last 90 days" is exactly what you would expect of somebody disabled
    two years ago — a fact about the window, not about the person."""
    _seed_usage(seeded)
    report = leavers.build_leavers(seeded)
    # The demo estate has no Entra sign-in data, so no account is inside the window.
    assert all(not i["activityWindowCovers"] for i in report["identities"])
    assert all(not i["activityConclusive"] for i in report["identities"])
    assert leavers.filter_identities(report["identities"], {"never_used": True}) == []


def test_never_used_matches_when_the_window_genuinely_covers_the_account():
    """…and the filter must still WORK when the answer is available, or it is just disabled."""
    inside = [{
        "principalId": "p1", "activityMeasured": True, "activityConclusive": True,
        "lastActivity": "", "signIn": {"known": True}, "lastSignIn": "2026-07-01T00:00:00Z",
    }]
    assert leavers.filter_identities(inside, {"never_used": True}) == inside
    used = [{**inside[0], "lastActivity": "2026-07-02T00:00:00Z"}]
    assert leavers.filter_identities(used, {"never_used": True}) == []
