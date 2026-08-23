"""Tests for the Entra ID App Registrations collector + facet aggregation."""
import asyncio

from app.identity import appregs


def _snap():
    return asyncio.run(appregs.collect_app_registrations(None, tenant_id="default"))


def test_demo_snapshot_shape():
    snap = _snap()
    assert snap["source"] == "demo_dummy_data"
    assert snap["connection_configured"] is False
    assert snap["summary"]["total"] == len(snap["apps"]) > 0
    # Sorted by display name (case-insensitive).
    names = [a["displayName"].lower() for a in snap["apps"]]
    assert names == sorted(names)


def test_permission_risk_tiers():
    assert appregs.permission_risk("Directory.ReadWrite.All") == "high"
    assert appregs.permission_risk("Directory.Read.All") == "medium"
    assert appregs.permission_risk("User.Read") == "low"


def test_app_row_counts_and_flags():
    apps = appregs.build_demo_app_registrations()
    by_name = {a["displayName"]: a for a in apps}

    # Ownerless + high-risk + expired credentials case.
    legacy = by_name["Legacy Migration Tool"]
    assert legacy["ownerless"] is True
    assert legacy["highRisk"] is True
    assert legacy["expiredCredentials"] == 2
    assert legacy["applicationPermissionsCount"] == 3
    assert legacy["delegatedPermissionsCount"] == 0

    # Mixed Application + Delegated permissions, secrets + cert.
    payments = by_name["Contoso Payments API"]
    assert payments["secretsCount"] == 1
    assert payments["certsCount"] == 1
    assert payments["applicationPermissionsCount"] == 3
    assert payments["delegatedPermissionsCount"] == 2
    assert payments["highRisk"] is True  # Directory.ReadWrite.All

    # Public client — no credentials.
    mobile = by_name["Field Service Mobile"]
    assert mobile["secretsCount"] == 0 and mobile["certsCount"] == 0
    assert mobile["nextExpiryDays"] is None
    assert mobile["highRisk"] is False


def test_aggregate_facets_and_summary():
    apps = appregs.build_demo_app_registrations()
    agg = appregs.aggregate(apps)

    # Audience facet totals reconcile with the app count.
    assert sum(f["count"] for f in agg["audiences"]) == len(apps)

    # Ownerless is represented in the owners facet.
    owners = {f["value"]: f["count"] for f in agg["owners"]}
    assert owners.get("(ownerless)") == agg["summary"]["ownerless"]

    # High-risk permissions surface in the permissions facet.
    perm_values = {f["value"] for f in agg["permissions"]}
    assert "Directory.ReadWrite.All" in perm_values

    s = agg["summary"]
    assert s["total"] == len(apps)
    assert s["withSecrets"] >= 1 and s["withCerts"] >= 1
    assert s["highRisk"] >= 1 and s["ownerless"] >= 1
    assert s["expired"] >= 1 and s["expiringSoon"] >= 1
    # Perm totals equal the sum of per-app counts.
    assert s["applicationPerms"] == sum(a["applicationPermissionsCount"] for a in apps)
    assert s["delegatedPerms"] == sum(a["delegatedPermissionsCount"] for a in apps)
    assert s["active"] + s["deactivated"] + s["notInstantiated"] + s["stateUnknown"] == len(apps)
    state_counts = {f["value"]: f["count"] for f in agg["enterpriseAppStates"]}
    assert set(state_counts) == set(appregs.ENTERPRISE_APP_STATES)
    assert sum(state_counts.values()) == len(apps)
    assert state_counts["deactivated"] == s["deactivated"] >= 1


def test_demo_signin_activity_covers_recent_dormant_and_stale():
    """The demo set has to exercise every rendered state, or the UI is untested by it."""
    snap = _snap()
    meta = snap["signin_activity"]
    assert meta["measured"] is True and meta["source"] == "demo"
    assert meta["credentials"]["measured"] is True

    by_name = {a["displayName"]: a for a in snap["apps"]}
    recent = by_name["Contoso Payments API"]
    assert recent["lastSignInKnown"] is True
    assert recent["lastSignIn"] and recent["lastSignInDays"] == 0

    # Dormant: measured, but nothing signed in. This must NOT look like "never".
    dormant = by_name["Legacy Migration Tool"]
    assert dormant["lastSignInKnown"] is True
    assert dormant["lastSignIn"] is None and dormant["lastSignInDays"] is None
    assert appregs.signin_bucket(dormant) == appregs.SIGNIN_BUCKET_NONE

    # Older than the report window is its own state, not "no sign-in".
    stale = by_name["Internal Wiki SSO"]
    assert appregs.signin_bucket(stale) == appregs.SIGNIN_BUCKET_OLD

    # Per-credential usage: the first credential is used, later ones are retirement candidates.
    gateway = by_name["Partner B2B Gateway"]
    assert [c["lastUsedKnown"] for c in gateway["credentials"]] == [True, True]
    assert all(c["lastUsed"] is None for c in gateway["credentials"])


def test_signin_buckets_partition_the_app_set():
    apps = appregs.build_demo_app_registrations()
    agg = appregs.aggregate(apps)
    buckets = {f["value"]: f["count"] for f in agg["signInActivity"]}

    assert list(buckets) == list(appregs.SIGNIN_BUCKETS)  # timeline order, not count order
    assert sum(buckets.values()) == len(apps)
    s = agg["summary"]
    assert s["signedIn30d"] + s["noRecentSignIn"] + s["signInNotMeasured"] == len(apps)
    assert s["signedIn7d"] <= s["signedIn30d"]
    # Non-vacuous: the dormant bucket is genuinely populated.
    assert s["noRecentSignIn"] >= 2


def test_unmeasured_activity_is_never_counted_as_dormant():
    apps = appregs.build_demo_app_registrations()
    for a in apps:
        a["lastSignInKnown"] = False
        a["lastSignIn"] = None
        a["lastSignInDays"] = None
    s = appregs.aggregate(apps)["summary"]
    assert s["signInNotMeasured"] == len(apps)
    assert s["noRecentSignIn"] == 0 and s["signedIn30d"] == 0


def test_cache_roundtrip(tmp_path, monkeypatch):
    from app.identity import appregs_cache

    monkeypatch.setattr(appregs_cache, "_CACHE_PATH", tmp_path / "appregs_cache.json")
    monkeypatch.setattr(appregs_cache, "_mem_cache", None)

    assert appregs_cache.get("t1", "c1") is None
    payload = {"apps": [], "summary": {"total": 0}}
    fetched_at = appregs_cache.set_("t1", "c1", payload)
    assert fetched_at
    hit = appregs_cache.get("t1", "c1")
    assert hit is not None
    assert hit["payload"] == payload
    assert hit["age_seconds"] >= 0


def test_workbook_multi_sheet():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.identity import appregs_export

    snap = _snap()
    content = appregs_export.to_workbook(snap)
    assert isinstance(content, bytes) and content[:2] == b"PK"  # xlsx is a zip

    wb = load_workbook(BytesIO(content), read_only=True)
    names = set(wb.sheetnames)
    assert {
        "Summary", "Applications", "Credentials", "API Permissions",
        "Owners", "High Risk", "Deactivated", "Permission Pivot",
    } <= names

    # Applications sheet: header + one row per app.
    ws = wb["Applications"]
    assert ws.max_row == len(snap["apps"]) + 1
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Enterprise app state" in headers
    assert "Service principal ID" in headers
    assert "Last sign-in" in headers and "Sign-in status" in headers

    # Credentials sheet: header + one row per credential across all apps.
    cred_total = sum(len(a.get("credentials") or []) for a in snap["apps"])
    assert wb["Credentials"].max_row == cred_total + 1
    cred_headers = [cell.value for cell in next(wb["Credentials"].iter_rows(min_row=1, max_row=1))]
    assert "Last used" in cred_headers and "Usage status" in cred_headers

    # API Permissions sheet: header + one row per granted permission.
    perm_total = sum(len(a.get("permissions") or []) for a in snap["apps"])
    assert wb["API Permissions"].max_row == perm_total + 1

    # High Risk sheet: header + only the flagged apps.
    hr = sum(1 for a in snap["apps"] if a.get("highRisk"))
    assert wb["High Risk"].max_row == hr + 1
    deactivated = sum(1 for a in snap["apps"] if a.get("enterpriseAppState") == "deactivated")
    assert wb["Deactivated"].max_row == deactivated + 1
    wb.close()


def test_workbook_neutralizes_formula_injection():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.identity import appregs_export

    snap = _snap()
    # Inject a formula-style display name into one app.
    snap["apps"][0]["displayName"] = "=cmd|'/c calc'!A1"
    content = appregs_export.to_workbook(snap)
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb["Applications"]
    # The cell must be neutralized to literal text (leading apostrophe / not a live formula).
    val = ws.cell(row=2, column=1).value
    assert str(val).startswith("'=") or not str(val).startswith("=")
    wb.close()


def test_date_columns_are_real_dates_so_excel_can_sort_them():
    """ISO text sorts lexicographically and offers no date filter; a typed cell does both."""
    import datetime as _dt
    from io import BytesIO

    from openpyxl import load_workbook

    from app.identity import appregs_export

    snap = _snap()
    wb = load_workbook(BytesIO(appregs_export.to_workbook(snap)))

    def column(ws, header):
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        return headers.index(header) + 1

    checked = 0
    for sheet, header in (("Applications", "Last sign-in"), ("Applications", "Created"),
                          ("Credentials", "Expires"), ("Credentials", "Last used")):
        ws = wb[sheet]
        ci = column(ws, header)
        populated = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ci)
            if cell.value is None or cell.value == "":
                continue
            populated += 1
            assert isinstance(cell.value, _dt.datetime), (
                f"{sheet}!{header} row {row} is {type(cell.value).__name__} "
                f"({cell.value!r}) — Excel cannot sort that as a date"
            )
            assert cell.value.tzinfo is None, "Excel cannot store a tz-aware datetime"
            assert cell.number_format == appregs_export.DATE_FORMAT
        # Non-vacuity: a column of entirely blank cells would prove nothing.
        assert populated, f"{sheet}!{header} had no populated cells to verify"
        checked += populated
    assert checked
    wb.close()


def test_unmeasured_signin_stays_blank_and_is_not_backdated_to_the_epoch():
    """Blank means "we did not measure"; a sentinel date would sort as 1970 and read as fact."""
    from io import BytesIO

    from openpyxl import load_workbook

    from app.identity import appregs_export

    snap = _snap()
    for a in snap["apps"]:
        a["lastSignInKnown"] = False
        a["lastSignIn"] = None
        a["lastSignInDays"] = None

    wb = load_workbook(BytesIO(appregs_export.to_workbook(snap)))
    ws = wb["Applications"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    date_ci = headers.index("Last sign-in") + 1
    status_ci = headers.index("Sign-in status") + 1
    for row in range(2, ws.max_row + 1):
        assert ws.cell(row=row, column=date_ci).value in (None, "")
        assert ws.cell(row=row, column=status_ci).value == "Not measured"
    wb.close()


def test_dt_parses_the_shapes_graph_actually_returns():
    import datetime as _dt

    from app.identity.appregs_export import _dt as parse

    # Graph returns a trailing 'Z'; fromisoformat could not read that before 3.11.
    assert parse("2026-08-15T09:22:41Z") == _dt.datetime(2026, 8, 15, 9, 22, 41)
    # An offset is normalised to UTC, then stripped (openpyxl rejects tz-aware).
    assert parse("2026-08-15T11:22:41+02:00") == _dt.datetime(2026, 8, 15, 9, 22, 41)
    assert parse("2026-08-15T09:22:41.123456Z") == _dt.datetime(2026, 8, 15, 9, 22, 41, 123456)
    assert parse("2026-08-15") == _dt.datetime(2026, 8, 15, 0, 0)
    # Absent or unparseable must be a blank cell, never a sentinel date.
    for junk in (None, "", "   ", "not a date", 12345, {}):
        assert parse(junk) is None


# --------------------------------------------------------------- attempt is not a sign-in
# Graph's sign-in reports count a REJECTED credential as activity: `lastSignInDateTime` is
# documented as the last attempt "either successfully or unsuccessfully", and
# `lastSuccessfulSignInDateTime` is the success-only companion. Reading the former as usage
# makes an application whose credential expired look like it was used yesterday, which is
# exactly the row a cleanup review must not skip.

def _row(**over):
    base = {
        "id": "1", "appId": "a", "displayName": "App", "credentials": [], "permissions": [],
        "owners": ["o"], "signInActivity": {"known": True},
    }
    base.update(over)
    return appregs._normalise_app(base)


def test_a_rejected_attempt_is_not_reported_as_a_sign_in():
    app = _row(signInActivity={
        "known": True, "last": "", "attempt": "2026-08-21T06:03:08Z",
    })
    assert app["lastSignIn"] is None, "an attempt must not populate the success field"
    assert app["lastSignInDays"] is None
    assert app["lastAttempt"] == "2026-08-21T06:03:08Z"
    assert app["lastAttemptDays"] is not None
    assert appregs.signin_bucket(app) == appregs.SIGNIN_BUCKET_FAILED


def test_a_successful_sign_in_is_still_reported_normally():
    app = _row(signInActivity={
        "known": True, "last": "2026-08-21T06:03:08Z", "attempt": "2026-08-21T06:03:08Z",
    })
    assert app["lastSignIn"] == "2026-08-21T06:03:08Z"
    assert appregs.signin_bucket(app) in (
        appregs.SIGNIN_BUCKET_RECENT, appregs.SIGNIN_BUCKET_WINDOW, appregs.SIGNIN_BUCKET_OLD,
    )


def test_no_activity_at_all_is_still_distinct_from_a_failed_attempt():
    app = _row(signInActivity={"known": True, "last": "", "attempt": ""})
    assert appregs.signin_bucket(app) == appregs.SIGNIN_BUCKET_NONE


def test_unreadable_report_outranks_everything_else():
    app = _row(signInActivity={"known": False, "last": "", "attempt": "2026-08-21T06:03:08Z"})
    assert appregs.signin_bucket(app) == appregs.SIGNIN_BUCKET_UNKNOWN


def test_the_collector_reads_the_successful_stamp_not_the_attempt():
    """The parsing layer, against the exact JSON shape Graph returns."""
    activity = {
        "lastSignInDateTime": "2026-08-21T06:03:08Z",
        "lastSuccessfulSignInDateTime": "2025-09-17T11:00:00Z",
    }
    assert appregs._newest_signin(activity) == "2026-08-21T06:03:08Z"
    assert appregs._newest_success(activity) == "2025-09-17T11:00:00Z"
    # The success property is not backfilled, so it can be absent while an attempt exists.
    only_attempt = {"lastSignInDateTime": "2026-08-21T06:03:08Z"}
    assert appregs._newest_signin(only_attempt) == "2026-08-21T06:03:08Z"
    assert appregs._newest_success(only_attempt) == ""


def test_failed_attempts_count_as_no_recent_sign_in_not_as_usage():
    recent = appregs._signed_in(1)
    apps = [
        _row(appId="ok", signInActivity={"known": True, "last": recent, "attempt": recent}),
        _row(appId="fail", signInActivity={"known": True, "last": "", "attempt": recent}),
    ]
    agg = appregs.aggregate(apps)
    assert agg["summary"]["signedIn7d"] == 1, "only the genuine sign-in counts"
    assert agg["summary"]["signedIn30d"] == 1
    assert agg["summary"]["noRecentSignIn"] == 1, "the rejected one is not usage"
    buckets = {f["value"]: f["count"] for f in agg["signInActivity"]}
    assert buckets[appregs.SIGNIN_BUCKET_FAILED] == 1
    assert buckets[appregs.SIGNIN_BUCKET_RECENT] == 1


def test_workbook_separates_the_failed_attempt_from_the_sign_in():
    import datetime as _dt
    from io import BytesIO

    from openpyxl import load_workbook

    from app.identity import appregs_export

    snap = _snap()
    target = snap["apps"][0]["appId"]
    for a in snap["apps"]:
        if a["appId"] == target:
            a["lastSignIn"], a["lastSignInDays"] = None, None
            a["lastAttempt"], a["lastAttemptDays"] = "2026-08-21T06:03:08Z", 1
            a["lastFailedSignIn"], a["lastFailedSignInDays"] = "2026-08-21T06:03:08Z", 1

    wb = load_workbook(BytesIO(appregs_export.to_workbook(snap)))
    ws = wb["Applications"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    row = next(r for r in range(2, ws.max_row + 1)
               if ws.cell(row=r, column=headers.index("App ID") + 1).value == target)

    assert ws.cell(row=row, column=headers.index("Last sign-in") + 1).value in (None, "")
    failed = ws.cell(row=row, column=headers.index("Last failed sign-in") + 1).value
    assert isinstance(failed, _dt.datetime) and failed == _dt.datetime(2026, 8, 21, 6, 3, 8)
    status = ws.cell(row=row, column=headers.index("Sign-in status") + 1).value
    assert "did not succeed" in status
    wb.close()


def test_the_row_derives_the_failed_sign_in_from_the_two_stamps():
    """End to end through `_normalise_app`, not just the helper."""
    failing = _row(signInActivity={
        "known": True, "last": "2025-09-17T11:00:00Z", "attempt": "2026-08-21T06:03:08Z",
    })
    assert failing["lastFailedSignIn"] == "2026-08-21T06:03:08Z"
    assert failing["lastFailedSignInDays"] is not None

    healthy = _row(signInActivity={
        "known": True, "last": "2026-08-21T06:03:08Z", "attempt": "2026-08-21T06:03:08Z",
    })
    assert healthy["lastFailedSignIn"] is None, "the last attempt WAS the success"
    assert healthy["lastSignIn"] == "2026-08-21T06:03:08Z"

