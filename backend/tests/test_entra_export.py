"""The Entra workbook export.

The behavior under test is almost entirely about honesty. A workbook is read offline, weeks
later, by somebody who cannot ask the tool a follow-up question — so a sheet that is empty
because a Graph permission was refused has to say so IN THE FILE, and a number that was
extrapolated from a sample has to carry that word next to it.
"""
from __future__ import annotations

import datetime as dt
import io
import re

import pytest
from openpyxl import load_workbook

from app.core import xlsx
from app.entra import export


def _wb(blob: bytes):
    return load_workbook(io.BytesIO(blob))


def _notes(wb) -> dict:
    """sheet name -> the Index note. Columns: Sheet, Section, Rows, Note."""
    return {r[0].value: r[3].value for r in wb["Index"].iter_rows(min_row=2)}


def _sections(wb) -> dict:
    return {r[0].value: r[1].value for r in wb["Index"].iter_rows(min_row=2)}


def _snapshot(**over):
    base = {
        "tenant_id": "t1",
        "loaded": True,
        "generated_at": "2026-08-02T00:00:00Z",
        "domains": {},
        "data": {},
        "licences": {},
        "permissions": {"granted": [], "granted_known": True, "domains": {}},
        "_analysis": {"score": {"score": 50, "coverage": 1.0, "pillars": []},
                      "findings": [], "by_signal": {}, "not_measured": {}, "errors": {}},
    }
    base.update(over)
    return base


# =========================================================================== blindness
def test_a_refused_domain_is_written_as_not_measured_not_as_an_empty_sheet():
    """The whole point. Entra domains go blind when a Graph scope is missing, and an empty
    sheet titled "Risky users" reads as "no risky users" — the opposite of the truth."""
    snap = _snapshot(
        domains={"risk": {"status": "blind", "missing_permissions": ["IdentityRiskEvent.Read.All"]}},
        data={"risk": {"risky_users": []}},
    )
    ws = _wb(export.to_workbook(snapshot=snap))["Sign-in summary"]
    assert ws.cell(row=2, column=1).value == "NOT MEASURED"
    detail = str(ws.cell(row=2, column=2).value)
    assert "blind" in detail
    assert "IdentityRiskEvent.Read.All" in detail, "the missing permission must be named"


@pytest.mark.parametrize("status", ["blind", "error", "not_collected", "unlicensed"])
def test_every_unreadable_status_blocks_the_sheet(status):
    snap = _snapshot(domains={"apps": {"status": status}}, data={"apps": {"service_principals": []}})
    ws = _wb(export.to_workbook(snapshot=snap))["Service principals"]
    assert ws.cell(row=2, column=1).value == "NOT MEASURED", status


def test_a_partial_domain_still_ships_its_rows_but_says_they_are_incomplete():
    """`partial` is not `blind`: the rows arrived, something alongside them was degraded.
    Suppressing them would lose real data; shipping them silently would overstate coverage."""
    snap = _snapshot(
        domains={"roles": {"status": "partial", "missing_permissions": ["RoleManagement.Read.All"]}},
        data={"roles": {"definitions": [{"display_name": "Global Administrator", "tier": "tier0"}],
                        "assignments": [], "group_derived": [], "eligible": []}},
    )
    blob = export.to_workbook(snapshot=snap)
    wb = _wb(blob)
    assert wb["Role definitions"].cell(row=2, column=1).value == "Global Administrator"
    assert "partial" in str(_notes(wb).get("Role definitions", ""))


def test_the_blind_spot_register_lists_every_domain():
    snap = _snapshot(domains={
        "ca": {"status": "ok"},
        "pim": {"status": "partial", "missing_permissions": ["RoleManagement.Read.Directory"]},
    })
    ws = _wb(export.to_workbook(snapshot=snap))["Coverage & blind spots"]
    names = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert {"ca", "pim"} <= names


def test_a_sampled_total_is_never_presented_as_a_count():
    """Sign-in totals are extrapolated when the collector sampled. A reader quoting "12
    failures" from a sampled window is quoting an estimate as a measurement."""
    snap = _snapshot(
        domains={"risk": {"status": "ok"}},
        data={"risk": {"signins": {"sampled": True, "total": 1000}, "risky_users": []}},
    )
    ws = _wb(export.to_workbook(snapshot=snap))["Sign-in summary"]
    rows = {ws.cell(row=r, column=1).value: str(ws.cell(row=r, column=2).value)
            for r in range(2, ws.max_row + 1)}
    assert "extrapolated" in rows["Sampled"].lower()


def test_an_empty_but_readable_domain_says_it_is_a_real_absence():
    """Zero lifecycle workflows on a READABLE governance domain is a configuration gap, not a
    blind spot. The two look identical in an empty sheet, so the note has to separate them."""
    snap = _snapshot(domains={"governance": {"status": "ok"}},
                     data={"governance": {"reviews": [], "packages": [], "assignments": [],
                                          "workflows": []}})
    wb = _wb(export.to_workbook(snapshot=snap))
    assert "not an absence of data" in str(_notes(wb).get("Lifecycle workflows", ""))


# =========================================================================== completeness
def test_nested_detail_becomes_its_own_sheet():
    """Credentials and grants are lists inside a row. Joined into one cell, an expiring secret
    is invisible; as rows, it is filterable."""
    snap = _snapshot(
        domains={"apps": {"status": "ok"}},
        data={"apps": {
            "service_principals": [{
                "display_name": "App A", "app_id": "a1",
                "credentials": [{"kind": "secret", "end": "2026-01-01", "expired": True, "days_left": -30}],
                "granted_app_permissions": [{"permission": "Directory.ReadWrite.All",
                                             "resource": "Microsoft Graph", "tier": "critical"}],
                "granted_delegated": [{"resource": "Microsoft Graph", "consent_type": "AllPrincipals",
                                       "scopes": ["User.Read", "Mail.Read"], "max_tier": "medium"}],
                "reply_urls": ["http://localhost"],
                "reply_url_risks": [{"uri": "http://localhost", "risk": "localhost"}],
                "owner_ids": ["u1"],
            }],
            "applications": [],
        }},
    )
    wb = _wb(export.to_workbook(snapshot=snap))
    perms = [[c.value for c in r] for r in wb["App permissions granted"].iter_rows(min_row=2)]
    # One row per SCOPE, not one per grant: a delegated grant of two scopes is two facts.
    assert {p[3] for p in perms} == {"Directory.ReadWrite.All", "User.Read", "Mail.Read"}
    assert any(p[6] == "AllPrincipals" for p in perms), "tenant-wide consent must be visible"
    creds = [[c.value for c in r] for r in wb["App credentials"].iter_rows(min_row=2)]
    assert creds[0][7] == "Yes" and creds[0][8] == -30
    urls = [[c.value for c in r] for r in wb["Redirect reply URLs"].iter_rows(min_row=2)]
    assert urls[0][5] == "localhost", "the reply-url risk must survive the join"
    assert wb["App owners"].cell(row=2, column=4).value == "u1"


def test_the_mfa_gap_is_every_enabled_user_without_a_method():
    """The screen caps this list at 500. An export that inherited the cap would be the same
    defect in a different wrapper."""
    users = [{"display_name": f"u{i}", "enabled": True, "mfa_registered": i % 2 == 0}
             for i in range(1200)]
    snap = _snapshot(domains={"people": {"status": "ok"}},
                     data={"people": {"users": users, "groups": []}})
    wb = _wb(export.to_workbook(snapshot=snap))
    assert wb["Users"].max_row - 1 == 1200
    assert wb["MFA registration gap"].max_row - 1 == 600


def test_the_index_reports_every_sheet_and_its_size():
    snap = _snapshot(domains={"people": {"status": "ok"}},
                     data={"people": {"users": [{"display_name": "a", "enabled": True}], "groups": []}})
    wb = _wb(export.to_workbook(snapshot=snap))
    assert wb.sheetnames[0] == "Index", "the contents page must come first in a 50-sheet file"
    listed = {r[0].value for r in wb["Index"].iter_rows(min_row=2)}
    assert listed >= set(wb.sheetnames) - {"Index", "Summary"}


def test_the_summary_warns_that_the_file_carries_personal_data():
    ws = _wb(export.to_workbook(snapshot=_snapshot()))["Summary"]
    text = " ".join(str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1))
    assert "PERSONAL DATA" in text


# =========================================================================== the xlsx helper
def test_a_formula_in_a_display_name_cannot_execute():
    """`=cmd|'/c calc'!A1` in a display name is a working RCE against the reviewer's
    workstation when the file is opened. Every cell is neutralised on the way in."""
    snap = _snapshot(
        domains={"people": {"status": "ok"}},
        data={"people": {"users": [{"display_name": "=cmd|'/c calc'!A1", "enabled": True}],
                         "groups": []}},
    )
    value = _wb(export.to_workbook(snapshot=snap))["Users"].cell(row=2, column=1).value
    assert str(value).startswith("'="), value


@pytest.mark.parametrize("raw", ["=1+1", "+1", "-1", "@SUM(A1)", " =1+1", "\t=1+1"])
def test_every_formula_trigger_is_neutralised(raw):
    assert str(xlsx.cell_safe(raw)).startswith("'")


def test_sheet_titles_are_legal_and_unique():
    """Excel rejects `[]:*?/\\` and truncates at 31 characters, and openpyxl silently renames a
    duplicate — so two long titles colliding would produce a sheet that is not what it says."""
    used: set[str] = set()
    a = xlsx.safe_title("Redirect / reply URLs", used)
    assert "/" not in a
    assert "  " not in a, "the sanitised character must not leave a gap that reads as a typo"
    long_a = xlsx.safe_title("A" * 40 + " one", used)
    long_b = xlsx.safe_title("A" * 40 + " two", used)
    assert len(long_a) <= 31 and len(long_b) <= 31
    assert long_a != long_b, "two titles truncating to the same 31 chars must be disambiguated"


def test_booleans_are_tri_state_so_no_is_not_confused_with_unknown():
    """"No", "not applicable" and "nobody looked" are three different answers on a review
    sheet. Rendering False as blank collapsed all three into the reassuring one."""
    assert xlsx.coerce(True) == "Yes"
    assert xlsx.coerce(False) == "No"
    assert xlsx.coerce(None) == ""
    assert xlsx.coerce(["a", "b"]) == "a, b"


def test_activation_history_comes_from_the_ledger_not_just_the_snapshot():
    """The snapshot holds only what the current collection saw. The screen merges it with a
    durable ledger that reaches past Graph's 30-day retention, and reading the snapshot alone
    lost 13 of 183 sessions on a live tenant \u2014 silently, which is the worst way to lose them."""
    snap = _snapshot(
        domains={"activations": {"status": "ok"}},
        data={"activations": {"sessions": [{"principal_name": "recent", "role_name": "GA"}]}},
    )
    merged = [{"principal_name": "recent", "role_name": "GA"},
              {"principal_name": "from the ledger", "role_name": "GA"}]
    wb = _wb(export.to_workbook(snapshot=snap, activations=merged))
    names = {r[2].value for r in wb["Activation sessions"].iter_rows(min_row=2)}
    assert names == {"recent", "from the ledger"}

    # And without the merge it must fall back rather than render nothing.
    wb2 = _wb(export.to_workbook(snapshot=snap))
    assert wb2["Activation sessions"].max_row - 1 == 1


# =========================================================================== tab colors
def _full_wb():
    """A workbook with every section populated, built from one readable snapshot."""
    snap = _snapshot(
        domains={d: {"status": "ok"} for d in
                 ("ca", "roles", "pim", "activations", "apps", "risk", "people", "governance")},
        data={
            "ca": {"policies": [{"display_name": "p", "conditions": {}}], "named_locations": [],
                   "auth_strengths": []},
            "_ca_analysis": {"coverage": {}, "conflicts": [], "breakglass": {}},
            "roles": {"definitions": [], "assignments": [], "group_derived": [], "eligible": []},
            "pim": {"policies": [], "group_eligibilities": []},
            "activations": {"sessions": []},
            "apps": {"service_principals": [], "applications": []},
            "risk": {"signins": {}, "risky_users": [], "risk_detections": [],
                     "risky_service_principals": [], "patterns": []},
            "people": {"users": [], "groups": []},
            "governance": {"reviews": [], "packages": [], "assignments": [], "workflows": []},
            "tenant": {},
            "_azure_link": {"available": False, "reason": "no Azure connection"},
        },
    )
    return _wb(export.to_workbook(snapshot=snap, escalations=[], setup_tiers=[], scanners=[]))


def test_every_sheet_from_one_parent_tab_shares_a_tab_colour():
    """The grouping a reader needs in a fifty-sheet file. Every Conditional Access sheet is the
    same color, and it is not the color any other section uses."""
    wb = _full_wb()
    by_section: dict[str, set[str]] = {}
    for name in wb.sheetnames:
        colour = wb[name].sheet_properties.tabColor
        section = _sections(wb).get(name)
        if section:
            by_section.setdefault(section, set()).add(str(colour.rgb) if colour else "")

    ca = by_section.get("Conditional Access")
    assert ca and len(ca) == 1, f"Conditional Access sheets disagree on colour: {ca}"

    # Every section is internally consistent...
    for section, colours in by_section.items():
        assert len(colours) == 1, f"{section} sheets disagree on colour: {colours}"
    # ...and no two sections share one, or the color would group the wrong things.
    flat = [next(iter(c)) for c in by_section.values()]
    assert len(flat) == len(set(flat)), f"two sections share a colour: {flat}"


def test_the_conditional_access_sheets_are_all_in_that_section():
    wb = _full_wb()
    sections = _sections(wb)
    ca_sheets = {n for n, sec in sections.items() if sec == "Conditional Access"}
    assert {"CA policies", "CA policy conditions", "CA conflicts"} <= ca_sheets


def test_the_index_names_the_section_as_well_as_colouring_it():
    """Color alone is not readable to everyone and does not survive a monochrome print."""
    wb = _full_wb()
    sections = _sections(wb)
    assert sections.get("CA policies") == "Conditional Access"
    assert sections.get("Users") == "Directory"
    assert sections.get("Findings") == "Findings & scanners"


def test_the_palette_covers_every_section_that_is_used():
    used = set(_sections(_full_wb()).values()) - {None, ""}
    assert used <= set(export.SECTION_COLOURS), f"section with no colour: {used - set(export.SECTION_COLOURS)}"


def test_tab_colours_are_opaque_or_excel_draws_nothing():
    """The bug this pins shipped once already.

    OOXML stores colors as aRGB and openpyxl pads a 6-digit value with ``00`` alpha \u2014 fully
    transparent. Every assertion about colors being present and distinct still passed, the
    round-trip read the value straight back, and Excel showed plain grey tabs. Only the alpha
    channel separates "purple" from "invisible"."""
    wb = _full_wb()
    for name in wb.sheetnames:
        colour = wb[name].sheet_properties.tabColor
        if colour is None:
            continue
        rgb = str(colour.rgb)
        assert len(rgb) == 8, f"{name}: {rgb} is not 8-digit ARGB"
        assert rgb[:2] == "FF", f"{name}: alpha {rgb[:2]} is transparent, Excel will draw nothing"


def test_the_colour_reaches_the_worksheet_xml_opaque():
    """Asserted on the FILE, not on the object model. The object model happily reports a color
    that the spreadsheet will never show."""
    import zipfile

    blob = export.to_workbook(
        snapshot=_snapshot(domains={"ca": {"status": "ok"}},
                           data={"ca": {"policies": [], "named_locations": [], "auth_strengths": []},
                                 "_ca_analysis": {}}),
    )
    z = zipfile.ZipFile(io.BytesIO(blob))
    seen = []
    for entry in z.namelist():
        if entry.startswith("xl/worksheets/sheet"):
            xml = z.read(entry).decode()
            i = xml.find('tabColor rgb="')
            if i >= 0:
                seen.append(xml[i + 14: i + 22])
    assert seen, "no tabColor reached the worksheet XML at all"
    assert all(v[:2] == "FF" for v in seen), f"transparent tab colours in the file: {set(seen)}"


@pytest.mark.parametrize(("given", "expected"), [
    ("7030A0", "FF7030A0"), ("#7030A0", "FF7030A0"), ("FF7030A0", "FF7030A0"), ("", ""),
])
def test_argb_normalisation(given, expected):
    assert xlsx.argb(given) == expected


# =========================================================================== dates
@pytest.mark.parametrize("raw", [
    "2026-06-26T16:34:22Z",                  # Graph's usual shape
    "2026-08-21T20:16:58.3437056+00:00",     # 7 fractional digits, from .NET
    "2026-08-21T19:31:27.7698413Z",
    "2021-10-21T14:49:06.15Z",
    "2026-06-26",
])
def test_every_shape_graph_emits_becomes_a_real_datetime(raw):
    got = xlsx.as_datetime(raw)
    assert isinstance(got, dt.datetime)
    assert got.tzinfo is None, "openpyxl refuses a tz-aware datetime outright"
    assert got.microsecond == 0


def test_an_offset_is_converted_not_discarded():
    """Dropping the tzinfo without normalising first would silently shift the clock."""
    assert xlsx.as_datetime("2026-06-26T18:00:00+02:00") == dt.datetime(2026, 6, 26, 16, 0, 0)


@pytest.mark.parametrize("raw", ["never", "not measured", "", None, 42, "2026-13-45"])
def test_a_value_that_is_not_a_date_is_left_alone_rather_than_vanishing(raw):
    """A column that quietly emptied itself is the failure this conversion exists to remove."""
    assert xlsx.as_datetime(raw) is None


def test_a_timestamp_column_is_a_date_in_the_sheet_not_a_string():
    """Text sorts lexically and only offers Excel's text filter, so "expired last month" is
    not a question a workbook of ISO strings can answer."""
    snap = _snapshot(
        domains={"people": {"status": "ok"}},
        data={"people": {"users": [{"display_name": "a", "enabled": True,
                                    "created_at": "2026-06-26T16:34:22Z"}], "groups": []}},
    )
    ws = _wb(export.to_workbook(snapshot=snap))["Users"]
    headers = [c.value for c in ws[1]]
    assert "Created (UTC)" in headers, "a converted column must say which zone it is in"
    cell = ws.cell(row=2, column=headers.index("Created (UTC)") + 1)
    assert isinstance(cell.value, dt.datetime)
    assert cell.number_format == xlsx.DATETIME_FMT


def test_an_unparseable_timestamp_survives_as_text_in_a_date_column():
    snap = _snapshot(
        domains={"people": {"status": "ok"}},
        data={"people": {"users": [{"display_name": "a", "enabled": True,
                                    "created_at": "unknown"}], "groups": []}},
    )
    ws = _wb(export.to_workbook(snapshot=snap))["Users"]
    headers = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=headers.index("Created (UTC)") + 1).value == "unknown"


def test_the_summary_timestamps_are_dates_too():
    snap = _snapshot(generated_at="2026-08-02T00:00:00Z", last_full="2026-08-01T23:00:00Z")
    ws = _wb(export.to_workbook(snapshot=snap))["Summary"]
    rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)}
    assert isinstance(rows["Snapshot generated (UTC)"], dt.datetime)
    assert isinstance(rows["Last full collection (UTC)"], dt.datetime)


def test_no_timestamp_anywhere_in_the_workbook_is_left_as_an_iso_string():
    """The regression guard for the whole change: one new sheet forgetting `dates=` puts an
    unsortable column back in front of a reviewer. The data-driven sheets are populated here
    on purpose — an empty sheet cannot fail this, and three of them did ship text timestamps.
    """
    iso = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}")
    snap = _snapshot(
        domains={"risk": {"status": "ok"}, "governance": {"status": "ok"},
                 "people": {"status": "ok"}},
        data={
            "people": {"users": [{"display_name": "a", "enabled": True,
                                  "created_at": "2026-01-01T00:00:00Z"}], "groups": []},
            "tenant": {},
            "risk": {"signins": {
                "window_start": "2026-08-22T02:15:48Z",
                "by_day": [{"day": "2026-08-22", "total": 10}],
                "by_app": [{"name": "app", "count": 3,
                            "last_seen": "2026-08-23T01:45:34Z"}],
            }, "risky_users": [{"name": "u", "last_updated": "2026-08-01T00:00:00Z"}],
                "risk_detections": [{"id": "d", "detected_at": "2026-08-01T00:00:00Z"}],
                "risky_service_principals": [{"name": "s",
                                              "last_updated": "2026-08-01T00:00:00Z"}],
                "patterns": [{"kind": "x", "count": 1, "first_seen": "2026-08-01T00:00:00Z"}]},
            "governance": {"reviews": [], "packages": [], "workflows": [],
                           "assignments": [{"id": "a1", "package_name": "P",
                                            "expires_at": "2022-04-12T17:50:50.42Z"}]},
        },
    )
    wb = _wb(export.to_workbook(snapshot=snap, escalations=[], setup_tiers=[], scanners=[]))
    offenders = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for i, v in enumerate(row):
                if isinstance(v, str) and iso.match(v):
                    offenders.append(f"{name}.{headers[i] if i < len(headers) else i}")
    assert not offenders, f"timestamps still exported as text: {sorted(set(offenders))}"


def test_a_discovered_column_of_timestamps_is_detected_by_value_not_by_name():
    """The sheets built from whatever keys the payload happens to carry cannot declare their
    date columns up front, so detection is on the values. Naming is never consulted — that is
    how a column of text ends up rendered as 1900-era dates."""
    heads, _rows, dates = export._explode([
        {"last_seen": "2026-08-23T01:45:34Z", "package": "Contributor Package", "count": 3},
        {"last_seen": "2026-08-22T04:26:57Z", "package": "Owner Package", "count": 4},
    ])
    assert dates == {"Last seen"}
    assert "Package" in heads and "Count" in heads


def test_a_column_that_is_only_sometimes_a_timestamp_is_left_as_text():
    """Half-converted is worse than unconverted: the text rows sort into a separate block."""
    _heads, _rows, dates = export._explode([
        {"when": "2026-08-23T01:45:34Z"}, {"when": "never"},
    ])
    assert dates == set()


def test_a_guest_sign_in_date_stays_a_date_and_the_sentinel_moves_to_its_own_column():
    """"Never signed in" and "never measured" must stay distinguishable, but neither belongs
    as a WORD in a date column — that is what made the column unsortable."""
    people = {"users": [
        {"display_name": "measured", "id": "g1", "user_type": "Guest",
         "mail": "a@partner.example", "upn": "a_partner.example#EXT#@contoso.onmicrosoft.com",
         "external_user_state": "Accepted", "created_at": "2026-01-01T00:00:00Z",
         "last_signin": "2026-06-01T00:00:00Z"},
    ], "groups": []}
    snap = _snapshot(domains={"people": {"status": "ok"}}, data={"people": people, "tenant": {}})
    ws = _wb(export.to_workbook(snapshot=snap))["Guests"]
    headers = [c.value for c in ws[1]]
    assert "Sign-in measured" in headers
    for label in ("Invited (UTC)", "Last human sign-in (UTC)"):
        assert label in headers
        for row in ws.iter_rows(min_row=2, values_only=True):
            value = row[headers.index(label)]
            assert value in (None, "") or isinstance(value, dt.datetime), value


# =========================================================================== dead columns
def test_the_cross_plane_sheet_carries_the_join_and_not_just_two_columns():
    """It read nine fields off the Azure link, which holds only the Azure half and under other
    names. Seven columns were blank on every row of the sheet whose whole point is the overlap."""
    snap = _snapshot(
        domains={"apps": {"status": "ok"}, "roles": {"status": "ok"}},
        data={
            "roles": {"definitions": [], "assignments": [], "group_derived": [], "eligible": []},
            "apps": {"applications": [], "service_principals": [{
                "display_name": "Powerful app", "object_id": "sp1", "app_id": "a1",
                "granted_app_permissions": [
                    {"permission": "Application.ReadWrite.All", "tier": "critical"}],
            }]},
            "_azure_link": {"available": True, "principals": {
                "sp1": {"name": "Powerful app", "powerful_roles": ["Owner"], "role_count": 4,
                        "broad_scopes": ["/subscriptions/s1"], "subscriptions": ["s1"]}}},
        },
    )
    ws = _wb(export.to_workbook(snapshot=snap))["Cross-plane power"]
    headers = [c.value for c in ws[1]]
    row = dict(zip(headers, [c.value for c in ws[2]]))
    assert row["Kind"] == "sp"
    assert row["Entra permissions"] == "Application.ReadWrite.All"
    assert row["Azure roles (powerful)"] == "Owner"
    assert row["Azure roles (all)"] == 4
    assert row["Broad scopes"] == "/subscriptions/s1"
    assert row["Both planes"] == "Yes"


def test_a_scanner_that_cannot_run_is_never_reported_as_runnable():
    """The column read `card['blocked'] or 'no'` against a card that never carried the key, so
    every scanner reported as fine \u2014 on the sheet that says which checks are switched off."""
    wb = _wb(export.to_workbook(
        snapshot=_snapshot(),
        scanners=[{"name": "Blocked one", "cadence": "daily", "signal_count": 3,
                   "blocked": "Domain 'apps' is blind."},
                  {"name": "Fine one", "cadence": "daily", "signal_count": 3, "blocked": ""}],
    ))
    ws = wb["Scanners"]
    headers = [c.value for c in ws[1]]
    rows = {r[0]: dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)}
    assert rows["Blocked one"]["Can run"] == "No"
    assert "blind" in rows["Blocked one"]["Why it cannot run"]
    assert rows["Fine one"]["Can run"] == "Yes"


def test_the_posture_sheet_shows_the_fields_a_pillar_actually_has():
    """`coverage` is not one of them, so that column was blank on every row."""
    snap = _snapshot()
    snap["_analysis"]["score"]["pillars"] = [
        {"key": "auth", "label": "Authentication", "score": None, "weight": 15,
         "state": "unmeasured", "reason": "Requires AuditLog.Read.All."},
    ]
    ws = _wb(export.to_workbook(snapshot=snap))["Posture score"]
    row = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))
    assert row["State"] == "unmeasured"
    assert "AuditLog.Read.All" in row["Why it is not complete"]


def test_no_configuration_is_stated_in_the_sheet_not_only_on_the_index():
    """A bare header row reads as a pass to anyone clicking through tabs."""
    snap = _snapshot(domains={"governance": {"status": "ok"}},
                     data={"governance": {"reviews": [], "packages": [], "assignments": [],
                                          "workflows": []}})
    ws = _wb(export.to_workbook(snapshot=snap))["Lifecycle workflows"]
    assert ws.cell(row=2, column=1).value == "NONE CONFIGURED"
    assert "not an absence of data" in str(ws.cell(row=2, column=4).value)


def test_the_mfa_gap_drops_the_columns_that_are_blank_by_construction():
    snap = _snapshot(domains={"people": {"status": "ok"}},
                     data={"people": {"users": [{"display_name": "a", "enabled": True}],
                                      "groups": []}})
    wb = _wb(export.to_workbook(snapshot=snap))
    gap = [c.value for c in wb["MFA registration gap"][1]]
    assert "MFA registered (UTC)" not in gap
    assert not ({"MFA registered", "MFA capable", "Passwordless capable", "Methods"} & set(gap))
    assert "MFA registered" in [c.value for c in wb["Users"][1]], "still meaningful on Users"


def test_no_sheet_ships_a_column_the_source_can_never_fill():
    """The four fixed here were blank on EVERY row of a fully populated workbook, because each
    read a key its payload has never carried. Asserted per-sheet against data that does supply
    the field: a blanket "no empty columns" sweep would only be measuring how complete the
    fixture is, and would fail for honest reasons the moment a sheet gained an optional field.
    """
    snap = _snapshot(
        domains={"apps": {"status": "ok"}, "roles": {"status": "ok"},
                 "people": {"status": "ok"}, "risk": {"status": "ok"}},
        data={
            "roles": {"definitions": [], "group_derived": [], "eligible": [],
                      "assignments": [{"principal_id": "u1", "principal_name": "Admin One",
                                       "role_name": "Global Administrator",
                                       "role_privileged": True}]},
            "people": {"users": [], "groups": []},
            "apps": {"applications": [], "service_principals": [{
                "display_name": "Powerful app", "object_id": "sp1", "app_id": "a1",
                "granted_app_permissions": [
                    {"permission": "Application.ReadWrite.All", "tier": "critical"}],
            }]},
            "risk": {"signins": {}, "risky_users": [], "risk_detections": [],
                     "risky_service_principals": [],
                     "patterns": [{"kind": "impossible_travel", "count": 146}]},
            "_azure_link": {"available": True, "principals": {
                "sp1": {"name": "Powerful app", "powerful_roles": ["Owner"], "role_count": 4,
                        "broad_scopes": ["/subscriptions/s1"], "subscriptions": ["s1"]},
                "u1": {"name": "Admin One", "powerful_roles": ["Owner"], "role_count": 2,
                       "broad_scopes": ["/subscriptions/s1"], "subscriptions": ["s1"]}}},
        },
    )
    snap["_analysis"]["score"]["pillars"] = [
        {"key": "auth", "label": "Authentication", "score": 40, "weight": 15,
         "state": "partial", "reason": "Requires AuditLog.Read.All."}]
    wb = _wb(export.to_workbook(
        snapshot=snap,
        scanners=[{"name": "s", "cadence": "daily", "severity_floor": "high", "signal_count": 3,
                   "blocked": "apps is blind",
                   "last_run": "2026-08-01T00:00:00Z", "last_counts": {"new": 1, "resolved": 2,
                                                                       "persisting": 3}}],
    ))
    for name in ("Cross-plane power", "Scanners", "Posture score", "Sign-in patterns"):
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows, f"{name} has no rows to judge"
        blank = [headers[i] for i in range(len(headers))
                 if headers[i] is not None
                 and all(r[i] in (None, "") for r in rows)]
        assert not blank, f"{name}: columns blank on every row: {blank}"


# =========================================================================== navigation
def test_the_index_entries_are_internal_links_not_file_links():
    """`target` would be written as an external relationship and Excel would try to open it as
    a file; `location` is what jumps within the workbook."""
    wb = _full_wb()
    cell = wb["Index"].cell(row=2, column=1)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target is None
    assert cell.hyperlink.location.startswith("'")


def test_the_header_row_holds_nothing_but_headers():
    """No navigation cell parked to the right of the data: it still lands in the header row,
    where anything reading the file programmatically picks it up as a column."""
    wb = _full_wb()
    for name in set(wb.sheetnames) - {"Summary"}:
        headers = [c.value for c in wb[name][1]]
        assert all(h not in (None, "") for h in headers), f"{name}: stray header cell {headers}"


def test_every_populated_sheet_is_a_real_excel_table():
    wb = _full_wb()
    for name in set(wb.sheetnames) - {"Index", "Summary"}:
        ws = wb[name]
        if ws.max_row > 1:
            assert ws.tables, f"{name} is a plain range, not a table"


def test_table_names_are_unique_across_the_workbook():
    """They are workbook-scoped; a collision makes Excel offer to repair the file."""
    wb = _full_wb()
    names = [t for name in wb.sheetnames for t in wb[name].tables]
    assert len(names) == len(set(names)), f"duplicate table names: {names}"


# =========================================================================== structure
def test_a_breakdown_becomes_one_sheet_per_dimension_with_real_columns():
    """Stacked in one Dimension/Key/Count/Detail grid they could not be pivoted, and every
    figure past the count was stringified into the Detail cell as `k=v; k=v`."""
    snap = _snapshot(
        domains={"risk": {"status": "ok"}},
        data={"risk": {"signins": {"by_day": [
            {"day": "2026-08-22", "total": 6532, "success": 3628, "failure": 2904},
        ]}, "risky_users": [], "risk_detections": [], "risky_service_principals": [],
            "patterns": []}},
    )
    wb = _wb(export.to_workbook(snapshot=snap))
    ws = wb["Sign-ins by day"]
    headers = [c.value for c in ws[1]]
    assert {"Total", "Success", "Failure"} <= set(headers)
    assert "Day (UTC)" in headers
    assert ws.cell(row=2, column=headers.index("Success") + 1).value == 3628


def test_a_breakdown_the_collector_did_not_return_says_so():
    snap = _snapshot(domains={"risk": {"status": "ok"}},
                     data={"risk": {"signins": {}, "risky_users": [], "risk_detections": [],
                                    "risky_service_principals": [], "patterns": []}})
    ws = _wb(export.to_workbook(snapshot=snap))["Sign-in breakdowns"]
    assert ws.cell(row=2, column=1).value == "NOT MEASURED"


def test_headers_use_the_products_spelling():
    """Column headers follow the product's spelling conventions. The underlying `licence_count`
    key is a wire name and is deliberately left alone."""
    wb = _full_wb()
    off_convention = {"Licences", "Organisation", "Colour", "Analyse"}
    for name in wb.sheetnames:
        headers = {str(c.value) for c in wb[name][1]}
        assert not (headers & off_convention), f"{name}: {headers & off_convention}"

