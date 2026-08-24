"""Data-plane access catalog and signals.

The behavior under test is a classification, so most of these assert on the tier a REAL Azure
role receives. Every role name and every dataAction string here was copied from a live tenant's
role catalog (981 definitions, 324 with dataActions) rather than from documentation — an
invented action string would make every test pass against a classifier that never runs.
"""
from __future__ import annotations

import pytest

from app.iam import dataplane as dp, schema
from app.iam.signal_defs import dp as dp_signals
from app.iam.signals import SignalContext, SignalUnavailable

# Real dataActions, verbatim from a live catalog.
KV_SECRETS_USER = ["Microsoft.KeyVault/vaults/secrets/getSecret/action"]
KV_ADMIN = ["Microsoft.KeyVault/vaults/*"]
KV_READER = ["Microsoft.KeyVault/vaults/*/read"]
BLOB_READER = ["Microsoft.Storage/storageAccounts/blobServices/containers/blobs/read"]
BLOB_OWNER = [
    "Microsoft.Storage/storageAccounts/blobServices/containers/blobs/*",
]
EH_RECEIVER = ["Microsoft.EventHub/*/receive/action"]
EH_SENDER = ["Microsoft.EventHub/*/send/action"]
AKS_CLUSTER_ADMIN = ["Microsoft.ContainerService/managedClusters/*"]
AKS_READER = [
    "Microsoft.ContainerService/managedClusters/apps/deployments/read",
    "Microsoft.ContainerService/managedClusters/batch/cronjobs/read",
    "Microsoft.ContainerService/managedClusters/secrets/read",
]
# Verbatim. The `query/action` entry is the one that matters: it is a READ expressed as an
# action, and it is what made the original classifier call this read-only role a write role.
TWINS_READER = [
    "Microsoft.DigitalTwins/digitaltwins/read",
    "Microsoft.DigitalTwins/digitaltwins/relationships/read",
    "Microsoft.DigitalTwins/eventroutes/read",
    "Microsoft.DigitalTwins/jobs/deletions/read",
    "Microsoft.DigitalTwins/jobs/import/read",
    "Microsoft.DigitalTwins/models/read",
    "Microsoft.DigitalTwins/query/action",
]
QUEUE_PROCESSOR = [
    "Microsoft.Storage/storageAccounts/queueServices/queues/messages/process/action",
    "Microsoft.Storage/storageAccounts/queueServices/queues/messages/read",
]


# =========================================================================== tiers
@pytest.mark.parametrize(
    ("name", "actions", "expected"),
    [
        # Reading a secret is an identity takeover, not a read.
        ("Key Vault Secrets User", KV_SECRETS_USER, dp.TIER_CREDENTIAL),
        ("Key Vault Administrator", KV_ADMIN, dp.TIER_CREDENTIAL),
        ("Storage Blob Data Reader", BLOB_READER, dp.TIER_READ),
        ("Storage Blob Data Owner", BLOB_OWNER, dp.TIER_WRITE),
        ("Azure Event Hubs Data Receiver", EH_RECEIVER, dp.TIER_READ),
        ("Azure Event Hubs Data Sender", EH_SENDER, dp.TIER_WRITE),
        ("Azure Kubernetes Service RBAC Cluster Admin", AKS_CLUSTER_ADMIN, dp.TIER_WRITE),
        ("Storage Queue Data Message Processor", QUEUE_PROCESSOR, dp.TIER_WRITE),
    ],
)
def test_tier_of_real_roles(name, actions, expected):
    assert dp.role_tier(name, actions) == expected


def test_key_vault_reader_cannot_open_anything():
    """`Microsoft.KeyVault/vaults/*/read` lists secret NAMES; reading a VALUE is
    `.../getSecret/action`. Classifying the two alike turns the harmless role into a critical
    credential finding, and a report that cries wolf on Key Vault Reader gets switched off
    before it reports Key Vault Administrator.

    Asserted through :func:`derive_tier` as well as :func:`role_tier`, because the catalog
    carries an explicit override for this role — testing only the public function would pass
    against a derivation that calls every Key Vault grant a credential."""
    assert dp.derive_tier(KV_READER) == dp.TIER_META
    assert dp.derive_tier(KV_SECRETS_USER) == dp.TIER_CREDENTIAL
    assert dp.role_tier("Key Vault Reader", KV_READER) == dp.TIER_META
    assert not dp.is_privileged_data_role("Key Vault Reader", KV_READER)


def test_a_read_expressed_as_an_action_is_still_a_read():
    """Azure models plenty of reads as `/action` — `query`, `receive`, `filter` — so the verb
    `action` alone carries no meaning. Treating every `/action` as a write reported `Azure
    Digital Twins Data Reader` and `Azure Event Hubs Data Receiver` as able to destroy data."""
    assert dp.derive_tier(["Microsoft.DigitalTwins/query/action"]) == dp.TIER_READ
    assert dp.derive_tier(EH_RECEIVER) == dp.TIER_READ
    assert dp.derive_tier(EH_SENDER) == dp.TIER_WRITE


def test_verbs_are_matched_as_segments_not_substrings():
    """`"/manage" in action` matched every single
    `Microsoft.ContainerService/managedClusters/...` string, so `Azure Kubernetes Service RBAC
    Reader` — 31 actions, every one of them a `/read` — was classified as able to destroy data.
    Any classifier that inspects the whole string rather than the final segment fails here."""
    assert dp.derive_tier(AKS_READER) == dp.TIER_READ
    assert dp.derive_tier(TWINS_READER) == dp.TIER_READ
    assert dp.role_tier("Azure Kubernetes Service RBAC Reader", AKS_READER) == dp.TIER_READ


def test_a_custom_role_is_judged_on_what_it_does_not_what_it_is_called():
    """Derivation, not a name list: a benign-sounding custom role carrying a Key Vault wildcard
    is a credential role."""
    assert dp.role_tier("Team Helper", KV_ADMIN) == dp.TIER_CREDENTIAL
    assert dp.role_tier("Emergency Superuser", BLOB_READER) == dp.TIER_READ


def test_a_role_with_no_data_actions_is_not_a_data_role():
    assert dp.role_tier("Reader", []) == dp.TIER_NONE
    assert dp.derive_tier(None) == dp.TIER_NONE


# =========================================================================== privileged flag
def test_privileged_now_catches_the_roles_the_name_test_missed():
    """On a real 981-role catalog the old test — dataActions plus "owner"/"contributor" in the
    name — missed 118 dangerous roles, including every Key Vault officer role and AKS cluster
    admin."""
    for name, actions in [
        ("Key Vault Administrator", KV_ADMIN),
        ("Key Vault Secrets User", KV_SECRETS_USER),
        ("Azure Kubernetes Service RBAC Cluster Admin", AKS_CLUSTER_ADMIN),
        ("Azure Event Hubs Data Sender", EH_SENDER),
    ]:
        assert schema.role_is_privileged(name, has_data_actions=True, data_actions=actions), name


def test_the_name_test_is_kept_as_well_never_swapped():
    """The two tests are unioned on purpose.

    The name test has false positives, and dropping it would have demoted `Log Analytics
    Contributor` and eight others from privileged — a REDUCTION in coverage dressed up as a
    precision fix. This flag drives the privileged counts and the PIM screens, where a false
    negative hides access and a false positive only adds noise."""
    # Read-tier dataActions: the catalog alone would say "not privileged".
    assert dp.role_tier("Log Analytics Contributor", BLOB_READER) == dp.TIER_READ
    assert schema.role_is_privileged(
        "Log Analytics Contributor", has_data_actions=True, data_actions=BLOB_READER
    )


def test_read_only_data_roles_are_not_privileged_by_the_catalogue():
    assert not dp.is_privileged_data_role("Storage Blob Data Reader", BLOB_READER)


# =========================================================================== the catalog
def test_every_blind_service_states_a_reason():
    """A blind spot without a stated reason is indistinguishable from a pass."""
    for svc in dp.UNREADABLE:
        assert svc.blind_reason.strip(), f"{svc.key} claims incomplete RBAC with no reason"


def test_the_services_that_azure_rbac_cannot_describe_are_marked():
    """These publish no ARM data-plane roles at all: their grants live in a service-native
    system. Marking them readable would let an empty result read as a clean one."""
    for key in ("sql", "aks", "cosmos", "databricks", "redis", "kusto", "postgres", "mysql"):
        assert not dp.SERVICE_BY_KEY[key].rbac_is_complete, key


def test_key_vault_is_readable_because_its_policies_are_collected():
    """The legacy access-policy model IS collected, so Key Vault is not a blind spot."""
    assert dp.SERVICE_BY_KEY["keyvault"].rbac_is_complete


def test_storage_accounts_map_to_every_data_service_they_host():
    services = {s.key for s in dp.services_for_type("microsoft.storage/storageaccounts")}
    assert {"blob", "files", "queue", "table"} <= services


# =========================================================================== signals
def _ctx(**kw) -> SignalContext:
    base = dict(tenant_id="t1", rows=[], kpis={}, scopes=[])
    base.update(kw)
    return SignalContext(**base)


def _grant(role: str, scope: str, scope_type: str, who: str = "alice") -> dict:
    return schema.make_row(
        surface=schema.SURFACE_AZURE_RBAC, effect=schema.EFFECT_ALLOW,
        assignmentState=schema.STATE_ACTIVE, accessPath=schema.PATH_DIRECT,
        principalId=who, effectivePrincipalId=who, effectivePrincipalName=who.title(),
        effectivePrincipalType="User", roleName=role, scope=scope, scopeType=scope_type,
        assignmentId=f"{who}-{role}-{scope}", principalExists=schema.EXISTS_TRUE,
    )


def _defs(*pairs):
    return {"role_defs": [{"roleName": n, "dataActions": a} for n, a in pairs]}


def test_credential_access_is_critical_at_a_wide_scope_and_error_at_a_narrow_one():
    """The scope multiplies it: a credential grant on a management group reaches every vault
    beneath it, including vaults that do not exist yet."""
    directory = _defs(("Key Vault Secrets User", KV_SECRETS_USER))
    wide = _ctx(
        rows=[_grant("Key Vault Secrets User", "/providers/Microsoft.Management/managementGroups/root",
                     schema.SCOPE_MANAGEMENT_GROUP)],
        directory=directory,
    )
    narrow = _ctx(
        rows=[_grant("Key Vault Secrets User", "/subscriptions/s1/resourceGroups/rg/providers/x/vaults/v",
                     schema.SCOPE_RESOURCE)],
        directory=directory,
    )
    assert dp_signals._credential_access(wide)[0].severity == "critical"
    assert dp_signals._credential_access(narrow)[0].severity == "error"


def test_a_metadata_role_is_not_a_credential_finding():
    ctx = _ctx(
        rows=[_grant("Key Vault Reader", "/subscriptions/s1", schema.SCOPE_SUBSCRIPTION)],
        directory=_defs(("Key Vault Reader", KV_READER)),
    )
    assert dp_signals._credential_access(ctx) == []


def test_without_the_role_catalogue_nothing_is_claimed():
    """No dataActions means no way to tell a data role from a control-plane one. Returning an
    empty list there would report "no data-plane exposure" from having never looked."""
    ctx = _ctx(rows=[_grant("Key Vault Secrets User", "/subscriptions/s1", schema.SCOPE_SUBSCRIPTION)],
               directory={})
    with pytest.raises(SignalUnavailable):
        dp_signals._credential_access(ctx)


def test_blind_spots_are_counted_by_RESOURCE_not_by_check():
    """The sweep emits one row per CHECK — four per storage account — so counting rows reports
    four times the estate. The first version also read the wrong field entirely (`type` instead
    of `resourceType`), matched nothing, and reported a confident ZERO blind spots on a tenant
    with 15 SQL servers and a Cosmos account."""
    rows = [
        {"resourceType": "microsoft.sql/servers", "resourceId": "/s/1", "resourceName": "sql-a", "key": "sql.auth"},
        {"resourceType": "microsoft.sql/servers", "resourceId": "/s/1", "resourceName": "sql-a", "key": "sql.other"},
        {"resourceType": "microsoft.sql/servers", "resourceId": "/s/2", "resourceName": "sql-b", "key": "sql.auth"},
    ]
    ctx = _ctx(bypass_rows=rows, bypass_assessed=2)
    found = [f for f in dp_signals._unreadable_authorization(ctx) if f.evidence["service"] == "sql"]
    assert len(found) == 1
    assert found[0].count == 2, "two distinct SQL servers, not three check rows"
    assert sorted(found[0].evidence["resources"]) == ["sql-a", "sql-b"]


def test_a_service_whose_access_we_can_read_is_not_reported_as_blind():
    ctx = _ctx(
        bypass_rows=[{"resourceType": "microsoft.eventhub/namespaces", "resourceId": "/e/1",
                      "resourceName": "eh", "key": "eventhub.local_auth"}],
        bypass_assessed=1,
    )
    assert dp_signals._unreadable_authorization(ctx) == []


def test_no_sweep_means_not_measured_never_zero():
    with pytest.raises(SignalUnavailable):
        dp_signals._unreadable_authorization(_ctx(bypass_rows=[], bypass_assessed=0))


# =========================================================================== wiring
def test_the_signals_are_reachable_through_the_registry():
    """Exercised through the REGISTERED specs, not the private functions.

    A signal whose evaluate is correct but which is not wired into the registry never runs, and
    a pillar that never runs reports nothing while looking configured."""
    from app.iam import signals as sig

    ids = {s.id for s in sig.all_signals() if s.pillar == "dp"}
    assert ids == {
        "dp.credential_store_access",
        "dp.write_wide_scope",
        "dp.authorization_not_readable",
    }

    by_id = {s.id: s for s in sig.all_signals()}
    ctx = _ctx(
        rows=[
            _grant("Key Vault Secrets User", "/subscriptions/s1", schema.SCOPE_SUBSCRIPTION),
            _grant("Storage Blob Data Owner", "/subscriptions/s1", schema.SCOPE_SUBSCRIPTION, who="bob"),
        ],
        directory=_defs(
            ("Key Vault Secrets User", KV_SECRETS_USER),
            ("Storage Blob Data Owner", BLOB_OWNER),
        ),
    )
    cred = by_id["dp.credential_store_access"].evaluate(ctx)
    assert [f.signal_id for f in cred] == ["dp.credential_store_access"]
    assert cred[0].severity == "critical"

    write = by_id["dp.write_wide_scope"].evaluate(ctx)
    assert [f.signal_id for f in write] == ["dp.write_wide_scope"]
    # The credential role must not also be counted as a plain write, or one grant becomes two
    # findings that a reader has to reconcile.
    assert write[0].evidence["roles"] == ["Storage Blob Data Owner"]


def test_the_data_plane_scanner_selects_the_pillar():
    from app.iam import scanners

    spec = scanners.get("iam.data_plane")
    assert spec is not None
    assert {s.id for s in spec.signals()} == {
        "dp.credential_store_access",
        "dp.write_wide_scope",
        "dp.authorization_not_readable",
    }


# =========================================================================== the workbook
def test_the_workbook_distinguishes_a_deny_from_a_grant():
    """A Deny assignment used to be exported into a sheet called "Effective Access" with no
    column to tell it apart from the 5,506 grants beside it. Anyone filtering or pivoting that
    sheet counted a CONTROL as access."""
    from app.iam import export

    assert "effect" in export._ACCESS_HEADERS
    # And it leads, so it cannot be missed by someone reading left to right.
    assert export._ACCESS_HEADERS[0] == "effect"


def test_the_workbook_shows_whether_the_principal_still_exists():
    """100 rows on a real tenant point at deleted principals. Without the column they read as
    live access, which is how an orphaned grant survives a review."""
    from app.iam import export

    assert "principalExists" in export._ACCESS_HEADERS


def test_the_workbook_carries_the_identifiers_needed_to_act_on_a_row():
    from app.iam import export

    assert "assignmentId" in export._ACCESS_HEADERS
    assert "roleDefinitionId" in export._ACCESS_HEADERS


def test_an_unmeasured_analysis_is_written_as_not_measured_not_as_an_empty_sheet():
    """An empty sheet titled "Shadow Access" reads as "no shadow access". The sweep not having
    run is a different statement and has to survive into the file."""
    import io

    from openpyxl import load_workbook

    from app.iam import export

    blob = export.to_workbook(
        rows=[], overview={"kpis": {}, "scopes": [], "collectors": []}, pivots={},
        pivot_labels={}, directory={},
        bypass={"never_loaded": True, "rows": []},
        rightsizing={"measured": False, "recommendations": [],
                     "limitations": ["Usage has not been collected for this tenant."]},
    )
    wb = load_workbook(io.BytesIO(blob))
    for title in ("Shadow Access", "Right-sizing"):
        ws = wb[title]
        assert ws.cell(row=2, column=1).value == "NOT MEASURED", title
        assert str(ws.cell(row=2, column=2).value).strip(), f"{title} states no reason"


def test_the_export_is_not_capped_at_the_ui_page_size():
    """`list_findings` caps a page at 500 so a browser is never handed 50k rows. The workbook
    passes cap=None: stopping at 500 of 1,167 findings is the same defect in a different
    wrapper."""
    import inspect

    from app.iam import findings as f

    assert "cap" in inspect.signature(f.list_findings).parameters
    src = inspect.getsource(f.list_findings)
    assert "if cap is None" in src


# =========================================================================== workbook shape
@pytest.fixture()
def isolated_cache(tmp_path, monkeypatch):
    """Point the cache index + blob dir at a tmp location, as test_iam.py does."""
    from app.iam import cache

    monkeypatch.setattr(cache, "_DATA", tmp_path)
    monkeypatch.setattr(cache, "_INDEX", tmp_path / "iam_cache.json")
    monkeypatch.setattr(cache, "_BLOBS", tmp_path / "iam")
    monkeypatch.setattr(cache, "_migrated", True)
    return tmp_path


def _demo_workbook(**over):
    """The IAM workbook built from the demo dataset."""
    import io

    from openpyxl import load_workbook

    from app.iam import cache, compose, demo, export, pivots

    demo.seed_demo("wbdemo")
    master = compose.build_master_rows("wbdemo")
    kwargs = dict(
        rows=master,
        overview=compose.compute_overview("wbdemo"),
        pivots=pivots.compute_pivots(master),
        pivot_labels=pivots.PIVOT_LABELS,
        directory=cache.read_directory("wbdemo"),
    )
    kwargs.update(over)
    return load_workbook(io.BytesIO(export.to_workbook(**kwargs)))


def test_every_timestamp_in_the_workbook_is_a_real_date(isolated_cache):
    """Text sorts lexically and only offers Excel's text filter, so "granted before 2023" is
    not a question a workbook of ISO strings can answer."""
    import datetime as dt
    import re

    iso = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}")
    wb = _demo_workbook()
    offenders = []
    for name in wb.sheetnames:
        ws = wb[name]
        heads = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for i, v in enumerate(row):
                if isinstance(v, str) and iso.match(v):
                    offenders.append(f"{name}.{heads[i] if i < len(heads) else i}")
    assert not offenders, f"timestamps exported as text: {sorted(set(offenders))}"

    ws = wb["Effective Access"]
    heads = [c.value for c in ws[1]]
    assert "assignmentCreatedOn (UTC)" in heads, "the lenses must say when access was granted"
    col = heads.index("assignmentCreatedOn (UTC)") + 1
    values = [ws.cell(r, col).value for r in range(2, ws.max_row + 1)]
    assert any(isinstance(v, dt.datetime) for v in values)


def test_the_workbook_speaks_one_yes_no_vocabulary(isolated_cache):
    """The collectors emit `true`/`false`, `Yes`/blank and `yes`/`no` for the same kind of
    fact, so filtering "not privileged" meant three different predicates in one file."""
    wb = _demo_workbook()
    allowed = {"Yes", "No", "Not measured", "n/a", "", None}
    for sheet, col in (("Effective Access", "principalExists"),
                       ("Effective Access", "roleIsPrivileged"),
                       ("Effective Access", "roleHasDataActions")):
        ws = wb[sheet]
        i = [c.value for c in ws[1]].index(col)
        seen = {r[i] for r in ws.iter_rows(min_row=2, values_only=True)}
        assert seen <= allowed, f"{sheet}.{col} speaks another dialect: {seen - allowed}"


def test_no_python_repr_ever_reaches_a_cell(isolated_cache):
    """`window` is a dict. It used to arrive through str(), printing the literal
    `{'days': 60, 'clamped': True}` into a review document."""
    wb = _demo_workbook(rightsizing={
        "measured": True,
        "recommendations": [{"principalName": "p", "window": {"days": 60, "clamped": True}}],
    })
    ws = wb["Right-sizing"]
    heads = [c.value for c in ws[1]]
    row = dict(zip(heads, [c.value for c in ws[2]]))
    assert row["Window (days)"] == 60
    assert row["Window clamped"] == "Yes"
    for name in wb.sheetnames:
        for r in wb[name].iter_rows(values_only=True):
            for v in r:
                assert not (isinstance(v, str) and v.startswith("{'")), f"{name}: {v[:60]}"


def test_a_count_that_could_not_be_measured_is_never_printed_as_zero(isolated_cache):
    """`PIM eligible: 0` beside a skipped PIM collector is the most misleading cell the file
    can print: it is not "nobody is eligible", it is "we were never allowed to look"."""
    wb = _demo_workbook(overview={
        "kpis": {"eligible": 0}, "scopes": [], "generated_at": "2026-08-19T18:09:24Z",
        "collectors": [
            {"collector": "PimEligibility", "scopeLabel": "s1", "status": "Skipped",
             "rowsAdded": 0, "message": "PIM is not licensed on this tenant."},
            {"collector": "ArgRoleAssignments", "scopeLabel": "s1", "status": "Succeeded",
             "rowsAdded": 12, "message": ""},
        ],
    })
    text = {str(r[0]): str(r[1]) for r in wb["Summary"].iter_rows(values_only=True)}
    assert "NOT MEASURED" in text["PIM eligible"]
    assert "not licensed" in text["PIM eligible"]
    assert "COVERAGE" in text, "the summary must point at the coverage register"

    # And the register itself, ahead of anything that looks like a complete answer.
    assert wb.sheetnames.index("Coverage & blind spots") < wb.sheetnames.index("Effective Access")
    cov = wb["Coverage & blind spots"]
    assert cov.cell(2, 2).value != "Succeeded", "what did not succeed sorts first"


def test_role_definition_counts_come_from_the_permission_lists(isolated_cache):
    """They were read as `actionsCount`/`dataActionsCount`, keys only the demo fixture writes.
    Against a real directory both columns were blank on every row."""
    wb = _demo_workbook(directory={"role_defs": [
        {"roleName": "Custom", "roleCategory": "ControlPlane", "roleIsPrivileged": False,
         "roleHasDataActions": True, "actions": ["a", "b", "c"], "dataActions": ["d"]},
    ]})
    ws = wb["Role Definitions"]
    row = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))
    assert row["Actions #"] == 3
    assert row["Data actions #"] == 1
    assert row["Privileged"] == "No"


def test_the_scope_column_is_a_name_not_a_second_copy_of_the_id(isolated_cache):
    """`scopeDisplayName` is set to the ARM id itself for subscription-scoped rows, so two
    60-wide columns held the same `/subscriptions/…` string."""
    from app.iam import export

    row = {"scopeDisplayName": "/subscriptions/s1", "scope": "/subscriptions/s1",
           "subscriptionName": "Prod", "resourceGroup": "rg1", "resourceName": ""}
    assert export._scope_label(row) == "Prod / rg1"
    # A real display name is kept, and an unnameable scope still falls back to its id.
    assert export._scope_label({"scopeDisplayName": "kv0", "scope": "/x"}) == "kv0"
    assert export._scope_label({"scope": "/subscriptions/s2"}) == "/subscriptions/s2"


def test_the_workbook_is_navigable_and_grouped_by_colour(isolated_cache):
    """A twenty-sheet file needs a contents page, and the tab colours have to be opaque or
    Excel draws nothing at all."""
    wb = _demo_workbook()
    assert wb.sheetnames[0] == "Index"
    listed = {r[0].value: r[1].value for r in wb["Index"].iter_rows(min_row=2)}
    assert listed, "the index is empty"

    by_section: dict[str, set[str]] = {}
    for name, section in listed.items():
        colour = wb[name].sheet_properties.tabColor
        rgb = str(colour.rgb) if colour else ""
        assert len(rgb) == 8 and rgb[:2] == "FF", f"{name}: {rgb} is transparent"
        by_section.setdefault(str(section), set()).add(rgb)
    for section, colours in by_section.items():
        assert len(colours) == 1, f"{section} sheets disagree on colour: {colours}"
    flat = [next(iter(c)) for c in by_section.values()]
    assert len(flat) == len(set(flat)), f"two sections share a colour: {flat}"

    # Every populated sheet is a real table, so filters and banding come for free.
    for name in listed:
        ws = wb[name]
        if ws.max_row > 1:
            assert ws.tables, f"{name} is a plain range, not a table"


def test_a_blind_spot_list_is_a_table_not_prose_in_a_grid(isolated_cache):
    wb = _demo_workbook(escalation={"paths": [], "limitations": ["Policy identities are not inventoried."]})
    ws = wb["Escalation - blind spots"]
    assert [c.value for c in ws[1]] == ["Limitation", "What this map cannot see"]
    assert ws.cell(2, 2).value == "Policy identities are not inventoried."

