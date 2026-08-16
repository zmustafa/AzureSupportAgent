"""Rich Backup Manager workbook, safe portal links, and snapshot-only export contract."""
from __future__ import annotations

import asyncio
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.api import backup_manager as api
from app.backup_manager import changes as change_ops
from app.backup_manager import drills as drill_ops
from app.backup_manager import export, service, snapshot as snapshot_store
from app.core import azure_connections
from app.core.azure_portal import portal_host, resource_url, resource_url_for_host, valid_resource_id
from app.core.security import Principal
from app.core.xlsx import WorkbookBuilder, hyperlink
from app.models import BackupDrill, BackupManagerChange

TENANT = "tenant-workbook"
CONNECTION_ID = "connection-workbook"
SUBSCRIPTION = "10000000-0000-0000-0000-000000000001"
ALL_PERMS = frozenset({"backup_manager.read"})


def _principal() -> Principal:
    return Principal("reviewer@example.test", "reviewer@example.test", TENANT, "operator", ALL_PERMS)


def _summary_values(workbook: Any) -> dict[str, Any]:
    return {
        str(row[0].value): row[1].value
        for row in workbook["Summary"].iter_rows(min_col=1, max_col=2)
        if row[0].value
    }


def _sheet_rows(workbook: Any, name: str) -> list[tuple[Any, ...]]:
    return list(workbook[name].iter_rows(min_row=2, values_only=True))


def _header_index(workbook: Any, name: str) -> dict[str, int]:
    return {str(cell.value): index for index, cell in enumerate(workbook[name][1], start=1)}


@pytest.fixture(autouse=True)
def isolated_snapshot_store(tmp_path, monkeypatch):
    monkeypatch.setattr(snapshot_store, "_PATH", tmp_path / "backup-manager-snapshots.json")
    monkeypatch.setattr(snapshot_store, "_locks", {})


@pytest.mark.asyncio
async def test_complete_workbook_contains_every_review_area_and_real_hyperlinks() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    first = snapshot["inventory"]["rows"][0]
    first.update({
        "friendly_name": "=unsafe()",
        "datasource_id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-demo/providers/Microsoft.Compute/virtualMachines/vm-demo",
        "id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-vault/providers/Microsoft.RecoveryServices/vaults/vault-demo/backupFabrics/Azure/protectionContainers/container/protectedItems/item",
        "vault_id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-vault/providers/Microsoft.RecoveryServices/vaults/vault-demo",
        "policy_id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-vault/providers/Microsoft.RecoveryServices/vaults/vault-demo/backupPolicies/policy-demo",
        "orphaned": False,
    })
    snapshot["age_seconds"] = 25

    content = export.to_workbook(
        snapshot=snapshot,
        changes=[],
        drills=[],
        portal_host="portal.azure.com",
        connection_label="Synthetic connection",
        ledger_generated_at="2026-08-12T00:00:00+00:00",
    )
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    expected = [
        "Summary", "Index", "Coverage & limitations", "Protected items", "RPO attainment",
        "Orphaned protection", "Backup jobs", "Failure clusters", "Chronic failures",
        "Job congestion", "Policies", "Duplicate policies", "Policy compliance",
        "Compliance tiers", "Vaults", "Vault posture", "Vault checks", "Vault capacity",
        "Protection gaps", "Coverage findings", "Native-only protections", "Replicated items",
        "Recovery plans", "Drill register", "Cost summary", "Cost by protected item",
        "Cost allocation", "Recoverable waste", "Managed changes",
    ]
    assert workbook.sheetnames == expected

    ws = workbook["Protected items"]
    headers = [cell.value for cell in ws[1]]
    name_cell = ws.cell(2, headers.index("Item") + 1)
    source_link = ws.cell(2, headers.index("Source portal") + 1)
    assert str(name_cell.value).startswith("'=")
    assert source_link.value == "Open source"
    assert source_link.hyperlink is not None
    assert source_link.hyperlink.target.startswith("https://portal.azure.com/#@/resource/subscriptions/")

    summary = _summary_values(workbook)
    assert summary["Connection"] == "Synthetic connection"
    assert summary["Snapshot notice"].startswith("This workbook reflects the last completed analysis")
    assert summary["Live ledgers read"] == "2026-08-12T00:00:00+00:00"

    limitations = list(workbook["Coverage & limitations"].iter_rows(values_only=True))
    assert any(row[1] == "demo" for row in limitations[1:])
    workbook.close()


@pytest.mark.asyncio
async def test_workbook_reconciles_every_snapshot_section() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["demo"] = False
    snapshot["age_seconds"] = 60
    snapshot["cost"]["allocation"] = {
        "currency": "EUR",
        "note": "Synthetic allocation",
        "rows": [
            {
                "name": "Item A", "instance_id": "item-a", "datasource_type": "vm",
                "vault_name": "Vault A", "vault_id": "", "allocated_cost": 4.25,
                "vault_total": 10.0, "weight": 0.425, "weight_basis": "consumed_gb",
            },
            {
                "name": "Item B", "instance_id": "item-b", "datasource_type": "vm",
                "vault_name": "Vault A", "vault_id": "", "allocated_cost": 5.75,
                "vault_total": 10.0, "weight": 0.575, "weight_basis": "consumed_gb",
            },
        ],
    }
    content = export.to_workbook(snapshot=snapshot, portal_host="portal.azure.com")
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)

    summary = _summary_values(workbook)
    protection = snapshot["summary"]["protection"]
    assert summary["Vaults"] == protection["vaults"]
    assert summary["Protected items"] == protection["protected_items"]
    assert summary["Policies"] == protection["policies"]
    assert summary["Jobs"] == snapshot["summary"]["jobs"]["total"]
    assert summary["Failed jobs"] == snapshot["summary"]["jobs"]["failed"]
    assert summary["Protection gaps"] == len(snapshot["gaps"]["gaps"])
    assert summary["RPO breached"] == snapshot["summary"]["rpo"]["breached"]
    assert summary["RPO at risk"] == snapshot["summary"]["rpo"]["at_risk"]
    assert summary["RPO unknown"] == snapshot["summary"]["rpo"]["unknown"]
    assert summary["Estimated monthly cost"] == snapshot["summary"]["cost"]["monthly_total"]
    assert summary["Estimated annual cost"] == snapshot["cost"]["annual_total"]
    assert summary["Currency"] == snapshot["cost"]["currency"]

    expected_counts = {
        "Protected items": len(snapshot["inventory"]["rows"]),
        "RPO attainment": len(snapshot["dr"]["rpo"]["rows"]),
        "Backup jobs": len(snapshot["jobs"]["rows"]),
        "Failure clusters": len(snapshot["job_analysis"]["clusters"]),
        "Chronic failures": len(snapshot["job_analysis"]["chronic"]),
        "Job congestion": len(snapshot["job_analysis"]["congestion"]),
        "Policies": len(snapshot["policies"]["policies"]),
        "Duplicate policies": len(snapshot["policies"]["duplicate_groups"]),
        "Policy compliance": len(snapshot["compliance"]["rows"]),
        "Compliance tiers": len(snapshot["compliance"]["tiers"]),
        "Vaults": len(snapshot["vaults"]["vaults"]),
        "Vault posture": len(snapshot["posture"]["vaults"]),
        "Vault capacity": len(snapshot["vaults"]["capacity"]),
        "Protection gaps": len(snapshot["gaps"]["gaps"]),
        "Coverage findings": len(snapshot["gaps"]["coverage_gaps"]),
        "Native-only protections": len(snapshot["gaps"]["native_only"]),
        "Replicated items": len(snapshot["dr"]["items"]),
        "Recovery plans": len(snapshot["dr"]["recovery_plans"]),
        "Cost by protected item": len(snapshot["cost"]["top_rows"]),
        "Cost allocation": 2,
        "Recoverable waste": len(snapshot["cost"]["waste"]["findings"]),
    }
    for sheet, expected_count in expected_counts.items():
        assert len(_sheet_rows(workbook, sheet)) == expected_count, sheet

    expected_checks = sum(len(vault["checks"]) for vault in snapshot["posture"]["vaults"])
    assert len(_sheet_rows(workbook, "Vault checks")) == expected_checks
    protected_ids = [row["id"] for row in snapshot["inventory"]["rows"]]
    headers = _header_index(workbook, "Protected items")
    exported_ids = [row[headers["Protected item id"] - 1] for row in _sheet_rows(workbook, "Protected items")]
    assert exported_ids == protected_ids

    congestion_headers = [cell.value for cell in workbook["Job congestion"][1]]
    assert congestion_headers == ["Hour", "Jobs", "Failed", "Average duration (s)"]
    assert all(not isinstance(value, (dict, list)) for row in _sheet_rows(workbook, "Job congestion") for value in row)
    allocation_headers = _header_index(workbook, "Cost allocation")
    allocated = sum(row[allocation_headers["Allocated actual"] - 1] for row in _sheet_rows(workbook, "Cost allocation"))
    assert allocated == 10.0
    workbook.close()


@pytest.mark.asyncio
async def test_limitations_make_failures_truncation_staleness_and_cost_blindness_explicit() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["demo"] = False
    snapshot["partial"] = True
    snapshot["age_seconds"] = export.SNAPSHOT_STALE_SECONDS + 1
    snapshot["errors"] = {
        "rsv_jobs": "Traceback at C:\\agent\\collector.py:42 access_token=do-not-export",
    }
    for section, rows_key in (("inventory", "rows"), ("jobs", "rows"), ("gaps", "gaps")):
        snapshot[section]["truncated"] = True
        snapshot[section]["total_count"] = len(snapshot[section][rows_key]) + 7
    snapshot["cost"].update({
        "rate_error": "pricing failed at /srv/app/pricing.py:9 bearer=do-not-export",
        "confidence": "partial",
        "unpriced_instances": 1,
        "instance_count": max(2, len(snapshot["cost"].get("top_rows") or []) + 1),
        "actuals": {
            "available": False,
            "reason": "password=do-not-export at C:\\agent\\cost.py:8",
        },
    })

    workbook = load_workbook(BytesIO(export.to_workbook(
        snapshot=snapshot, changes_truncated=True, changes=[{"id": "change-1"}],
    )), read_only=False, data_only=False)
    rows = _sheet_rows(workbook, "Coverage & limitations")
    states = {str(row[1]) for row in rows}
    assert {"failed", "partial", "truncated", "unpriced", "stale"} <= states
    assert any(row[0] == "analysis" and row[1] == "partial" for row in rows)
    assert any(row[0] == "snapshot" and row[1] == "stale" for row in rows)
    for source in ("inventory", "jobs", "gaps"):
        row = next(item for item in rows if item[0] == source and item[1] == "truncated")
        section = snapshot[source]
        key = "gaps" if source == "gaps" else "rows"
        assert row[3] == len(section[key])
        assert row[4] == section["total_count"]
    flattened = " ".join(str(value or "") for row in rows for value in row)
    assert "NOT MEASURED" in flattened
    assert "do-not-export" not in flattened
    assert "collector.py" not in flattened
    assert "pricing.py" not in flattened
    assert "cost.py" not in flattened
    assert max(len(str(value or "")) for row in rows for value in row) <= 1000
    workbook.close()


@pytest.mark.asyncio
async def test_public_ledgers_reconcile_without_private_change_fields() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    now = datetime(2026, 8, 12, tzinfo=timezone.utc)
    change = BackupManagerChange(
        id="change-public", tenant_id=TENANT, connection_id=CONNECTION_ID,
        target_type="vault", target_id=f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg/providers/Microsoft.RecoveryServices/vaults/vault-a",
        operation="update", status="pending", risk="medium", summary_json={"description": "Public summary"},
        desired_encrypted="enc:v1:private-desired", before_encrypted="enc:v1:private-before",
        after_encrypted="enc:v1:private-after", expected_state_hash="private-hash",
        requested_by="operator@example.test", requested_at=now, operation_url="https://management.azure.test/private-job",
        depends_on=[], poll_attempts=2,
    )
    drill = BackupDrill(
        id="drill-public", tenant_id=TENANT, connection_id=CONNECTION_ID, name="Synthetic drill",
        kind="test_failover", scope_kind="subscription", scope_id=SUBSCRIPTION,
        target_id=change.target_id, target_name="Synthetic target", status="passed", cadence_days=90,
        executed_at=now, executed_by="operator@example.test", outcome_notes="Synthetic evidence",
        rto_minutes=18, created_by="operator@example.test", created_at=now,
    )
    workbook = load_workbook(BytesIO(export.to_workbook(
        snapshot=snapshot,
        changes=[change_ops.public_change(change)],
        drills=[drill_ops.public_drill(drill, now=now)],
        portal_host="portal.azure.com",
    )), read_only=False, data_only=False)
    assert len(_sheet_rows(workbook, "Managed changes")) == 1
    assert len(_sheet_rows(workbook, "Drill register")) == 1
    all_values = " ".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    for forbidden in (
        "private-desired", "private-before", "private-after", "private-hash",
        "management.azure.test/private-job", "operation_url", "desired_encrypted", "before_encrypted",
    ):
        assert forbidden not in all_values
    assert "Public summary" in all_values
    assert "Synthetic drill" in all_values
    workbook.close()


@pytest.mark.asyncio
async def test_orphaned_workbook_rows_keep_only_defensible_links() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    item = snapshot["inventory"]["rows"][0]
    item.update({
        "orphaned": True,
        "datasource_id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-source/providers/Microsoft.Compute/virtualMachines/deleted-vm",
        "id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-vault/providers/Microsoft.RecoveryServices/vaults/vault-a/backupFabrics/Azure/protectionContainers/c/protectedItems/i",
        "vault_id": f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-vault/providers/Microsoft.RecoveryServices/vaults/vault-a",
    })
    workbook = load_workbook(BytesIO(export.to_workbook(
        snapshot=snapshot, portal_host="portal.azure.com",
    )), read_only=False, data_only=False)
    headers = _header_index(workbook, "Protected items")
    row_number = next(
        row
        for row in range(2, workbook["Protected items"].max_row + 1)
        if workbook["Protected items"].cell(row, headers["Protected item id"]).value == item["id"]
    )
    source = workbook["Protected items"].cell(row_number, headers["Source portal"])
    protected = workbook["Protected items"].cell(row_number, headers["Protected item portal"])
    vault = workbook["Protected items"].cell(row_number, headers["Vault portal"])
    assert source.value == "Source deleted"
    assert source.hyperlink is None
    assert protected.hyperlink is not None
    assert vault.hyperlink is not None
    orphan_headers = _header_index(workbook, "Orphaned protection")
    orphan_source = workbook["Orphaned protection"].cell(2, orphan_headers["Source"])
    assert orphan_source.value == "Source deleted"
    assert orphan_source.hyperlink is None
    index_rows = _sheet_rows(workbook, "Index")
    orphan_index = next(row for row in index_rows if row[0] == "Orphaned protection")
    assert "portal-only" in str(orphan_index[3]).lower()
    workbook.close()


def test_workbook_builder_hyperlink_is_relationship_not_formula() -> None:
    builder = WorkbookBuilder()
    builder.section("Test", "44546A")
    builder.first_sheet("Summary").append(["Summary"])
    builder.sheet("Links", ["Resource"], [[hyperlink("Open", "https://portal.azure.com/#view/test")]])
    workbook = load_workbook(BytesIO(builder.to_bytes()), read_only=False, data_only=False)
    cell = workbook["Links"]["A2"]
    assert cell.value == "Open"
    assert cell.data_type != "f"
    assert cell.hyperlink.target == "https://portal.azure.com/#view/test"
    workbook.close()


@pytest.mark.parametrize("value", ["=cmd()", "+cmd()", "-1+2", "@SUM(A1:A2)", "  =cmd()", "\t@cmd()"])
def test_workbook_builder_neutralizes_every_formula_prefix(value: str) -> None:
    builder = WorkbookBuilder()
    builder.section("Test", "44546A")
    builder.first_sheet("Summary").append(["Summary"])
    builder.sheet("Values", ["Value"], [[value]])
    workbook = load_workbook(BytesIO(builder.to_bytes()), read_only=False, data_only=False)
    cell = workbook["Values"]["A2"]
    assert str(cell.value).startswith("'")
    assert cell.data_type != "f"
    workbook.close()


@pytest.mark.asyncio
async def test_summary_front_matter_is_formula_safe() -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["scope"]["scope_name"] = "=unsafe-scope()"
    workbook = load_workbook(BytesIO(export.to_workbook(
        snapshot=snapshot, connection_label="@unsafe-connection",
    )), read_only=False, data_only=False)
    summary = _summary_values(workbook)
    assert summary["Scope"] == "'=unsafe-scope()"
    assert summary["Connection"] == "'@unsafe-connection"
    workbook.close()


@pytest.mark.parametrize("value", ["=cmd()", "+cmd()", "-1+2", "@SUM(A1:A2)", "  =cmd()", "\t@cmd()"])
def test_existing_csv_is_formula_injection_safe(value: str) -> None:
    text = export.export("gaps", [{"resource_name": value, "severity": "critical"}])
    assert f"'{value}" in text


@pytest.mark.parametrize(
    ("cloud", "host"),
    [
        ("AzureCloud", "portal.azure.com"),
        ("AzureUSGovernment", "portal.azure.us"),
        ("AzureChinaCloud", "portal.azure.cn"),
    ],
)
def test_portal_links_follow_trusted_connection_cloud(cloud: str, host: str) -> None:
    rid = f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-demo/providers/Microsoft.Storage/storageAccounts/account-demo"
    assert portal_host({"azure_cloud": cloud}) == host
    assert resource_url(rid, {"azure_cloud": cloud}).startswith(f"https://{host}/#@/resource/")


def test_portal_links_accept_mixed_case_and_encode_safe_resource_names() -> None:
    rid = f"/Subscriptions/{SUBSCRIPTION.upper()}/resourceGroups/RG Synthetic/providers/Microsoft.Storage/storageAccounts/account (one)"
    url = resource_url_for_host(rid, "portal.azure.com")
    assert valid_resource_id(rid) == rid
    assert url.startswith("https://portal.azure.com/#@/resource/Subscriptions/")
    assert "RG%20Synthetic" in url
    assert "account%20(one)" in url


@pytest.mark.parametrize(
    "value",
    [
        "https://evil.example/subscriptions/10000000-0000-0000-0000-000000000001",
        f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg?redirect=https://evil.example",
        f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg#fragment",
        f"/subscriptions/{SUBSCRIPTION}\r/resourceGroups/rg",
        "/subscriptions/not-a-guid/resourceGroups/rg",
        f"https://user@evil.example/subscriptions/{SUBSCRIPTION}",
        "javascript:alert(1)",
    ],
)
def test_portal_links_reject_untrusted_or_malformed_targets(value: str) -> None:
    assert valid_resource_id(value) == ""
    assert resource_url_for_host(value, "portal.azure.com") == ""


def test_unknown_explicit_cloud_fails_closed() -> None:
    rid = f"/subscriptions/{SUBSCRIPTION}/resourceGroups/rg-demo"
    assert portal_host({"azure_cloud": "UnknownCloud"}) == ""
    assert resource_url(rid, {"azure_cloud": "UnknownCloud"}) == ""


@pytest.mark.parametrize("value", [None, "", "   "])
def test_missing_portal_targets_produce_no_link(value: Any) -> None:
    assert valid_resource_id(value) == ""
    assert resource_url_for_host(value, "portal.azure.com") == ""


class _ScalarResult:
    def scalars(self):
        return []


class _Db:
    async def execute(self, _statement: Any) -> _ScalarResult:
        return _ScalarResult()


@pytest.mark.asyncio
async def test_workbook_endpoint_reads_completed_snapshot_without_azure(monkeypatch) -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["demo"] = False
    snapshot_store.write_snapshot(TENANT, CONNECTION_ID, "subscription", SUBSCRIPTION, snapshot)
    monkeypatch.setattr(api, "_connection", lambda *_args, **_kwargs: {
        "id": CONNECTION_ID, "display_name": "Synthetic connection", "azure_cloud": "AzureCloud",
    })

    async def no_azure(*_args: Any, **_kwargs: Any) -> Any:
        raise AssertionError("workbook export must not collect Azure data")

    async def no_drills(*_args: Any, **_kwargs: Any) -> list[Any]:
        return []

    async def no_changes(*_args: Any, **_kwargs: Any) -> int:
        return 0

    monkeypatch.setattr(api, "_estate", no_azure)
    monkeypatch.setattr(api.inventory_ops, "collect_estate", no_azure)
    monkeypatch.setattr(api.drill_ops, "list_drills", no_drills)
    monkeypatch.setattr(api, "_actionable_changes", no_changes)

    response = await api.export_workbook(
        connection_id=CONNECTION_ID,
        workload_id="",
        subscription_id=SUBSCRIPTION,
        management_group_id="",
        principal=_principal(),
        db=_Db(),
    )
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert bytes(response.body).startswith(b"PK")
    assert "backup-manager-review-" in response.headers["content-disposition"]
    assert response.headers["cache-control"] == "no-store"
    workbook = load_workbook(BytesIO(bytes(response.body)), read_only=False, data_only=False)
    assert _summary_values(workbook)["Connection"] == "Synthetic connection"
    workbook.close()


@pytest.mark.asyncio
async def test_workbook_endpoint_requires_a_completed_snapshot(monkeypatch) -> None:
    monkeypatch.setattr(api, "_connection", lambda *_args, **_kwargs: {"id": CONNECTION_ID})
    with pytest.raises(HTTPException) as excinfo:
        await api.export_workbook(
            connection_id=CONNECTION_ID,
            workload_id="",
            subscription_id=SUBSCRIPTION,
            management_group_id="",
            principal=_principal(),
            db=_Db(),
        )
    assert excinfo.value.status_code == 409
    assert "Analyze backups first" in str(excinfo.value.detail)


def test_requested_unknown_connection_does_not_fall_back(monkeypatch) -> None:
    monkeypatch.setattr(azure_connections, "get_connection", lambda _connection_id: None)
    with pytest.raises(LookupError):
        service.resolve_selected_connection("missing-connection")


def test_workload_connection_mismatch_is_rejected_before_collection(monkeypatch) -> None:
    monkeypatch.setattr(
        service,
        "workload_context",
        lambda _workload_id: ({"id": "workload-a", "connection_id": "connection-b"}, set(), set()),
    )
    monkeypatch.setattr(
        azure_connections,
        "get_connection",
        lambda connection_id: {"id": connection_id, "disabled": False},
    )
    with pytest.raises(ValueError, match="different Azure connection"):
        service.resolve_selected_connection("connection-a", "workload-a")


@pytest.mark.asyncio
async def test_workbook_endpoint_maps_unknown_requested_connection_to_404(monkeypatch) -> None:
    def missing(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        raise LookupError("Azure connection 'missing-connection' was not found.")

    monkeypatch.setattr(api.service, "resolve_selected_connection", missing)
    with pytest.raises(HTTPException) as excinfo:
        await api.export_workbook(
            connection_id="missing-connection",
            workload_id="",
            subscription_id=SUBSCRIPTION,
            management_group_id="",
            principal=_principal(),
            db=_Db(),
        )
    assert excinfo.value.status_code == 404


@pytest.mark.asyncio
async def test_workbook_endpoint_isolates_tenant_connection_and_scope(monkeypatch) -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["demo"] = False
    other_tenant = "tenant-other"
    snapshot_store.write_snapshot(other_tenant, CONNECTION_ID, "subscription", SUBSCRIPTION, snapshot)
    monkeypatch.setattr(api, "_connection", lambda *_args, **_kwargs: {
        "id": CONNECTION_ID, "display_name": "Synthetic connection", "azure_cloud": "AzureCloud",
    })

    with pytest.raises(HTTPException) as tenant_error:
        await api.export_workbook(
            connection_id=CONNECTION_ID, workload_id="", subscription_id=SUBSCRIPTION,
            management_group_id="", principal=_principal(), db=_Db(),
        )
    assert tenant_error.value.status_code == 409

    snapshot_store.write_snapshot(TENANT, "connection-other", "subscription", SUBSCRIPTION, snapshot)
    with pytest.raises(HTTPException) as connection_error:
        await api.export_workbook(
            connection_id=CONNECTION_ID, workload_id="", subscription_id=SUBSCRIPTION,
            management_group_id="", principal=_principal(), db=_Db(),
        )
    assert connection_error.value.status_code == 409

    snapshot_store.write_snapshot(TENANT, CONNECTION_ID, "subscription", "20000000-0000-0000-0000-000000000002", snapshot)
    with pytest.raises(HTTPException) as scope_error:
        await api.export_workbook(
            connection_id=CONNECTION_ID, workload_id="", subscription_id=SUBSCRIPTION,
            management_group_id="", principal=_principal(), db=_Db(),
        )
    assert scope_error.value.status_code == 409


@pytest.mark.asyncio
async def test_demo_workbook_endpoint_uses_only_synthetic_snapshot(monkeypatch) -> None:
    async def no_provider(*_args: Any, **_kwargs: Any) -> Any:
        raise AssertionError("demo workbook export must not call an Azure provider")

    monkeypatch.setattr(api, "_estate", no_provider)
    monkeypatch.setattr(api.inventory_ops, "collect_estate", no_provider)
    response = await api.export_workbook(
        connection_id="",
        workload_id="demo-zava-shoes-website",
        subscription_id="",
        management_group_id="",
        principal=_principal(),
        db=_Db(),
    )
    workbook = load_workbook(BytesIO(bytes(response.body)), read_only=False, data_only=False)
    assert _summary_values(workbook)["Demo data"] is True
    assert any(row[1] == "demo" for row in _sheet_rows(workbook, "Coverage & limitations"))
    workbook.close()


@pytest.mark.asyncio
async def test_workbook_endpoint_offloads_formatter_and_preserves_exact_scope(monkeypatch) -> None:
    snapshot = await api._demo_snapshot("demo-zava-shoes-website")
    snapshot["demo"] = False
    seen: dict[str, Any] = {}

    def read_snapshot(tenant: str, connection: str, kind: str, scope_id: str) -> dict[str, Any]:
        seen["snapshot_key"] = (tenant, connection, kind, scope_id)
        return snapshot

    async def fake_to_thread(function: Any, /, *args: Any, **kwargs: Any) -> Any:
        seen["thread_function"] = function
        seen["thread_kwargs"] = kwargs
        await asyncio.sleep(0)
        return function(*args, **kwargs)

    async def no_changes(*_args: Any, **_kwargs: Any) -> int:
        return 0

    monkeypatch.setattr(api, "_connection", lambda *_args, **_kwargs: {
        "id": CONNECTION_ID, "display_name": "Synthetic connection", "azure_cloud": "AzureCloud",
    })
    monkeypatch.setattr(api.snapshot_store, "read_snapshot", read_snapshot)
    monkeypatch.setattr(api, "_actionable_changes", no_changes)
    monkeypatch.setattr(api.asyncio, "to_thread", fake_to_thread)

    response = await api.export_workbook(
        connection_id=CONNECTION_ID, workload_id="", subscription_id=SUBSCRIPTION,
        management_group_id="", principal=_principal(), db=_Db(),
    )
    assert bytes(response.body).startswith(b"PK")
    assert seen["snapshot_key"] == (TENANT, CONNECTION_ID, "subscription", SUBSCRIPTION)
    assert seen["thread_function"] is export.to_workbook
    assert seen["thread_kwargs"]["connection_label"] == "Synthetic connection"
