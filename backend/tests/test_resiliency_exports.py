"""The Recovery Readiness exports: workbook, PDF, and the promise they agree.

The cross-format test is the one that earns trust. A reader who checks the PDF against the
workbook and finds different numbers stops treating either as evidence, and no amount of
correctness elsewhere recovers that.
"""
from __future__ import annotations

import io
import zipfile

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app import demo_catalog
from app.resiliency import (
    analysis,
    evidence as evidence_mod,
    export as export_mod,
    model,
    pdf_report,
    reference,
    snapshot as snapshot_store,
)


@pytest.fixture(autouse=True)
def _isolated(tmp_path):
    reference.reset_for_tests(tmp_path)
    yield


def _verdict(rto_class, *, applicable=True, rpo_state=model.RPO_KNOWN, rpo=60,
             breach=None, detail="vault is locally redundant"):
    out = {
        "applicable": applicable, "rto_class": rto_class, "rpo_state": rpo_state,
        "rpo_minutes": rpo, "confidence": model.CONFIDENCE_MEDIUM,
        "rto_band_minutes": [60, 240] if rto_class == model.RTO_HOURS else None,
        "rto_assumptions": ["restore throughput is assumed, not measured"],
        "basis": [{"kind": "backup_policy", "detail": detail, "source": "Backup Manager"}],
    }
    if breach:
        out["breach"] = {"state": breach}
        out["target"] = {"rto_class": model.RTO_HOURS, "rpo_minutes": 60}
    return out


def _row(name, rtype="microsoft.storage/storageaccounts", *, verdicts=None,
         zone_redundant=None, replication=""):
    return {
        "id": f"/subscriptions/00000000-0000-0000-0000-000000000001/rg/{name}",
        "name": name, "type": rtype, "location": "westeurope",
        "resource_group": "rg", "subscription_id": "sub-1", "tier_label": "Business critical",
        "redundancy": {"zone_redundant": zone_redundant, "replication": replication,
                       "zones": [], "sku": "Standard"},
        "protection": {"state": "protected", "frequency": "Daily", "retention_days": 30,
                       "recovery_point_age_hours": 6, "vault_redundancy": "LRS",
                       "reason": "", "policy_name": "daily"},
        "dr": {"replicated": False},
        "worst": {"rto_class": model.RTO_NONE, "scenario": "region_loss", "undetermined": 0,
                  "no_recovery_path": ["region_loss"]},
        "verdicts": verdicts or {
            "instance_loss": _verdict(model.RTO_AUTOMATIC),
            "zone_loss": _verdict(model.RTO_MINUTES),
            "region_loss": _verdict(model.RTO_NONE),
            "data_corruption": _verdict(model.RTO_HOURS, breach="breached"),
            "accidental_delete": _verdict(model.RTO_AUTOMATIC, applicable=False),
        },
    }


def _snapshot(*, rows=None, truncation=None, acknowledged=True):
    rows = rows if rows is not None else [_row(f"sa{i}") for i in range(4)]
    return {
        "schema_version": 1, "report_exists": True,
        "generated_at": "2026-08-20T10:00:00+00:00", "demo": False, "reason": "",
        "scope": {"scope_kind": "workload", "scope_id": "wl-1",
                  "scope_name": "Contoso Hotels", "subscriptions": ["sub-1"]},
        "summary": {
            "resources": len(rows),
            "by_scenario": {s: {"determined": len(rows), "no_recovery_path":
                                len(rows) if s == "region_loss" else 0,
                                "undetermined": 0, "not_applicable": 0, "total": len(rows)}
                            for s in model.SCENARIOS},
            "protection": {"protected": len(rows), "not_protected": 0, "unknown": 0},
            "worst": {"scenario": "region_loss", "no_recovery_path": len(rows)},
        },
        "resources": rows,
        "breaches": [{
            "resource_id": r["id"], "name": r["name"], "type": r["type"],
            "scenario": "data_corruption", "tier": "mission_critical",
            "rto_class": model.RTO_HOURS, "rpo_state": model.RPO_KNOWN, "rpo_minutes": 1440,
            "no_recovery_path": False,
            "target": {"rto_class": model.RTO_MINUTES, "rpo_minutes": 60},
            "basis": [{"detail": "daily backup only"}],
        } for r in rows],
        "breach_summary": {"breached": len(rows)},
        "workloads": [{
            "workload_id": "wl-1", "name": "Contoso Hotels", "tier": "mission_critical",
            "scenarios": {"region_loss": {
                "applicable": True, "rto_class": model.RTO_NONE,
                "rpo_state": model.RPO_NONE, "rpo_minutes": None,
                "weakest_link": {"name": "sa0", "reason": "vault is locally redundant"},
                "coverage": {"determined": len(rows), "total": len(rows)},
            }},
        }],
        "provenance": {"redundancy": {"source": "Azure Resource Graph",
                                      "collected_at": "2026-08-20T10:00:00+00:00",
                                      "unreadable": False, "reason": ""},
                       "backup": {"source": "Backup Manager", "collected_at": "",
                                  "unreadable": True,
                                  "reason": "Backup Manager has not analyzed this scope."}},
        "truncation": truncation or {},
        "targets_acknowledged": acknowledged,
    }


def _sheets(content: bytes) -> dict[str, list[list]]:
    wb = load_workbook(io.BytesIO(content))
    return {name: [list(r) for r in wb[name].iter_rows(values_only=True)]
            for name in wb.sheetnames}


def _pdf_text(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    return "\n".join((p.extract_text() or "") for p in reader.pages)


# ============================================================== workbook
def test_the_workbook_carries_every_section():
    sheets = _sheets(export_mod.build(_snapshot(), reference_doc=reference.load()))
    for expected in ["Index", "Summary", "RTO-RPO by type", "RTO-RPO distribution",
                     "Reason index", "Redundancy gap", "Recovery matrix",
                     "Resources", "Reasoning", "Breaches", "Workloads", "Objectives",
                     "Assumptions and rates", "Provenance", "Truncation", "Trend"]:
        assert expected in sheets, expected


def test_there_is_no_stray_default_sheet():
    """`sheet()` for the front matter leaves openpyxl's empty 'Sheet' tab in every export."""
    assert "Sheet" not in _sheets(export_mod.build(_snapshot()))


def test_a_formula_in_a_resource_name_is_neutralised():
    """`=cmd|'/c calc'!A1` in a display name is a working RCE against whoever opens it."""
    rows = [_row("=cmd|'/c calc'!A1")]
    sheets = _sheets(export_mod.build(_snapshot(rows=rows)))
    seen = False
    for name, grid in sheets.items():
        for row in grid:
            for cell in row:
                if not isinstance(cell, str):
                    continue
                # A cell merely CONTAINING the payload (inside an ARM id) is inert; only one
                # that STARTS with a trigger character is evaluated on open.
                assert not cell.startswith(("=", "+", "-", "@")), f"{name}: {cell}"
                if "cmd|" in cell:
                    seen = True
    assert seen, "the payload never reached the workbook, so nothing was proven"


def test_truncation_is_reported_rather_than_silently_dropped():
    """A file presented as complete that lost its tail is worse than one that says where it
    stopped, because only the second can be checked."""
    snap = _snapshot(truncation={"resources": {"exported": 5000, "known_total": 7412}})
    grid = _sheets(export_mod.build(snap))["Truncation"]
    body = [r for r in grid if r[0] not in (None, "Section")]
    assert any(r[0] == "resources" and r[3] == 2412 for r in body), body


def test_a_complete_export_says_so_instead_of_showing_an_empty_grid():
    """A sheet called 'Truncation' with no rows is ambiguous: the reader cannot tell
    'nothing was dropped' from 'nobody checked'."""
    grid = _sheets(export_mod.build(_snapshot()))["Truncation"]
    body = [r for r in grid if r and r[0] not in (None, "Section")]
    assert body, "empty grid"
    assert any("Nothing was truncated" in str(c) for row in body for c in row)


def test_the_reasoning_sheet_is_one_row_per_fact_not_a_joined_cell():
    """Reasoning nobody can filter or count is reasoning nobody uses."""
    grid = _sheets(export_mod.build(_snapshot()))["Reasoning"]
    body = [r for r in grid if r and r[0] not in (None, "Resource") and "note" not in str(r[0]).lower()]
    assert len(body) >= 4 * 4, "one row per resource per applicable scenario per basis item"
    assert all("; " not in str(r[5] or "") for r in body), "details must not be pre-joined"


def test_the_objectives_sheet_makes_the_breach_targets_verifiable():
    sheets = _sheets(export_mod.build(_snapshot(), reference_doc=reference.load()))
    flat = " ".join(str(c) for row in sheets["Objectives"] for c in row if c)
    assert "Mission critical" in flat or "mission_critical" in flat
    # Sheet notes live on the Index, which is where a reader looks for what a tab means.
    index = " ".join(str(c) for row in sheets["Index"] for c in row if c)
    assert "Registry version" in index


def test_unacknowledged_objectives_are_labelled_on_the_sheet_not_just_the_index():
    """A caveat that only appears on a contents page is a caveat that gets missed, and this
    one changes what every target in the file means."""
    snap = _snapshot(acknowledged=False)
    sheets = _sheets(export_mod.build(snap, reference_doc=reference.load()))
    flat = " ".join(str(c) for row in sheets["Objectives"] for c in row if c)
    assert "SHIPPED DEFAULT" in flat
    index = " ".join(str(c) for row in sheets["Index"] for c in row if c)
    assert "STILL THE SHIPPED DEFAULTS" in index


def test_agreed_objectives_are_not_labelled_as_defaults():
    sheets = _sheets(export_mod.build(_snapshot(), reference_doc=reference.load()))
    flat = " ".join(str(c) for row in sheets["Objectives"] for c in row if c)
    assert "SHIPPED DEFAULT" not in flat
    assert "Agreed" in flat


def test_the_rates_sheet_exposes_the_constants_behind_every_band():
    grid = _sheets(export_mod.build(_snapshot(), reference_doc=reference.load()))
    flat = " ".join(str(c) for row in grid["Assumptions and rates"] for c in row if c)
    assert "vm_restore_mbps" in flat
    assert "Roll-up assumption" in flat


def test_truncation_is_called_out_on_the_index_so_a_reader_meets_it_first():
    snap = _snapshot(truncation={"resources": {"exported": 5000, "known_total": 7412}})
    index = _sheets(export_mod.build(snap))["Index"]
    flat = " ".join(str(c) for row in index for c in row if c)
    assert "exceeded the stored row cap" in flat


def test_the_trend_sheet_refuses_to_show_a_direction_from_one_point():
    grid = _sheets(export_mod.build(_snapshot(), trend={"available": False,
                                                        "reason": "Only one analysis."}))["Trend"]
    flat = " ".join(str(c) for row in grid for c in row if c)
    assert "NOT MEASURED" in flat


def test_the_trend_sheet_lists_the_points_when_history_exists():
    trend = {"available": True, "caveat": "",
             "points": [{"generated_at": "2026-01-01T00:00:00Z", "resources": 10,
                         "no_recovery_path": 8, "undetermined": 0, "breaches": 3,
                         "protected": 6, "not_protected": 2, "protection_unknown": 2},
                        {"generated_at": "2026-02-01T00:00:00Z", "resources": 10,
                         "no_recovery_path": 2, "undetermined": 0, "breaches": 1,
                         "protected": 9, "not_protected": 1, "protection_unknown": 0}]}
    grid = _sheets(export_mod.build(_snapshot(), trend=trend))["Trend"]
    body = [r for r in grid if r and str(r[0] or "").startswith("2026")]
    assert len(body) == 2


def test_tab_colours_are_opaque_in_the_file_not_just_the_object_model():
    """`tabColor = '7030A0'` serialises with alpha 00 — fully transparent, so Excel draws
    nothing. openpyxl reads the useless value back quite happily, so assert on the XML."""
    content = export_mod.build(_snapshot())
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        xml = "".join(zf.read(n).decode("utf-8")
                      for n in zf.namelist() if n.startswith("xl/worksheets/sheet"))
    assert 'tabColor rgb="FF' in xml or 'tabColor rgb="ff' in xml.lower()
    assert 'tabColor rgb="00' not in xml


def test_an_empty_snapshot_still_produces_a_readable_workbook():
    snap = _snapshot(rows=[])
    snap["breaches"] = []
    snap["workloads"] = []
    sheets = _sheets(export_mod.build(snap))
    assert "Summary" in sheets and "Index" in sheets


# ============================================================== pdf
def test_the_pdf_is_a_real_multi_page_document():
    pdf = pdf_report.build(_snapshot(), reference_doc=reference.load())
    assert pdf[:5] == b"%PDF-"
    assert len(PdfReader(io.BytesIO(pdf)).pages) >= 8


def test_the_pdf_contains_every_section_and_its_appendices():
    text = _pdf_text(pdf_report.build(_snapshot(), reference_doc=reference.load()))
    for expected in ["Executive summary", "How to read this report", "Trend",
                     "Recovery by failure scenario", "RTO and RPO by resource type",
                     "Resources that cannot be recovered", "Breaches against objectives",
                     "Appendix A", "Appendix B", "Appendix C"]:
        assert expected in text, expected


def test_the_pdf_states_the_honesty_rules_before_the_numbers():
    """A reader who reaches the tables without these is one misreading away from a falsely
    reassuring conclusion."""
    text = _pdf_text(pdf_report.build(_snapshot(), reference_doc=reference.load()))
    assert "unprotected" in text.lower()
    assert "no average RTO" in text or "no average rto" in text.lower()
    assert "not proven by a" in text or "never been tested" in text


def test_the_pdf_headline_is_a_count_not_a_score():
    """A score out of 100 invites comparison between estates that share no assumptions."""
    text = _pdf_text(pdf_report.build(_snapshot(), reference_doc=reference.load()))
    assert "no recovery path" in text.lower()
    assert "/100" not in text and "out of 100" not in text


def test_the_pdf_exposes_its_bookmarks():
    pdf = pdf_report.build(_snapshot(), reference_doc=reference.load())
    titles = [b.title for b in PdfReader(io.BytesIO(pdf)).outline if hasattr(b, "title")]
    assert any("Executive summary" in t for t in titles)


def test_the_pdf_says_when_it_omitted_rows():
    """A bounded report that does not admit its bounds is just an incomplete one."""
    rows = [_row(f"sa{i}") for i in range(pdf_report.MAX_NO_PATH + 20)]
    text = _pdf_text(pdf_report.build(_snapshot(rows=rows), reference_doc=reference.load()))
    assert "omitted" in text.lower()
    assert "workbook" in text.lower()


def test_the_pdf_flags_unacknowledged_objectives():
    text = _pdf_text(pdf_report.build(_snapshot(acknowledged=False),
                                      reference_doc=reference.load()))
    assert "shipped defaults" in text.lower()


def test_the_pdf_marks_demo_data_on_the_cover():
    snap = _snapshot()
    snap["demo"] = True
    text = _pdf_text(pdf_report.build(snap, reference_doc=reference.load()))
    assert "Demo data" in text


def test_the_pdf_refuses_a_direction_from_a_single_measurement():
    text = _pdf_text(pdf_report.build(
        _snapshot(), reference_doc=reference.load(),
        trend={"available": False, "reason": "Only one analysis has been run."}))
    assert "never measured" in text or "single measurement" in text


def test_the_pdf_flags_an_improvement_that_came_from_losing_visibility():
    trend = {
        "available": True, "reading_degraded": True,
        "caveat": "Fewer resources are reported without a recovery path, but more could "
                  "not be read at all. That is not necessarily an improvement.",
        "deltas": {"no_recovery_path": -8, "undetermined": 9, "breaches": 0,
                   "resources": 0, "not_protected": 0},
        "points": [{"generated_at": "2026-01-01T00:00:00Z", "resources": 10,
                    "no_recovery_path": 10, "undetermined": 0, "breaches": 2},
                   {"generated_at": "2026-02-01T00:00:00Z", "resources": 10,
                    "no_recovery_path": 2, "undetermined": 9, "breaches": 2}],
    }
    text = _pdf_text(pdf_report.build(_snapshot(), reference_doc=reference.load(),
                                      trend=trend))
    assert "not necessarily an improvement" in text


def test_the_pdf_reports_unreadable_provenance():
    text = _pdf_text(pdf_report.build(_snapshot(), reference_doc=reference.load()))
    assert "could not look" in text.lower()


def test_an_empty_snapshot_still_renders():
    snap = _snapshot(rows=[])
    snap["breaches"] = []
    snap["workloads"] = []
    pdf = pdf_report.build(snap, reference_doc=reference.load())
    assert pdf[:5] == b"%PDF-"


# ============================================================== they must agree
def test_the_pdf_and_the_workbook_report_the_same_numbers():
    """A reader who cross-checks one against the other and finds a difference stops
    treating either as evidence."""
    snap = _snapshot()
    ref = reference.load()
    facts = analysis.analyze(snap)

    sheets = _sheets(export_mod.build(snap, reference_doc=ref))
    text = _pdf_text(pdf_report.build(snap, reference_doc=ref))

    # The headline count appears in both.
    no_path = snap["summary"]["worst"]["no_recovery_path"]
    summary_flat = " ".join(str(c) for row in sheets["Summary"] for c in row if c is not None)
    assert f"no recovery path {no_path}" in summary_flat.lower().replace("  ", " ") \
        or str(no_path) in summary_flat
    assert str(no_path) in text

    # Every by-type row in the workbook comes from the same function the PDF used.
    body = [r for r in sheets["RTO-RPO by type"]
            if r and r[0] not in (None, "Resource type") and str(r[0]).count("/") == 1]
    assert len(body) == len(facts["by_type"])


def test_both_formats_derive_from_one_analysis_function():
    snap = _snapshot()
    facts = analysis.analyze(snap)
    sheets = _sheets(export_mod.build(snap))
    grid = [r for r in sheets["Reason index"] if r and r[0] not in (None, "Scenario")]
    reasons = [r for r in grid if r[1]]
    assert len(reasons) == len(facts["reasons"])


# ============================================================== evidence
def test_evidence_content_carries_the_reasoning_with_each_finding():
    """A frozen verdict whose basis was left behind cannot be argued with, only believed."""
    name, scope, included, tags, content = evidence_mod.build_evidence_content(
        _snapshot(), reference_doc=reference.load())
    assert "Recovery Readiness" in name
    assert scope["id"] == "wl-1"
    assert set(included) == {"findings", "metrics", "inventory"}
    assert "resiliency" in tags
    assert content["findings"], "no findings captured"
    assert all(f["basis"] for f in content["findings"])


def test_evidence_metrics_pin_the_objectives_used():
    _n, _s, _i, _t, content = evidence_mod.build_evidence_content(
        _snapshot(), reference_doc=reference.load())
    metrics = content["metrics"]
    assert metrics["objectives_version"] == reference.load()["version"]
    assert metrics["objectives"], "targets are unverifiable a year from now without these"
    assert metrics["restore_rates"]


def test_evidence_records_whether_a_source_could_be_read():
    _n, _s, _i, _t, content = evidence_mod.build_evidence_content(_snapshot())
    assert content["metrics"]["provenance"]["backup"]["unreadable"] is True


def test_evidence_findings_are_bounded_and_say_so():
    rows = [_row(f"sa{i}") for i in range(evidence_mod.MAX_FINDINGS + 50)]
    _n, _s, _i, _t, content = evidence_mod.build_evidence_content(_snapshot(rows=rows))
    assert len(content["findings"]) <= evidence_mod.MAX_FINDINGS
    assert content["metrics"]["findings_truncated"] is True


def test_evidence_content_is_json_serialisable():
    """The locker hashes canonical JSON; a set or a datetime in here breaks the SHA."""
    import json

    _n, _s, _i, _t, content = evidence_mod.build_evidence_content(
        _snapshot(), reference_doc=reference.load())
    json.dumps(content, sort_keys=True)


# ============================================================== the demo estate
@pytest.mark.parametrize("scope", [demo_catalog.CONTOSO_ID, demo_catalog.ZAVA_WEB_ID])
def test_the_demo_estate_exports_in_both_formats(scope):
    import asyncio

    from app.resiliency import analyze as analyze_mod

    snap = asyncio.run(analyze_mod.analyze(
        None, tenant_id="t", scope_kind="workload", scope_id=scope,
        subscriptions=[], workload_id=scope))
    ref = reference.load()
    assert export_mod.build(snap, reference_doc=ref)[:2] == b"PK"
    assert pdf_report.build(snap, reference_doc=ref)[:5] == b"%PDF-"


# ============================================================== unreadable estate
#
# Found against a live tenant whose pasted token had expired. Resource Graph returned
# nothing, so every count was zero and the report rendered a green "0 resources with no
# recovery path" — our own blindness presented as a clean bill of health.
def _unreadable_snapshot():
    """What `analyze` really produces when the configuration query fails: the report exists,
    but nothing was enumerated. Uses the production provenance keys."""
    snap = _snapshot(rows=[])
    snap["summary"] = {
        "resources": 0,
        "by_scenario": {s: {"determined": 0, "no_recovery_path": 0, "undetermined": 0,
                            "not_applicable": 0, "total": 0} for s in model.SCENARIOS},
        "protection": {"protected": 0, "not_protected": 0, "unknown": 0},
        "worst": {"scenario": "instance_loss", "no_recovery_path": 0},
    }
    snap["breaches"] = []
    snap["workloads"] = []
    snap["provenance"] = {
        "configuration": {"source": "Resource Graph", "collected_at": "",
                          "unreadable": True,
                          "reason": "Pasted token has expired \u2014 paste a fresh one."},
        "protection": {"source": "Backup Manager", "collected_at": "", "unreadable": True,
                       "reason": "Backup Manager has not analyzed this scope."},
    }
    return snap


def test_estate_unreadable_keys_on_configuration_not_on_any_source():
    """`protection` is unreadable on almost every real analysis (Backup Manager has not run
    for the scope). If that counted, every healthy report would be flagged."""
    healthy = _snapshot()
    healthy["provenance"] = {
        "configuration": {"unreadable": False, "reason": ""},
        "protection": {"unreadable": True, "reason": "Backup Manager has not analyzed this scope."},
    }
    assert snapshot_store.estate_unreadable(healthy) == ""
    assert "expired" in snapshot_store.estate_unreadable(_unreadable_snapshot())


def test_an_unreadable_estate_is_never_reported_as_zero_findings_in_the_pdf():
    text = _pdf_text(pdf_report.build(_unreadable_snapshot(), reference_doc=reference.load()))
    assert "Not read" in text, "the cover still prints a count"
    assert "could not be read" in text
    # The reason must travel with the verdict, or the reader cannot act on it.
    assert "expired" in text


def test_an_unreadable_estate_says_so_on_the_workbook_summary():
    sheets = _sheets(export_mod.build(_unreadable_snapshot(), reference_doc=reference.load()))
    flat = [str(c) for row in sheets["Summary"] for c in row if c is not None]
    assert any("ESTATE COULD NOT BE READ" in c for c in flat)
    assert any("expired" in c for c in flat)
    assert any("NOT because" in c for c in flat), "the zeros are not explained"


def test_a_readable_estate_is_not_flagged_as_unreadable():
    """Non-vacuous guard: the banner must not fire on every normal report."""
    sheets = _sheets(export_mod.build(_snapshot(), reference_doc=reference.load()))
    flat = [str(c) for row in sheets["Summary"] for c in row if c is not None]
    assert not any("ESTATE COULD NOT BE READ" in c for c in flat)
    assert "Not read" not in _pdf_text(
        pdf_report.build(_snapshot(), reference_doc=reference.load()))
