from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.alert_analysis import collector
from app.api import alert_analysis as alert_analysis_api
from app.alert_analysis.collector import compute_analysis
from app.core.genjob import JobRegistry
from app.core.security import Principal
from app.alert_analysis.decisions import apply_decisions
from app.alert_analysis.demo import build_demo_snapshot
from app.alert_analysis.export import to_csv, to_workbook
from app.alert_analysis.iac import generate_review_artifact
from app.alert_analysis import plans


def _resource(name: str = "vm1") -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/Microsoft.Compute/virtualMachines/{name}",
        "name": name,
        "type": "microsoft.compute/virtualmachines",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "location": "eastus",
    }


def _group(name: str, email: str = "oncall@example.com", *, enabled: bool = True) -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.insights/actionGroups/{name}",
        "name": name,
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "enabled": enabled,
            "emailReceivers": [{"name": "mail", "emailAddress": email}],
        },
    }


def _rule(
    name: str,
    *,
    threshold: float = 80,
    severity: int = 2,
    group: str = "ag1",
    enabled: bool = True,
    window: str = "PT5M",
) -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.insights/metricAlerts/{name}",
        "name": name,
        "type": "microsoft.insights/metricalerts",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "enabled": enabled,
            "severity": severity,
            "scopes": [_resource()["id"]],
            "windowSize": window,
            "evaluationFrequency": "PT1M",
            "criteria": {
                "allOf": [
                    {
                        "metricName": "Percentage CPU",
                        "operator": "GreaterThan",
                        "threshold": threshold,
                        "timeAggregation": "Average",
                    }
                ]
            },
            "actions": [{"actionGroupId": _group(group)["id"]}] if group else [],
        },
    }


def _analyze(rules: list[dict], groups: list[dict] | None = None) -> dict:
    return compute_analysis(
        [_resource()],
        rules,
        groups if groups is not None else [_group("ag1")],
        scope_kind="workload",
        scope_id="wl1",
        scope_name="Workload One",
        tolerance_pct=10,
    )


def test_cross_subscription_action_group_is_unresolved_not_missing() -> None:
    rule = _rule("cross-sub")
    remote = "/subscriptions/sub2/resourceGroups/remote/providers/microsoft.insights/actionGroups/remote-ag"
    rule["properties"]["actions"] = [{"actionGroupId": remote}]
    snapshot = _analyze([rule], [])
    analyzed = snapshot["rules"][0]
    assert analyzed["missing_action_group_ids"] == []
    assert analyzed["cross_subscription_action_group_ids"] == [remote.lower()]
    assert "missing_action_group" not in analyzed["issues"]
    assert "unresolved_action_group_access" in analyzed["issues"]
    assert any(gap["type"] == "unresolved_action_group_access" for gap in snapshot["gaps"])


def test_exact_duplicate_and_shared_recipient_are_flagged() -> None:
    snapshot = _analyze(
        [_rule("cpu-a", group="ag1"), _rule("cpu-b", group="ag2")],
        [_group("ag1"), _group("ag2")],
    )
    assert snapshot["kpis"]["overlap_groups"] == 1
    overlap = snapshot["overlaps"][0]
    assert overlap["type"] == "exact"
    assert overlap["confidence"] == "high"
    assert overlap["notification_overlap"] is True
    assert overlap["shared_recipient_count"] == 1


def test_near_duplicate_is_flagged_but_layered_severity_is_not() -> None:
    near = _analyze([_rule("cpu-80", threshold=80), _rule("cpu-85", threshold=85)])
    assert any(item["type"] == "near" for item in near["overlaps"])

    layered = _analyze(
        [_rule("cpu-warning", threshold=80, severity=2), _rule("cpu-critical", threshold=95, severity=0)]
    )
    assert layered["overlaps"] == []


def test_disjoint_metric_dimensions_are_not_overlaps() -> None:
    prod = _rule("cpu-prod")
    stage = _rule("cpu-stage")
    prod["properties"]["criteria"]["allOf"][0]["dimensions"] = [{"name": "environment", "operator": "Include", "values": ["prod"]}]
    stage["properties"]["criteria"]["allOf"][0]["dimensions"] = [{"name": "environment", "operator": "Include", "values": ["stage"]}]
    assert _analyze([prod, stage])["overlaps"] == []


def test_semantically_equivalent_promql_is_near_overlap() -> None:
    resource = _resource()
    def prom(name: str, expression: str) -> dict:
        return {
            "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.alertsmanagement/prometheusrulegroups/{name}",
            "name": name, "type": "microsoft.alertsmanagement/prometheusrulegroups", "subscriptionId": "sub1", "resourceGroup": "rg1",
            "properties": {"scopes": [resource["id"]], "interval": "PT1M", "rules": [{"alert": name, "expression": expression, "for": "PT5M", "severity": 2, "actions": [{"actionGroupId": _group("ag1")["id"]}]}]},
        }
    snapshot = compute_analysis([resource], [prom("a", 'sum(rate(requests_total{job="api",code="500"}[5m])) > 10'), prom("b", 'sum(rate(requests_total{code="500", job="api"}[15m])) > 12')], [_group("ag1")], scope_kind="workload", scope_id="wl1", scope_name="One", tolerance_pct=20)
    assert any(item["type"] == "near" for item in snapshot["overlaps"])


def test_routing_gaps_and_full_recipient_destinations() -> None:
    snapshot = _analyze([_rule("no-route", group=""), _rule("disabled", enabled=False)])
    gap_types = {gap["type"] for gap in snapshot["gaps"]}
    assert "no_action_group" in gap_types
    assert "disabled_rule" in gap_types
    assert snapshot["recipients"][0]["destination"] == "oncall@example.com"
    assert snapshot["recipients"][0]["masked"] == "oncall@example.com"
    serialized = str(snapshot)
    assert "oncall@example.com" in serialized


def test_single_rule_duplicate_recipient_paths_are_notification_overlap() -> None:
    rule = _rule("fanout", group="ag1")
    rule["properties"]["actions"] = [
        {"actionGroupId": _group("ag1")["id"]},
        {"actionGroupId": _group("ag2")["id"]},
    ]
    snapshot = _analyze([rule], [_group("ag1"), _group("ag2")])
    assert snapshot["overlaps"][0]["type"] == "notification"
    assert snapshot["overlaps"][0]["notification_overlap"] is True


def test_orphaned_and_receiverless_action_group_is_a_gap() -> None:
    empty_group = _group("unused", enabled=False)
    empty_group["properties"]["emailReceivers"] = []
    snapshot = _analyze([_rule("healthy")], [_group("ag1"), empty_group])
    gap_types = {gap["type"] for gap in snapshot["gaps"]}
    assert "action_group_no_receivers" in gap_types
    assert "orphaned_action_group" in gap_types


def test_parent_scope_alert_targets_workload_resource() -> None:
    rule = _rule("subscription-cpu")
    rule["properties"]["scopes"] = ["/subscriptions/sub1"]
    snapshot = _analyze([rule])
    assert snapshot["rules"][0]["effective_target_count"] == 1


def test_csv_contains_finding_columns_and_neutralizes_formulas() -> None:
    snapshot = _analyze([_rule("=malicious", group="")])
    text = to_csv(snapshot)
    assert "finding_status,finding_type,risk_level" in text
    assert "'=malicious" in text


def test_demo_snapshot_has_overlap_gap_and_full_recipient() -> None:
    snapshot = build_demo_snapshot("demo-amba-coverage")
    assert snapshot["demo"] is True
    assert snapshot["kpis"]["total_rules"] > 0
    assert snapshot["kpis"]["overlap_groups"] > 0
    assert snapshot["kpis"]["gap_count"] > 0
    assert "platform@contoso.example" in str(snapshot)


def test_smart_detector_and_prometheus_rules_are_normalized() -> None:
    resource = _resource()
    smart = {
        "id": "/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.alertsmanagement/smartdetectoralertrules/anomaly",
        "name": "anomaly",
        "type": "microsoft.alertsmanagement/smartdetectoralertrules",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "enabled": True,
            "severity": 1,
            "scope": [resource["id"]],
            "detector": {"id": "providers/templates/FailureAnomaliesDetector"},
            "actionGroups": [{"actionGroupId": _group("ag1")["id"]}],
        },
    }
    prometheus = {
        "id": "/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.alertsmanagement/prometheusrulegroups/prom",
        "name": "prom",
        "type": "microsoft.alertsmanagement/prometheusrulegroups",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "scopes": [resource["id"]],
            "interval": "PT1M",
            "rules": [{
                "alert": "HighRate",
                "expr": "rate(requests_total[5m]) > 100",
                "for": "PT5M",
                "labels": {"severity": "warning"},
                "actions": [{"actionGroupId": _group("ag1")["id"]}],
            }],
        },
    }
    snapshot = compute_analysis(
        [resource], [smart, prometheus], [_group("ag1")],
        scope_kind="workload", scope_id="wl1", scope_name="One",
    )
    assert snapshot["kpis"]["smart_detector_rules"] == 1
    assert snapshot["kpis"]["prometheus_rules"] == 1
    prom_rule = next(rule for rule in snapshot["rules"] if "prometheus" in rule["type"])
    assert prom_rule["conditions"][0]["signal_type"] == "prometheus"
    assert prom_rule["conditions"][0]["threshold"] == 100


@pytest.mark.asyncio
async def test_management_group_scope_expands_to_subscriptions(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, object] = {}

    async def subscriptions(_connection, mg_id):
        assert mg_id == "mg-root"
        return ["sub1", "sub2"]

    async def resources(predicates, _connection):
        captured["predicates"] = predicates
        return []

    async def empty(*_args, **_kwargs):
        return []

    monkeypatch.setattr("app.workloads.discovery.subscriptions_under_mg", subscriptions)
    monkeypatch.setattr("app.amba.collector._query_resources", resources)
    monkeypatch.setattr(collector, "_query_alerts", empty)
    monkeypatch.setattr(collector, "_query_action_groups", empty)
    monkeypatch.setattr(collector, "_query_firings", empty)
    snapshot = await collector.collect_analysis(
        {"id": "conn"}, scope_kind="management_group", scope_id="mg-root", workload=None,
    )
    assert snapshot["scope_kind"] == "management_group"
    assert "sub1" in str(captured["predicates"])
    assert "sub2" in str(captured["predicates"])


@pytest.mark.asyncio
async def test_collector_emits_truthful_detailed_progress(monkeypatch: pytest.MonkeyPatch) -> None:
    async def resources(_predicates, _connection):
        return [_resource()]

    async def alerts(_subscriptions, _connection):
        return [_rule("cpu")]

    async def groups(_subscriptions, _connection):
        return [_group("ag1")]

    async def firings(_subscriptions, _connection):
        return []

    messages: list[str] = []

    async def progress(_phase: str, message: str) -> None:
        messages.append(message)

    monkeypatch.setattr("app.amba.collector._query_resources", resources)
    monkeypatch.setattr(collector, "_query_alerts", alerts)
    monkeypatch.setattr(collector, "_query_action_groups", groups)
    monkeypatch.setattr(collector, "_query_firings", firings)
    snapshot = await collector.collect_analysis(
        {"id": "conn"}, scope_kind="subscription", scope_id="sub1", workload=None,
        progress=progress,
    )
    joined = "\n".join(messages).lower()
    assert snapshot["kpis"]["total_rules"] == 1
    assert "resolving subscription scope" in joined
    assert "scope resolved to 1 subscription" in joined
    assert "launching resource inventory query" in joined
    assert "received 1 resource row" in joined
    assert "received 1 alert rule row" in joined
    assert "normalizing" in joined
    assert "overlap" in joined
    assert "gaps" in joined
    assert "estimated rule costs" in joined


@pytest.mark.asyncio
async def test_refresh_job_is_idempotent_reconnectable_and_request_detached(monkeypatch: pytest.MonkeyPatch) -> None:
    registry = JobRegistry("alert-analysis-test")
    monkeypatch.setattr(alert_analysis_api, "_refresh_jobs", registry)
    monkeypatch.setattr(alert_analysis_api, "_effective_connection_id", lambda *_args: "conn1")
    release = __import__("asyncio").Event()
    persisted: list[dict] = []
    invalidated: list[tuple[str, str, str | None]] = []

    async def snapshot(_principal, scope_kind, scope_id, *, force, connection_id, progress=None):
        assert force and scope_kind == "subscription" and scope_id == "sub1" and connection_id == "conn1"
        assert progress is not None
        await progress("scope", "Resolving subscription scope…")
        await progress("query", "Launching resource inventory, rules, action groups, and firings queries…")
        await release.wait()
        await progress("compute", "Computing overlap, gap, and cost findings…")
        return {
            "report_exists": True, "demo": False, "partial": False,
            "rationalization_score": 90,
            "kpis": {"total_rules": 2, "overlap_groups": 1, "gap_count": 0, "resources_evaluated": 1},
        }

    class FakeSession:
        async def __aenter__(self):
            return self

        async def __aexit__(self, *_args):
            return None

    async def persist(result, _principal, _scope_kind, _scope_id, _db, progress=None):
        persisted.append(result)
        if progress:
            await progress("save", "Saved snapshot, run, trend, and audit record.")

    async def invalidate(_principal, scope_kind, scope_id, connection_id):
        invalidated.append((scope_kind, scope_id, connection_id))

    monkeypatch.setattr(alert_analysis_api, "_snapshot", snapshot)
    monkeypatch.setattr(alert_analysis_api, "_persist_refresh", persist)
    monkeypatch.setattr(alert_analysis_api, "_invalidate_live_inventory", invalidate)
    monkeypatch.setattr(alert_analysis_api, "SessionLocal", FakeSession)
    principal = Principal(
        subject="operator", email="operator@example.com", tenant_id="tenant1", role="operator",
        permissions=frozenset({"alert_analysis.read"}),
    )

    first = await alert_analysis_api.refresh_start(
        subscription_id="sub1", workload_id=None, management_group_id=None,
        connection_id="conn1", principal=principal,
    )
    second = await alert_analysis_api.refresh_start(
        subscription_id="sub1", workload_id=None, management_group_id=None,
        connection_id="conn1", principal=principal,
    )
    assert first["job"]["id"] == second["job"]["id"]
    assert first["job"]["status"] == "running"

    # A later GET is a reconnect: it replays progress without owning/cancelling the task.
    await __import__("asyncio").sleep(0)
    reconnected = await alert_analysis_api.refresh_job(
        subscription_id="sub1", workload_id=None, management_group_id=None,
        connection_id="conn1", principal=principal,
    )
    assert reconnected["job"]["status"] == "running"
    assert reconnected["progress"]
    assert persisted == []

    # Returning from both request handlers did not cancel the detached server task.
    release.set()
    for _ in range(20):
        await __import__("asyncio").sleep(0)
        completed = await alert_analysis_api.refresh_job(
            subscription_id="sub1", workload_id=None, management_group_id=None,
            connection_id="conn1", principal=principal,
        )
        if completed["job"]["status"] == "done":
            break
    assert completed["job"]["status"] == "done"
    assert completed["result"]["rationalization_score"] == 90
    assert persisted == [completed["result"]]
    assert invalidated == [("subscription", "sub1", "conn1")]
    assert any(line["phase"] == "done" for line in completed["progress"])


def test_xlsx_has_phase_two_sheets_highlights_and_full_recipients() -> None:
    snapshot = build_demo_snapshot("demo-amba-coverage")
    content = to_workbook(snapshot, [{"at": "2026-07-10T00:00:00Z", "pct": 81, "extra": snapshot["kpis"], "demo": True}])
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == [
        "Summary", "Rules", "ActionGroups", "Overlaps", "Gaps",
        "SmartDetector", "Prometheus", "Trends",
    ]
    assert workbook["Rules"].freeze_panes == "A2"
    assert workbook["Overlaps"]["A2"].fill.fgColor.rgb is not None
    assert "platform@contoso.example" in str([
        cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row
    ])


def test_firing_history_enriches_rule_volume() -> None:
    rule = _rule("cpu")
    snapshot = compute_analysis(
        [_resource()], [rule], [_group("ag1")],
        [
            {"essentials": {"alertRule": rule["id"], "firedDateTime": "2099-01-01T00:00:00+00:00"}},
            {"essentials": {"alertRule": rule["id"], "firedDateTime": "2099-01-02T00:00:00+00:00"}},
        ],
        scope_kind="workload", scope_id="wl1", scope_name="One",
    )
    assert snapshot["rules"][0]["firing_7d"] == 2
    assert snapshot["rules"][0]["firing_30d"] == 2
    assert snapshot["kpis"]["firings_30d"] == 2


def test_keep_and_exempt_decisions_reduce_actionable_findings() -> None:
    snapshot = _analyze([_rule("cpu-a"), _rule("cpu-b")])
    overlap_id = snapshot["overlaps"][0]["id"]
    decided = apply_decisions(
        snapshot,
        [{
            "id": f"overlap:{overlap_id}", "target_type": "overlap", "target_id": overlap_id,
            "action": "dismiss_finding", "reason": "Intentional layered ownership",
        }],
    )
    assert decided["kpis"]["accepted_findings"] == 1
    assert decided["kpis"]["actionable_overlap_groups"] == 0


def test_remediation_artifact_is_valid_noop_existing_only() -> None:
    snapshot = _analyze([_rule("cpu-a"), _rule("cpu-b")])
    snapshot["active_overlaps"] = snapshot["overlaps"]
    snapshot["active_gaps"] = snapshot["gaps"]
    artifact, actions = generate_review_artifact(snapshot)
    assert actions
    assert "existing =" in artifact
    assert "SAFE BY DESIGN" in artifact
    assert "az resource delete" not in artifact
    assert "resourceGroup()." not in artifact


def test_plan_requires_human_decision_and_never_executes(tmp_path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(plans, "_PATH", tmp_path / "plans.json")
    plan = plans.create_plan(
        tenant_id="t1", connection_id="c1", scope_kind="workload", scope_id="w1",
        scope_name="One", requested_by="user", artifact="targetScope = 'resourceGroup'\n",
        actions=[{"action": "review_gap"}],
    )
    assert plan["status"] == "pending"
    assert "no endpoint that executes" in plan["safety"]
    approved = plans.decide_plan("t1", plan["id"], "approved", "approver", "Reviewed")
    assert approved and approved["status"] == "approved"
    assert plans.decide_plan("t1", plan["id"], "rejected", "other") is None
