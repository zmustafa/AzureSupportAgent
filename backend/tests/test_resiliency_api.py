"""The Recovery Readiness API, in-process over the demo estate.

Calling handlers directly bypasses FastAPI's dependency resolution, so every ``Query``
default has to be passed explicitly or the parameter arrives as a ``Query`` object.
"""
from __future__ import annotations

import asyncio
import io

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app import demo_catalog
from app.api import resiliency as api
from app.resiliency import analyze as analyze_mod
from app.resiliency import join, model, reference, snapshot as snapshot_store

CONTOSO = demo_catalog.CONTOSO_ID


class _FakeDB:
    def __init__(self) -> None:
        self.rows: list[object] = []

    def add(self, row: object) -> None:
        self.rows.append(row)

    async def commit(self) -> None:
        pass


class _Principal:
    tenant_id = "t-demo"
    subject = "dev"


@pytest.fixture(autouse=True)
def _isolated(tmp_path, monkeypatch):
    snapshot_store.set_path_for_tests(tmp_path / "resiliency_snapshot.json")
    reference.reset_for_tests(tmp_path)
    import app.core.azure_connections as ac

    demo_connection = {"id": "conn-demo", "display_name": "Demo", "tenant_id": "t-demo"}
    # `resolve_selected_connection` imports these inside the function, so patching the
    # module attributes is what takes effect.
    monkeypatch.setattr(ac, "get_connection", lambda cid: demo_connection)
    monkeypatch.setattr(ac, "resolve_connection", lambda cid: demo_connection)
    yield
    snapshot_store.clear()


def _run(coro):
    return asyncio.run(coro)


def _scope() -> api.ScopeParams:
    return api.ScopeParams(workload_id=CONTOSO, connection_id="conn-demo")


def _analyse() -> dict:
    snap = _run(analyze_mod.analyze(
        {"id": "conn-demo"}, tenant_id="t-demo", scope_kind="workload", scope_id=CONTOSO,
        subscriptions=[], workload_id=CONTOSO))
    snapshot_store.write("t-demo", "conn-demo", "workload", CONTOSO, snap)
    return snap


# ============================================================== the gate
def test_an_unanalysed_scope_returns_the_shell_not_a_404():
    """'Nothing has been analyzed' is a state the UI renders, not an error."""
    body = _run(api.get_snapshot(scope=_scope(), principal=_Principal()))
    assert body["report_exists"] is False
    # Every section present and empty so the UI needs no null guards.
    assert set(body) >= {"summary", "resources", "breaches", "workloads", "provenance"}
    assert body["resources"] == []


def test_reads_never_touch_azure(monkeypatch):
    """GETs are cache-only. Proven by a spy, not by inspection."""
    called = []

    async def _boom(*_a, **_kw):
        called.append(1)
        return [], {}, "should not be called"

    from app.backup_manager import service

    monkeypatch.setattr(service, "arg_safe_detailed", _boom)
    _run(api.get_snapshot(scope=_scope(), principal=_Principal()))
    _run(api.summary(scope=_scope(), principal=_Principal()))
    _run(api.resources(scope=_scope(), scenario=None, state=None, search=None,
                       offset=0, limit=200, principal=_Principal()))
    assert called == []


# ============================================================== the vocabulary
def test_meta_publishes_the_scenarios_and_which_ones_redundancy_helps():
    body = _run(api.meta(_principal=_Principal()))
    ids = [s["id"] for s in body["scenarios"]]
    assert ids == list(model.SCENARIOS)
    helps = {s["id"]: s["redundancy_helps"] for s in body["scenarios"]}
    assert helps[model.SCENARIO_ZONE_LOSS] is True
    assert helps[model.SCENARIO_DATA_CORRUPTION] is False
    assert helps[model.SCENARIO_ACCIDENTAL_DELETE] is False


def test_meta_labels_no_recovery_path_distinctly_from_unknown():
    """These must never read alike; the wording comes from the server so the client cannot
    invent its own."""
    labels = {c["id"]: c["label"] for c in _run(api.meta(_principal=_Principal()))["rto_classes"]}
    assert labels[model.RTO_NONE] != labels[model.RTO_UNKNOWN]
    assert "No recovery" in labels[model.RTO_NONE]


# ============================================================== analysis
def test_the_demo_analysis_produces_a_full_matrix():
    snap = _analyse()
    assert snap["report_exists"] is True
    assert snap["demo"] is True
    assert snap["summary"]["resources"] > 0
    for row in snap["resources"]:
        assert set(row["verdicts"]) == set(model.SCENARIOS)


def test_the_analysis_names_a_resource_with_no_recovery_path():
    snap = _analyse()
    culprits = {r["name"] for r in snap["resources"]
                if r["worst"]["rto_class"] == model.RTO_NONE}
    assert "contoso-pms-vm" in culprits


def test_targets_are_applied_and_the_tier_is_reported():
    snap = _analyse()
    row = next(r for r in snap["resources"] if r["name"] == "contoso-guests-cosmos")
    assert row["tier"] == "mission_critical"
    assert row["tier_source"]
    corruption = row["verdicts"][model.SCENARIO_DATA_CORRUPTION]
    assert corruption["target"]
    assert corruption["breach"]["state"] == "breached", "24h RPO against a 4h objective"


def test_the_workload_rolls_up_and_names_its_weakest_link():
    snap = _analyse()
    workload = snap["workloads"][0]
    delete = workload["scenarios"][model.SCENARIO_ACCIDENTAL_DELETE]
    assert delete["rto_class"] == model.RTO_NONE
    assert delete["weakest_link"]["name"] in {"contoso-pms-vm", "contoso-pms-vm-datadisk"}
    assert delete["assumptions"]


def test_bands_are_attached_and_carry_their_assumptions():
    snap = _analyse()
    banded = [v for r in snap["resources"] for v in r["verdicts"].values()
              if v.get("rto_band_minutes")]
    assert banded, "day_plus verdicts should carry a duration band"
    for v in banded:
        assert v["rto_assumptions"], "a band without assumptions is a bare number"
        assert any("unverified" in a for a in v["rto_assumptions"])
        low, high = v["rto_band_minutes"]
        assert high > low


def test_no_band_is_attached_where_there_is_nothing_to_time():
    snap = _analyse()
    for row in snap["resources"]:
        for v in row["verdicts"].values():
            if v["rto_class"] in (model.RTO_NONE, model.RTO_UNKNOWN, model.RTO_AUTOMATIC):
                assert v["rto_band_minutes"] is None


# ============================================================== filters
def test_resources_can_be_filtered_to_the_resources_with_no_path():
    _analyse()
    body = _run(api.resources(scope=_scope(), scenario=None, state="no_path", search=None,
                              offset=0, limit=200, principal=_Principal()))
    assert body["total"] >= 1
    assert all(any(v["rto_class"] == model.RTO_NONE for v in r["verdicts"].values())
               for r in body["rows"])


def test_a_resource_detail_is_addressable_by_its_arm_id():
    snap = _analyse()
    target = snap["resources"][0]
    body = _run(api.resource_detail(resource_id=target["id"], scope=_scope(),
                                    principal=_Principal()))
    assert body["resource"]["id"] == target["id"]


def test_a_resource_detail_works_without_the_leading_slash():
    """A FastAPI path parameter arrives with the leading slash consumed by the route, so the
    client sends `subscriptions/...`. Rejecting that 404s every drawer."""
    snap = _analyse()
    target = snap["resources"][0]
    body = _run(api.resource_detail(resource_id=target["id"].lstrip("/"), scope=_scope(),
                                    principal=_Principal()))
    assert body["resource"]["id"] == target["id"]


def test_an_unknown_resource_is_a_404():
    _analyse()
    with pytest.raises(HTTPException) as excinfo:
        _run(api.resource_detail(resource_id="/subscriptions/x/nope", scope=_scope(),
                                 principal=_Principal()))
    assert excinfo.value.status_code == 404


# ============================================================== the registry
def test_the_reference_is_readable_by_any_reader():
    doc = _run(api.get_reference(_principal=_Principal()))
    assert doc["restore_rates"]["vm_restore_mbps"] > 0
    assert len(doc["tiers"]) == 4


def test_an_invalid_rate_is_rejected_by_name_rather_than_coerced_silently():
    body = api.ReferenceBody(restore_rates={"vm_restore_mbps": -5, "nonsense": 10})
    result = _run(api.put_reference(body=body, principal=_Principal(), db=_FakeDB()))
    assert any("vm_restore_mbps" in r for r in result["rejected"])
    assert any("nonsense" in r for r in result["rejected"])
    assert result["reference"]["restore_rates"]["vm_restore_mbps"] >= 1


def test_saving_the_reference_bumps_the_version():
    before = _run(api.get_reference(_principal=_Principal()))["version"]
    _run(api.put_reference(body=api.ReferenceBody(mechanism_minutes={"asr_failover": 45}),
                           principal=_Principal(), db=_FakeDB()))
    after = _run(api.get_reference(_principal=_Principal()))
    assert after["version"] == before + 1
    assert after["mechanism_minutes"]["asr_failover"] == 45


# ============================================================== portal links
def test_a_real_scope_carries_a_portal_host():
    """The UI cannot invent the host: a guess sends a sovereign customer to the wrong cloud."""
    body = _run(api.get_snapshot(scope=_scope(), principal=_Principal()))
    assert body["portal_host"] == "portal.azure.com"


def test_demo_data_gets_no_portal_host():
    """Demo ids are GUID-shaped enough to build a URL, so the link must be withheld here —
    it would open a 404 in the reader's own tenant and look like broken data."""
    _analyse()
    body = _run(api.get_snapshot(scope=_scope(), principal=_Principal()))
    assert body["demo"] is True
    assert body["portal_host"] == ""


def test_the_portal_host_follows_the_connection_cloud_not_the_stored_snapshot(monkeypatch):
    """Resolved per read. A host frozen into the snapshot goes stale when a tenant moves."""
    import app.core.azure_connections as ac

    gov = {"id": "conn-demo", "display_name": "Gov", "tenant_id": "t-demo",
           "azure_cloud": "AzureUSGovernment"}
    monkeypatch.setattr(ac, "get_connection", lambda cid: gov)
    monkeypatch.setattr(ac, "resolve_connection", lambda cid: gov)
    body = _run(api.get_snapshot(
        scope=api.ScopeParams(subscription_id="sub-1", connection_id="conn-demo"),
        principal=_Principal()))
    assert body["portal_host"] == "portal.azure.us"


def test_an_unrecognised_cloud_yields_no_link_rather_than_the_public_portal(monkeypatch):
    import app.core.azure_connections as ac

    weird = {"id": "conn-demo", "tenant_id": "t-demo", "azure_cloud": "AzureMarsCloud"}
    monkeypatch.setattr(ac, "get_connection", lambda cid: weird)
    monkeypatch.setattr(ac, "resolve_connection", lambda cid: weird)
    body = _run(api.get_snapshot(
        scope=api.ScopeParams(subscription_id="sub-1", connection_id="conn-demo"),
        principal=_Principal()))
    assert body["portal_host"] == ""


# ============================================================== export
def test_exporting_default_objectives_is_refused_until_they_are_acknowledged():
    """A watermark is easy to crop. A number handed to an auditor must have been agreed."""
    _analyse()
    with pytest.raises(HTTPException) as excinfo:
        _run(api.export(scope=_scope(), principal=_Principal(), db=_FakeDB()))
    assert excinfo.value.status_code == 409
    assert "Acknowledge" in excinfo.value.detail


def test_the_pdf_is_refused_on_the_same_terms_as_the_workbook():
    """One gate for both formats. A duplicated refusal is how one copy quietly drifts."""
    _analyse()
    with pytest.raises(HTTPException) as excinfo:
        _run(api.export(scope=_scope(), format="pdf", principal=_Principal(), db=_FakeDB()))
    assert excinfo.value.status_code == 409


def _acknowledge():
    _run(api.put_reference(body=api.ReferenceBody(targets_acknowledged=True),
                           principal=_Principal(), db=_FakeDB()))


def test_an_acknowledged_export_carries_the_analysis_and_the_trust_sheets():
    _acknowledge()
    _analyse()
    resp = _run(api.export(scope=_scope(), principal=_Principal(), db=_FakeDB()))
    names = load_workbook(io.BytesIO(resp.body), read_only=True).sheetnames
    for expected in ["Index", "Summary", "RTO-RPO by type", "Reason index", "Reasoning",
                     "Objectives", "Assumptions and rates", "Provenance", "Truncation"]:
        assert expected in names, expected


def test_the_format_parameter_selects_the_builder():
    _acknowledge()
    _analyse()
    xlsx = _run(api.export(scope=_scope(), format="xlsx", principal=_Principal(),
                           db=_FakeDB()))
    pdf = _run(api.export(scope=_scope(), format="pdf", principal=_Principal(),
                          db=_FakeDB()))
    assert xlsx.media_type == api.XLSX_MEDIA
    assert bytes(xlsx.body)[:2] == b"PK"
    assert pdf.media_type == "application/pdf"
    assert bytes(pdf.body)[:5] == b"%PDF-"


def test_the_filename_names_the_scope_and_the_date():
    """A constant filename means three subscriptions collide in one Downloads folder."""
    _acknowledge()
    _analyse()
    resp = _run(api.export(scope=_scope(), principal=_Principal(), db=_FakeDB()))
    disposition = resp.headers["content-disposition"]
    assert "recovery-readiness-" in disposition
    assert disposition.endswith('.xlsx"')
    assert "recovery_readiness.xlsx" not in disposition


def test_exporting_before_analysis_is_refused():
    with pytest.raises(HTTPException) as excinfo:
        _run(api.export(scope=_scope(), principal=_Principal(), db=_FakeDB()))
    assert excinfo.value.status_code == 400


# ============================================================== analysis lens
def test_the_analysis_endpoint_answers_the_shell_before_an_analysis():
    body = _run(api.analysis(scope=_scope(), principal=_Principal()))
    assert body["report_exists"] is False
    assert body["by_type"] == [] and body["reasons"] == []


def test_the_analysis_endpoint_groups_by_type_and_names_the_reason():
    _analyse()
    body = _run(api.analysis(scope=_scope(), principal=_Principal()))
    assert body["report_exists"] is True
    assert body["by_type"], "no per-type analysis"
    assert all("dominant_reason" in e for e in body["by_type"])
    assert set(body["rto_distribution"]) == set(model.SCENARIOS)


def test_the_screen_and_the_workbook_read_the_same_function():
    """If the lens and the export each aggregated their own way, the first reader to
    cross-check them would stop trusting both."""
    from app.resiliency import analysis as analysis_mod

    _analyse()
    body = _run(api.analysis(scope=_scope(), principal=_Principal()))
    snap = _run(api.get_snapshot(scope=_scope(), principal=_Principal()))
    assert body["by_type"] == analysis_mod.analyze(snap)["by_type"]


# ============================================================== trend
def test_the_trend_refuses_a_direction_before_two_analyses():
    body = _run(api.trend(scope=_scope(), principal=_Principal()))
    assert body["available"] is False


def test_a_demo_analysis_records_no_history():
    """A synthetic trend printed beside real numbers is the kind of thing that gets quoted."""
    from app.resiliency import history as history_store

    _analyse()
    assert history_store.read("t-demo", "conn-demo", "workload", CONTOSO) == []


# ============================================================== evidence
def test_evidence_is_refused_until_the_objectives_are_agreed():
    """An artifact that outlives the screen must not quote numbers nobody agreed to."""
    _analyse()
    with pytest.raises(HTTPException) as excinfo:
        _run(api.save_evidence(body=api.EvidenceBody(), scope=_scope(),
                               principal=_Principal(), db=_FakeDB()))
    assert excinfo.value.status_code == 409


def test_evidence_captures_the_analysis_as_content(tmp_path, monkeypatch):
    import app.evidence.registry as registry

    # No `raising=False`: a wrong attribute name here would silently write into the real
    # evidence store instead of the temp one.
    monkeypatch.setattr(registry, "_INDEX", tmp_path / "index.json")
    monkeypatch.setattr(registry, "_BLOB_DIR", tmp_path / "blobs")
    _acknowledge()
    _analyse()
    body = _run(api.save_evidence(body=api.EvidenceBody(), scope=_scope(),
                                  principal=_Principal(), db=_FakeDB()))
    meta = body["evidence"]
    assert body["ok"] is True
    assert meta["sha256"], "an evidence entry without a hash proves nothing"
    assert set(meta["included"]) == {"findings", "metrics", "inventory"}
    assert meta["demo"] is True
    # The capture must have landed in the temp store, not the real one.
    assert (tmp_path / "index.json").exists()


def test_evidence_content_can_be_reread_and_still_hashes_the_same(tmp_path, monkeypatch):
    """Immutability is the whole point: a stored verdict that cannot be re-verified is a
    claim, not evidence."""
    import app.evidence.registry as registry

    monkeypatch.setattr(registry, "_INDEX", tmp_path / "index.json")
    monkeypatch.setattr(registry, "_BLOB_DIR", tmp_path / "blobs")
    _acknowledge()
    _analyse()
    meta = _run(api.save_evidence(body=api.EvidenceBody(), scope=_scope(),
                                  principal=_Principal(), db=_FakeDB()))["evidence"]
    assert registry.verify_sha(meta) is True
    content = registry.get_content(meta["id"])
    assert content["metrics"]["objectives_version"]
    assert content["findings"]


# ============================================================== degraded sources
def test_a_missing_backup_estate_degrades_to_unknown_not_unprotected(monkeypatch):
    """The single most likely way this feature does damage."""
    from app.backup_manager import snapshot as bm_snapshot

    monkeypatch.setattr(bm_snapshot, "read_snapshot", lambda *a, **k: None)
    snap = _run(analyze_mod.analyze(
        {"id": "conn-demo"}, tenant_id="t-demo", scope_kind="subscription",
        scope_id="sub-1", subscriptions=[], workload_id=""))
    # No config either (no live Azure), but the contract still has to hold.
    assert snap["provenance"]["protection"]["unreadable"] is True
    assert "not a statement" in snap["provenance"]["protection"]["reason"]


def test_the_join_never_invents_a_resource_id():
    assert join.normalize_resource_id("garbage") == ""
