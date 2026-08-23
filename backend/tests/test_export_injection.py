"""Spreadsheet / CSV formula-injection guard.

Excel and LibreOffice treat a cell beginning ``= + - @`` as a formula, so
``=cmd|'/c calc'!A1`` in an Azure resource name, an Entra display name, a policy description
or an LLM-generated FMEA row is a working code-execution vector against the workstation of
whoever opens the export. ``app/core/xlsx.py::cell_safe`` neutralises it; this module exists
to make sure every writer actually reaches it, today and for every export added later.

Two complementary halves:

* ``test_*_neutralises_*`` drive real builders end to end and assert the payload comes back
  quoted.
* ``test_no_module_writes_cells_without_the_shared_guard`` is structural: it fails when a NEW
  module starts writing cells without routing through ``app.core.xlsx``. The behavioural half
  cannot do that, because nobody remembers to add a fixture for the export they just wrote.
"""
from __future__ import annotations

import ast
import csv
import io
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

APP = Path(__file__).resolve().parents[1] / "app"

# Classic payloads. The first is the remote-code-execution one named in core/xlsx.py.
PAYLOADS = [
    "=cmd|'/c calc'!A1",
    "+1+1",
    "-1+1",
    "@SUM(A1)",
    "\t=1+1",          # leading whitespace still triggers the parser
    "\r\n=HYPERLINK(\"http://evil\")",
]


TRIGGERS = ("=", "+", "-", "@")


def _cells(sheet) -> list[str]:
    return [str(c.value) for row in sheet.iter_rows() for c in row if isinstance(c.value, str)]


def _assert_neutralised(values: list[str], payload: str, *, allow_formulas: tuple[str, ...] = ()) -> None:
    """No emitted cell may BEGIN with a formula trigger.

    A payload embedded mid-string (``"FMEA — =cmd|..."``) is inert - the spreadsheet only
    evaluates a cell whose FIRST character is a trigger - so requiring an apostrophe on every
    occurrence would fail on harmless cells and train people to weaken the test.
    ``allow_formulas`` carries the deliberate formulas a builder is supposed to emit.
    """
    assert any(payload.lstrip("\t\r\n ") in v for v in values), (
        f"payload {payload!r} never reached the output - the fixture is not exercising it"
    )
    for v in values:
        stripped = v.lstrip("\t\r\n ")
        if not stripped or stripped[0] not in TRIGGERS:
            continue
        if any(stripped.startswith(f) for f in allow_formulas):
            continue
        raise AssertionError(f"cell would evaluate as a formula: {v!r}")


@pytest.mark.parametrize("payload", PAYLOADS)
def test_policy_workbook_neutralises_formula_injection(payload):
    from app.policy import xlsx_export

    content = xlsx_export.build_workbook([
        {"name": "Assignments", "columns": ["Name", payload], "rows": [[payload, "ok"]]},
    ])
    wb = load_workbook(io.BytesIO(content), read_only=True)
    _assert_neutralised(_cells(wb["Assignments"]), payload)
    wb.close()


@pytest.mark.parametrize("payload", PAYLOADS)
def test_changeexplorer_csv_neutralises_formula_injection(payload):
    from app.changeexplorer import export

    text = export.to_csv([{"resourceName": payload, "actor": payload, "riskLabel": "High"}])
    values = [c for row in csv.reader(io.StringIO(text)) for c in row]
    _assert_neutralised(values, payload)


@pytest.mark.parametrize("payload", PAYLOADS)
def test_fmea_workbook_neutralises_formula_injection(payload):
    from app.fmea.excel import build_fmea_xlsx

    doc = {
        "status": "draft",
        "tables": [{
            "name": "Ingress",
            "rows": [{
                "item": payload, "function": payload, "failure_mode": payload,
                "effects": payload, "causes": payload,
                "severity": 5, "occurrence": 4, "detection": 3,
            }],
        }],
    }
    content = build_fmea_xlsx(doc, workload_name=payload)
    wb = load_workbook(io.BytesIO(content), read_only=True)
    values = [v for name in wb.sheetnames for v in _cells(wb[name])]
    # The RPN column is a deliberate Excel formula - see the dedicated test below.
    _assert_neutralised(values, payload, allow_formulas=("=IF(OR(",))
    wb.close()


def test_fmea_keeps_its_intentional_rpn_formula():
    """The guard must not neutralise the RPN column - that formula is deliberate.

    Wrapping every cell indiscriminately would turn it into literal text and silently break
    the computed column, which is exactly the kind of "fix" this test exists to prevent.
    """
    from app.fmea.excel import build_fmea_xlsx

    doc = {
        "status": "draft",
        "tables": [{"name": "Ingress", "rows": [
            {"item": "A", "severity": 5, "occurrence": 4, "detection": 3},
        ]}],
    }
    wb = load_workbook(io.BytesIO(build_fmea_xlsx(doc, "wl")), read_only=True)
    formulas = [v for name in wb.sheetnames for v in _cells(wb[name]) if v.startswith("=IF(OR(")]
    assert formulas, "the RPN formula disappeared - it must stay a live Excel formula"
    wb.close()


@pytest.mark.parametrize("payload", PAYLOADS)
def test_shared_builder_neutralises_formula_injection(payload):
    from app.core.xlsx import WorkbookBuilder

    b = WorkbookBuilder()
    b.sheet("Data", ["Name"], [[payload]])
    wb = load_workbook(io.BytesIO(b.to_bytes()), read_only=True)
    _assert_neutralised(_cells(wb["Data"]), payload)
    wb.close()


@pytest.mark.parametrize("payload", PAYLOADS)
def test_api_fmea_csv_neutralises_formula_injection(payload):
    """This export lives in a ROUTER, not an export module - which is exactly why two audits
    that only inspected ``*export*.py`` missed it."""
    from app.api import fmea as fmea_api

    row = {key: payload for key, _label in fmea_api._CSV_COLUMNS}
    assert str(fmea_api._csv_cell(payload)).startswith("'")
    assert all(str(fmea_api._csv_cell(v)).startswith("'") for v in row.values())


def test_consolidating_the_guard_did_not_strip_what_wrapped_it():
    """Three helpers do MORE than guard. Collapsing them into plain ``cell_safe`` would be a
    silent regression, and one of them would be a SECURITY regression.

    This test exists because an equivalence probe declared `_safe_scalar` "identical to
    cell_safe" - it only looked identical because no test input contained a URL query string,
    so the redaction branch was never exercised.
    """
    from app.alert_analysis.export import _safe as flatten_and_guard
    from app.alerts_manager.activity_export import _safe_scalar as redact_and_guard
    from app.ownership.sheet import _csv_safe as stringify_and_guard

    # Secret redaction: a query string here can carry a token.
    assert redact_and_guard("https://example.com/x?sig=SECRET&k=v") == "https://example.com/x?<redacted>"
    assert redact_and_guard("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"

    # Container flattening.
    assert flatten_and_guard(["a", "b"]) == "a; b"
    assert flatten_and_guard("=1+1") == "'=1+1"

    # Stringification - this sheet round-trips every column through a CSV importer as text.
    assert stringify_and_guard(0) == "0"
    assert stringify_and_guard("=1+1") == "'=1+1"


def test_the_two_pure_guards_are_the_shared_definition():
    """`iam` and `appregs` had byte-identical copies; they are now aliases, not duplicates."""
    from app.core.xlsx import cell_safe
    from app.iam.export import _csv_safe as iam_guard
    from app.identity.appregs_export import _csv_safe as appregs_guard

    assert iam_guard is cell_safe
    assert appregs_guard is cell_safe


# --------------------------------------------------------------------------- structural guard
# Modules allowed to write cells directly. Each entry needs a reason; the list may only shrink.
_ALLOWED_DIRECT_WRITERS = {
    "app/core/xlsx.py": "the shared builder - it IS the guard",
}

_WRITE_CALL = re.compile(r"\bws\.append\(|\bws\.cell\(|\bwriter\.writerow\(|\bw\.writerow\(")
# Any route to a real guard: the shared module, or one of the audited local helpers.
_REACHES_GUARD = re.compile(r"app\.core\.xlsx|\b_csv_safe\b|\bcell_safe\b|\b_safe_scalar\b")
# A local helper that strips leading whitespace then prefixes an apostrophe is the same guard.
_LOCAL_GUARD = re.compile(r"lstrip\(\s*[\"']\\t\\r\\n", re.S)


def _modules_writing_cells() -> list[tuple[str, str]]:
    out = []
    for path in sorted(APP.rglob("*.py")):
        if "__pycache__" in str(path):
            continue
        src = path.read_text(encoding="utf-8", errors="ignore")
        if _WRITE_CALL.search(src) or ("openpyxl" in src and "Workbook(" in src):
            out.append((path.relative_to(APP.parent).as_posix(), src))
    return out


def test_no_module_writes_cells_without_the_shared_guard():
    """Every spreadsheet/CSV writer must reach app.core.xlsx (or carry the same guard inline).

    This is the check that survives staff turnover: a new export added next quarter fails here
    without anyone having to remember this file exists.
    """
    offenders = []
    for rel, src in _modules_writing_cells():
        if rel in _ALLOWED_DIRECT_WRITERS:
            continue
        if _REACHES_GUARD.search(src) or _LOCAL_GUARD.search(src):
            continue
        offenders.append(rel)

    assert not offenders, (
        "these modules build spreadsheet/CSV cells without a formula-injection guard - "
        "route them through app.core.xlsx.cell_safe / coerce / WorkbookBuilder: "
        + ", ".join(offenders)
    )


def test_the_structural_guard_is_not_vacuous(tmp_path):
    """Prove the detector fires on a module that writes cells with no guard."""
    bad = tmp_path / "rogue_export.py"
    bad.write_text("def to_workbook(rows):\n    ws.append([r for r in rows])\n", encoding="utf-8")
    src = bad.read_text(encoding="utf-8")
    assert _WRITE_CALL.search(src), "detector missed a raw ws.append - it would never fire"
    assert not _REACHES_GUARD.search(src)
    assert not _LOCAL_GUARD.search(src)


def test_every_writer_module_parses_and_is_discoverable():
    """Guards the guard: if discovery silently returns nothing, the structural test passes vacuously."""
    found = _modules_writing_cells()
    assert len(found) >= 10, f"discovery found only {len(found)} writers - the pattern has drifted"
    names = {rel for rel, _ in found}
    for expected in ("app/core/xlsx.py", "app/fmea/excel.py", "app/policy/xlsx_export.py"):
        assert expected in names, f"{expected} was not discovered as a cell writer"
    for _, src in found:
        ast.parse(src)
