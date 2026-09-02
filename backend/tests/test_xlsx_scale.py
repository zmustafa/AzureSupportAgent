"""Scale guard for the shared workbook builder.

`WorkbookBuilder.sheet` used to derive each row's index from `ws.max_row`. In openpyxl 3.1
that property is `max(self._cells)[0]` — a full scan of every cell written so far — so asking
it once per row made writing a sheet O(rows^2 x cols).

Measured before the fix, at 25 columns: 2,000 rows took 0.60 s, 4,000 took 2.29 s and 8,000
took 8.84 s (a clean 4x per doubling). On the live tenant the workbook's big sheets pushed the
whole request past the 240 s Azure Container Apps ingress timeout, so the browser got a 504 and
the frozen event loop starved the database pool for every other request in the process.

The timing half of this file is a smoke check with a deliberately generous bound; the
call-counting half is the deterministic one and is what should fail if this regresses.
"""
from __future__ import annotations

import time

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.xlsx import WorkbookBuilder

HEADERS = [f"col{i}" for i in range(25)]


def _rows(n: int) -> list[list[str]]:
    return [[f"v{r}-{c}" for c in range(25)] for r in range(n)]


def test_the_writer_never_rescans_the_sheet_per_row(monkeypatch):
    """The deterministic guard: `max_row` reads must not scale with the row count.

    A timing assertion alone would be tuned away on a slow CI box. This one fails the moment
    somebody reintroduces a per-row `ws.max_row`, whatever the machine is doing."""
    calls = {"n": 0}
    original = Worksheet.max_row.fget

    def counting(self):
        calls["n"] += 1
        return original(self)

    monkeypatch.setattr(Worksheet, "max_row", property(counting))

    WorkbookBuilder().sheet("small", HEADERS, _rows(20))
    small = calls["n"]
    calls["n"] = 0
    WorkbookBuilder().sheet("large", HEADERS, _rows(400))
    large = calls["n"]

    assert large == small, (
        f"writing 400 rows read max_row {large} times against {small} for 20 rows — the cost "
        "is proportional to the data again, which is the O(rows^2) defect that took the live "
        "site down."
    )


def test_writing_rows_stays_roughly_linear():
    """Quadratic growth is ~16x per 4x of rows; linear is ~4x. The bound sits between them."""
    def elapsed(n: int) -> float:
        rows = _rows(n)
        builder = WorkbookBuilder()
        started = time.perf_counter()
        builder.sheet("bench", HEADERS, rows)
        return time.perf_counter() - started

    elapsed(500)  # warm the import/allocator so the first sample is not the outlier
    base = elapsed(1000)
    scaled = elapsed(4000)
    ratio = scaled / max(base, 1e-4)
    assert ratio < 10, (
        f"4x the rows cost {ratio:.1f}x the time (expected ~4x). Before the fix this was ~16x."
    )


@pytest.mark.parametrize("count", [0, 1, 5, 300])
def test_every_row_is_written_exactly_once_and_in_order(count: int):
    """The fix swapped a value read back from the sheet for a counted one, so the thing that
    must be proven is that they still agree — a row silently overwriting its predecessor would
    lose data while leaving a perfectly valid workbook behind."""
    import io

    rows = _rows(count)
    builder = WorkbookBuilder()
    builder.sheet("data", HEADERS, rows)
    stream = io.BytesIO(builder.to_bytes())
    ws = load_workbook(stream)["data"]

    assert ws.max_row == count + 1, "row count on the sheet does not match the input"
    assert [c.value for c in ws[1]] == HEADERS
    for index, expected in enumerate(rows):
        assert [c.value for c in ws[index + 2]] == expected, f"row {index} landed wrong"


def test_short_and_blank_rows_still_advance_the_cursor():
    """A row shorter than the header, or entirely blank, must still occupy its own line.

    This is the case where a counted index and a read-back index could genuinely diverge, so it
    is the one worth pinning."""
    import io

    rows = [
        ["first"] + [""] * 24,
        ["short"],                 # deliberately fewer cells than there are headers
        [""] * 25,                 # entirely blank
        ["last"] + [""] * 24,
    ]
    builder = WorkbookBuilder()
    builder.sheet("ragged", HEADERS, rows)
    ws = load_workbook(io.BytesIO(builder.to_bytes()))["ragged"]

    assert ws.max_row == 5
    assert ws.cell(row=2, column=1).value == "first"
    assert ws.cell(row=3, column=1).value == "short"
    assert ws.cell(row=4, column=1).value in (None, "")
    assert ws.cell(row=5, column=1).value == "last"


def test_the_index_sheet_links_to_the_row_it_describes():
    """`index_sheet` had the same per-row `max_row` read. It is a small sheet, so the cost never
    mattered — but the hyperlink is written to whatever row that returned, so a wrong answer
    there points every contents-page link at the wrong sheet."""
    import io

    builder = WorkbookBuilder()
    builder.section("Alpha", "7030A0")
    builder.sheet("One", ["a"], [["x"]])
    builder.section("Beta", "0F6CBD")
    builder.sheet("Two", ["a"], [["y"]])
    builder.sheet("Three", ["a"], [["z"]])
    builder.index_sheet()

    ws = load_workbook(io.BytesIO(builder.to_bytes()))["Index"]
    assert [ws.cell(row=r, column=1).value for r in (2, 3, 4)] == ["One", "Two", "Three"]
    assert [ws.cell(row=r, column=2).value for r in (2, 3, 4)] == ["Alpha", "Beta", "Beta"]
    for row, name in ((2, "One"), (3, "Two"), (4, "Three")):
        link = ws.cell(row=row, column=1).hyperlink
        assert link is not None and link.location == f"'{name}'!A1"
