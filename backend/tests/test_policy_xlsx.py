"""Tests for the generic policy xlsx export (tables + pivot outline)."""
from __future__ import annotations

from io import BytesIO


def test_build_workbook_multi_sheet_and_outline():
    from app.policy.xlsx_export import build_workbook
    from openpyxl import load_workbook

    data = build_workbook([
        {"name": "Raw data", "columns": ["A", "B"], "rows": [["x", 1], ["y", 2]]},
        {
            "name": "Pivot",
            "columns": ["Group", "Waiver", "Total"],
            "rows": [["Scope A", 3, 3], ["    Sub1", 2, 2], ["    Sub2", 1, 1], ["Grand total", 3, 3]],
            "outline_levels": [0, 1, 1, 0],
        },
    ])
    assert isinstance(data, bytes) and len(data) > 0
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Raw data", "Pivot"]
    pv = wb["Pivot"]
    # Header + outline grouping on the child rows.
    assert pv["A1"].value == "Group"
    assert pv.row_dimensions[3].outline_level == 1
    assert pv.row_dimensions[4].outline_level == 1
    assert pv.row_dimensions[2].outline_level == 0  # group header, not grouped
    # Raw sheet has an auto-filter for sorting in Excel.
    assert wb["Raw data"].auto_filter.ref is not None


def test_build_workbook_sanitizes_titles_and_empty():
    from app.policy.xlsx_export import build_workbook
    from openpyxl import load_workbook

    data = build_workbook([{"name": "Bad:/\\?*[]Name longer than thirty-one characters here", "columns": ["X"], "rows": []}])
    wb = load_workbook(BytesIO(data))
    assert len(wb.sheetnames[0]) <= 31
    assert ":" not in wb.sheetnames[0]
