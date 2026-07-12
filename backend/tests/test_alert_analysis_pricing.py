from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.alert_analysis.collector import SNAPSHOT_SCHEMA_VERSION, compute_analysis, empty_snapshot
from app.alert_analysis.demo import build_demo_snapshot
from app.alert_analysis.export import to_csv, to_workbook
from app.alert_analysis.pricing import CATALOG_VERSION, estimate_rule_cost, pricing_catalog


def _resource(name: str) -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/Microsoft.Compute/virtualMachines/{name}",
        "name": name,
        "type": "microsoft.compute/virtualmachines",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
    }


def _metric(name: str, scopes: list[str], *, enabled: bool = True, dimensions: list[dict] | None = None) -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.insights/metricAlerts/{name}",
        "name": name,
        "type": "microsoft.insights/metricalerts",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "enabled": enabled,
            "severity": 2,
            "scopes": scopes,
            "evaluationFrequency": "PT1M",
            "windowSize": "PT5M",
            "criteria": {
                "allOf": [{
                    "metricName": "Percentage CPU",
                    "operator": "GreaterThan",
                    "threshold": 80,
                    "timeAggregation": "Average",
                    "dimensions": dimensions or [],
                }]
            },
        },
    }


def _log_rule(name: str, scope: str, frequency: str = "PT5M") -> dict:
    return {
        "id": f"/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.insights/scheduledQueryRules/{name}",
        "name": name,
        "type": "microsoft.insights/scheduledqueryrules",
        "subscriptionId": "sub1",
        "resourceGroup": "rg1",
        "properties": {
            "enabled": True,
            "severity": 2,
            "scopes": [scope],
            "evaluationFrequency": frequency,
            "windowSize": "PT15M",
            "query": "Heartbeat | count",
            "criteria": {"allOf": [{"metricMeasureColumn": "Computer", "operator": "GreaterThan", "threshold": 0}]},
        },
    }


def test_metric_uses_effective_targets_and_bounded_dimensions() -> None:
    resources = [_resource("vm1"), _resource("vm2")]
    rule = _metric(
        "cpu",
        ["/subscriptions/sub1/resourceGroups/rg1"],
        dimensions=[{"name": "core", "operator": "Include", "values": ["0", "1", "2"]}],
    )
    snapshot = compute_analysis(
        resources, [rule], [], scope_kind="subscription", scope_id="sub1", scope_name="Sub One"
    )
    cost = snapshot["rules"][0]["cost"]
    assert cost["family"] == "metric"
    assert cost["status"] == "estimated"
    assert cost["confidence"] == "medium"
    assert cost["components"][0]["quantity"] == 6
    assert cost["monthly_usd"] == cost["monthly_min_usd"] == cost["monthly_max_usd"] == 0.6


def test_metric_unbounded_dimension_is_an_explicit_range() -> None:
    estimate = estimate_rule_cost(
        {
            "type": "microsoft.insights/metricalerts",
            "enabled": True,
            "effective_target_count": 2,
            "conditions": [{"dimensions": ["instance:include:*"]}],
        }
    )
    assert estimate["status"] == "range_estimate"
    assert estimate["confidence"] == "low"
    assert estimate["monthly_min_usd"] == 0.2
    assert estimate["monthly_usd"] == 1.0
    assert estimate["monthly_max_usd"] == 4.0
    assert any("unbounded observed cardinality" in item for item in estimate["assumptions"])


def test_metric_without_resolved_targets_is_unknown_not_fabricated() -> None:
    estimate = estimate_rule_cost(
        {"type": "microsoft.insights/metricalerts", "enabled": True, "effective_target_count": 0, "conditions": [{}]}
    )
    assert estimate["status"] == "unknown"
    assert estimate["monthly_usd"] is None
    assert estimate["monthly_max_usd"] is None


def test_log_frequency_tiers_are_configurable_and_unsupported_is_unknown() -> None:
    base = {"type": "microsoft.insights/scheduledqueryrules", "enabled": True}
    five = estimate_rule_cost({**base, "conditions": [{"frequency": "PT5M"}]})
    assert five["status"] == "partial_estimate"
    assert five["monthly_usd"] == five["monthly_min_usd"] == 1.5
    assert five["monthly_max_usd"] is None
    assert five["components"][1]["unit_price_usd"] == 0.15

    custom = estimate_rule_cost(
        {**base, "conditions": [{"frequency": "PT30M"}]},
        {"version": "tenant-v1", "log": {"frequency_tiers": {"PT30M": {"first_time_series_monthly": 0.25, "additional_time_series_monthly": 0.02}}}},
    )
    assert custom["catalog_version"] == "tenant-v1"
    assert custom["monthly_usd"] == 0.25

    unsupported = estimate_rule_cost({**base, "conditions": [{"frequency": "PT30M"}]})
    assert unsupported["status"] == "unknown"
    assert unsupported["monthly_usd"] is None


def test_direct_free_and_unknown_families_do_not_invent_precision() -> None:
    activity = estimate_rule_cost({"type": "microsoft.insights/activitylogalerts"})
    prometheus = estimate_rule_cost({"type": "microsoft.alertsmanagement/prometheusrules"})
    smart = estimate_rule_cost({"type": "microsoft.alertsmanagement/smartdetectoralertrules"})
    assert activity["status"] == prometheus["status"] == "direct_free"
    assert activity["monthly_usd"] == prometheus["monthly_usd"] == 0.0
    assert "ingestion" in " ".join(prometheus["assumptions"]).lower()
    assert smart["status"] == "unknown"
    assert smart["monthly_usd"] is None


def test_summary_separates_current_disabled_unknown_and_families() -> None:
    resource = _resource("vm1")
    rules = [
        _metric("enabled", [resource["id"]]),
        _metric("disabled", [resource["id"]], enabled=False),
        _log_rule("logs", resource["id"]),
        {
            "id": "/subscriptions/sub1/resourceGroups/rg1/providers/microsoft.alertsmanagement/smartdetectoralertrules/smart",
            "name": "smart",
            "type": "microsoft.alertsmanagement/smartdetectoralertrules",
            "properties": {"enabled": True, "scope": [resource["id"]], "detector": {"id": "FailureAnomaliesDetector"}},
        },
    ]
    snapshot = compute_analysis(
        [resource], rules, [], scope_kind="workload", scope_id="wl1", scope_name="One"
    )
    summary = snapshot["cost_summary"]
    assert snapshot["schema_version"] == SNAPSHOT_SCHEMA_VERSION == 3
    assert summary["catalog_version"] == CATALOG_VERSION
    assert summary["currency"] == "USD"
    assert summary["monthly_usd"] == 1.6
    assert summary["monthly_min_usd"] == 1.6
    assert summary["monthly_max_usd"] is None
    assert summary["potential_disabled_monthly"] == 0.1
    assert summary["priced_count"] == 3
    assert summary["unknown_count"] == 1
    assert set(summary["by_family"]) == {"log", "metric", "smart_detector"}
    assert summary["top_rules"][0]["rule_name"] == "logs"


def test_empty_demo_csv_and_xlsx_include_cost_contract() -> None:
    empty = empty_snapshot("workload", "wl1")
    assert empty["schema_version"] == 3
    assert empty["cost_summary"]["monthly_usd"] == 0.0
    assert empty["cost_summary"]["by_family"] == {}

    demo = build_demo_snapshot("demo-amba-coverage")
    assert demo["cost_summary"]["priced_count"] > 0
    assert all("cost" in rule for rule in demo["rules"])

    csv_rows = list(csv.DictReader(StringIO(to_csv(demo))))
    assert csv_rows[0]["row_kind"] == "cost_summary"
    assert "cost_monthly_min_usd" in csv_rows[0]
    rule_row = next(row for row in csv_rows if row["row_kind"] == "rule")
    assert rule_row["cost_catalog_version"] == CATALOG_VERSION
    assert rule_row["cost_status"]

    workbook = load_workbook(BytesIO(to_workbook(demo)))
    summary_values = {workbook["Summary"][f"A{row}"].value: workbook["Summary"][f"B{row}"].value for row in range(2, workbook["Summary"].max_row + 1)}
    assert summary_values["Cost catalog"] == CATALOG_VERSION
    assert "cost_monthly_usd" in [cell.value for cell in workbook["Rules"][1]]


def test_catalog_override_does_not_mutate_default() -> None:
    custom = pricing_catalog({"metric": {"static_per_time_series_monthly": 9.0}})
    assert custom["metric"]["static_per_time_series_monthly"] == 9.0
    assert pricing_catalog()["metric"]["static_per_time_series_monthly"] == 0.10
