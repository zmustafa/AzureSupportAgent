"""Live end-to-end check of the disabled-access report and its exports.

The one thing this script exists to prove is the property no unit test can: that the FILE
somebody downloads contains exactly the rows the SCREEN they downloaded it from was showing.
The main access export shipped that regression once — a CSV taken with the privileged lens
active silently contained every row — so every filter is exercised against both surfaces and
the counts are compared.

Usage:  .venv\\Scripts\\python.exe scripts\\iam_disabled_live.py [connection_id]
"""
from __future__ import annotations

import csv
import http.cookiejar
import io
import json
import sys
import urllib.parse
import urllib.request

BASE = "http://127.0.0.1:8000/api"
JAR = http.cookiejar.CookieJar()
OPENER = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(JAR))

FAILURES: list[str] = []


def check(label: str, ok: bool, detail: str = "") -> None:
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}{(' — ' + detail) if detail else ''}")
    if not ok:
        FAILURES.append(label)


def get(path: str, body: dict | None = None):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(BASE + path, data=data)
    if data:
        req.add_header("Content-Type", "application/json")
        req.add_header("Sec-Fetch-Site", "same-origin")  # the app rejects cookie-bearing
        req.add_header("Origin", "http://127.0.0.1:8000")  # cross-origin state changes
    with OPENER.open(req, timeout=180) as resp:
        return json.loads(resp.read() or b"{}")


def raw(path: str) -> tuple[bytes, str]:
    req = urllib.request.Request(BASE + path)
    with OPENER.open(req, timeout=180) as resp:
        return resp.read(), resp.headers.get("Content-Disposition", "")


def csv_rows(body: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(body.decode("utf-8-sig"))))


def qs(**kw) -> str:
    parts = {k: ("true" if v is True else str(v)) for k, v in kw.items() if v not in (None, "", False)}
    return ("?" + urllib.parse.urlencode(parts)) if parts else ""


def main() -> int:
    selector = sys.argv[1] if len(sys.argv) > 1 else ""
    get("/auth/login", {"username": "admin", "password": "admin"})

    conn = ""
    if selector:
        # Resolve by display name as well as id so a live run never has to name a tenant here.
        for c in get("/azure/connections")["connections"]:
            if c["id"] == selector or str(c.get("display_name", "")).lower() == selector.lower():
                conn = c["id"]
                break
        if not conn:
            print(f"No connection matches {selector!r}.")
            return 1
    cq = {"connection_id": conn} if conn else {}

    if conn:
        # NEVER seed over a real tenant. The demo rows would be indistinguishable from live
        # findings in the very screen this script is verifying.
        print("\n=== live connection: demo seed skipped ===")
    else:
        print("\n=== seed the demo estate ===")
        seeded = get("/iam/demo/seed", {})
        check("demo seeded", bool(seeded.get("ok")), f"{seeded.get('scopes')} scopes")

    print("\n=== GET /iam/leavers ===")
    rep = get("/iam/leavers" + qs(**cq))
    check("measured", rep["measured"] is True)
    check("identities returned", len(rep["identities"]) > 0, str(len(rep["identities"])))
    d = rep["denominator"]
    check(
        "denominator published",
        d["principals_with_access"] > 0 and d["state_resolved"] > 0,
        f"{d['state_resolved']}/{d['principals_with_access']} checked, "
        f"{d['state_unknown']} unknown, {d['not_applicable']} n/a",
    )
    check("two tiers only", set(rep["tiers"]) == {"live_now", "restorable"}, str(list(rep["tiers"])))
    check("limitations stated", len(rep["limitations"]) > 0, f"{len(rep['limitations'])} stated")
    for i in rep["identities"]:
        print(
            f"       {i['displayName']:22} {i['tier']:11} grants={i['grants']} "
            f"priv={i['privilegedGrants']} group={i['groupGrants']} owns={len(i['ownedServicePrincipals'])}"
        )

    print("\n=== KPIs carry the gate ===")
    k = get("/iam/overview" + qs(**cq))["kpis"]
    check("account_state_collected true", k["account_state_collected"] is True)
    check(
        "disabled KPIs agree with the report",
        k["disabled_principals"] == rep["totals"]["identities"]
        and k["disabled_assignments"] == rep["totals"]["grants"],
        f"{k['disabled_principals']} principals / {k['disabled_assignments']} grants",
    )

    print("\n=== findings ===")
    # Filter to the pillar. An UNFILTERED page is capped, so on a real 5,500-row tenant the
    # disabled findings sit beyond it and "not in the page" would be reported as "signal did
    # not fire" — a false alarm about the very thing being verified.
    f = get("/iam/findings" + qs(pillar="hyg", limit=500, **cq))
    ids = {x["signal_id"] for x in f["findings"]}
    for sid in (
        "hyg.disabled_principal_access", "hyg.disabled_privileged_access",
        "hyg.disabled_via_group", "hyg.disabled_owns_credential", "hyg.disabled_pim_eligible",
    ):
        check(f"{sid} produced findings", sid in ids)
    # The recycle-bin signal legitimately produces nothing when the bin holds nobody with
    # access, so the assertion is REGISTRATION, not output. Asserting output here would make
    # the script fail on a well-kept tenant.
    registry = {s["id"] for s in get("/iam/signals" + qs(**cq))["signals"]}
    check("hyg.deleted_principal_restorable registered", "hyg.deleted_principal_restorable" in registry)

    print("\n=== recycle bin (soft-deleted, restorable) ===")
    soft = [i for i in rep["identities"] if i.get("softDeleted")]
    check(
        "soft-deleted holders are flagged, not folded into plain orphans",
        len(soft) == rep["totals"].get("soft_deleted", 0),
        f"{len(soft)} flagged",
    )
    check(
        "a soft-deleted holder carries its deletion date",
        all(i.get("deletedDateTime") for i in soft) if soft else True,
    )

    print("\n=== enrichment shape ===")
    sample = rep["identities"][0] if rep["identities"] else {}
    check("structured resources published", bool(sample.get("resources")),
          f"{len(sample.get('resources') or [])} scope(s) on the first identity")
    check("grant detail published", bool(sample.get("grantDetail")))
    check("all four sign-in kinds present",
          set(sample.get("signIn", {})) >= {"interactive", "nonInteractive", "successful", "servicePrincipal", "known"})
    check("dormancy bucket present", bool(sample.get("dormancyBucket")),
          str(sample.get("dormancyBucket")))
    check("dormancy labels published", bool(rep.get("dormancy_labels")))
    unmeasured = [i for i in rep["identities"] if not i.get("activityMeasured")]
    check("unmeasured usage never claims 'never used'",
          all(not i.get("lastActivity") for i in unmeasured))
    check("grant age published", any(i.get("oldestGrantAt") for i in rep["identities"]))

    print("\n=== server-side count maps ===")
    counts = rep.get("counts") or {}
    n = len(rep["identities"])
    for dim in ("tier", "principal_type", "on_prem", "dormancy"):
        total = sum((counts.get(dim) or {}).values())
        check(f"counts.{dim} sums to the population", total == n, f"{total} vs {n}")
    facets = rep.get("facets") or {}
    check("facets published for the dropdowns",
          all(k in facets for k in ("subscriptions", "roles", "planes", "groups", "signin_kinds")))


    print("\n=== THE SCREEN AND THE FILE MUST AGREE ===")
    # Every filter, applied to both surfaces, compared.
    cases = [
        ("unfiltered", {}),
        ("tier=live_now", {"tier": "live_now"}),
        ("tier=restorable", {"tier": "restorable"}),
        ("privileged_only", {"privileged_only": True}),
        ("principal_type=User", {"principal_type": "User"}),
        ("principal_type=ServicePrincipal", {"principal_type": "ServicePrincipal"}),
        ("on_prem=cloud", {"on_prem": "cloud"}),
        ("on_prem=onprem", {"on_prem": "onprem"}),
        ("on_prem=unknown", {"on_prem": "unknown"}),
        ("via_group_only", {"via_group_only": True}),
        ("soft_deleted", {"soft_deleted": True}),
        ("has_owned_sp", {"has_owned_sp": True}),
        ("pim_eligible", {"pim_eligible": True}),
        ("dormancy=unknown", {"dormancy": "unknown"}),
        ("dormancy=over_2y+interactive", {"dormancy": "over_2y", "signin_kind": "interactive"}),
        ("search=nothingmatchesthis", {"search": "nothingmatchesthis"}),
    ]
    for label, filt in cases:
        screen = get("/iam/leavers" + qs(**filt, **cq))["identities"]
        body, disp = raw("/iam/leavers/export" + qs(**filt, **cq, fmt="csv", shape="identities"))
        rows = csv_rows(body)
        check(
            f"{label}: screen {len(screen)} == csv {len(rows)}",
            len(screen) == len(rows),
            disp,
        )

    print("\n=== the grant shape is filtered by the same selection ===")
    all_grants = csv_rows(raw("/iam/leavers/export" + qs(**cq, fmt="csv", shape="grants"))[0])
    live_grants = csv_rows(
        raw("/iam/leavers/export" + qs(tier="live_now", **cq, fmt="csv", shape="grants"))[0]
    )
    check(
        "grant export narrows with the tier filter",
        0 < len(live_grants) < len(all_grants),
        f"{len(live_grants)} of {len(all_grants)}",
    )
    check(
        "every exported grant is held by a disabled principal",
        all(r["principalAccountEnabled"] == "false" for r in all_grants),
    )
    check("no deny row is exported as access", all(r["effect"] != "Deny" for r in all_grants))
    check(
        "grant csv carries the full schema",
        "principalAccountEnabled" in all_grants[0] and "roleName" in all_grants[0],
        f"{len(all_grants[0])} columns",
    )

    print("\n=== disabled_only lens on the MAIN access grid + export ===")
    grid = get("/iam/access" + qs(tab="all", disabled_only=True, limit=500, **cq))
    grid_csv = csv_rows(raw("/iam/export" + qs(fmt="csv", tab="all", disabled_only=True, **cq))[0])
    check(
        "access grid and its csv agree",
        grid["total"] == len(grid_csv) == len(all_grants),
        f"grid {grid['total']} / csv {len(grid_csv)} / leavers {len(all_grants)}",
    )
    unfiltered = get("/iam/access" + qs(tab="all", limit=1, **cq))["total"]
    check("the lens actually narrows", grid["total"] < unfiltered, f"{grid['total']} of {unfiltered}")

    print("\n=== XLSX workbook ===")
    book, disp = raw("/iam/leavers/export" + qs(**cq, fmt="xlsx", privileged_only=True))
    check("xlsx returned", book[:2] == b"PK" and len(book) > 5000, f"{len(book)} bytes")
    check("filename is dated", "iam-disabled-access-" in disp and ".xlsx" in disp, disp)
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(book))
        expected = ["Summary", "Identities", "Grants", "Via groups", "Owns credentials",
                    "Resources", "Not measured"]
        check("all six sheets present", wb.sheetnames == expected, str(wb.sheetnames))
        text = " ".join(
            str(c or "") for row in wb["Summary"].iter_rows(values_only=True) for c in row
        )
        check("summary carries the denominator", "DENOMINATOR" in text)
        check("summary records the applied filter", "privileged_only" in text)
        nm = " ".join(
            str(c or "") for row in wb["Not measured"].iter_rows(values_only=True) for c in row
        )
        check("not-measured sheet is populated", len(nm) > 40)
    except ImportError:
        print("  (openpyxl unavailable — skipping sheet inspection)")

    print("\n=== the main review workbook gained the sheet ===")
    book2, _ = raw("/iam/export/workbook" + qs(**cq))
    try:
        import openpyxl

        wb2 = openpyxl.load_workbook(io.BytesIO(book2))
        check("Disabled Access sheet present", "Disabled Access" in wb2.sheetnames)
    except ImportError:
        pass

    print("\n=== who_can_access flags the disabled ===")
    # Discovered from the data rather than hardcoded: a scope id in a script is both a tenant
    # identifier and a guarantee the check goes vacuous the day the fixture changes.
    scope_ids = sorted({
        r["scope"] for r in all_grants
        if r.get("scope", "").count("/") == 2 and r["scope"].startswith("/subscriptions/")
    })
    if scope_ids:
        ra = get("/iam/resource/access-summary" + qs(resource_id=scope_ids[0], **cq))
        states = {p["principalName"]: p.get("principalAccountEnabled") for p in ra.get("principals", [])}
        check(
            "resource summary carries account state",
            any(v == "false" for v in states.values()),
            f"{sum(1 for v in states.values() if v == 'false')} disabled of {len(states)}",
        )
    else:
        print("  (no subscription-scoped disabled grant to query)")

    print("\n=== campaign selector ===")
    before = len(get("/iam/campaigns" + qs(**cq))["campaigns"])
    made = get(
        "/iam/campaigns" + qs(**cq),
        {"name": "live check — disabled accounts", "selector": {"kind": "disabled"}},
    )["campaign"]
    after = get("/iam/campaigns" + qs(**cq))["campaigns"]
    check("campaign created from the disabled selector", len(after) == before + 1)
    total = int((made.get("stats") or {}).get("total") or 0)
    # Not equal to the grant count: one principal holding the same role at the same scope
    # through two groups is ONE review decision, and used to be a 500. It must be > 0 and
    # never MORE than the grants it came from.
    check("campaign has items", total > 0, f"{total} items from {len(all_grants)} grants")
    check("duplicates were folded, not multiplied", total <= len(all_grants))

    # The defect this catches: the selector used to ignore the screen's filters entirely, so
    # "review these 7 PIM-eligible people" silently opened a campaign over all 78.
    narrow = get(
        "/iam/campaigns" + qs(**cq),
        {
            "name": "live check — disabled + pim only",
            "selector": {"kind": "disabled", "pim_eligible": True},
        },
    )["campaign"]
    n_stats = narrow.get("stats") or {}
    on_screen = len(get("/iam/leavers" + qs(pim_eligible="true", **cq))["identities"])
    check(
        "campaign scope matches what the screen showed",
        int(n_stats.get("scoped_principals") or 0) == on_screen,
        f"campaign {n_stats.get('scoped_principals')} == screen {on_screen}",
    )
    check(
        "the narrowed campaign really is smaller",
        int(n_stats.get("total") or 0) < total,
        f"{n_stats.get('total')} < {total}",
    )
    check("the campaign records the filter it was cut from", bool(n_stats.get("scope_filter")))

    print("\n=== remediation preview (read-only, no campaign) ===")
    prev = get("/iam/leavers/remediation" + qs(**cq, fmt="az", tier="live_now"))
    script = prev.get("script") or ""
    check("preview returned a script", bool(script.strip()))
    check("it is labelled as not run by the product", "NOT RUN BY THE PRODUCT" in script)
    check("it carries a rollback", "ROLLBACK" in script)
    check(
        "it covers only the filtered slice",
        int(prev.get("action_count") or 0) <= len(all_grants),
        f"{prev.get('action_count')} actions",
    )
    check("no secret or token appears in the script", "Bearer" not in script)

    print("\n=== every step targets the API that actually governs the access ===")
    full = get("/iam/leavers/remediation" + qs(**cq, fmt="az"))
    planes = full.get("planes") or {}
    for p, n in sorted(planes.items(), key=lambda kv: -kv[1]):
        print(f"       {n:5}  {p}")
    print(f"       {full.get('grants')} grants -> {full.get('action_count')} steps")
    # Reported from a real run: `az role assignment delete --assignee <user>` printed "No matched
    # assignments were found to delete", and `--role 'Global Reader'` failed with "Role doesn't
    # exist". Both were the SAME defect — one Azure-RBAC template used for every plane.
    lines = [
        ln.strip() for ln in full["script"].splitlines()
        if ln.strip() and not ln.strip().startswith("#")
    ]
    rbac_lines = [ln for ln in lines if "role assignment " in ln]
    check(
        "no ARM role verb is aimed at anything but Azure RBAC",
        len(rbac_lines) <= 2 * int(planes.get("azure_rbac") or 0),
        f"{len(rbac_lines)} ARM lines for {planes.get('azure_rbac', 0)} RBAC step(s) (incl. rollbacks)",
    )
    check(
        "group-derived access removes the MEMBERSHIP",
        ("az ad group member remove" in full["script"]) == bool(planes.get("group_membership")),
    )
    check(
        "directory roles go to Graph, not ARM",
        ("graph.microsoft.com" in full["script"]) == bool(planes.get("entra_directory_role")),
    )
    check(
        "duplicate memberships are folded into one step",
        int(full.get("action_count") or 0) <= int(full.get("grants") or 0),
        f"{full.get('action_count')} steps from {full.get('grants')} grants",
    )
    if len(planes) > 1:
        check("the script warns the steps are not interchangeable", "spans more than one API" in full["script"])
    check(
        "every non-comment line belongs to a step, never a stray dry run",
        all(not ln.startswith("az ad group member check") for ln in lines),
    )

    print("\n=== 'never used' is not claimed from data that cannot support it ===")
    u = rep["usage"]
    identities = rep["identities"]
    totals = rep["totals"]
    print(f"  usage: available={u['available']} truncated={u['truncated']} window={u['window_days']}d")
    conclusive = sum(1 for i in identities if i.get("activityConclusive"))
    print(f"  identities where a 'never used' verdict is even possible: {conclusive} of {len(identities)}")
    if u["truncated"]:
        # Measured for real: 11 subscriptions hit the 6 MB Activity Log cap in one 90-day sweep.
        check(
            "a truncated sweep yields NO 'never used' verdicts",
            conclusive == 0 and int(totals.get("never_used") or 0) == 0,
            "the prefix is not evidence of disuse",
        )
        check("and the report says so", any("6 MB" in l for l in rep["limitations"]))
    else:
        check(
            "'never used' is only claimed where the window covers the account",
            all(i["activityWindowCovers"] for i in identities if i.get("activityConclusive")),
        )
    check(
        "no identity is both 'conclusive' and unmeasured",
        not [i for i in identities if i.get("activityConclusive") and not i.get("activityMeasured")],
    )

    print("\n" + ("=" * 60))
    if FAILURES:
        print(f"{len(FAILURES)} CHECK(S) FAILED:")
        for x in FAILURES:
            print("  -", x)
        return 1
    print("ALL LIVE CHECKS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
